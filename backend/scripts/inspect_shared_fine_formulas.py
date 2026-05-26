"""Inspect formulas in the shared QBD women's fine table workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


SHARED_DIR = Path(r"\\192.168.10.229\运营组资料\影刀\千百度精细表")
FILE_KEYWORD = "千百度女鞋精细数新5.26"
SKU = "QT653891S30"


def _norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    candidates = [
        path
        for path in SHARED_DIR.glob("*")
        if not path.name.startswith("~$")
        and path.suffix.lower() in {".xlsx", ".xlsm"}
        and FILE_KEYWORD in path.name
    ]
    if not candidates:
        print(f"未找到文件: {SHARED_DIR}\\*{FILE_KEYWORD}*")
        return

    path = sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)[0]
    print(f"文件: {path}")

    workbook = load_workbook(path, data_only=False, read_only=False, keep_vba=True)
    try:
        worksheet = workbook.active
        print("Sheets:")
        for sheet_name in workbook.sheetnames:
            print(f"- {sheet_name} [{sheet_name.encode('unicode_escape').decode('ascii')}]")
        print(f"Sheet: {worksheet.title}")
        print(f"max_row={worksheet.max_row}, max_column={worksheet.max_column}")

        headers = [_norm(cell.value) for cell in worksheet[1]]
        interesting_cols = []
        for index, header in enumerate(headers, start=1):
            if any(keyword in header for keyword in ("其他", "原始", "3天", "7天", "15天", "30天", "日均")):
                interesting_cols.append(index)

        print("相关列:")
        for index in interesting_cols:
            print(f"{index}: {headers[index - 1]}")

        rows_to_check = range(max(1, worksheet.max_row - 5), worksheet.max_row + 1)
        print("最后几行公式:")
        for row_index in rows_to_check:
            print(f"row {row_index}:")
            for col_index in interesting_cols:
                cell = worksheet.cell(row=row_index, column=col_index)
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    print(f"  {cell.coordinate} {headers[col_index - 1]} = {value}")

        print("QT653891S30 行相关公式:")
        for row in worksheet.iter_rows(min_row=2):
            values = [_norm(cell.value) for cell in row[:3]]
            if "QT653891S30" in values:
                row_index = row[0].row
                print(f"row {row_index}")
                for col_index in interesting_cols:
                    cell = worksheet.cell(row=row_index, column=col_index)
                    value = cell.value
                    print(f"  {cell.coordinate} {headers[col_index - 1]}: {value}")
                break

        detail_sheet_name = next((name for name in workbook.sheetnames if name == "辅助的"), None)
        if detail_sheet_name is not None:
            detail_sheet = workbook[detail_sheet_name]
            print(f"精细数 sheet: {detail_sheet_name} [{detail_sheet_name.encode('unicode_escape').decode('ascii')}]")
            print(f"max_row={detail_sheet.max_row}, max_column={detail_sheet.max_column}")
            detail_headers = [_norm(cell.value) for cell in detail_sheet[1]]
            for index, header in enumerate(detail_headers, start=1):
                if index <= 12 or any(keyword in header for keyword in ("货号", "原始", "3", "7", "15", "30", "销量", "数量")):
                    print(f"{index}: {header}")

            print("精细数最后几行 E-H/J/S/AB/AK 公式:")
            cols = [3, 5, 6, 7, 8, 10, 19, 28, 37]
            for row_index in range(max(1, detail_sheet.max_row - 5), detail_sheet.max_row + 1):
                print(f"row {row_index}:")
                for col_index in cols:
                    cell = detail_sheet.cell(row=row_index, column=col_index)
                    value = cell.value
                    header = detail_headers[col_index - 1] if col_index - 1 < len(detail_headers) else ""
                    if value not in (None, ""):
                        print(f"  {cell.coordinate} {header}: {value}")

            print(f"精细数里包含 {SKU} 的行:")
            hit_count = 0
            for row in detail_sheet.iter_rows(min_row=2):
                values = [_norm(cell.value) for cell in row[:12]]
                if SKU in values:
                    hit_count += 1
                    print(f"row {row[0].row}:")
                    for col_index in cols:
                        cell = detail_sheet.cell(row=row[0].row, column=col_index)
                        header = detail_headers[col_index - 1] if col_index - 1 < len(detail_headers) else ""
                        print(f"  {cell.coordinate} {header}: {cell.value}")
                    if hit_count >= 10:
                        break
            print(f"hit_count_shown={hit_count}")

        for source_sheet_name in ("3聚水潭", "7聚水潭", "15聚水潭和15罗盘", "月聚水潭"):
            if source_sheet_name not in workbook.sheetnames:
                continue
            source_sheet = workbook[source_sheet_name]
            print(f"{source_sheet_name} source sheet [{source_sheet_name.encode('unicode_escape').decode('ascii')}]:")
            print(f"max_row={source_sheet.max_row}, max_column={source_sheet.max_column}")
            headers = [_norm(cell.value) for cell in source_sheet[1]]
            for index in (1, 2, 3, 4, 5, 10, 19, 28, 37, 50, 51):
                if index <= source_sheet.max_column:
                    print(f"  header {index}: {headers[index - 1] if index - 1 < len(headers) else ''}")
            for row_index in range(max(1, source_sheet.max_row - 3), source_sheet.max_row + 1):
                print(f"  row {row_index}:")
                for col_index in (1, 2, 3, 4, 5, 10, 19, 28, 37, 50, 51):
                    if col_index <= source_sheet.max_column:
                        cell = source_sheet.cell(row=row_index, column=col_index)
                        if cell.value not in (None, ""):
                            print(f"    {cell.coordinate}: {cell.value}")
            print(f"  rows containing {SKU}:")
            shown = 0
            for row in source_sheet.iter_rows(min_row=2):
                values = [_norm(cell.value) for cell in row[:5]]
                if SKU in values:
                    shown += 1
                    print(f"  row {row[0].row}:")
                    for col_index in (1, 2, 3, 4, 5, 10, 19, 28, 37, 50, 51):
                        if col_index <= source_sheet.max_column:
                            cell = source_sheet.cell(row=row[0].row, column=col_index)
                            print(f"    {cell.coordinate}: {cell.value}")
                    if shown >= 5:
                        break
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
