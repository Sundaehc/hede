"""Align archive color codes and product models with the overview workbook."""
from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, update

from config import load_settings
from domain.fields import PRODUCT_FIELDS
from domain.schema import PRODUCT_ARCHIVE_TABLES
from storage.product_repository import ProductRepository
from transform.rows import normalize_admin_field, normalize_cell, normalize_header


DEFAULT_SOURCE_PATH = Path.home() / "Desktop" / "总览商品信息档案 (1).xlsx"
BRAND_MAP = {
    "千百度男鞋": "cbanner_mens",
    "千百度女鞋": "cbanner_womens",
    "烟斗": "yandou",
    "伊伴": "eblan",
}
MATCH_HEADERS = ("货号", "原始货号")
ARCHIVE_KEY_FIELDS = {"image_path", "sku", "original_sku"}
SYNCABLE_FIELDS = tuple(field.name for field in PRODUCT_FIELDS if field.name not in ARCHIVE_KEY_FIELDS)
SOURCE_LABELS_BY_TARGET = {
    field.name: field.all_labels
    for field in PRODUCT_FIELDS
    if field.name not in ARCHIVE_KEY_FIELDS
}
FIELD_MAP = {
    label: field_name
    for field_name, labels in SOURCE_LABELS_BY_TARGET.items()
    for label in labels
}
FIELD_LABEL_BY_TARGET = {
    field.name: field.label
    for field in PRODUCT_FIELDS
    if field.name not in ARCHIVE_KEY_FIELDS
}
DEFAULT_FIELDS = ("color_code", "product_model")


def text_value(value: object) -> str:
    normalized = normalize_cell(value)
    return str(normalized).strip() if normalized is not None else ""


def find_headers(sheet, source_fields: tuple[str, ...]) -> tuple[int, dict[str, int]]:
    for row_index in range(1, min(sheet.max_row, 30) + 1):
        values = next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        headers = {
            normalize_header(value): index
            for index, value in enumerate(values)
            if normalize_header(value)
        }
        has_match_headers = {*MATCH_HEADERS, "品牌"}.issubset(headers)
        has_source_fields = all(
            any(label in headers for label in SOURCE_LABELS_BY_TARGET[field])
            for field in source_fields
        )
        if has_match_headers and has_source_fields:
            return row_index, headers
    raise ValueError(f"{sheet.title} 未找到对齐所需表头")


def read_source(
    path: Path,
    fields: tuple[str, ...],
) -> tuple[dict[str, dict[str, dict[str, dict[str, set[str]]]]], int]:
    if not path.exists():
        raise FileNotFoundError(f"未找到来源文件：{path}")
    source: dict[str, dict[str, dict[str, dict[str, set[str]]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    )
    unclassified = 0
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        source_fields = fields
        for sheet in workbook.worksheets:
            header_row, headers = find_headers(sheet, source_fields)
            source_headers = {
                field: next(label for label in SOURCE_LABELS_BY_TARGET[field] if label in headers)
                for field in fields
            }
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(value is not None for value in row):
                    continue
                brand = BRAND_MAP.get(text_value(row[headers["品牌"]]))
                if brand is None:
                    unclassified += 1
                    continue
                field_values = {
                    field: normalize_admin_field(field, row[headers[source_headers[field]]])
                    for field in fields
                    if headers[source_headers[field]] < len(row)
                    and normalize_admin_field(field, row[headers[source_headers[field]]]) is not None
                    and str(normalize_admin_field(field, row[headers[source_headers[field]]])).strip()
                }
                for header in MATCH_HEADERS:
                    code = text_value(row[headers[header]]) if headers[header] < len(row) else ""
                    if not code:
                        continue
                    source[brand][header][code]
                    for field, value in field_values.items():
                        source[brand][header][code][field].add(value)
    finally:
        workbook.close()
    return {
        brand: {
            header: {code: dict(fields) for code, fields in codes.items()}
            for header, codes in headers.items()
        }
        for brand, headers in source.items()
    }, unclassified


def build_updates(
    repository: ProductRepository,
    brand: str,
    source,
    fields: tuple[str, ...],
    *,
    only_fill_blanks: bool,
):
    table = PRODUCT_ARCHIVE_TABLES[brand]
    updates: list[tuple[int, dict[str, str]]] = []
    stats = Counter()
    with repository.engine.connect() as connection:
        rows = connection.execute(
            select(table.c.id, table.c.sku, table.c.original_sku, *(table.c[field] for field in fields))
        ).mappings()
        for row in rows:
            stats["total"] += 1
            codes = (
                ("货号", text_value(row["sku"])),
                ("原始货号", text_value(row["original_sku"])),
            )
            changes: dict[str, str] = {}
            for field in fields:
                source_value = None
                selected_code = None
                for code_type, code in codes:
                    if not code or code not in source.get(code_type, {}):
                        continue
                    candidates = source[code_type][code].get(field, set())
                    selected_code = code
                    if len(candidates) > 1:
                        stats[f"{field}_conflict"] += 1
                        break
                    if len(candidates) == 1:
                        source_value = next(iter(candidates))
                    else:
                        stats[f"{field}_source_blank"] += 1
                    break
                if selected_code is None or source_value is None:
                    continue
                stats[f"{field}_matched"] += 1
                # Compare using the same field normalizer as the source so numeric
                # values such as Decimal("190.0") and the database value 190.00
                # are treated as equal.
                current_value = normalize_admin_field(field, row[field])
                if only_fill_blanks and current_value:
                    stats[f"{field}_skipped_existing"] += 1
                elif current_value == source_value:
                    stats[f"{field}_same"] += 1
                else:
                    changes[field] = source_value
                    stats[f"{field}_different"] += 1
            if changes:
                updates.append((int(row["id"]), changes))
    return updates, stats


def apply_updates(repository: ProductRepository, brand: str, updates: list[tuple[int, dict[str, str]]]) -> None:
    if not updates:
        return
    table = PRODUCT_ARCHIVE_TABLES[brand]
    with repository.engine.begin() as connection:
        for product_id, changes in updates:
            connection.execute(update(table).where(table.c.id == product_id).values(**changes))


def main() -> None:
    parser = argparse.ArgumentParser(description="按总览商品信息档案对齐商品字段")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH)
    parser.add_argument(
        "--fields",
        nargs="+",
        choices=SYNCABLE_FIELDS,
        default=DEFAULT_FIELDS,
        help="需要对齐的档案字段，默认颜色代码和产品型号",
    )
    parser.add_argument("--all-fields", action="store_true", help="对齐总览表中的全部可维护档案字段")
    parser.add_argument("--only-fill-blanks", action="store_true", help="仅填补档案中的空值")
    parser.add_argument("--apply", action="store_true", help="写入数据库；默认仅预览")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    fields = SYNCABLE_FIELDS if args.all_fields else tuple(args.fields)
    source, unclassified = read_source(args.source, fields)
    repository = ProductRepository(load_settings().database_url)
    print("模式：" + ("正式对齐" if args.apply else "预览（未写入数据库）"))
    print(f"来源：{args.source.name}，未识别品牌行：{unclassified}")

    total_updates = 0
    for brand in BRAND_MAP.values():
        updates, stats = build_updates(
            repository,
            brand,
            source.get(brand, {}),
            fields,
            only_fill_blanks=args.only_fill_blanks,
        )
        summaries = [
            f"{FIELD_LABEL_BY_TARGET[field]}匹配 {stats[f'{field}_matched']}，"
            f"一致 {stats[f'{field}_same']}，待对齐 {stats[f'{field}_different']}，"
            f"冲突 {stats[f'{field}_conflict']}，已有跳过 {stats[f'{field}_skipped_existing']}"
            for field in fields
        ]
        print(f"{brand}：{'；'.join(summaries)}；涉及商品 {len(updates)} 条")
        if args.apply:
            apply_updates(repository, brand, updates)
        total_updates += len(updates)
    print(("已对齐" if args.apply else "待对齐") + f"商品记录：{total_updates} 条")


if __name__ == "__main__":
    main()
