"""Import Smiley analysis workbook summary sheet into its own table.

Run:
    python -m scripts.import_smiley_fine_table --replace
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, insert, select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.fine_table_snapshot_schema import (
    FINE_TABLE_SNAPSHOT_BATCH_TABLE,
    fine_table_snapshot_row_table_for_date,
    fine_table_snapshot_year_table_exists,
)
from domain.smiley_schema import SMILEY_FINE_TABLE
from fileio.image_matcher import ImageMatcher
from scripts.import_fine_table_history_snapshots import (
    empty_payload,
    header_index,
    normalize_header,
    normalize_text,
    status_key,
    to_float,
    to_int,
    to_json_value,
)
from storage.inventory_repository import InventoryRepository


DEFAULT_ROOT = Path(r"\\Hede\运营组资料\影刀\笑脸分析表")
DEFAULT_IMAGE_ROOT = Path(r"\\Hede\图片\产品45主图随时更新\45主图\笑脸45度图")
DEFAULT_BRAND = "smiley"
DEFAULT_LEGACY_SNAPSHOT_BRAND = "xiaolian"
DEFAULT_SHEET = "汇总"
ROW_CHUNK_SIZE = 1000

SIZE_HEADERS = ("35-36", "37-38", "39-40", "35", "36", "37", "38", "39", "40", "41", "42", "43", "44")


def default_file_for_date(root: Path, snapshot_date: date) -> Path:
    return root / f"笑脸分析表{snapshot_date.month}.{snapshot_date.day}.xlsx"


def row_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def build_indexes(headers: list[str]) -> dict[str, int | None]:
    indexes = {
        "image_code": header_index(headers, "图片"),
        "sku": header_index(headers, "货号"),
        "factory_code": header_index(headers, "工厂"),
        "factory_sku": header_index(headers, "工厂货号"),
        "market_price": header_index(headers, "吊牌价"),
        "cost": header_index(headers, "成本"),
        "product_name": header_index(headers, "品名"),
        "barcode": header_index(headers, "国际码"),
        "execution_standard": header_index(headers, "执行标准"),
        "insole_material": header_index(headers, "鞋垫材质"),
        "outsole_material": header_index(headers, "大底材质"),
        "lining_material": header_index(headers, "内里材质"),
        "upper_material": header_index(headers, "鞋面材质"),
        "shoe_box_spec": header_index(headers, "鞋盒规格"),
        "accessories": header_index(headers, "辅料"),
        "first_order_time": header_index(headers, "首单日期"),
        "stock_qty": header_index(headers, "聚水潭库存"),
        "inbound_qty": header_index(headers, "在途数量"),
        "warehouse_stock": header_index(headers, "进货仓库存"),
        "available_stock": header_index(headers, "可用数库存"),
        "daily_sales_total": header_index(headers, "日总销量"),
        "total_3d_sales": header_index(headers, "3天总销量"),
        "total_7d_sales": header_index(headers, "7天总销量"),
        "total_15d_sales": header_index(headers, "15天总销量"),
        "total_30d_sales": header_index(headers, "30天总销量"),
        "dewu_30d_sales": header_index(headers, "30天得物销量"),
        "douyin_30d_sales": header_index(headers, "30天抖音销量"),
        "pdd_30d_sales": header_index(headers, "30天拼多多销量"),
        "tmall_30d_sales": header_index(headers, "30天天猫销量"),
        "other_platform_30d_sales": header_index(headers, "30天其他平台销量"),
        "dewu_return_rate": header_index(headers, "得物退货率"),
        "douyin_return_rate": header_index(headers, "抖音退货率"),
        "tmall_return_rate": header_index(headers, "天猫退货率"),
        "return_rate": header_index(headers, "综合退货率"),
        "remark": header_index(headers, "备注"),
        "season_category": header_index(headers, "季节分类"),
    }
    for size_header in SIZE_HEADERS:
        indexes[f"size_{size_header}"] = header_index(headers, size_header)
    return indexes


def normalize_date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = normalize_text(value)
    if not text:
        return None
    if " " in text:
        return text.split(" ", 1)[0]
    return text


def parse_date_value(value: Any) -> date | None:
    text = normalize_date_text(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def build_shop_sales(row: tuple[Any, ...], indexes: dict[str, int | None]) -> list[dict[str, Any]]:
    shops = (
        ("得物", "dewu_30d_sales"),
        ("抖音", "douyin_30d_sales"),
        ("拼多多", "pdd_30d_sales"),
        ("天猫", "tmall_30d_sales"),
        ("其他平台", "other_platform_30d_sales"),
    )
    result = []
    for shop_name, key in shops:
        quantity = to_int(row_value(row, indexes[key]))
        if quantity:
            result.append({"shop_name": shop_name, "quantity": quantity})
    return result


def build_return_rates(row: tuple[Any, ...], indexes: dict[str, int | None]) -> dict[str, float | None]:
    return {
        "得物": to_float(row_value(row, indexes["dewu_return_rate"])),
        "抖音": to_float(row_value(row, indexes["douyin_return_rate"])),
        "天猫": to_float(row_value(row, indexes["tmall_return_rate"])),
        "综合": to_float(row_value(row, indexes["return_rate"])),
    }


def build_payload(
    *,
    row: tuple[Any, ...],
    row_number: int,
    headers: list[str],
    indexes: dict[str, int | None],
    source_file: Path,
    sheet_name: str,
    brand: str,
    image_matcher: ImageMatcher,
    include_extra_fields: bool,
) -> dict[str, Any] | None:
    sku = normalize_text(row_value(row, indexes["sku"]))
    image_code = normalize_text(row_value(row, indexes["image_code"]))
    original_sku = sku or image_code
    if not sku and not image_code:
        return None

    image_path = image_matcher.find(sku) or image_matcher.find(image_code) or image_matcher.find(original_sku)
    size_stock = {
        size_header: to_int(row_value(row, indexes[f"size_{size_header}"]))
        for size_header in SIZE_HEADERS
    }
    stock_qty = to_int(row_value(row, indexes["stock_qty"]))
    inbound_qty = to_int(row_value(row, indexes["inbound_qty"]))
    total_3d_sales = to_int(row_value(row, indexes["total_3d_sales"]))
    total_7d_sales = to_int(row_value(row, indexes["total_7d_sales"]))
    total_15d_sales = to_int(row_value(row, indexes["total_15d_sales"]))
    total_30d_sales = to_int(row_value(row, indexes["total_30d_sales"]))

    payload = empty_payload()
    payload.update(
        {
            "id": -row_number,
            "brand": brand,
            "image_path": image_path,
            "sku": sku or None,
            "original_sku": original_sku or None,
            "factory_code": normalize_text(row_value(row, indexes["factory_code"])) or None,
            "factory_sku": normalize_text(row_value(row, indexes["factory_sku"])) or None,
            "barcode": normalize_text(row_value(row, indexes["barcode"])) or None,
            "cost": normalize_text(row_value(row, indexes["cost"])) or None,
            "latest_purchase_price": to_float(row_value(row, indexes["cost"])),
            "product_name": normalize_text(row_value(row, indexes["product_name"])) or None,
            "product_model": normalize_text(row_value(row, indexes["product_name"])) or None,
            "execution_standard": normalize_text(row_value(row, indexes["execution_standard"])) or None,
            "insole_material": normalize_text(row_value(row, indexes["insole_material"])) or None,
            "outsole_material": normalize_text(row_value(row, indexes["outsole_material"])) or None,
            "lining_material": normalize_text(row_value(row, indexes["lining_material"])) or None,
            "upper_material": normalize_text(row_value(row, indexes["upper_material"])) or None,
            "shoe_box_spec": normalize_text(row_value(row, indexes["shoe_box_spec"])) or None,
            "accessories": normalize_text(row_value(row, indexes["accessories"])) or None,
            "first_order_time": normalize_date_text(row_value(row, indexes["first_order_time"])),
            "season_category": normalize_text(row_value(row, indexes["season_category"])) or None,
            "market_price": to_float(row_value(row, indexes["market_price"])),
            "stock_qty": stock_qty,
            "size_stock": size_stock,
            "inbound_qty": inbound_qty,
            "warehouse_stock": to_int(row_value(row, indexes["warehouse_stock"])),
            "available_stock": to_int(row_value(row, indexes["available_stock"])),
            "daily_sales_total": to_int(row_value(row, indexes["daily_sales_total"])),
            "vip_daily_average_sales": to_float(row_value(row, indexes["daily_sales_total"])) or 0,
            "other_3d_sales": total_3d_sales,
            "other_7d_sales": total_7d_sales,
            "other_15d_sales": total_15d_sales,
            "other_30d_sales": total_30d_sales,
            "original_other_3d_sales": total_3d_sales,
            "original_other_7d_sales": total_7d_sales,
            "original_other_15d_sales": total_15d_sales,
            "original_other_30d_sales": total_30d_sales,
            "shop_30d_sales": build_shop_sales(row, indexes),
            "return_rates": build_return_rates(row, indexes),
            "remark": normalize_text(row_value(row, indexes["remark"])) or None,
            "projected_15d_stock": stock_qty + inbound_qty,
            "source_workbook": str(source_file),
            "source_sheet": sheet_name,
            "source_row_number": str(row_number),
        }
    )
    payload["status_key"] = status_key(payload.get("goods_status"))
    if include_extra_fields:
        payload["extra_fields"] = {
            header or f"column_{index + 1}": to_json_value(row_value(row, index))
            for index, header in enumerate(headers)
        }
    return payload


def read_payloads(
    *,
    file_path: Path,
    sheet_name: str,
    brand: str,
    image_root: Path,
    include_extra_fields: bool,
) -> tuple[list[dict[str, Any]], int]:
    image_matcher = ImageMatcher(image_root)
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            return [], len(image_matcher.index)
        headers = [normalize_header(value) for value in raw_headers]
        indexes = build_indexes(headers)
        payloads: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows, start=2):
            payload = build_payload(
                row=row,
                row_number=row_number,
                headers=headers,
                indexes=indexes,
                source_file=file_path,
                sheet_name=sheet_name,
                brand=brand,
                image_matcher=image_matcher,
                include_extra_fields=include_extra_fields,
            )
            if payload is not None:
                payloads.append(payload)
        return payloads, len(image_matcher.index)
    finally:
        workbook.close()


def row_for_table(payload: dict[str, Any], snapshot_date: date) -> dict[str, Any]:
    return {
        "snapshot_date": snapshot_date,
        "source_workbook": str(payload["source_workbook"]),
        "source_sheet": str(payload["source_sheet"]),
        "source_row_number": int(payload["source_row_number"]),
        "image_path": payload.get("image_path"),
        "sku": str(payload.get("sku") or "").strip(),
        "original_sku": payload.get("original_sku"),
        "factory_code": payload.get("factory_code"),
        "factory_sku": payload.get("factory_sku"),
        "market_price": payload.get("market_price"),
        "cost": to_float(payload.get("cost")),
        "product_name": payload.get("product_name"),
        "barcode": payload.get("barcode"),
        "execution_standard": payload.get("execution_standard"),
        "insole_material": payload.get("insole_material"),
        "outsole_material": payload.get("outsole_material"),
        "lining_material": payload.get("lining_material"),
        "upper_material": payload.get("upper_material"),
        "shoe_box_spec": payload.get("shoe_box_spec"),
        "accessories": payload.get("accessories"),
        "first_order_date": parse_date_value(payload.get("first_order_time")),
        "season_category": payload.get("season_category"),
        "stock_qty": int(payload.get("stock_qty") or 0),
        "inbound_qty": int(payload.get("inbound_qty") or 0),
        "warehouse_stock": int(payload.get("warehouse_stock") or 0),
        "available_stock": int(payload.get("available_stock") or 0),
        "daily_sales_total": int(payload.get("daily_sales_total") or 0),
        "total_3d_sales": int(payload.get("other_3d_sales") or 0),
        "total_7d_sales": int(payload.get("other_7d_sales") or 0),
        "total_15d_sales": int(payload.get("other_15d_sales") or 0),
        "total_30d_sales": int(payload.get("other_30d_sales") or 0),
        "shop_sales": payload.get("shop_30d_sales") or [],
        "size_stock": payload.get("size_stock") or {},
        "return_rates": payload.get("return_rates") or {},
        "remark": payload.get("remark"),
        "raw_payload": payload.get("extra_fields") or {},
    }


def write_smiley_table(
    repository: InventoryRepository,
    *,
    snapshot_date: date,
    payloads: list[dict[str, Any]],
    replace: bool,
) -> str:
    rows = [row_for_table(payload, snapshot_date) for payload in payloads if payload.get("sku")]
    SMILEY_FINE_TABLE.create(repository.engine, checkfirst=True)
    if not rows:
        return "empty"

    with repository.engine.begin() as connection:
        if replace:
            connection.execute(
                delete(SMILEY_FINE_TABLE)
                .where(SMILEY_FINE_TABLE.c.snapshot_date == snapshot_date)
            )
            for start in range(0, len(rows), ROW_CHUNK_SIZE):
                connection.execute(insert(SMILEY_FINE_TABLE), rows[start:start + ROW_CHUNK_SIZE])
            return "replaced"

        update_columns = [
            column.name
            for column in SMILEY_FINE_TABLE.columns
            if column.name not in ("id", "snapshot_date", "sku", "created_at")
        ]
        for start in range(0, len(rows), ROW_CHUNK_SIZE):
            statement = pg_insert(SMILEY_FINE_TABLE).values(rows[start:start + ROW_CHUNK_SIZE])
            excluded = statement.excluded
            statement = statement.on_conflict_do_update(
                index_elements=["snapshot_date", "sku"],
                set_={column: getattr(excluded, column) for column in update_columns},
            )
            connection.execute(statement)
    return "upserted"


def delete_legacy_snapshot(repository: InventoryRepository, *, brand: str, snapshot_date: date) -> int:
    if not fine_table_snapshot_year_table_exists(repository.engine, snapshot_date):
        return 0

    snapshot_row_table = fine_table_snapshot_row_table_for_date(snapshot_date)
    with repository.engine.begin() as connection:
        batch_ids = [
            int(row["id"])
            for row in connection.execute(
                select(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.id)
                .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.brand == brand)
                .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.snapshot_date == snapshot_date)
            ).mappings()
        ]
        if not batch_ids:
            return 0
        connection.execute(
            delete(snapshot_row_table)
            .where(snapshot_row_table.c.batch_id.in_(batch_ids))
        )
        connection.execute(
            delete(FINE_TABLE_SNAPSHOT_BATCH_TABLE)
            .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.id.in_(batch_ids))
        )
    return len(batch_ids)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Smiley summary sheet into smiley_fine_table")
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT, help="笑脸分析表目录；未传 --file 时按 snapshot date 自动找文件")
    parser.add_argument("--file", type=Path, default=None, help="指定单个笑脸分析表文件；默认使用当天文件")
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--image-root", type=Path, default=DEFAULT_IMAGE_ROOT)
    parser.add_argument("--snapshot-date", type=date.fromisoformat, default=None, help="默认当天日期")
    parser.add_argument("--brand", default=DEFAULT_BRAND)
    parser.add_argument("--legacy-snapshot-brand", default=DEFAULT_LEGACY_SNAPSHOT_BRAND)
    parser.add_argument("--replace", action="store_true")
    parser.add_argument("--include-extra-fields", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--delete-legacy-snapshot", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    snapshot_date = args.snapshot_date or date.today()
    file_path = args.file or default_file_for_date(args.root, snapshot_date)

    payloads, image_index_size = read_payloads(
        file_path=file_path,
        sheet_name=args.sheet,
        brand=args.brand,
        image_root=args.image_root,
        include_extra_fields=args.include_extra_fields,
    )
    matched_images = sum(1 for payload in payloads if payload.get("image_path"))
    print(
        f"file={file_path} sheet={args.sheet} brand={args.brand} "
        f"snapshot_date={snapshot_date} rows={len(payloads)} "
        f"image_index={image_index_size} matched_images={matched_images}",
        flush=True,
    )
    if args.dry_run:
        size_counts: dict[str, int] = defaultdict(int)
        for payload in payloads:
            if payload.get("season_category"):
                size_counts[str(payload["season_category"])] += 1
        for key in sorted(size_counts):
            print(f"season[{key}]={size_counts[key]}")
        return

    settings = load_settings()
    assert settings.database_url is not None
    repository = InventoryRepository(settings.database_url)
    result = write_smiley_table(
        repository,
        snapshot_date=snapshot_date,
        payloads=payloads,
        replace=args.replace,
    )
    deleted_legacy = (
        delete_legacy_snapshot(repository, brand=args.legacy_snapshot_brand, snapshot_date=snapshot_date)
        if args.delete_legacy_snapshot
        else 0
    )
    print(f"result={result} rows={len(payloads)} deleted_legacy_batches={deleted_legacy}")


if __name__ == "__main__":
    main()
