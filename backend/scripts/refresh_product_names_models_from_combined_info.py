"""Backfill archive product names and models from the combined footwear workbook.

The workbook is matched by both ``货号`` and ``原始货号``.  Rows are routed to
their product archive using the workbook's brand and primary supplier.  A
conflicting source value is skipped instead of being selected arbitrarily.

Run a preview first:
    uv run python -m scripts.refresh_product_names_models_from_combined_info

Apply the proposed changes:
    uv run python -m scripts.refresh_product_names_models_from_combined_info --apply
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, update

from config import load_settings
from domain.gj_brand import infer_gj_fine_table_brand
from domain.schema import PRODUCT_ARCHIVE_TABLES
from storage.product_repository import ProductRepository
from transform.rows import normalize_cell, normalize_header


DEFAULT_SOURCE_PATH = Path.home() / "Desktop" / "男女鞋合并商品信息.xlsx"
MATCH_HEADERS = ("货号", "原始货号")
SOURCE_FIELDS = {
    "品名": "product_name",
    "产品型号": "product_model",
}
TARGET_BRANDS = ("cbanner_mens", "cbanner_womens", "yandou", "eblan", "smiley")


@dataclass(frozen=True)
class SourceData:
    values: dict[str, dict[str, dict[str, set[str]]]]
    total_rows: int
    usable_rows: int
    unclassified_rows: int


def text_value(value: object) -> str:
    normalized = normalize_cell(value)
    return str(normalized).strip() if normalized is not None else ""


def find_headers(sheet) -> dict[str, int]:
    required = {*MATCH_HEADERS, *SOURCE_FIELDS, "品牌", "主供应商"}
    for row_index in range(1, min(sheet.max_row, 30) + 1):
        headers = {
            normalize_header(value): index
            for index, value in enumerate(next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True)))
            if normalize_header(value)
        }
        if required.issubset(headers):
            return headers
    raise ValueError(f"{sheet.title} 未找到所需表头：{', '.join(sorted(required))}")


def read_source(path: Path) -> SourceData:
    if not path.exists():
        raise FileNotFoundError(f"未找到来源文件：{path}")

    values: dict[str, dict[str, dict[str, set[str]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(set))
    )
    total_rows = 0
    usable_rows = 0
    unclassified_rows = 0
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            headers = find_headers(sheet)
            header_row = min(
                row_index
                for row_index in range(1, min(sheet.max_row, 30) + 1)
                if {
                    normalize_header(value)
                    for value in next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
                    if normalize_header(value)
                }.issuperset({*MATCH_HEADERS, *SOURCE_FIELDS, "品牌", "主供应商"})
            )
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(value is not None for value in row):
                    continue
                total_rows += 1
                codes = {
                    text_value(row[headers[header]])
                    for header in MATCH_HEADERS
                    if headers[header] < len(row) and text_value(row[headers[header]])
                }
                field_values = {
                    target: text_value(row[headers[source]])
                    for source, target in SOURCE_FIELDS.items()
                    if headers[source] < len(row) and text_value(row[headers[source]])
                }
                if not codes or not field_values:
                    continue

                brand = infer_gj_fine_table_brand({
                    "brand": text_value(row[headers["品牌"]]),
                    "primary_supplier": text_value(row[headers["主供应商"]]),
                })
                if brand not in TARGET_BRANDS:
                    unclassified_rows += 1
                    continue

                usable_rows += 1
                for code in codes:
                    for field, value in field_values.items():
                        values[brand][code][field].add(value)
    finally:
        workbook.close()

    return SourceData(
        values={
            brand: {code: dict(field_values) for code, field_values in code_values.items()}
            for brand, code_values in values.items()
        },
        total_rows=total_rows,
        usable_rows=usable_rows,
        unclassified_rows=unclassified_rows,
    )


def resolved_source_values(
    source_values: dict[str, dict[str, set[str]]],
) -> tuple[dict[str, dict[str, str]], int]:
    resolved: dict[str, dict[str, str]] = {}
    conflicts = 0
    for code, field_values in source_values.items():
        values: dict[str, str] = {}
        for field, candidates in field_values.items():
            if len(candidates) == 1:
                values[field] = next(iter(candidates))
            elif candidates:
                conflicts += 1
        if values:
            resolved[code] = values
    return resolved, conflicts


def build_updates(
    repository: ProductRepository,
    brand: str,
    source_values: dict[str, dict[str, set[str]]],
) -> tuple[list[tuple[int, dict[str, str]]], dict[str, int]]:
    resolved_values, source_conflicts = resolved_source_values(source_values)
    table = PRODUCT_ARCHIVE_TABLES[brand]
    updates: list[tuple[int, dict[str, str]]] = []
    matched_products = 0
    target_conflicts = 0
    unchanged_fields = 0

    with repository.engine.connect() as connection:
        rows = connection.execute(
            select(table.c.id, table.c.sku, table.c.original_sku, table.c.product_name, table.c.product_model)
        ).mappings()
        for row in rows:
            codes = {text_value(row["sku"]), text_value(row["original_sku"])} - {""}
            candidates: dict[str, set[str]] = defaultdict(set)
            for code in codes:
                for field, value in resolved_values.get(code, {}).items():
                    candidates[field].add(value)
            if not candidates:
                continue

            matched_products += 1
            changes: dict[str, str] = {}
            for field, values in candidates.items():
                if len(values) != 1:
                    target_conflicts += 1
                    continue
                value = next(iter(values))
                if text_value(row[field]) == value:
                    unchanged_fields += 1
                    continue
                changes[field] = value
            if changes:
                updates.append((int(row["id"]), changes))

    return updates, {
        "source_codes": len(resolved_values),
        "source_conflicts": source_conflicts,
        "matched_products": matched_products,
        "target_conflicts": target_conflicts,
        "unchanged_fields": unchanged_fields,
    }


def apply_updates(repository: ProductRepository, brand: str, updates_to_apply: list[tuple[int, dict[str, str]]]) -> None:
    if not updates_to_apply:
        return
    table = PRODUCT_ARCHIVE_TABLES[brand]
    with repository.engine.begin() as connection:
        for product_id, changes in updates_to_apply:
            connection.execute(update(table).where(table.c.id == product_id).values(**changes))


def main() -> None:
    parser = argparse.ArgumentParser(description="从男女鞋合并商品信息回填商品档案品名和产品型号")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH, help="来源 Excel 路径")
    parser.add_argument("--apply", action="store_true", help="执行数据库更新；未传入时仅预览")
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")
    source = read_source(args.source)
    repository = ProductRepository(load_settings().database_url)
    print("模式：" + ("正式回填" if args.apply else "预览（未写入数据库）"))
    print(
        f"来源：{args.source.name}，数据行 {source.total_rows}，有效行 {source.usable_rows}，"
        f"无法归类 {source.unclassified_rows} 行"
    )

    total_updates = 0
    for brand in TARGET_BRANDS:
        updates_to_apply, stats = build_updates(repository, brand, source.values.get(brand, {}))
        field_counts = {
            field: sum(field in changes for _, changes in updates_to_apply)
            for field in SOURCE_FIELDS.values()
        }
        print(
            f"{brand}：来源货号 {stats['source_codes']} 个，匹配商品 {stats['matched_products']} 条，"
            f"待更新 {len(updates_to_apply)} 条，品名 {field_counts['product_name']} 项，"
            f"产品型号 {field_counts['product_model']} 项，来源冲突 {stats['source_conflicts']} 项，"
            f"匹配冲突 {stats['target_conflicts']} 项"
        )
        if args.apply:
            apply_updates(repository, brand, updates_to_apply)
        total_updates += len(updates_to_apply)

    print(("已回填" if args.apply else "待回填") + f"商品记录：{total_updates} 条")


if __name__ == "__main__":
    main()
