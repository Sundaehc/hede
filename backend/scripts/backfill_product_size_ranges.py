"""Import product-code to size-group mappings and fill blank product archive size ranges."""
from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.product_size_group_mapping_schema import PRODUCT_SIZE_GROUP_MAPPINGS_TABLE
from domain.schema import PRODUCT_TABLES
from domain.size_group_schema import SIZE_GROUPS_TABLE
from storage.db import Database


DEFAULT_SOURCE_FILE = Path.home() / "Desktop" / "货号对应的尺码段名称8.5.xlsx"
CODE_HEADER = "款式编码"
SIZE_GROUP_HEADER = "系统尺码名称"


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_size_group_mappings(path: Path) -> tuple[dict[str, str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = [_text(value) for value in next(worksheet.iter_rows(values_only=True), ())]
        try:
            code_index = headers.index(CODE_HEADER)
            size_group_index = headers.index(SIZE_GROUP_HEADER)
        except ValueError as error:
            raise ValueError(f"Excel 需要包含 {CODE_HEADER} 和 {SIZE_GROUP_HEADER} 两列") from error

        groups_by_code: dict[str, set[str]] = defaultdict(set)
        skipped = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            code = _text(row[code_index] if code_index < len(row) else None)
            size_group = _text(row[size_group_index] if size_group_index < len(row) else None)
            if not code or not size_group or size_group == "0":
                skipped += 1
                continue
            groups_by_code[code].add(size_group)

        conflicts = {
            code: sorted(size_groups)
            for code, size_groups in groups_by_code.items()
            if len(size_groups) > 1
        }
        if conflicts:
            examples = ", ".join(f"{code}: {' / '.join(size_groups)}" for code, size_groups in list(conflicts.items())[:5])
            raise ValueError(f"同一款式编码存在多个尺码段：{examples}")

        return {
            code: next(iter(size_groups))
            for code, size_groups in groups_by_code.items()
        }, skipped
    finally:
        workbook.close()


def _blank_size_range(table):
    return (table.c.size_range.is_(None)) | (func.btrim(table.c.size_range) == "")


def backfill_size_ranges(*, source_file: Path, dry_run: bool) -> dict[str, object]:
    if not source_file.is_file():
        raise FileNotFoundError(f"找不到文件：{source_file}")

    mappings, skipped_rows = read_size_group_mappings(source_file)
    database = Database(load_settings().database_url)
    engine = database._require_engine()
    PRODUCT_SIZE_GROUP_MAPPINGS_TABLE.create(engine, checkfirst=True)

    with engine.connect() as connection:
        size_group_names = set(connection.execute(select(SIZE_GROUPS_TABLE.c.name)).scalars())
    missing_groups = sorted(set(mappings.values()) - size_group_names)
    if missing_groups:
        raise ValueError(f"以下尺码段尚未在尺码组管理中配置：{', '.join(missing_groups)}")

    source_rows = [
        {
            "product_code": code,
            "size_group_name": size_group_name,
            "source_workbook": source_file.name,
            "source_sheet": "汇总",
            "source_row_number": "",
        }
        for code, size_group_name in mappings.items()
    ]
    results: dict[str, object] = {
        "source_file": str(source_file),
        "mapping_count": len(mappings),
        "skipped_rows": skipped_rows,
        "missing_groups": missing_groups,
        "updated_by_brand": {},
        "dry_run": dry_run,
    }
    if dry_run:
        return results

    with engine.begin() as connection:
        for start in range(0, len(source_rows), 1_000):
            statement = pg_insert(PRODUCT_SIZE_GROUP_MAPPINGS_TABLE).values(source_rows[start:start + 1_000])
            excluded = statement.excluded
            connection.execute(
                statement.on_conflict_do_update(
                    index_elements=["product_code"],
                    set_={
                        "size_group_name": excluded.size_group_name,
                        "source_workbook": excluded.source_workbook,
                        "source_sheet": excluded.source_sheet,
                        "source_row_number": excluded.source_row_number,
                        "updated_at": func.date_trunc("minute", func.now()),
                    },
                )
            )

        for brand, table in PRODUCT_TABLES.items():
            original_size_range = select(PRODUCT_SIZE_GROUP_MAPPINGS_TABLE.c.size_group_name).where(
                PRODUCT_SIZE_GROUP_MAPPINGS_TABLE.c.product_code == table.c.original_sku
            ).scalar_subquery()
            sku_size_range = select(PRODUCT_SIZE_GROUP_MAPPINGS_TABLE.c.size_group_name).where(
                PRODUCT_SIZE_GROUP_MAPPINGS_TABLE.c.product_code == table.c.sku
            ).scalar_subquery()
            mapped_size_range = func.coalesce(original_size_range, sku_size_range)
            result = connection.execute(
                update(table)
                .where(_blank_size_range(table))
                .where(mapped_size_range.is_not(None))
                .values(size_range=mapped_size_range, updated_at=func.date_trunc("minute", func.now()))
            )
            results["updated_by_brand"][brand] = result.rowcount or 0

    return results


def main() -> None:
    parser = argparse.ArgumentParser(description="导入货号与尺码段映射并回填商品信息档案")
    parser.add_argument("--file", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    print(backfill_size_ranges(source_file=args.file, dry_run=args.dry_run))


if __name__ == "__main__":
    main()
