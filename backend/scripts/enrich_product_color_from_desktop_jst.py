from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from config import load_settings
from domain.sources import TABLE_NAMES
from storage.product_repository import ProductRepository


sys.stdout.reconfigure(encoding="utf-8")

DEFAULT_FILE = Path.home() / "Desktop" / "聚水潭货号颜色颜色代码汇总.xlsx"
LOOKUP_COLUMNS = {
    "sku": ("货号", "商品编码", "款式编码", "原始货号"),
    "color": ("颜色", "颜色名称", "颜色及规格", "新色"),
    "color_code": ("颜色代码", "颜色条码", "色号"),
}


@dataclass
class ColorInfo:
    sku: str
    color: str | None = None
    color_code: str | None = None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: object) -> str:
    return _cell_text(value).replace("\n", "").replace("\r", "")


def _column_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    normalized_candidates = {_normalize_header(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if header in normalized_candidates:
            return index
    return None


def _read_color_map(file_path: Path) -> tuple[dict[str, ColorInfo], list[str]]:
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    color_map: dict[str, ColorInfo] = {}
    sheet_summaries: list[str] = []

    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if header_row is None:
                continue

            headers = [_normalize_header(value) for value in header_row]
            sku_index = _column_index(headers, LOOKUP_COLUMNS["sku"])
            color_index = _column_index(headers, LOOKUP_COLUMNS["color"])
            color_code_index = _column_index(headers, LOOKUP_COLUMNS["color_code"])
            if sku_index is None or (color_index is None and color_code_index is None):
                sheet_summaries.append(f"{sheet_name}: 跳过，未识别到货号/颜色列")
                continue

            sheet_rows = 0
            sheet_valid = 0
            for row in rows:
                sheet_rows += 1
                sku = _cell_text(row[sku_index] if sku_index < len(row) else None)
                if not sku:
                    continue

                color = _cell_text(row[color_index] if color_index is not None and color_index < len(row) else None)
                color_code = _cell_text(row[color_code_index] if color_code_index is not None and color_code_index < len(row) else None)
                if not color and not color_code:
                    continue

                current = color_map.get(sku)
                if current is None:
                    color_map[sku] = ColorInfo(
                        sku=sku,
                        color=color or None,
                        color_code=color_code or None,
                    )
                else:
                    if color and not current.color:
                        current.color = color
                    if color_code and not current.color_code:
                        current.color_code = color_code
                sheet_valid += 1

            sheet_summaries.append(
                f"{sheet_name}: 读取 {sheet_rows} 行，可用 {sheet_valid} 行，货号列={headers[sku_index]}"
            )
    finally:
        workbook.close()

    return color_map, sheet_summaries


def _count_existing_values(items: list[dict[str, object]]) -> tuple[int, int]:
    color_count = sum(1 for item in items if _cell_text(item.get("color")))
    color_code_count = sum(1 for item in items if _cell_text(item.get("color_code")))
    return color_count, color_code_count


def run(file_path: Path, *, apply: bool) -> None:
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    color_map, sheet_summaries = _read_color_map(file_path)
    print(f"文件: {file_path}")
    for summary in sheet_summaries:
        print(f"  {summary}")
    print(f"可匹配货号数: {len(color_map)}")
    print(f"模式: {'执行更新' if apply else '预览，不写入数据库'}")

    settings = load_settings()
    repository = ProductRepository(settings.database_url)
    total_matched = 0
    total_updates = 0

    for brand, table_name in TABLE_NAMES.items():
        result = repository.list_products(brand, query=None, page=1, page_size=1_000_000)
        items = list(result["items"])
        existing_color, existing_color_code = _count_existing_values(items)

        updates: list[tuple[int, dict[str, str]]] = []
        for item in items:
            sku = _cell_text(item.get("sku"))
            original_sku = _cell_text(item.get("original_sku"))
            info = color_map.get(sku) or color_map.get(original_sku)
            if info is None:
                continue

            payload: dict[str, str] = {}
            if info.color and _cell_text(item.get("color")) != info.color:
                payload["color"] = info.color
            if info.color_code and _cell_text(item.get("color_code")) != info.color_code:
                payload["color_code"] = info.color_code
            if payload:
                updates.append((int(item["id"]), payload))

        total_matched += len(updates)
        print(
            f"\n{brand} ({table_name}): 总数 {len(items)}，已有颜色 {existing_color}，"
            f"已有颜色代码 {existing_color_code}，待更新 {len(updates)}"
        )
        for item_id, payload in updates[:5]:
            print(f"  示例 id={item_id}: {payload}")

        if apply and updates:
            with repository.engine.begin() as connection:
                for item_id, payload in updates:
                    set_sql = ", ".join(f"{column} = :{column}" for column in payload)
                    params = {"id": item_id, **payload}
                    connection.execute(
                        text(f"UPDATE {table_name} SET {set_sql}, updated_at = date_trunc('minute', now()) WHERE id = :id"),
                        params,
                    )
            total_updates += len(updates)

    print(f"\n匹配需更新记录: {total_matched}")
    if apply:
        print(f"已更新记录: {total_updates}")
    else:
        print("未写入数据库；确认无误后运行：python -m scripts.enrich_product_color_from_desktop_jst --apply")


def main() -> int:
    parser = argparse.ArgumentParser(description="用桌面聚水潭货号颜色颜色代码汇总补全商品档案颜色信息")
    parser.add_argument("--file", type=Path, default=DEFAULT_FILE)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    run(args.file, apply=args.apply)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
