"""Fill blank product archive fields from a complete overview workbook.

The source is matched by ``货号`` and ``原始货号`` within the source brand.
Only blank archive fields are updated, so existing manual values are retained.

Preview:
    uv run python -m scripts.backfill_product_archive_from_overview

Apply:
    uv run python -m scripts.backfill_product_archive_from_overview --apply
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, update

from config import load_settings
from domain.schema import PRODUCT_ARCHIVE_TABLES
from domain.sources import COLUMN_ALIASES
from storage.product_repository import ProductRepository
from transform.rows import normalize_admin_field, normalize_cell, normalize_header


DEFAULT_SOURCE_PATH = Path.home() / "Desktop" / "总览商品信息档案 (1).xlsx"
BRAND_MAP = {
    "千百度男鞋": "cbanner_mens",
    "千百度女鞋": "cbanner_womens",
    "烟斗": "yandou",
    "伊伴": "eblan",
}
MATCH_HEADERS = ("货号", "原始货号")


def text_value(value: object) -> str:
    normalized = normalize_cell(value)
    return str(normalized).strip() if normalized is not None else ""


def find_header_row(sheet) -> tuple[int, dict[str, int]]:
    required = {*MATCH_HEADERS, "品牌"}
    for row_index in range(1, min(sheet.max_row, 30) + 1):
        values = next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        headers = {
            normalize_header(value): index
            for index, value in enumerate(values)
            if normalize_header(value)
        }
        if required.issubset(headers):
            return row_index, headers
    raise ValueError(f"{sheet.title} 未找到商品档案表头")


def read_source(path: Path) -> tuple[dict[str, dict[str, dict[str, set[object]]]], int, int]:
    if not path.exists():
        raise FileNotFoundError(f"未找到来源文件：{path}")

    values_by_brand: dict[str, dict[str, dict[str, set[object]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(set))
    )
    total_rows = 0
    skipped_brands = 0
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            header_row, headers = find_header_row(sheet)
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(value is not None for value in row):
                    continue
                total_rows += 1
                brand_label = text_value(row[headers["品牌"]]) if headers["品牌"] < len(row) else ""
                brand = BRAND_MAP.get(brand_label)
                if brand is None:
                    skipped_brands += 1
                    continue

                codes = {
                    text_value(row[headers[header]])
                    for header in MATCH_HEADERS
                    if headers[header] < len(row) and text_value(row[headers[header]])
                }
                if not codes:
                    continue

                for header, index in headers.items():
                    field = COLUMN_ALIASES.get(header)
                    if not field or field not in PRODUCT_ARCHIVE_TABLES[brand].c:
                        continue
                    value = row[index] if index < len(row) else None
                    normalized = normalize_admin_field(field, value)
                    if normalized is None or not str(normalized).strip():
                        continue
                    for code in codes:
                        values_by_brand[brand][code][field].add(normalized)
    finally:
        workbook.close()

    return (
        {
            brand: {code: dict(fields) for code, fields in code_values.items()}
            for brand, code_values in values_by_brand.items()
        },
        total_rows,
        skipped_brands,
    )


def build_updates(repository: ProductRepository, brand: str, source: dict[str, dict[str, set[object]]]):
    table = PRODUCT_ARCHIVE_TABLES[brand]
    updates: list[tuple[int, dict[str, object]]] = []
    field_counts: dict[str, int] = defaultdict(int)
    conflicts = 0
    matched = 0

    with repository.engine.connect() as connection:
        rows = connection.execute(select(table)).mappings()
        for row in rows:
            codes = {text_value(row["sku"]), text_value(row["original_sku"])} - {""}
            candidates: dict[str, set[object]] = defaultdict(set)
            for code in codes:
                for field, values in source.get(code, {}).items():
                    candidates[field].update(values)
            if not candidates:
                continue
            matched += 1

            changes: dict[str, object] = {}
            for field, values in candidates.items():
                if len(values) != 1:
                    conflicts += 1
                    continue
                current = row[field]
                if current is not None and str(current).strip():
                    continue
                changes[field] = next(iter(values))
                field_counts[field] += 1
            if changes:
                updates.append((int(row["id"]), changes))

    return updates, matched, conflicts, dict(field_counts)


def apply_updates(repository: ProductRepository, brand: str, updates: list[tuple[int, dict[str, object]]]) -> None:
    if not updates:
        return
    table = PRODUCT_ARCHIVE_TABLES[brand]
    with repository.engine.begin() as connection:
        for product_id, changes in updates:
            connection.execute(update(table).where(table.c.id == product_id).values(**changes))


def main() -> None:
    parser = argparse.ArgumentParser(description="根据总览商品信息档案补充商品档案空字段")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH)
    parser.add_argument("--apply", action="store_true", help="写入数据库；默认仅预览")
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")
    source, total_rows, skipped_brands = read_source(args.source)
    repository = ProductRepository(load_settings().database_url)
    print("模式：" + ("正式回填" if args.apply else "预览（未写入数据库）"))
    print(f"来源：{args.source.name}，数据行 {total_rows}，未识别品牌行 {skipped_brands}")

    total_updates = 0
    for brand in BRAND_MAP.values():
        updates, matched, conflicts, field_counts = build_updates(repository, brand, source.get(brand, {}))
        print(
            f"{brand}：匹配 {matched} 条，待补充 {len(updates)} 条，"
            f"字段 {field_counts}，冲突字段 {conflicts}"
        )
        if args.apply:
            apply_updates(repository, brand, updates)
        total_updates += len(updates)
    print(("已补充" if args.apply else "待补充") + f"商品记录：{total_updates} 条")


if __name__ == "__main__":
    main()
