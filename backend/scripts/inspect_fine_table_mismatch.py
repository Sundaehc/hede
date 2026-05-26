"""Inspect a SKU mismatch between shared fine table and local fine-table data."""

from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from config import load_settings


SKU = "QT653891S30"
SHARED_DIR = Path(r"\\192.168.10.229\运营组资料\影刀\千百度精细表")
FILE_KEYWORDS = ("千百度女鞋精细数新5.26",)


def _norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def inspect_shared_file() -> None:
    print("=== shared fine table ===")
    print(f"dir: {SHARED_DIR}")
    print(f"exists: {SHARED_DIR.exists()}")
    candidates = [
        path
        for path in SHARED_DIR.glob("*")
        if not path.name.startswith("~$")
        and path.suffix.lower() in {".xlsx", ".xlsm"}
        and all(keyword in path.name for keyword in FILE_KEYWORDS)
    ]
    print("candidates:")
    for path in candidates:
        print(f"- {path.name}")
    if not candidates:
        return

    path = sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)[0]
    print(f"using: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = [_norm(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        print("headers:")
        for index, name in enumerate(header, start=1):
            if name:
                print(f"{index}: {name}")

        hits = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            values = [_norm(value) for value in row]
            if SKU in values:
                hits.append((row_number, values))

        print(f"rows containing {SKU}: {len(hits)}")
        for row_number, values in hits[:5]:
            print(f"row: {row_number}")
            for index, value in enumerate(values, start=1):
                if value:
                    label = header[index - 1] if index - 1 < len(header) else ""
                    if SKU in value or "原始" in label or "其他" in label or "3天" in label or "7天" in label or "15天" in label or "30天" in label:
                        print(f"  {index} {label}: {value}")
    finally:
        workbook.close()


def inspect_database() -> None:
    print("=== local jst_monthly_orders ===")
    settings = load_settings()
    assert settings.database_url is not None
    engine = create_engine(settings.database_url, future=True)

    with engine.connect() as conn:
        summary = conn.execute(text("""
            select
                count(*) as rows,
                coalesce(sum(quantity), 0) as qty,
                min(order_time_at) as min_time,
                max(order_time_at) as max_time
            from jst_monthly_orders
            where style_code = :sku
              and order_time_at >= timestamp '2026-05-23'
              and order_time_at < timestamp '2026-05-26'
              and shop_name not like '%唯品%'
        """), {"sku": SKU}).mappings().one()
        print(f"3d summary by style_code: {dict(summary)}")

        by_status = conn.execute(text("""
            select status, count(*) as rows, coalesce(sum(quantity), 0) as qty
            from jst_monthly_orders
            where style_code = :sku
              and order_time_at >= timestamp '2026-05-23'
              and order_time_at < timestamp '2026-05-26'
              and shop_name not like '%唯品%'
            group by status
            order by qty desc
        """), {"sku": SKU}).mappings().all()
        print("by status:")
        for row in by_status:
            item = dict(row)
            item["status_escape"] = str(item.get("status")).encode("unicode_escape").decode("ascii")
            print(item)

        by_day_status = conn.execute(text("""
            select date(order_time_at) as order_day, status, count(*) as rows, coalesce(sum(quantity), 0) as qty
            from jst_monthly_orders
            where style_code = :sku
              and order_time_at >= timestamp '2026-05-23'
              and order_time_at < timestamp '2026-05-26'
              and shop_name not like '%唯品%'
            group by date(order_time_at), status
            order by order_day, qty desc
        """), {"sku": SKU}).mappings().all()
        print("by day/status:")
        for row in by_day_status:
            item = dict(row)
            item["status_escape"] = str(item.get("status")).encode("unicode_escape").decode("ascii")
            print(item)

        active = conn.execute(text("""
            select
                coalesce(sum(case when status not in ('取消', '异常') then quantity else 0 end), 0) as no_cancel_abnormal,
                coalesce(sum(case when status in ('已发货', '已完成', '已付款待审核', '待审核') then quantity else 0 end), 0) as active_like,
                coalesce(sum(case when status = '已发货' then quantity else 0 end), 0) as shipped
            from jst_monthly_orders
            where style_code = :sku
              and order_time_at >= timestamp '2026-05-23'
              and order_time_at < timestamp '2026-05-26'
              and shop_name not like '%唯品%'
        """), {"sku": SKU}).mappings().one()
        print(f"filtered totals: {dict(active)}")

        by_shop = conn.execute(text("""
            select shop_name, count(*) as rows, coalesce(sum(quantity), 0) as qty
            from jst_monthly_orders
            where style_code = :sku
              and order_time_at >= timestamp '2026-05-23'
              and order_time_at < timestamp '2026-05-26'
              and shop_name not like '%唯品%'
            group by shop_name
            order by qty desc
            limit 20
        """), {"sku": SKU}).mappings().all()
        print("by shop:")
        for row in by_shop:
            print(dict(row))

        samples = conn.execute(text("""
            select internal_order_id, status, shop_name, style_code, product_code, quantity, order_time_at
            from jst_monthly_orders
            where style_code = :sku
              and order_time_at >= timestamp '2026-05-23'
              and order_time_at < timestamp '2026-05-26'
              and shop_name not like '%唯品%'
            order by order_time_at desc
            limit 20
        """), {"sku": SKU}).mappings().all()
        print("samples:")
        for row in samples:
            print(dict(row))


def main() -> None:
    inspect_shared_file()
    inspect_database()


if __name__ == "__main__":
    main()
