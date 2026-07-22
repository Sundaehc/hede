"""Import 2022+ product-goods order aggregates from Hede product workbooks."""

from __future__ import annotations

import argparse
from collections import defaultdict
from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, delete, inspect
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.product_goods_historical_orders_schema import (
    HISTORICAL_ORDER_START_YEAR,
    ensure_product_goods_historical_orders_table,
    product_goods_historical_orders_table_for_year,
)
from storage.date_normalization import parse_date


SHEET_NAME = "订单数据"
DATE_HEADERS = ("日期", "录单日期")
SKU_HEADERS = ("原货号", "原始货号")
ORDER_HEADERS = ("订单", "订货数量")
CHANNEL_HEADERS = ("渠道",)
YEAR_HEADERS = ("年",)
MONTH_HEADERS = ("月",)


def _text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _integer(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(Decimal(str(value).replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def _header_index(headers: tuple[object, ...], candidates: tuple[str, ...]) -> int | None:
    normalized = [str(value or "").strip().replace("\n", "") for value in headers]
    for candidate in candidates:
        if candidate in normalized:
            return normalized.index(candidate)
    return None


def _row_value(values: tuple[object, ...], index: int | None) -> object:
    return values[index] if index is not None and index < len(values) else None


def _order_date(values: tuple[object, ...], indexes: dict[str, int | None]) -> date | None:
    year = _integer(_row_value(values, indexes["year"]))
    month = _integer(_row_value(values, indexes["month"]))
    value = parse_date(_row_value(values, indexes["date"]))
    if year is not None and month is not None and 1 <= month <= 12:
        if value is not None and value.year == year and value.month == month:
            return value
        return date(year, month, 1)
    return value


def iter_rows(path: Path, *, brand: str, sheet_name: str = SHEET_NAME) -> Iterator[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"未找到 sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return
        indexes = {
            "date": _header_index(headers, DATE_HEADERS),
            "sku": _header_index(headers, SKU_HEADERS),
            "order": _header_index(headers, ORDER_HEADERS),
            "channel": _header_index(headers, CHANNEL_HEADERS),
            "year": _header_index(headers, YEAR_HEADERS),
            "month": _header_index(headers, MONTH_HEADERS),
        }
        if indexes["sku"] is None or indexes["order"] is None:
            raise ValueError("订单数据页未找到原货号或订单数量列")
        for row_number, values in enumerate(rows, start=2):
            original_sku = _text(_row_value(values, indexes["sku"]))
            order_quantity = _integer(_row_value(values, indexes["order"]))
            order_date = _order_date(values, indexes)
            if not original_sku or order_quantity is None or order_date is None:
                continue
            if order_date.year < HISTORICAL_ORDER_START_YEAR:
                continue
            yield {
                "brand": brand,
                "order_year": order_date.year,
                "order_date": order_date,
                "original_sku": original_sku,
                "channel": _text(_row_value(values, indexes["channel"])),
                "order_quantity": order_quantity,
                "source_workbook": path.name,
                "source_sheet": sheet_name,
                "source_row_number": row_number,
            }
    finally:
        workbook.close()


def import_rows(path: Path, *, brand: str, sheet_name: str = SHEET_NAME) -> dict[str, object]:
    settings = load_settings(require_database=True)
    assert settings.database_url is not None
    engine = create_engine(settings.database_url, future=True)
    tables: dict[int, object] = {}
    read = 0
    written = 0
    counts_by_year: dict[int, int] = defaultdict(int)
    chunks: dict[int, list[dict[str, object]]] = defaultdict(list)

    def write_chunk(year: int) -> int:
        chunk = chunks[year]
        if not chunk:
            return 0
        table = tables.setdefault(year, ensure_product_goods_historical_orders_table(engine, year))
        payload = [{key: value for key, value in row.items() if key != "order_year"} for row in chunk]
        with engine.begin() as connection:
            statement = pg_insert(table).values(payload)
            statement = statement.on_conflict_do_update(
                constraint=f"uq_product_goods_historical_orders_{year:04d}_source_row",
                set_={
                    "brand": statement.excluded.brand,
                    "order_date": statement.excluded.order_date,
                    "original_sku": statement.excluded.original_sku,
                    "channel": statement.excluded.channel,
                    "order_quantity": statement.excluded.order_quantity,
                },
            )
            connection.execute(statement)
        length = len(payload)
        chunks[year] = []
        return length

    for row in iter_rows(path, brand=brand, sheet_name=sheet_name):
        read += 1
        year = int(row["order_year"])
        counts_by_year[year] += 1
        chunks[year].append(row)
        if len(chunks[year]) >= 1_000:
            written += write_chunk(year)
    for year in list(chunks):
        written += write_chunk(year)
    return {"source": str(path), "read": read, "written": written, "years": dict(sorted(counts_by_year.items()))}


def replace_order_date_rows(
    rows: list[dict[str, object]],
    *,
    brand: str,
    order_date: date,
) -> dict[str, object]:
    if not rows:
        raise ValueError("订单数据为空")
    if any(row["brand"] != brand or row["order_date"] != order_date for row in rows):
        raise ValueError("订单数据品牌或日期不一致")

    settings = load_settings(require_database=True)
    assert settings.database_url is not None
    engine = create_engine(settings.database_url, future=True)
    table = ensure_product_goods_historical_orders_table(engine, order_date.year)
    payload = [{key: value for key, value in row.items() if key != "order_year"} for row in rows]

    with engine.begin() as connection:
        deleted = int(
            connection.execute(
                delete(table).where(
                    (table.c.brand == brand)
                    & (table.c.order_date == order_date)
                )
            ).rowcount
            or 0
        )
        statement = pg_insert(table).values(payload)
        statement = statement.on_conflict_do_update(
            constraint=f"uq_product_goods_historical_orders_{order_date.year:04d}_source_row",
            set_={
                "brand": statement.excluded.brand,
                "order_date": statement.excluded.order_date,
                "original_sku": statement.excluded.original_sku,
                "channel": statement.excluded.channel,
                "order_quantity": statement.excluded.order_quantity,
            },
        )
        connection.execute(statement)

    return {
        "brand": brand,
        "order_date": order_date.isoformat(),
        "deleted": deleted,
        "written": len(payload),
    }


def delete_source_rows(*, brand: str, source_workbook: str) -> int:
    settings = load_settings(require_database=True)
    assert settings.database_url is not None
    engine = create_engine(settings.database_url, future=True)
    deleted = 0
    with engine.begin() as connection:
        inspector = inspect(connection)
        for year in range(HISTORICAL_ORDER_START_YEAR, date.today().year + 1):
            table = product_goods_historical_orders_table_for_year(year)
            if not inspector.has_table(table.name):
                continue
            deleted += int(
                connection.execute(
                    delete(table).where(
                        (table.c.brand == brand)
                        & (table.c.source_workbook == source_workbook)
                    )
                ).rowcount
                or 0
            )
    return deleted


def main() -> None:
    parser = argparse.ArgumentParser(description="Import 2022+ product-goods historical order aggregates")
    parser.add_argument("path", type=Path)
    parser.add_argument("--brand", required=True, choices=("cbanner_mens", "cbanner_womens", "eblan"))
    parser.add_argument("--sheet", default=SHEET_NAME)
    args = parser.parse_args()
    print(import_rows(args.path, brand=args.brand, sheet_name=args.sheet))


if __name__ == "__main__":
    main()
