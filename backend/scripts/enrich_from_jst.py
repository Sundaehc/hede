"""Read 聚水潭 Excel files and update 5 new columns in the product database.

New columns: supplier_name, color_code, size_range, product_model, launch_date

Match strategy: use 原始货号 (or 款式编码/货号 as fallback) to match
against sku and original_sku in the database.
"""
from __future__ import annotations

import os
import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from config import load_settings
from domain.sources import TABLE_NAMES
from storage.product_repository import ProductRepository
from sqlalchemy import text

JST_DIR = "//192.168.10.229/运营组资料/9商品组（卢嘉诚）/商品档案/商品基础信息/商品资料档案汇总/新建文件夹"

# Map filename -> list of brand groups
FILE_BRAND_MAP = {
    "千百度聚水潭商品表最新.xlsx": ["cbanner_mens", "cbanner_womens"],
    "千百度福建聚水潭商品表（最新）.xlsx": ["cbanner_mens"],
    "京东分销聚水潭商品表最新.xlsx": ["cbanner_mens"],
    "千百度拖鞋聚水潭商品表.xlsx": ["cbanner_mens", "cbanner_womens"],
    "千百度女鞋聚水潭商品表最新 - 副本.xlsx": ["cbanner_womens"],
    "千百度女鞋R组聚水潭商品表最新.xlsx": ["cbanner_womens"],
    "烟斗聚水潭商品表（最新）.xlsx": ["yandou"],
    "烟斗福建聚水潭商品表（最新）.xlsx": ["yandou"],
    "伊伴聚水潭商品表（最新）.xlsx": ["eblan"],
    "伊伴女鞋聚水潭商品表（最新）.xlsx": ["eblan"],
    "伊伴福建聚水潭商品表（最新）.xlsx": ["eblan"],
    "伊伴女鞋桐乡聚水潭商品表.xlsx": ["eblan"],
}

# Column name mapping from Excel headers to our DB field names
FIELD_MAP = {
    "供应商名": "supplier_name",
    "颜色代码": "color_code",
    "尺码段": "size_range",
    "产品型号": "product_model",
    "上市时间": "launch_date",
}

# Possible match key column names in Excel
MATCH_KEY_NAMES = {"原始货号", "款式编码", "货号"}


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_rows_from_sheet(ws) -> list[dict]:
    """Extract rows from a worksheet, returning list of dicts keyed by header."""
    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        return []

    headers = [normalize_header(h) for h in header_row]
    rows = []
    for row in iterator:
        row_dict = {}
        for idx, cell_value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                if cell_value is not None:
                    row_dict[headers[idx]] = str(cell_value).strip() if not isinstance(cell_value, str) else cell_value.strip()
                else:
                    row_dict[headers[idx]] = None
        rows.append(row_dict)
    return rows


def get_match_key(row: dict) -> str | None:
    """Get the match key (original_sku equivalent) from a row."""
    for key in MATCH_KEY_NAMES:
        val = row.get(key)
        if val and str(val).strip():
            return str(val).strip()
    return None


def main():
    settings = load_settings()
    repo = ProductRepository(settings.database_url)

    # Collect all enrichment data: match_key -> {field: value}
    # We build per-brand-group maps
    enrich_data: dict[str, dict[str, dict[str, str]]] = {
        brand: {} for brands in FILE_BRAND_MAP.values() for brand in brands
    }

    for filename, brand_groups in FILE_BRAND_MAP.items():
        fpath = os.path.join(JST_DIR, filename)
        if not os.path.exists(fpath):
            print(f"  SKIP (not found): {filename}")
            continue

        print(f"Reading: {filename}")
        try:
            wb = load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        # Prefer "汇总" sheet; if not found, try the first sheet with match keys
        sheets_to_read = []
        if "汇总" in wb.sheetnames:
            sheets_to_read = ["汇总"]
        else:
            sheets_to_read = wb.sheetnames[:1]

        for sheet_name in sheets_to_read:
            ws = wb[sheet_name]
            rows = extract_rows_from_sheet(ws)
            print(f"  [{sheet_name}] {len(rows)} rows")

            for row in rows:
                match_key = get_match_key(row)
                if not match_key:
                    continue

                fields = {}
                for cn_name, db_field in FIELD_MAP.items():
                    val = row.get(cn_name)
                    if val and str(val).strip():
                        fields[db_field] = str(val).strip()

                if not fields:
                    continue

                for brand in brand_groups:
                    if match_key not in enrich_data[brand]:
                        enrich_data[brand][match_key] = fields

        wb.close()

        wb.close()

    # Now update database records
    total_updated = 0
    for brand, key_map in enrich_data.items():
        if not key_map:
            print(f"\n{brand}: no enrichment data")
            continue

        print(f"\n{brand}: {len(key_map)} unique keys to match")

        # Get all products for this brand
        all_products = repo.list_products(brand, query=None, page=1, page_size=1_000_000)
        items = all_products["items"]
        print(f"  Total products in DB: {len(items)}")

        brand_updated = 0
        with repo.engine.begin() as conn:
            for item in items:
                sku_val = str(item.get("sku") or "").strip()
                orig_val = str(item.get("original_sku") or "").strip()

                # Try matching by original_sku first, then sku
                fields = None
                if orig_val in key_map:
                    fields = key_map[orig_val]
                elif sku_val in key_map:
                    fields = key_map[sku_val]

                if not fields:
                    continue

                # Build UPDATE SET clause
                set_parts = []
                params = {"_id": item["id"]}
                for db_field, value in fields.items():
                    # Handle datetime values for launch_date
                    if db_field == "launch_date" and " " in value:
                        value = value.split(" ")[0]  # Take date part only
                    set_parts.append(f"{db_field} = :{db_field}")
                    params[db_field] = value

                if not set_parts:
                    continue

                table_name = TABLE_NAMES[brand]
                sql = f"UPDATE {table_name} SET {', '.join(set_parts)} WHERE id = :_id"
                conn.execute(text(sql), params)
                brand_updated += 1

        total_updated += brand_updated
        print(f"  Updated: {brand_updated}")

    print(f"\nDone. Total updated: {total_updated}")


if __name__ == "__main__":
    main()
