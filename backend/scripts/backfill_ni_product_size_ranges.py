"""Fill NI product size groups from the Desktop NI style-size summary workbook.

The workbook's size detail is matched to existing size-group items exactly.
Missing NI-specific groups are created once, and every resolved product code is
stored in ``product_size_group_mappings`` so future product syncs keep the same
manual size group.
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import insert, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.product_size_group_mapping_schema import PRODUCT_SIZE_GROUP_MAPPINGS_TABLE
from domain.schema import PRODUCT_ARCHIVE_TABLES
from domain.size_group_schema import SIZE_GROUP_ITEMS_TABLE, SIZE_GROUPS_TABLE
from storage.db import Database


DEFAULT_SOURCE_FILE = Path.home() / "Desktop" / "NI款式尺码段汇总.xlsx"
SUMMARY_SHEET_NAME = "款式尺码段汇总"
CODE_HEADER = "款式编码"
SIZE_RANGE_HEADER = "尺码段"
SIZE_DETAIL_HEADER = "尺码明细"


@dataclass(frozen=True)
class NiSizeRangeSource:
    code: str
    size_range: str
    size_labels: tuple[str, ...]
    row_number: int


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _size_labels(value: object) -> tuple[str, ...]:
    normalized = _text(value).replace(",", "、")
    return tuple(item.strip() for item in normalized.split("、") if item and item.strip())


def read_ni_size_ranges(source_file: Path) -> list[NiSizeRangeSource]:
    if not source_file.is_file():
        raise FileNotFoundError(f"找不到 NI 尺码汇总文件：{source_file}")

    workbook = load_workbook(source_file, read_only=True, data_only=True)
    try:
        if SUMMARY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"未找到工作表：{SUMMARY_SHEET_NAME}")
        worksheet = workbook[SUMMARY_SHEET_NAME]
        header_row_number = None
        headers: dict[str, int] = {}
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            candidate = {_text(value): index for index, value in enumerate(row) if _text(value)}
            if {CODE_HEADER, SIZE_RANGE_HEADER, SIZE_DETAIL_HEADER}.issubset(candidate):
                header_row_number = row_number
                headers = candidate
                break
            if row_number >= 50:
                break
        if header_row_number is None:
            raise ValueError("未找到款式编码、尺码段、尺码明细表头")

        source_by_code: dict[str, NiSizeRangeSource] = {}
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            code = _text(row[headers[CODE_HEADER]] if headers[CODE_HEADER] < len(row) else None)
            size_range = _text(row[headers[SIZE_RANGE_HEADER]] if headers[SIZE_RANGE_HEADER] < len(row) else None)
            size_labels = _size_labels(row[headers[SIZE_DETAIL_HEADER]] if headers[SIZE_DETAIL_HEADER] < len(row) else None)
            if not code or not size_labels:
                continue
            item = NiSizeRangeSource(code, size_range, size_labels, row_number)
            previous = source_by_code.get(code)
            if previous is not None and previous.size_labels != item.size_labels:
                raise ValueError(f"款式编码 {code} 存在不一致的尺码明细")
            source_by_code[code] = item
        return list(source_by_code.values())
    finally:
        workbook.close()


def _group_name_for_new_ni_size_range(size_range: str, size_labels: tuple[str, ...]) -> str:
    label = size_range or "-".join(size_labels)
    return f"NI尺码段{label}"


def _load_groups_by_size_labels(connection) -> dict[tuple[str, ...], list[str]]:
    rows = connection.execute(
        select(SIZE_GROUPS_TABLE.c.name, SIZE_GROUP_ITEMS_TABLE.c.size_name)
        .select_from(
            SIZE_GROUPS_TABLE.join(
                SIZE_GROUP_ITEMS_TABLE,
                SIZE_GROUP_ITEMS_TABLE.c.size_group_id == SIZE_GROUPS_TABLE.c.id,
            )
        )
        .order_by(SIZE_GROUPS_TABLE.c.name, SIZE_GROUP_ITEMS_TABLE.c.sort_order, SIZE_GROUP_ITEMS_TABLE.c.id)
    ).mappings()
    labels_by_name: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        labels_by_name[_text(row["name"])].append(_text(row["size_name"]))
    groups: dict[tuple[str, ...], list[str]] = defaultdict(list)
    for name, labels in labels_by_name.items():
        groups[tuple(labels)].append(name)
    return groups


def _prefer_ni_group_name(names: list[str]) -> str:
    return min(names, key=lambda name: (not name.startswith("NI"), name))


def backfill_ni_size_ranges(*, source_file: Path, apply: bool) -> dict[str, object]:
    source_rows = read_ni_size_ranges(source_file)
    database = Database(load_settings(require_database=True).database_url)
    engine = database._require_engine()
    for table in (SIZE_GROUPS_TABLE, SIZE_GROUP_ITEMS_TABLE, PRODUCT_SIZE_GROUP_MAPPINGS_TABLE, PRODUCT_ARCHIVE_TABLES["ni"]):
        table.create(engine, checkfirst=True)

    signatures = {row.size_labels for row in source_rows}
    with engine.connect() as connection:
        groups_by_labels = _load_groups_by_size_labels(connection)
    missing_signatures = sorted(signatures - set(groups_by_labels), key=lambda labels: (len(labels), labels))
    new_groups = {
        labels: _group_name_for_new_ni_size_range(
            next(row.size_range for row in source_rows if row.size_labels == labels),
            labels,
        )
        for labels in missing_signatures
    }
    resolved_groups = {
        labels: _prefer_ni_group_name(names)
        for labels, names in groups_by_labels.items()
        if labels in signatures
    }
    resolved_groups.update(new_groups)
    resolved_by_code = {row.code: resolved_groups[row.size_labels] for row in source_rows}

    summary: dict[str, object] = {
        "source_file": str(source_file),
        "source_codes": len(source_rows),
        "size_combinations": len(signatures),
        "created_groups": list(new_groups.values()),
        "resolved_groups": {"、".join(labels): name for labels, name in resolved_groups.items()},
        "mapped_products": 0,
        "apply": apply,
    }
    if not apply:
        return summary

    with engine.begin() as connection:
        for labels, name in new_groups.items():
            size_group_id = connection.execute(
                insert(SIZE_GROUPS_TABLE).values(name=name).returning(SIZE_GROUPS_TABLE.c.id)
            ).scalar_one()
            connection.execute(
                insert(SIZE_GROUP_ITEMS_TABLE),
                [
                    {
                        "size_group_id": size_group_id,
                        "size_name": size_name,
                        "barcode": size_name,
                        "sort_order": index,
                    }
                    for index, size_name in enumerate(labels, start=1)
                ],
            )

        mapping_rows = [
            {
                "product_code": row.code,
                "size_group_name": resolved_by_code[row.code],
                "source_workbook": source_file.name,
                "source_sheet": SUMMARY_SHEET_NAME,
                "source_row_number": str(row.row_number),
            }
            for row in source_rows
        ]
        statement = pg_insert(PRODUCT_SIZE_GROUP_MAPPINGS_TABLE).values(mapping_rows)
        excluded = statement.excluded
        connection.execute(
            statement.on_conflict_do_update(
                index_elements=["product_code"],
                set_={
                    "size_group_name": excluded.size_group_name,
                    "source_workbook": excluded.source_workbook,
                    "source_sheet": excluded.source_sheet,
                    "source_row_number": excluded.source_row_number,
                },
            )
        )

        ni_table = PRODUCT_ARCHIVE_TABLES["ni"]
        mapped_products = 0
        for row in source_rows:
            result = connection.execute(
                update(ni_table)
                .where((ni_table.c.sku == row.code) | (ni_table.c.original_sku == row.code))
                .values(size_range=resolved_by_code[row.code])
            )
            mapped_products += result.rowcount or 0
        summary["mapped_products"] = mapped_products

    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="从 NI 款式尺码段汇总回填 NI 商品档案尺码组")
    parser.add_argument("--file", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--apply", action="store_true", help="写入尺码组、映射和商品档案；未传入时仅预览")
    args = parser.parse_args()
    print(backfill_ni_size_ranges(source_file=args.file, apply=args.apply))


if __name__ == "__main__":
    main()
