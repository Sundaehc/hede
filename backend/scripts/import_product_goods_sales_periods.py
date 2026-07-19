"""Import annual and monthly sales matrices from a product-goods workbook."""
from __future__ import annotations

import argparse
import re
from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.product_goods_sales_period_schema import PRODUCT_GOODS_SALES_PERIODS_TABLE, ensure_product_goods_sales_periods_table


YEAR_PATTERN = re.compile(r"(?<!\d)(20\d{2})(?:年)?(?:年度)?(?:销(?:售|量))?(?!\d)")
MONTH_PATTERN = re.compile(r"(?<!\d)(\d{2,4})[-/]([1-9]|1[0-2])(?!\d)")
GOODS_CODE_HEADERS = ("货号", "商品货号", "商品编码", "SKU", "sku")
STYLE_CODE_HEADERS = ("款号", "款式编码", "原始款号", "原货号")


def _text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _integer(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(Decimal(str(value).replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def _header_index(headers: tuple[object, ...], names: tuple[str, ...]) -> int | None:
    normalized = [str(value or "").strip().replace("\n", "") for value in headers]
    for name in names:
        if name in normalized:
            return normalized.index(name)
    return None


def _period_columns(headers: tuple[object, ...]) -> list[tuple[str, date, int]]:
    periods: list[tuple[str, date, int]] = []
    for index, header in enumerate(headers):
        if not isinstance(header, str):
            continue
        label = header.strip()
        year_match = YEAR_PATTERN.fullmatch(label)
        if year_match:
            periods.append(("year", date(int(year_match.group(1)), 1, 1), index))
            continue
        month_match = MONTH_PATTERN.fullmatch(label)
        if not month_match:
            continue
        raw_year = int(month_match.group(1))
        year = raw_year if raw_year >= 2000 else 2000 + raw_year
        periods.append(("month", date(year, int(month_match.group(2)), 1), index))
    return periods


def _find_matrix_header(sheet) -> tuple[int, tuple[object, ...], list[tuple[str, date, int]]]:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        periods = _period_columns(row)
        if len(periods) >= 12:
            return row_number, row, periods
    raise ValueError(f"未在 {sheet.title} 找到年度/月度销量矩阵表头")


def _source_as_of_date(headers: tuple[object, ...]) -> date | None:
    dates = [
        value.date() if isinstance(value, datetime) else value
        for value in headers
        if isinstance(value, (date, datetime))
    ]
    return max(dates, default=None)


def iter_rows(path: Path, *, brand: str, sheet_name: str = "商品明细表") -> Iterator[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"未找到 sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        header_row, headers, periods = _find_matrix_header(sheet)
        goods_code_index = _header_index(headers, GOODS_CODE_HEADERS)
        style_code_index = _header_index(headers, STYLE_CODE_HEADERS)
        if goods_code_index is None:
            raise ValueError("未找到货号列")
        source_as_of_date = _source_as_of_date(headers)
        for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            product_code = _text(values[goods_code_index] if goods_code_index < len(values) else None)
            if not product_code:
                continue
            style_code = _text(values[style_code_index] if style_code_index is not None and style_code_index < len(values) else None)
            for period_type, period_start, column_index in periods:
                sales_quantity = _integer(values[column_index] if column_index < len(values) else None)
                if sales_quantity is None:
                    continue
                yield {
                    "brand": brand,
                    "product_code": product_code,
                    "style_code": style_code,
                    "period_type": period_type,
                    "period_start": period_start,
                    "sales_quantity": sales_quantity,
                    "source_as_of_date": source_as_of_date,
                    "source_workbook": path.name,
                    "source_sheet": sheet.title,
                    "source_row_number": row_number,
                }
    finally:
        workbook.close()


def import_rows(path: Path, *, brand: str, sheet_name: str = "商品明细表") -> dict[str, int]:
    settings = load_settings(require_database=True)
    assert settings.database_url is not None
    engine = create_engine(settings.database_url, future=True)
    ensure_product_goods_sales_periods_table(engine)
    read = 0
    written = 0
    chunk: list[dict[str, object]] = []

    def save_chunk(rows: list[dict[str, object]]) -> int:
        if not rows:
            return 0
        with engine.begin() as connection:
            statement = pg_insert(PRODUCT_GOODS_SALES_PERIODS_TABLE).values(rows)
            statement = statement.on_conflict_do_update(
                constraint="uq_product_goods_sales_period_source",
                set_={
                    "brand": statement.excluded.brand,
                    "product_code": statement.excluded.product_code,
                    "style_code": statement.excluded.style_code,
                    "sales_quantity": statement.excluded.sales_quantity,
                    "source_as_of_date": statement.excluded.source_as_of_date,
                },
            )
            connection.execute(statement)
        return len(rows)

    for row in iter_rows(path, brand=brand, sheet_name=sheet_name):
        read += 1
        chunk.append(row)
        if len(chunk) >= 2_000:
            written += save_chunk(chunk)
            chunk = []
    written += save_chunk(chunk)
    return {"read": read, "written": written}


def main() -> None:
    parser = argparse.ArgumentParser(description="Import annual and monthly product-goods sales matrix")
    parser.add_argument("path", type=Path)
    parser.add_argument("--brand", required=True)
    parser.add_argument("--sheet", default="商品明细表")
    args = parser.parse_args()
    result = import_rows(args.path, brand=args.brand, sheet_name=args.sheet)
    print(f"Read {result['read']} period values; wrote {result['written']} rows")


if __name__ == "__main__":
    main()
