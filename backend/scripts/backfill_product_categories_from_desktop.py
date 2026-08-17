"""Backfill product archive categories from the desktop style-code workbook.

Preview changes:
    uv run python -m scripts.backfill_product_categories_from_desktop

Apply changes:
    uv run python -m scripts.backfill_product_categories_from_desktop --apply
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from itertools import islice
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import bindparam, func, inspect, select, text, update

from config import load_settings
from storage.product_repository import ProductRepository
from transform.rows import normalize_cell, normalize_header


DEFAULT_SOURCE_PATH = Path.home() / "Desktop" / "款式编码分类汇总.xlsx"
REQUIRED_HEADERS = ("款式编码", "分类")
BRAND_CATEGORY_HINTS = {
    "cbanner_mens": "男鞋",
    "cbanner_womens": "女鞋",
    "yandou": "男鞋",
    "eblan": "女鞋",
}


@dataclass(frozen=True)
class SourceData:
    categories_by_code: dict[str, set[str]]
    sheet_name: str
    header_row: int
    data_rows: int


def normalized_text(value: object) -> str:
    normalized = normalize_cell(value)
    return str(normalized).strip() if normalized is not None else ""


def normalized_code(value: object) -> str:
    return normalized_text(value).upper()


def find_headers(sheet) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(islice(sheet.iter_rows(values_only=True), 30), start=1):
        headers = {
            normalize_header(value): index
            for index, value in enumerate(row)
            if normalize_header(value)
        }
        if all(header in headers for header in REQUIRED_HEADERS):
            return row_index, headers
    return None


def read_source(path: Path) -> SourceData:
    if not path.exists():
        raise FileNotFoundError(f"未找到来源文件：{path}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            header_result = find_headers(sheet)
            if header_result is None:
                continue
            header_row, headers = header_result
            categories_by_code: dict[str, set[str]] = defaultdict(set)
            data_rows = 0
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                code = normalized_code(row[headers["款式编码"]]) if headers["款式编码"] < len(row) else ""
                category = normalized_text(row[headers["分类"]]) if headers["分类"] < len(row) else ""
                if not code or not category:
                    continue
                categories_by_code[code].add(category)
                data_rows += 1
            return SourceData(
                categories_by_code=dict(categories_by_code),
                sheet_name=sheet.title,
                header_row=header_row,
                data_rows=data_rows,
            )
    finally:
        workbook.close()

    raise ValueError(f"未找到同时包含{'、'.join(REQUIRED_HEADERS)}的工作表")


def resolve_category(categories: set[str], brand: str) -> str | None:
    if len(categories) == 1:
        return next(iter(categories))
    hint = BRAND_CATEGORY_HINTS.get(brand)
    return hint if hint in categories else None


def category_column_exists(repository: ProductRepository, table_name: str) -> bool:
    with repository.engine.connect() as connection:
        return any(column["name"] == "category" for column in inspect(connection).get_columns(table_name))


def build_updates(
    repository: ProductRepository,
    brand: str,
    source: SourceData,
) -> tuple[list[dict[str, object]], dict[str, int], set[str]]:
    table = repository._table_for_brand(brand)
    has_category = category_column_exists(repository, table.name)
    selected_columns = [table.c.id, table.c.sku, table.c.original_sku]
    if has_category:
        selected_columns.append(table.c.category)

    updates: list[dict[str, object]] = []
    ambiguous_codes: set[str] = set()
    matched = 0
    unchanged = 0
    total = 0
    matched_by_original_sku = 0

    with repository.engine.connect() as connection:
        for row in connection.execute(select(*selected_columns)).mappings():
            total += 1
            sku = normalized_code(row.get("sku"))
            original_sku = normalized_code(row.get("original_sku"))
            categories = source.categories_by_code.get(sku)
            if not categories and original_sku:
                categories = source.categories_by_code.get(original_sku)
                if categories:
                    matched_by_original_sku += 1
            if not categories:
                continue

            matched += 1
            category = resolve_category(categories, brand)
            if category is None:
                ambiguous_codes.add(sku or original_sku)
                continue
            if normalized_text(row.get("category")) == category:
                unchanged += 1
                continue
            updates.append({"_product_id": int(row["id"]), "_category": category})

    return updates, {
        "total": total,
        "matched": matched,
        "matched_by_original_sku": matched_by_original_sku,
        "unchanged": unchanged,
        "updates": len(updates),
    }, ambiguous_codes


def ensure_category_columns(repository: ProductRepository, brands: list[str]) -> None:
    with repository.engine.begin() as connection:
        preparer = connection.dialect.identifier_preparer
        for brand in brands:
            table = repository._table_for_brand(brand)
            table_name = preparer.quote(table.name)
            connection.execute(text(f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS category TEXT"))


def apply_updates(repository: ProductRepository, brand: str, updates_to_apply: list[dict[str, object]]) -> None:
    if not updates_to_apply:
        return
    table = repository._table_for_brand(brand)
    statement = (
        update(table)
        .where(table.c.id == bindparam("_product_id"))
        .values(
            category=bindparam("_category"),
            updated_at=func.date_trunc("minute", func.now()),
        )
    )
    with repository.engine.begin() as connection:
        connection.execute(statement, updates_to_apply)


def main() -> None:
    parser = argparse.ArgumentParser(description="根据款式编码分类汇总回填商品档案分类")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH, help="来源 Excel 路径")
    parser.add_argument("--apply", action="store_true", help="执行数据库更新；未传入时仅预览")
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")
    source = read_source(args.source)
    repository = ProductRepository(load_settings().database_url)
    brands = repository.product_archive_brands()
    if args.apply:
        ensure_category_columns(repository, brands)

    category_counts = Counter(
        category
        for categories in source.categories_by_code.values()
        for category in categories
    )
    conflicting_source_codes = sum(len(categories) > 1 for categories in source.categories_by_code.values())
    print("模式：" + ("正式回填" if args.apply else "预览（未写入数据库）"))
    print(
        f"来源：{args.source.name} / {source.sheet_name}，表头第 {source.header_row} 行，"
        f"数据 {source.data_rows} 行，款式编码 {len(source.categories_by_code)} 个，"
        f"分类 {dict(category_counts)}，多分类款式 {conflicting_source_codes} 个"
    )

    total_updates = 0
    all_ambiguous_codes: set[str] = set()
    for brand in brands:
        updates_to_apply, stats, ambiguous_codes = build_updates(repository, brand, source)
        if args.apply:
            apply_updates(repository, brand, updates_to_apply)
        total_updates += len(updates_to_apply)
        all_ambiguous_codes.update(ambiguous_codes)
        print(
            f"{brand}：档案 {stats['total']} 条，命中 {stats['matched']} 条，"
            f"原始货号命中 {stats['matched_by_original_sku']} 条，"
            f"{'已更新' if args.apply else '待更新'} {stats['updates']} 条，"
            f"无需更新 {stats['unchanged']} 条，无法判定 {len(ambiguous_codes)} 条"
        )

    print(("已回填" if args.apply else "待回填") + f"商品记录：{total_updates} 条")
    if all_ambiguous_codes:
        print("未自动处理的多分类款式：" + "、".join(sorted(all_ambiguous_codes)))


if __name__ == "__main__":
    main()
