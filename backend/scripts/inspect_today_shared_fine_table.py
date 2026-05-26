"""Inspect today's shared JST/VIP files and optional fine-table workbook.

Run from backend:
    python -m scripts.inspect_today_shared_fine_table
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from config import load_settings


def _safe_list(path: Path) -> list[Path]:
    if not path.exists():
        return []
    return sorted(path.iterdir(), key=lambda item: item.name)


def _preview_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(min_row=1, max_row=5, values_only=True))
        return {
            "sheet": worksheet.title,
            "rows": rows,
        }
    finally:
        workbook.close()


def main() -> None:
    settings = load_settings(require_database=False)
    assert settings.jst_stock_root is not None, "JST_STOCK_ROOT is required"

    today_name = date.today().strftime("%m.%d")
    other_platform_dir = settings.jst_stock_root.parent / "其他平台" / today_name
    print(f"目录: {other_platform_dir}")
    print(f"存在: {other_platform_dir.exists()}")

    files = _safe_list(other_platform_dir)
    print("文件:")
    for file_path in files:
        print(f"- {file_path.name} [{file_path.name.encode('unicode_escape').decode('ascii')}]")
        if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
            preview = _preview_workbook(file_path)
            print(f"  sheet: {preview['sheet']} [{str(preview['sheet']).encode('unicode_escape').decode('ascii')}]")
            for row in preview["rows"]:
                encoded = tuple(
                    None if value is None else str(value).encode("unicode_escape").decode("ascii")
                    for value in row
                )
                print(f"  {encoded}")

    fine_files = [
        file_path
        for file_path in files
        if file_path.suffix.lower() in {".xlsx", ".xlsm"} and "精细" in file_path.stem
    ]
    print("精细表候选:")
    for file_path in fine_files:
        print(f"- {file_path.name} [{file_path.name.encode('unicode_escape').decode('ascii')}]")
        preview = _preview_workbook(file_path)
        print(f"  sheet: {preview['sheet']}")
        for row in preview["rows"]:
            print(f"  {row}")


if __name__ == "__main__":
    main()
