"""Import color barcode workbooks into color_barcodes.

Run:
    python -m scripts.import_color_barcodes --replace
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, insert
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.color_barcode_schema import COLOR_BARCODE_TABLE
from storage.db import Database
from transform.rows import normalize_cell, normalize_header


DEFAULT_ROOT = Path(r"\\Creation\超级共享\影刀技术开发部共享\最新颜色表")
DEFAULT_SHEET = "导出数据"
ROW_CHUNK_SIZE = 1000
WORKBOOK_SPECS = (
    ("cbanner_mens", "男鞋_可读.xlsx"),
    ("cbanner_womens", "女鞋_可读.xlsx"),
)


def _cell_text(value: object) -> str:
    return str(normalize_cell(value) or "").strip()


def _find_header_indexes(rows: list[tuple[Any, ...]]) -> tuple[int, int, int] | None:
    for row_index, row in enumerate(rows[:30]):
        headers = [normalize_header(value) for value in row]
        code_index = next((index for index, value in enumerate(headers) if value == "颜色条码"), None)
        name_index = next((index for index, value in enumerate(headers) if value == "颜色名称"), None)
        if code_index is not None and name_index is not None:
            return row_index, code_index, name_index
    return None


def read_color_barcode_rows(path: Path, *, brand: str, sheet_name: str = DEFAULT_SHEET) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        header_indexes = _find_header_indexes(rows)
        if header_indexes is None:
            raise ValueError(f"未找到颜色条码表头: {path}")
        header_row_index, code_index, name_index = header_indexes
        headers = [normalize_header(value) for value in rows[header_row_index]]
        payloads: list[dict[str, object]] = []
        for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            color_barcode = _cell_text(row[code_index] if code_index < len(row) else None)
            color_name = _cell_text(row[name_index] if name_index < len(row) else None)
            if not color_barcode or not color_name:
                continue
            row_brand = "smiley" if "笑脸" in color_name else brand
            raw_payload = {
                headers[index] or f"column_{index + 1}": _cell_text(value)
                for index, value in enumerate(row)
                if index < len(headers)
            }
            payloads.append(
                {
                    "brand": row_brand,
                    "color_barcode": color_barcode,
                    "color_name": color_name,
                    "source_workbook": str(path),
                    "source_sheet": worksheet.title,
                    "source_row_number": str(row_number),
                    "raw_payload": raw_payload,
                }
            )
        return payloads
    finally:
        workbook.close()


def dedupe_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    latest_by_key: dict[tuple[str, str], dict[str, object]] = {}
    for row in rows:
        key = (str(row["brand"]), str(row["color_barcode"]))
        latest_by_key[key] = row
    return list(latest_by_key.values())


def import_color_barcodes(root: Path, *, sheet_name: str = DEFAULT_SHEET, replace: bool = False) -> dict[str, int]:
    rows: list[dict[str, object]] = []
    for brand, filename in WORKBOOK_SPECS:
        rows.extend(read_color_barcode_rows(root / filename, brand=brand, sheet_name=sheet_name))
    rows = dedupe_rows(rows)

    settings = load_settings()
    database = Database(settings.database_url)
    database.create_tables()
    table = COLOR_BARCODE_TABLE
    with database._require_engine().begin() as connection:
        table.create(connection, checkfirst=True)
        if replace:
            connection.execute(delete(table))
        if rows:
            update_columns = [
                column.name
                for column in table.columns
                if column.name not in ("id", "brand", "color_barcode", "created_at")
            ]
            for start in range(0, len(rows), ROW_CHUNK_SIZE):
                statement = pg_insert(table).values(rows[start:start + ROW_CHUNK_SIZE])
                excluded = statement.excluded
                statement = statement.on_conflict_do_update(
                    index_elements=["brand", "color_barcode"],
                    set_={
                        **{column: getattr(excluded, column) for column in update_columns},
                        "updated_at": func.date_trunc("minute", func.now()),
                    },
                )
                connection.execute(statement)
    return {"rows": len(rows)}


def main() -> None:
    parser = argparse.ArgumentParser(description="导入颜色条码表")
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--replace", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows = []
    for brand, filename in WORKBOOK_SPECS:
        brand_rows = read_color_barcode_rows(args.root / filename, brand=brand, sheet_name=args.sheet)
        rows.extend(brand_rows)
        print(f"brand={brand} file={filename} rows={len(brand_rows)}")
    rows = dedupe_rows(rows)
    print(f"total={len(rows)}")
    if args.dry_run:
        return

    result = import_color_barcodes(args.root, sheet_name=args.sheet, replace=args.replace)
    print(f"imported={result['rows']}")


if __name__ == "__main__":
    main()
