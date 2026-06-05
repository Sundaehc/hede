from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, insert, select, update

from config import load_settings
from domain.fine_table_snapshot_schema import (
    FINE_TABLE_SNAPSHOT_BATCH_TABLE,
    FINE_TABLE_SNAPSHOT_ROW_TABLE,
    ensure_fine_table_snapshot_row_table,
)
from storage.inventory_repository import InventoryRepository


sys.stdout.reconfigure(encoding="utf-8")

DEFAULT_ROOTS = (
    Path(r"\\192.168.10.229\运营组资料\影刀\千百度精细表"),
    Path(r"\\192.168.10.229\运营组资料\影刀\烟斗精细表"),
    Path(r"\\192.168.10.229\运营组资料\影刀\伊伴精细表"),
)
SHEET_NAME = "精细表"
ROW_CHUNK_SIZE = 1000
BRAND_MENS = "cbanner_mens"
BRAND_WOMENS = "cbanner_womens"
BRAND_YANDOU = "yandou"
BRAND_EBLAN = "eblan"
SIZE_LABELS = [
    "34/220",
    "35/225",
    "36/230",
    "37/235",
    "38/240",
    "39/245",
    "40/250",
    "41/255",
    "42/260",
    "43/265",
    "44/270",
    "45/275",
    "46/280",
    "47/285",
]


def log(message: str) -> None:
    print(message, flush=True)


@dataclass(frozen=True)
class FineTableFile:
    path: Path
    snapshot_date: date
    brand: str | None
    scope: str
    last_write_time: datetime


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def to_json_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        value = value.strip().replace("%", "")
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> int:
    numeric = to_float(value)
    return 0 if numeric is None else int(round(numeric))


def normalize_header(value: Any) -> str:
    return normalize_text(value).replace("\n", "").replace(" ", "")


def parse_month_day_suffix(suffix: str) -> tuple[int, int] | None:
    month_text, dot, rest = suffix.partition(".")
    if not dot:
        return None
    day_text = ""
    for char in rest:
        if char.isdigit():
            day_text += char
        else:
            break
    if not month_text.isdigit() or not day_text:
        return None
    return int(month_text), int(day_text)


def parse_file(root: Path, path: Path) -> FineTableFile | None:
    name = path.stem
    brand: str | None = None
    suffix: str | None = None

    cbanner_marker = "精细数新"
    if name.startswith("千百度") and cbanner_marker in name:
        brand_part = name[: name.index(cbanner_marker)]
        if "男鞋" in brand_part:
            brand = BRAND_MENS
        elif "女鞋" in brand_part:
            brand = BRAND_WOMENS
        suffix = name.split(cbanner_marker, 1)[1]
    elif name.startswith("精细数据新"):
        brand = BRAND_YANDOU
        suffix = name.split("精细数据新", 1)[1]
    elif name.startswith("伊伴精细数"):
        brand = BRAND_EBLAN
        suffix = name.split("伊伴精细数", 1)[1]
    if suffix is None:
        return None

    parsed_date = parse_month_day_suffix(suffix)
    if parsed_date is None:
        return None
    month, day = parsed_date

    relative = path.relative_to(root)
    scope = relative.parts[0] if len(relative.parts) > 1 else "根目录"
    year = int(scope[:4]) if scope.endswith("年") and scope[:4].isdigit() else 2026
    try:
        snapshot_date = date(year, month, day)
    except ValueError:
        return None
    return FineTableFile(
        path=path,
        snapshot_date=snapshot_date,
        brand=brand,
        scope=scope,
        last_write_time=datetime.fromtimestamp(path.stat().st_mtime),
    )


def scan_files(root: Path) -> list[FineTableFile]:
    files: list[FineTableFile] = []
    scanned = 0
    for path in root.rglob("*"):
        scanned += 1
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in {".xlsm", ".xlsx", ".xls"}:
            continue
        parsed = parse_file(root, path)
        if parsed is not None:
            files.append(parsed)
            if len(files) % 100 == 0:
                log(f"scanned_paths={scanned} matched_files={len(files)}")
    return files


def scan_roots(roots: list[Path]) -> list[FineTableFile]:
    files: list[FineTableFile] = []
    for root in roots:
        log(f"scan_root={root}")
        files.extend(scan_files(root))
    return files


def choose_files(files: list[FineTableFile]) -> list[FineTableFile]:
    grouped: dict[tuple[str, date, str], FineTableFile] = {}
    for item in files:
        brand_key = item.brand or "combined"
        key = (item.scope, item.snapshot_date, brand_key)
        existing = grouped.get(key)
        if existing is None or item.last_write_time > existing.last_write_time:
            grouped[key] = item
    return sorted(grouped.values(), key=lambda item: (item.snapshot_date, item.scope, item.brand or "combined"))


def header_index(headers: list[str], *names: str) -> int | None:
    normalized_names = {normalize_header(name) for name in names}
    return next((index for index, value in enumerate(headers) if value in normalized_names), None)


def row_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def status_key(status: str | None) -> str:
    if not status:
        return "unknown"
    if "下线" in status or "下架" in status:
        return "offline"
    if "部分" in status:
        return "partial"
    if "上线" in status:
        return "online"
    return "unknown"


def empty_payload() -> dict[str, Any]:
    return {
        "id": 0,
        "brand": BRAND_MENS,
        "image_path": None,
        "image_url": None,
        "sku": None,
        "original_sku": None,
        "group_name": None,
        "cost": None,
        "factory_sku": None,
        "factory_code": None,
        "color": None,
        "season_category": None,
        "year": None,
        "upper_material": None,
        "lining_material": None,
        "outsole_material": None,
        "insole_material": None,
        "execution_standard": None,
        "heel_height": None,
        "shoe_width": None,
        "shoe_length": None,
        "shaft_circumference": None,
        "shaft_height": None,
        "internal_height_increase": None,
        "internal_height_note": None,
        "upper_height": None,
        "toe_shape": None,
        "closure_type": None,
        "shoe_box_spec": None,
        "first_order_time": None,
        "size_range": None,
        "product_model": None,
        "supplier_name": None,
        "color_code": None,
        "launch_date": None,
        "source_workbook": "",
        "source_sheet": SHEET_NAME,
        "source_row_number": "",
        "goods_id": None,
        "p_spu": None,
        "style_code": None,
        "category_l3": None,
        "product_name": None,
        "main_style": None,
        "goods_status": None,
        "status_key": "unknown",
        "sales_tag": None,
        "goods_tag": None,
        "latest_purchase_price": None,
        "final_price": None,
        "vip_price": None,
        "market_price": None,
        "price_band": None,
        "activity_profit": None,
        "margin_rate": None,
        "discount_rate": None,
        "vip_1d_sales": 0,
        "vip_3d_sales": 0,
        "vip_7d_sales": 0,
        "vip_15d_sales": 0,
        "vip_30d_sales": 0,
        "vip_3d_uv": 0,
        "vip_7d_uv": 0,
        "vip_30d_uv": 0,
        "vip_3d_ctr": None,
        "vip_7d_ctr": None,
        "vip_30d_ctr": None,
        "vip_3d_conversion": None,
        "vip_7d_conversion": None,
        "vip_30d_conversion": None,
        "vip_3d_sales_change_rate": None,
        "vip_3d_uv_change_rate": None,
        "vip_3d_ctr_change_rate": None,
        "vip_3d_conversion_change_rate": None,
        "vip_7d_sales_change_rate": None,
        "vip_7d_uv_change_rate": None,
        "vip_7d_ctr_change_rate": None,
        "vip_7d_conversion_change_rate": None,
        "vip_30d_reject_count": 0,
        "vip_30d_reject_rate": None,
        "vip_daily_average_sales": 0,
        "other_3d_sales": 0,
        "other_7d_sales": 0,
        "other_15d_sales": 0,
        "other_30d_sales": 0,
        "original_other_3d_sales": 0,
        "original_other_7d_sales": 0,
        "original_other_15d_sales": 0,
        "original_other_30d_sales": 0,
        "shop_30d_sales": [],
        "stock_qty": 0,
        "size_stock": {label: 0 for label in SIZE_LABELS},
        "purchase_diff": 0,
        "inbound_qty": 0,
        "defect_stock": 0,
        "original_defect_stock": 0,
        "original_inbound_qty": 0,
        "original_order_in_transit_stock": 0,
        "original_defect_in_transit_stock": 0,
        "off_shelf_stock": 0,
        "order_occupy_stock": 0,
        "defect_in_transit_stock": 0,
        "projected_15d_stock": 0,
        "daily_sales": [],
        "extra_fields": None,
    }


def build_indexes(headers: list[str]) -> dict[str, int | None]:
    return {
        "sku": header_index(headers, "货号"),
        "original_sku": header_index(headers, "原始货号"),
        "style_code": header_index(headers, "款号"),
        "goods_id": header_index(headers, "商品ID(MID)", "商品ID"),
        "p_spu": header_index(headers, "P-SPU"),
        "execution_standard": header_index(headers, "执行标"),
        "goods_tag": header_index(headers, "小灯塔"),
        "factory_code": header_index(headers, "工厂代码"),
        "factory_sku": header_index(headers, "工厂货号"),
        "insole_material": header_index(headers, "现管家婆鞋垫材质", "鞋垫材质"),
        "outsole_material": header_index(headers, "现管家婆大底材质", "大底材质"),
        "lining_material": header_index(headers, "现管家婆内里材质", "内里材质"),
        "upper_material": header_index(headers, "现管家婆鞋面材质", "鞋面材质"),
        "shoe_box_spec": header_index(headers, "鞋盒规格"),
        "group_name": header_index(headers, "组别"),
        "cost": header_index(headers, "成本"),
        "product_name": header_index(headers, "品名"),
        "final_price": header_index(headers, "今日常态到手价"),
        "vip_price": header_index(headers, "唯品价"),
        "market_price": header_index(headers, "市场价"),
        "discount_rate": header_index(headers, "打折率"),
        "activity_profit": header_index(headers, "活动毛利"),
        "margin_rate": header_index(headers, "活动毛利率"),
        "category_l3": header_index(headers, "三级分类"),
        "season_category": header_index(headers, "季节分类"),
        "year": header_index(headers, "季节"),
        "main_style": header_index(headers, "主款式"),
        "first_order_time": header_index(headers, "首单日期"),
        "goods_status": header_index(headers, "最新上下线状态"),
        "sales_tag": header_index(headers, "畅销度"),
        "projected_5d_stock": header_index(headers, "现有五天后预计库存（不加未到）"),
        "vip_daily_average_sales": header_index(headers, "唯品日常日均"),
        "vip_projected_15d_sales": header_index(headers, "唯品预计15天销量"),
        "other_daily_average_sales": header_index(headers, "其他平台日均销量"),
        "other_projected_15d_sales": header_index(headers, "其他平台15天预计销量"),
        "original_other_3d_sales": header_index(headers, "其他平台3天原始货号汇总"),
        "original_other_7d_sales": header_index(headers, "其他平台7天原始货号汇总"),
        "original_other_15d_sales": header_index(headers, "其他平台15天原始货号汇总"),
        "original_other_30d_sales": header_index(headers, "其他平台30天原始货号汇总"),
        "stock_qty": header_index(headers, "聚水潭库存"),
        "original_defect_stock": header_index(headers, "原始货号次品仓汇总"),
        "original_inbound_qty": header_index(headers, "原始货号采购在途数量汇总"),
        "original_order_in_transit_stock": header_index(headers, "原始货号已下订单未到数量汇总"),
        "original_defect_in_transit_stock": header_index(headers, "原始货号打次未到數量汇总", "原始货号打次未到数量汇总"),
        "defect_stock": header_index(headers, "次品仓库存"),
        "inbound_qty": header_index(headers, "采购在途数量"),
        "order_occupy_stock": header_index(headers, "已下订单未到数量"),
        "defect_in_transit_stock": header_index(headers, "打次未到數量", "打次未到数量"),
        "off_shelf_stock": header_index(headers, "下架仓商品数量"),
        **{f"size_{label}": header_index(headers, label) for label in SIZE_LABELS},
    }


def infer_row_brand(file_brand: str | None, group_name: str | None) -> str:
    if file_brand is not None:
        return file_brand
    if group_name and "女鞋" in group_name:
        return BRAND_WOMENS
    return BRAND_MENS


def build_payload(
    *,
    file: FineTableFile,
    row: tuple[Any, ...],
    row_number: int,
    indexes: dict[str, int | None],
    include_extra_fields: bool,
    headers: list[str],
) -> tuple[str, dict[str, Any]] | None:
    sku = normalize_text(row_value(row, indexes["sku"]))
    original_sku = normalize_text(row_value(row, indexes["original_sku"])) or sku
    if not sku and not original_sku:
        return None

    group_name = normalize_text(row_value(row, indexes["group_name"])) or None
    brand = infer_row_brand(file.brand, group_name)
    payload = empty_payload()
    payload.update(
        {
            "id": -row_number,
            "brand": brand,
            "sku": sku or None,
            "original_sku": original_sku or None,
            "source_workbook": str(file.path),
            "source_sheet": SHEET_NAME,
            "source_row_number": str(row_number),
            "group_name": group_name,
            "factory_code": normalize_text(row_value(row, indexes["factory_code"])) or None,
            "factory_sku": normalize_text(row_value(row, indexes["factory_sku"])) or None,
            "style_code": normalize_text(row_value(row, indexes["style_code"])) or None,
            "goods_id": normalize_text(row_value(row, indexes["goods_id"])) or None,
            "p_spu": normalize_text(row_value(row, indexes["p_spu"])) or None,
            "execution_standard": normalize_text(row_value(row, indexes["execution_standard"])) or None,
            "goods_tag": normalize_text(row_value(row, indexes["goods_tag"])) or None,
            "insole_material": normalize_text(row_value(row, indexes["insole_material"])) or None,
            "outsole_material": normalize_text(row_value(row, indexes["outsole_material"])) or None,
            "lining_material": normalize_text(row_value(row, indexes["lining_material"])) or None,
            "upper_material": normalize_text(row_value(row, indexes["upper_material"])) or None,
            "shoe_box_spec": normalize_text(row_value(row, indexes["shoe_box_spec"])) or None,
            "cost": normalize_text(row_value(row, indexes["cost"])) or None,
            "product_name": normalize_text(row_value(row, indexes["product_name"])) or None,
            "category_l3": normalize_text(row_value(row, indexes["category_l3"])) or None,
            "season_category": normalize_text(row_value(row, indexes["season_category"])) or None,
            "year": normalize_text(row_value(row, indexes["year"])) or None,
            "main_style": normalize_text(row_value(row, indexes["main_style"])) or None,
            "first_order_time": normalize_text(row_value(row, indexes["first_order_time"])) or None,
            "goods_status": normalize_text(row_value(row, indexes["goods_status"])) or None,
            "sales_tag": normalize_text(row_value(row, indexes["sales_tag"])) or None,
            "latest_purchase_price": to_float(row_value(row, indexes["cost"])),
            "final_price": to_float(row_value(row, indexes["final_price"])),
            "vip_price": to_float(row_value(row, indexes["vip_price"])),
            "market_price": to_float(row_value(row, indexes["market_price"])),
            "discount_rate": to_float(row_value(row, indexes["discount_rate"])),
            "activity_profit": to_float(row_value(row, indexes["activity_profit"])),
            "margin_rate": to_float(row_value(row, indexes["margin_rate"])),
            "vip_daily_average_sales": to_float(row_value(row, indexes["vip_daily_average_sales"])) or 0,
            "other_3d_sales": to_int(row_value(row, indexes["original_other_3d_sales"])),
            "other_7d_sales": to_int(row_value(row, indexes["original_other_7d_sales"])),
            "other_15d_sales": to_int(row_value(row, indexes["original_other_15d_sales"])),
            "other_30d_sales": to_int(row_value(row, indexes["original_other_30d_sales"])),
            "original_other_3d_sales": to_int(row_value(row, indexes["original_other_3d_sales"])),
            "original_other_7d_sales": to_int(row_value(row, indexes["original_other_7d_sales"])),
            "original_other_15d_sales": to_int(row_value(row, indexes["original_other_15d_sales"])),
            "original_other_30d_sales": to_int(row_value(row, indexes["original_other_30d_sales"])),
            "stock_qty": to_int(row_value(row, indexes["stock_qty"])),
            "original_defect_stock": to_int(row_value(row, indexes["original_defect_stock"])),
            "original_inbound_qty": to_int(row_value(row, indexes["original_inbound_qty"])),
            "original_order_in_transit_stock": to_int(row_value(row, indexes["original_order_in_transit_stock"])),
            "original_defect_in_transit_stock": to_int(row_value(row, indexes["original_defect_in_transit_stock"])),
            "defect_stock": to_int(row_value(row, indexes["defect_stock"])),
            "inbound_qty": to_int(row_value(row, indexes["inbound_qty"])),
            "order_occupy_stock": to_int(row_value(row, indexes["order_occupy_stock"])),
            "defect_in_transit_stock": to_int(row_value(row, indexes["defect_in_transit_stock"])),
            "off_shelf_stock": to_int(row_value(row, indexes["off_shelf_stock"])),
        }
    )
    payload["status_key"] = status_key(payload["goods_status"])
    payload["projected_15d_stock"] = payload["stock_qty"] + payload["inbound_qty"]
    payload["size_stock"] = {
        label: to_int(row_value(row, indexes[f"size_{label}"]))
        for label in SIZE_LABELS
    }
    if include_extra_fields:
        payload["extra_fields"] = {
            header or f"column_{index + 1}": to_json_value(row_value(row, index))
            for index, header in enumerate(headers)
        }
    return brand, payload


def read_file_payloads(
    file: FineTableFile,
    *,
    include_extra_fields: bool,
) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(file.path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            return {}
        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            return {}
        headers = [normalize_header(value) for value in raw_headers]
        indexes = build_indexes(headers)
        by_brand: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row_number, row in enumerate(rows, start=2):
            if row_number == 2:
                continue
            result = build_payload(
                file=file,
                row=row,
                row_number=row_number,
                indexes=indexes,
                include_extra_fields=include_extra_fields,
                headers=headers,
            )
            if result is None:
                continue
            brand, payload = result
            by_brand[brand].append(payload)
        return by_brand
    finally:
        workbook.close()


def existing_batch_ids(connection, brand: str, snapshot_date: date) -> list[int]:
    return [
        int(row["id"])
        for row in connection.execute(
            select(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.id)
            .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.brand == brand)
            .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.snapshot_date == snapshot_date)
        ).mappings()
    ]


def expected_file_brands(file: FineTableFile) -> list[str]:
    if file.brand is not None:
        return [file.brand]
    return [BRAND_MENS, BRAND_WOMENS]


def file_snapshots_exist(repository: InventoryRepository, file: FineTableFile) -> bool:
    with repository.engine.connect() as connection:
        return all(
            bool(existing_batch_ids(connection, brand, file.snapshot_date))
            for brand in expected_file_brands(file)
        )


def write_snapshot(
    repository: InventoryRepository,
    *,
    brand: str,
    snapshot_date: date,
    payloads: list[dict[str, Any]],
    replace: bool,
) -> str:
    if not payloads:
        return "empty"
    snapshot_row_table = ensure_fine_table_snapshot_row_table(repository.engine, snapshot_date)
    with repository.engine.begin() as connection:
        existing_ids = existing_batch_ids(connection, brand, snapshot_date)
        if existing_ids and not replace:
            return "skipped_existing"
        if existing_ids and replace:
            connection.execute(
                delete(snapshot_row_table)
                .where(snapshot_row_table.c.batch_id.in_(existing_ids))
            )
            connection.execute(
                delete(FINE_TABLE_SNAPSHOT_ROW_TABLE)
                .where(FINE_TABLE_SNAPSHOT_ROW_TABLE.c.batch_id.in_(existing_ids))
            )
            connection.execute(
                delete(FINE_TABLE_SNAPSHOT_BATCH_TABLE)
                .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.id.in_(existing_ids))
            )
        batch = connection.execute(
            insert(FINE_TABLE_SNAPSHOT_BATCH_TABLE)
            .values(
                brand=brand,
                snapshot_date=snapshot_date,
                total_rows=len(payloads),
                latest_order_date=snapshot_date,
            )
            .returning(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.id)
        ).mappings().one()
        batch_id = int(batch["id"])
        rows = [
            {
                "batch_id": batch_id,
                "sku": str(payload.get("sku") or "").strip() or None,
                "original_sku": str(payload.get("original_sku") or "").strip() or None,
                "row_index": index,
                "payload": payload,
            }
            for index, payload in enumerate(payloads, start=1)
        ]
        for start in range(0, len(rows), ROW_CHUNK_SIZE):
            connection.execute(insert(snapshot_row_table), rows[start:start + ROW_CHUNK_SIZE])
        connection.execute(
            update(FINE_TABLE_SNAPSHOT_BATCH_TABLE)
            .where(FINE_TABLE_SNAPSHOT_BATCH_TABLE.c.id == batch_id)
            .values(total_rows=len(rows))
        )
    return "imported"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", action="append", default=None, help="历史精细表共享目录，可传多次；默认扫描千百度、烟斗、伊伴")
    parser.add_argument("--dry-run", action="store_true", help="只扫描文件，不写入数据库")
    parser.add_argument("--replace", action="store_true", help="覆盖已存在的同品牌同日期快照")
    parser.add_argument("--limit", type=int, default=0, help="限制导入文件数量，用于试跑")
    parser.add_argument("--include-extra-fields", action="store_true", help="把整行 Excel 原始字段也写入 extra_fields")
    args = parser.parse_args()

    roots = [Path(value) for value in args.root] if args.root else list(DEFAULT_ROOTS)
    files = choose_files(scan_roots(roots))
    if args.limit > 0:
        files = files[: args.limit]
    log(f"selected_files={len(files)}")

    if args.dry_run:
        by_scope_brand: dict[str, int] = defaultdict(int)
        for item in files:
            by_scope_brand[f"{item.scope}:{item.brand or 'combined'}"] += 1
        for key in sorted(by_scope_brand):
            log(f"{key}={by_scope_brand[key]}")
        return

    settings = load_settings()
    repository = InventoryRepository(settings.database_url)
    FINE_TABLE_SNAPSHOT_BATCH_TABLE.create(repository.engine, checkfirst=True)

    summary: dict[str, int] = defaultdict(int)
    for file_index, item in enumerate(files, start=1):
        if not args.replace and file_snapshots_exist(repository, item):
            skipped_brands = ",".join(expected_file_brands(item))
            summary["skipped_existing"] += len(expected_file_brands(item))
            log(f"[{file_index}/{len(files)}] skip existing {item.snapshot_date} {skipped_brands} {item.path}")
            continue
        log(f"[{file_index}/{len(files)}] reading {item.snapshot_date} {item.brand or 'combined'} {item.path}")
        by_brand = read_file_payloads(item, include_extra_fields=args.include_extra_fields)
        for brand, payloads in by_brand.items():
            result = write_snapshot(
                repository,
                brand=brand,
                snapshot_date=item.snapshot_date,
                payloads=payloads,
                replace=args.replace,
            )
            summary[result] += 1
            log(f"  {brand} rows={len(payloads)} result={result}")

    log("summary")
    for key in sorted(summary):
        log(f"{key}={summary[key]}")


if __name__ == "__main__":
    main()
