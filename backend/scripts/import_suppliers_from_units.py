from __future__ import annotations

import argparse
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from sqlalchemy import insert, select, update

from config import load_settings
from domain.inventory_schema import SUPPLIER_TABLE
from storage.inventory_repository import InventoryRepository
from transform.rows import normalize_cell, normalize_header


DEFAULT_FILE_PREFIX = "往来单位信息数据-2026_06_04-13_55_08"
SUPPORTED_SUFFIXES = (".xlsx", ".xlsm", ".xls")
CODE_HEADERS = ("单位编号", "编号", "单位代码")
NAME_HEADERS = ("单位全名", "单位名称", "全名")


def _desktop_candidates() -> list[Path]:
    candidates = [Path.home() / "Desktop", Path.home() / "桌面"]
    admin_home = Path("C:/Users/Administrator")
    candidates.extend([admin_home / "Desktop", admin_home / "桌面"])
    return candidates


def resolve_source(path: str | None) -> Path:
    if path:
        source = Path(path)
        if source.is_file():
            return source
        for suffix in SUPPORTED_SUFFIXES:
            candidate = Path(f"{source}{suffix}")
            if candidate.is_file():
                return candidate
        raise FileNotFoundError(f"未找到文件: {path}")

    for desktop in _desktop_candidates():
        if not desktop.exists():
            continue
        for suffix in SUPPORTED_SUFFIXES:
            candidate = desktop / f"{DEFAULT_FILE_PREFIX}{suffix}"
            if candidate.is_file():
                return candidate
        matches = [
            item
            for item in desktop.iterdir()
            if item.is_file()
            and item.suffix.lower() in SUPPORTED_SUFFIXES
            and item.stem.startswith(DEFAULT_FILE_PREFIX)
        ]
        if matches:
            return max(matches, key=lambda item: item.stat().st_mtime)
    raise FileNotFoundError(f"桌面未找到 {DEFAULT_FILE_PREFIX}")


def _find_header_indexes(headers: list[str]) -> tuple[int, int] | None:
    code_index = next((index for index, value in enumerate(headers) if value in CODE_HEADERS), None)
    name_index = next((index for index, value in enumerate(headers) if value in NAME_HEADERS), None)
    if code_index is None or name_index is None:
        return None
    return code_index, name_index


def _get_value(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def _normalize_record(code: object, name: object) -> dict[str, str] | None:
    factory_code = normalize_cell(code)
    supplier_name = normalize_cell(name)
    if not factory_code or not supplier_name:
        return None
    return {"factory_code": str(factory_code), "name": str(supplier_name)}


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        header_indexes: tuple[int, int] | None = None
        for row_number, row in enumerate(rows, start=1):
            headers = [normalize_header(value) for value in row]
            header_indexes = _find_header_indexes(headers)
            if header_indexes is not None:
                break
            if row_number >= 30:
                return []
        if header_indexes is None:
            return []

        code_index, name_index = header_indexes
        records: list[dict[str, str]] = []
        for row in rows:
            record = _normalize_record(_get_value(row, code_index), _get_value(row, name_index))
            if record is not None:
                records.append(record)
        return records
    finally:
        workbook.close()


def _read_xls(path: Path) -> list[dict[str, str]]:
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_by_index(0)
    header_indexes: tuple[int, int] | None = None
    data_start_row = 0
    for row_index in range(min(worksheet.nrows, 30)):
        headers = [normalize_header(value) for value in worksheet.row_values(row_index)]
        header_indexes = _find_header_indexes(headers)
        if header_indexes is not None:
            data_start_row = row_index + 1
            break
    if header_indexes is None:
        return []

    code_index, name_index = header_indexes
    records: list[dict[str, str]] = []
    for row_index in range(data_start_row, worksheet.nrows):
        row = tuple(worksheet.row_values(row_index))
        record = _normalize_record(_get_value(row, code_index), _get_value(row, name_index))
        if record is not None:
            records.append(record)
    return records


def read_unit_supplier_records(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".xls":
        records = _read_xls(path)
    else:
        records = _read_xlsx(path)

    deduped: dict[str, dict[str, str]] = {}
    for record in records:
        deduped[record["factory_code"]] = record
    return list(deduped.values())


def import_suppliers(path: Path) -> dict[str, int]:
    records = read_unit_supplier_records(path)
    settings = load_settings()
    repository = InventoryRepository(settings.database_url)

    inserted = 0
    updated = 0
    skipped = 0
    with repository.engine.begin() as connection:
        existing_rows = [dict(row) for row in connection.execute(select(SUPPLIER_TABLE)).mappings()]
        by_code = {
            str(row["factory_code"]): row
            for row in existing_rows
            if row.get("factory_code")
        }
        by_name = {str(row["name"]): row for row in existing_rows if row.get("name")}

        for record in records:
            existing = by_code.get(record["factory_code"]) or by_name.get(record["name"])
            if existing is None:
                connection.execute(insert(SUPPLIER_TABLE).values(**record))
                inserted += 1
                continue

            payload = {
                key: value
                for key, value in record.items()
                if existing.get(key) != value
            }
            if payload:
                connection.execute(
                    update(SUPPLIER_TABLE)
                    .where(SUPPLIER_TABLE.c.id == existing["id"])
                    .values(**payload)
                )
                updated += 1
            else:
                skipped += 1

    return {"read": len(records), "inserted": inserted, "updated": updated, "skipped": skipped}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("path", nargs="?", help="往来单位信息数据 Excel 路径")
    parser.add_argument("--inspect", action="store_true", help="打印工作簿前几行用于识别表头")
    args = parser.parse_args()
    source = resolve_source(args.path)
    if args.inspect:
        if source.suffix.lower() == ".xls":
            workbook = xlrd.open_workbook(source)
            print(f"source={source}")
            print(f"sheets={workbook.sheet_names()}")
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                print(f"sheet[{sheet_index}] name={sheet.name!r} rows={sheet.nrows} cols={sheet.ncols}")
            worksheet = workbook.sheet_by_index(0)
            for row_index in range(min(15, worksheet.nrows)):
                values = [
                    str(value).encode("unicode_escape").decode("ascii")
                    for value in worksheet.row_values(row_index)[:20]
                ]
                print(row_index, values)
        else:
            workbook = load_workbook(source, data_only=True, read_only=True)
            try:
                print(f"source={source}")
                print(f"sheets={workbook.sheetnames}")
                worksheet = workbook[workbook.sheetnames[0]]
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    values = [
                        str(value).encode("unicode_escape").decode("ascii")
                        for value in row[:20]
                    ]
                    print(row_index, values)
                    if row_index >= 15:
                        break
            finally:
                workbook.close()
        return
    result = import_suppliers(source)
    print(f"source={source}")
    print(
        f"read={result['read']}, inserted={result['inserted']}, "
        f"updated={result['updated']}, skipped={result['skipped']}"
    )


if __name__ == "__main__":
    main()
