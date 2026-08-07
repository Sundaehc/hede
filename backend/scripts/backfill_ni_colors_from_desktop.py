"""Backfill NI product colors and color codes from desktop workbooks.

This is a manual import. It does not create a scheduled task and only updates
the NI product archive plus the NI color-code lookup used by product editing.
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.color_barcode_schema import COLOR_BARCODE_TABLE
from domain.schema import NI_PRODUCT_TABLE


DEFAULT_PRODUCT_FILE = Path.home() / "Desktop" / "NI货号对应的颜色8.7.xlsx"
DEFAULT_COLOR_CODE_FILE = Path.home() / "Desktop" / "NI颜色名对应的颜色代码8.7.xls"


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_product_colors(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [_text(value).replace("\n", "").replace("\r", "") for value in next(rows, ())]
        sku_index = next((index for index, value in enumerate(headers) if value in {"款式编码", "货号", "商品编码"}), None)
        color_index = next((index for index, value in enumerate(headers) if value in {"颜色名称", "颜色"}), None)
        if sku_index is None or color_index is None:
            raise ValueError(f"未找到款式编码/颜色名称表头：{path}")

        result: dict[str, str] = {}
        for row in rows:
            sku = _text(row[sku_index] if sku_index < len(row) else None)
            color = _text(row[color_index] if color_index < len(row) else None)
            if not sku or not color:
                continue
            previous = result.get(sku)
            if previous and previous != color:
                raise ValueError(f"货号对应多个颜色：{sku} -> {previous} / {color}")
            result[sku] = color
        return result
    finally:
        workbook.close()


def _read_xls_with_excel(path: Path) -> list[dict[str, str]]:
    if __import__("platform").system() != "Windows":
        raise ValueError(f"无法读取旧版 Excel 文件：{path}")

    powershell = r"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$workbook = $null
try {
  $workbook = $excel.Workbooks.Open($env:NI_COLOR_CODE_FILE, 0, $true)
  $worksheet = $workbook.Worksheets.Item(1)
  $range = $worksheet.UsedRange
  $result = @()
  for ($row = 1; $row -le $range.Rows.Count; $row++) {
    $colorCode = [string]$worksheet.Cells.Item($row, 2).Text
    $colorName = [string]$worksheet.Cells.Item($row, 3).Text
    if ($colorCode.Trim() -and $colorName.Trim() -and $colorCode.Trim() -ne '颜色条码' -and $colorName.Trim() -ne '颜色名称') {
      $result += [pscustomobject]@{
        color_code = $colorCode.Trim()
        color_name = $colorName.Trim()
        row_number = [string]$row
      }
    }
  }
  ConvertTo-Json -InputObject @($result) -Compress
}
finally {
  if ($workbook) { $workbook.Close($false) }
  $excel.Quit()
  [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
}
"""
    environment = os.environ.copy()
    environment["NI_COLOR_CODE_FILE"] = str(path)
    completed = subprocess.run(
        ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", powershell],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        env=environment,
    )
    payload = json.loads(completed.stdout.strip() or "[]")
    return [
        {
            "color_code": _text(item.get("color_code")),
            "color_name": _text(item.get("color_name")),
            "row_number": _text(item.get("row_number")),
        }
        for item in payload
        if _text(item.get("color_code")) and _text(item.get("color_name"))
    ]


def _read_color_codes(path: Path) -> list[dict[str, str]]:
    try:
        import xlrd

        workbook = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheet = workbook.sheet_by_index(0)
            if sheet.nrows:
                rows = [sheet.row_values(index) for index in range(sheet.nrows)]
                header_index = next((index for index, row in enumerate(rows[:30]) if "颜色条码" in row and "颜色名称" in row), None)
                if header_index is not None:
                    code_index = rows[header_index].index("颜色条码")
                    name_index = rows[header_index].index("颜色名称")
                    return [
                        {"color_code": _text(row[code_index]), "color_name": _text(row[name_index]), "row_number": str(index + 1)}
                        for index, row in enumerate(rows[header_index + 1 :], start=header_index + 1)
                        if _text(row[code_index]) and _text(row[name_index])
                    ]
        finally:
            release = getattr(workbook, "release_resources", None)
            if release:
                release()
    except Exception:
        pass
    return _read_xls_with_excel(path)


def _dedupe_color_codes(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    by_name: dict[str, dict[str, str]] = {}
    by_code: dict[str, str] = {}
    for row in rows:
        code = row["color_code"]
        name = row["color_name"]
        if name in by_name and by_name[name]["color_code"] != code:
            raise ValueError(f"颜色名称对应多个颜色代码：{name}")
        if code in by_code and by_code[code] != name:
            raise ValueError(f"颜色代码对应多个颜色名称：{code}")
        by_name[name] = row
        by_code[code] = name
    return by_name


def run(product_file: Path, color_code_file: Path, *, apply: bool) -> None:
    product_colors = _read_product_colors(product_file)
    color_code_rows = _read_color_codes(color_code_file)
    color_codes = _dedupe_color_codes(color_code_rows)
    missing_codes = sorted({color for color in product_colors.values() if color not in color_codes})

    engine = create_engine(load_settings().database_url)
    with engine.connect() as connection:
        products = list(connection.execute(
            select(NI_PRODUCT_TABLE.c.id, NI_PRODUCT_TABLE.c.sku, NI_PRODUCT_TABLE.c.color, NI_PRODUCT_TABLE.c.color_code)
        ).mappings())

    updates: list[dict[str, object]] = []
    unmatched_products: list[str] = []
    for row in products:
        sku = _text(row["sku"])
        color = product_colors.get(sku)
        if not color:
            unmatched_products.append(sku)
            continue
        updates.append({
            "id": row["id"],
            "color": color,
            "color_code": color_codes.get(color, {}).get("color_code", ""),
        })

    print(f"货号颜色表：{len(product_colors)} 条")
    print(f"颜色代码表：{len(color_codes)} 条")
    print(f"NI 商品总数：{len(products)} 条")
    print(f"可更新：{len(updates)} 条")
    print(f"未匹配货号：{len(unmatched_products)} 条")
    print(f"未匹配颜色代码：{', '.join(missing_codes) if missing_codes else '无'}")
    if not apply:
        print("预览模式，未写入数据库。")
        return

    source_workbook = str(color_code_file)
    with engine.begin() as connection:
        for item in updates:
            connection.execute(
                update(NI_PRODUCT_TABLE)
                .where(NI_PRODUCT_TABLE.c.id == item["id"])
                .values(color=item["color"], color_code=item["color_code"], updated_at=func.date_trunc("minute", func.now()))
            )

        barcode_rows = [
            {
                "brand": "ni",
                "color_barcode": row["color_code"],
                "color_name": row["color_name"],
                "source_workbook": source_workbook,
                "source_sheet": "导出数据",
                "source_row_number": row["row_number"],
                "raw_payload": row,
            }
            for row in color_code_rows
        ]
        statement = pg_insert(COLOR_BARCODE_TABLE).values(barcode_rows)
        statement = statement.on_conflict_do_update(
            index_elements=["brand", "color_barcode"],
            set_={
                "color_name": statement.excluded.color_name,
                "source_workbook": statement.excluded.source_workbook,
                "source_sheet": statement.excluded.source_sheet,
                "source_row_number": statement.excluded.source_row_number,
                "raw_payload": statement.excluded.raw_payload,
                "updated_at": func.date_trunc("minute", func.now()),
            },
        )
        connection.execute(statement)
    print(f"已更新 NI 商品：{len(updates)} 条；已保存颜色代码映射：{len(color_code_rows)} 条")


def main() -> int:
    parser = argparse.ArgumentParser(description="从桌面 Excel 回填 NI 商品颜色和颜色代码")
    parser.add_argument("--product-file", type=Path, default=DEFAULT_PRODUCT_FILE)
    parser.add_argument("--color-code-file", type=Path, default=DEFAULT_COLOR_CODE_FILE)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    run(args.product_file, args.color_code_file, apply=args.apply)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
