from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path
from typing import Any

import orjson
from openpyxl import load_workbook
from sqlalchemy import create_engine, func
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.fields import GJ_MERGED_PRODUCT_INFO_FIELDS, alias_map
from domain.gj_schema import GJ_MERGED_PRODUCT_INFO_TABLE
from storage.date_normalization import parse_date
from transform.rows import normalize_cell, normalize_header


SOURCE_FILE_KEYWORD = "男女鞋合并商品信息"
HEADER_ROW = 4


def _json_serializer(value: object) -> str:
    return orjson.dumps(value).decode("utf-8")


def _to_int(value: object) -> int | None:
    normalized = normalize_cell(value)
    if normalized is None:
        return None
    try:
        return int(float(str(normalized).replace(",", "")))
    except (TypeError, ValueError):
        return None


def _source_file(source_dir: Path) -> Path:
    candidates = [
        path
        for path in sorted(source_dir.glob("*"))
        if SOURCE_FILE_KEYWORD in path.name
        and not path.name.startswith("~$")
        and path.suffix.lower() in {".xlsx", ".xlsm"}
    ]
    if not candidates:
        raise FileNotFoundError(f"未找到 {SOURCE_FILE_KEYWORD}: {source_dir}")
    return candidates[0]


def _read_rows(file_path: Path, source_date: str) -> list[dict[str, Any]]:
    column_aliases = alias_map(GJ_MERGED_PRODUCT_INFO_FIELDS)
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    assert ws is not None
    sheet_title = ws.title

    header_values = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True), None)
    if header_values is None:
        wb.close()
        return []

    headers = [normalize_header(value) for value in header_values]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1):
        raw_payload = {
            headers[index]: normalize_cell(values[index])
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        if not any(value is not None for value in raw_payload.values()):
            continue

        record: dict[str, Any] = {
            "source_date": source_date,
            "source_date_value": parse_date(source_date),
            "source_workbook": file_path.stem,
            "source_sheet": sheet_title,
            "source_row_number": str(row_number),
            "raw_payload": raw_payload,
        }
        extra_fields: dict[str, object] = {}
        for header, value in raw_payload.items():
            target = column_aliases.get(header)
            if target is None:
                if value is not None:
                    extra_fields[header] = value
                continue
            record[target] = _to_int(value) if target == "row_no" else normalize_cell(value)

        goods_code = str(record.get("goods_code") or "").strip()
        if not goods_code:
            continue
        record["goods_code"] = goods_code
        record["extra_fields"] = extra_fields or None
        rows.append(record)

    wb.close()
    return rows


def import_gj_merged_product_info(database_url: str, source_dir: Path) -> dict[str, object]:
    file_path = _source_file(source_dir)
    source_date = source_dir.name
    rows = _read_rows(file_path, source_date)

    deduped_by_goods_code: dict[str, dict[str, Any]] = {}
    for row in rows:
        deduped_by_goods_code[str(row["goods_code"])] = row
    payload = list(deduped_by_goods_code.values())

    engine = create_engine(database_url, future=True, json_serializer=_json_serializer)
    update_columns = [
        column.name
        for column in GJ_MERGED_PRODUCT_INFO_TABLE.columns
        if column.name not in ("id", "source_date", "goods_code", "created_at")
    ]

    with engine.begin() as conn:
        for index in range(0, len(payload), 1000):
            chunk = payload[index:index + 1000]
            if not chunk:
                continue
            stmt = pg_insert(GJ_MERGED_PRODUCT_INFO_TABLE).values(chunk)
            excluded = stmt.excluded
            set_values = {column: getattr(excluded, column) for column in update_columns}
            set_values["updated_at"] = func.date_trunc("minute", func.now())
            stmt = stmt.on_conflict_do_update(
                index_elements=["source_date", "goods_code"],
                set_=set_values,
            )
            conn.execute(stmt)

    return {
        "source_dir": str(source_dir),
        "source_file": file_path.name,
        "source_date": source_date,
        "read_rows": len(rows),
        "imported": len(payload),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="导入管家婆男女鞋合并商品信息")
    parser.add_argument("--source-dir", type=Path, default=None)
    args = parser.parse_args()

    settings = load_settings(require_database=True)
    assert settings.database_url is not None
    assert settings.jst_price_root is not None, "JST_PRICE_ROOT is required in .env"
    source_dir = args.source_dir or settings.jst_price_root / date.today().strftime("%Y-%m-%d")
    result = import_gj_merged_product_info(settings.database_url, source_dir)
    print(result)


if __name__ == "__main__":
    main()
