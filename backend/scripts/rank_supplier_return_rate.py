from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from config import load_settings


DEFAULT_AFTERSALE_FILE = Path(r"\\192.168.10.229\运营组资料\影刀\商品库存\售后（退货退款）.xlsx")
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parents[2] / "outputs"


@dataclass(frozen=True)
class ReturnStat:
    supplier: str
    returned_qty: int
    sold_qty: int
    return_rate: float
    sku_count: int
    unmatched_returned_qty: int = 0


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_int(value: object) -> int:
    text_value = normalize_text(value).replace(",", "")
    if not text_value:
        return 0
    try:
        return int(float(text_value))
    except ValueError:
        return 0


def parse_excel_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text_value = normalize_text(value)
    if not text_value:
        return None
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue
    return None


def header_indexes(headers: tuple[object, ...]) -> dict[str, int]:
    return {
        normalize_text(header): index
        for index, header in enumerate(headers)
        if normalize_text(header)
    }


def read_aftersale_returns(
    file_path: Path,
    *,
    start_date: date | None,
    end_date: date | None,
) -> tuple[dict[str, int], dict[str, object]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None:
        wb.close()
        return {}, {"read_rows": 0, "used_rows": 0}

    indexes = header_indexes(headers)
    required_headers = ("原始货号", "实退数量", "订单日期")
    missing = [name for name in required_headers if name not in indexes]
    if missing:
        wb.close()
        raise ValueError(f"售后表缺少字段: {', '.join(missing)}")

    returns_by_original_code: dict[str, int] = defaultdict(int)
    read_rows = 0
    used_rows = 0
    skipped_out_of_range = 0
    for row in rows:
        read_rows += 1
        order_date = parse_excel_date(row[indexes["订单日期"]])
        if start_date is not None and (order_date is None or order_date < start_date):
            skipped_out_of_range += 1
            continue
        if end_date is not None and (order_date is None or order_date > end_date):
            skipped_out_of_range += 1
            continue

        original_code = normalize_text(row[indexes["原始货号"]])
        returned_qty = to_int(row[indexes["实退数量"]])
        if not original_code or returned_qty <= 0:
            continue
        returns_by_original_code[original_code] += returned_qty
        used_rows += 1

    wb.close()
    return dict(returns_by_original_code), {
        "read_rows": read_rows,
        "used_rows": used_rows,
        "skipped_out_of_range": skipped_out_of_range,
        "source_sheet": ws.title,
    }


def fetch_order_date_range(database_url: str) -> tuple[date | None, date | None]:
    engine = create_engine(database_url, future=True)
    with engine.connect() as conn:
        row = conn.execute(
            text("select min(order_time_at)::date, max(order_time_at)::date from jst_monthly_orders")
        ).one()
    return row[0], row[1]


def fetch_latest_supplier_map(database_url: str) -> tuple[str, dict[str, str]]:
    engine = create_engine(database_url, future=True)
    with engine.connect() as conn:
        source_date = conn.execute(
            text("select max(source_date) from gj_merged_product_info")
        ).scalar_one()
        if not source_date:
            return "", {}
        rows = conn.execute(
            text(
                """
                select distinct on (original_goods_code)
                    original_goods_code,
                    primary_supplier
                from gj_merged_product_info
                where source_date = :source_date
                  and coalesce(original_goods_code, '') <> ''
                  and coalesce(primary_supplier, '') <> ''
                order by original_goods_code, id desc
                """
            ),
            {"source_date": source_date},
        )
        supplier_by_original_code = {
            str(row.original_goods_code).strip(): str(row.primary_supplier).strip()
            for row in rows
            if str(row.original_goods_code or "").strip()
        }
    return str(source_date), supplier_by_original_code


def fetch_sales_by_supplier(
    database_url: str,
    supplier_by_original_code: dict[str, str],
    *,
    start_date: date | None,
    end_date: date | None,
    include_cancelled: bool,
) -> tuple[dict[str, int], dict[str, int]]:
    conditions = ["coalesce(style_code, '') <> ''"]
    params: dict[str, object] = {}
    if start_date is not None:
        conditions.append("order_time_at::date >= :start_date")
        params["start_date"] = start_date
    if end_date is not None:
        conditions.append("order_time_at::date <= :end_date")
        params["end_date"] = end_date
    if not include_cancelled:
        conditions.append("coalesce(status, '') <> '取消'")

    query = text(
        f"""
        select style_code, coalesce(sum(quantity), 0)::int as sold_qty
        from jst_monthly_orders
        where {" and ".join(conditions)}
        group by style_code
        """
    )
    sales_by_supplier: dict[str, int] = defaultdict(int)
    sales_by_original_code: dict[str, int] = {}
    engine = create_engine(database_url, future=True)
    with engine.connect() as conn:
        for row in conn.execute(query, params):
            original_code = normalize_text(row.style_code)
            sold_qty = int(row.sold_qty or 0)
            if sold_qty <= 0:
                continue
            sales_by_original_code[original_code] = sold_qty
            supplier = supplier_by_original_code.get(original_code)
            if supplier:
                sales_by_supplier[supplier] += sold_qty
    return dict(sales_by_supplier), sales_by_original_code


def rank_suppliers(
    returns_by_original_code: dict[str, int],
    sales_by_supplier: dict[str, int],
    supplier_by_original_code: dict[str, str],
) -> tuple[list[ReturnStat], dict[str, object]]:
    returns_by_supplier: dict[str, int] = defaultdict(int)
    sku_codes_by_supplier: dict[str, set[str]] = defaultdict(set)
    unmatched_returned_qty = 0
    unmatched_sku_count = 0
    for original_code, returned_qty in returns_by_original_code.items():
        supplier = supplier_by_original_code.get(original_code)
        if not supplier:
            unmatched_returned_qty += returned_qty
            unmatched_sku_count += 1
            continue
        returns_by_supplier[supplier] += returned_qty
        sku_codes_by_supplier[supplier].add(original_code)

    stats: list[ReturnStat] = []
    for supplier, returned_qty in returns_by_supplier.items():
        sold_qty = sales_by_supplier.get(supplier, 0)
        return_rate = returned_qty / sold_qty if sold_qty else 0.0
        stats.append(
            ReturnStat(
                supplier=supplier,
                returned_qty=returned_qty,
                sold_qty=sold_qty,
                return_rate=return_rate,
                sku_count=len(sku_codes_by_supplier[supplier]),
            )
        )

    stats.sort(key=lambda item: (item.return_rate, item.returned_qty), reverse=True)
    return stats, {
        "unmatched_returned_qty": unmatched_returned_qty,
        "unmatched_sku_count": unmatched_sku_count,
    }


def write_csv(stats: list[ReturnStat], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["排名", "供应商", "退货率", "实退数量", "订单数量", "退货货号数"])
        for rank, item in enumerate(stats, start=1):
            writer.writerow([
                rank,
                item.supplier,
                f"{item.return_rate:.2%}" if item.sold_qty else "",
                item.returned_qty,
                item.sold_qty,
                item.sku_count,
            ])


def main() -> None:
    parser = argparse.ArgumentParser(description="按供应商统计售后退货率排行榜")
    parser.add_argument("--aftersale-file", type=Path, default=DEFAULT_AFTERSALE_FILE)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--top", type=int, default=30)
    parser.add_argument("--start-date", type=lambda value: datetime.strptime(value, "%Y-%m-%d").date(), default=None)
    parser.add_argument("--end-date", type=lambda value: datetime.strptime(value, "%Y-%m-%d").date(), default=None)
    parser.add_argument("--include-cancelled", action="store_true")
    args = parser.parse_args()

    settings = load_settings(require_database=True)
    assert settings.database_url is not None

    db_start_date, db_end_date = fetch_order_date_range(settings.database_url)
    start_date = args.start_date or db_start_date
    end_date = args.end_date or db_end_date

    source_date, supplier_by_original_code = fetch_latest_supplier_map(settings.database_url)
    returns_by_original_code, read_meta = read_aftersale_returns(
        args.aftersale_file,
        start_date=start_date,
        end_date=end_date,
    )
    sales_by_supplier, sales_by_original_code = fetch_sales_by_supplier(
        settings.database_url,
        supplier_by_original_code,
        start_date=start_date,
        end_date=end_date,
        include_cancelled=args.include_cancelled,
    )
    stats, unmatched_meta = rank_suppliers(
        returns_by_original_code,
        sales_by_supplier,
        supplier_by_original_code,
    )

    output_path = args.output or DEFAULT_OUTPUT_DIR / (
        f"supplier_return_rate_rank_{date.today().isoformat()}.csv"
    )
    write_csv(stats, output_path)

    print(f"售后文件: {args.aftersale_file}")
    print(f"统计区间: {start_date} ~ {end_date}")
    print(f"供应商映射批次: gj_merged_product_info {source_date}")
    print(f"售后读取行数: {read_meta['read_rows']}, 使用行数: {read_meta['used_rows']}, 区间外跳过: {read_meta['skipped_out_of_range']}")
    print(f"售后原始货号数: {len(returns_by_original_code)}, 订单原始货号数: {len(sales_by_original_code)}")
    print(f"未匹配供应商退货数: {unmatched_meta['unmatched_returned_qty']}, 未匹配货号数: {unmatched_meta['unmatched_sku_count']}")
    print(f"输出文件: {output_path}")
    print()
    print("排名\t供应商\t退货率\t实退数量\t订单数量\t退货货号数")
    for rank, item in enumerate(stats[: args.top], start=1):
        rate = f"{item.return_rate:.2%}" if item.sold_qty else "无销量"
        print(f"{rank}\t{item.supplier}\t{rate}\t{item.returned_qty}\t{item.sold_qty}\t{item.sku_count}")


if __name__ == "__main__":
    main()
