"""Import immutable 2024/2025 historical sales from a product-goods workbook."""
from __future__ import annotations

import argparse
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from collections.abc import Iterator

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.dialects.postgresql import insert as pg_insert

from config import load_settings
from domain.product_goods_historical_sales_schema import ensure_product_goods_historical_sales_table
from storage.date_normalization import parse_date


SHEET_YEARS = {"2024年销量": 2024, "2025年销量": 2025}
HISTORICAL_YEARS = frozenset(SHEET_YEARS.values())
HEADER_ALIASES = {
    "channel": ("渠道", "销售渠道"),
    "sales_date": ("日期", "销售日期", "月份"),
    "style_code": ("款式编码", "款号"),
    "product_code": ("商品编码", "货号"),
    "sales_quantity": ("销售数量", "销量", "销售量"),
    "sales_amount": ("销售金额", "销售额"),
    "original_sku": ("原始货号", "原始款号", "原货号"),
    "size": ("尺码", "规格"),
    "color": ("颜色", "颜色名称"),
}


def _text(value: object) -> str | None:
    value = str(value or "").strip()
    return value or None


def _integer(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(Decimal(str(value).replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def _decimal(value: object) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", ""))
    except InvalidOperation:
        return None


def _headers(values: tuple[object, ...]) -> dict[str, int]:
    normalized = [str(value or "").strip().replace("\n", "") for value in values]
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        if field == "channel":
            channel_indexes = [index for index, value in enumerate(normalized) if value == "渠道"]
            index = channel_indexes[-1] if channel_indexes else next(
                (index for index, value in enumerate(normalized) if value == "销售渠道"),
                None,
            )
        else:
            index = next(
                (index for index, value in enumerate(normalized) if value in aliases),
                None,
            )
        if index is None and field in {"channel", "sales_date", "style_code", "product_code", "sales_quantity"}:
            raise ValueError(f"未找到历史销量列: {aliases[0]}")
        if index is not None:
            result[field] = index
    return result


def _value(row: tuple[object, ...], indexes: dict[str, int], field: str) -> object:
    index = indexes.get(field)
    return row[index] if index is not None and index < len(row) else None


def iter_rows(path: Path, *, brand: str) -> Iterator[tuple[dict[str, object], str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        named_year_sheets = {
            sheet_name: sales_year
            for sheet_name, sales_year in SHEET_YEARS.items()
            if sheet_name in workbook.sheetnames
        }
        if named_year_sheets:
            source_sheets: dict[str, int | None] = named_year_sheets
        elif "销售数据" in workbook.sheetnames:
            source_sheets = {"销售数据": None}
        else:
            raise ValueError("未找到 2024年销量、2025年销量或销售数据 sheet")
        for sheet_name, expected_year in source_sheets.items():
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            header = next(iterator, None)
            if header is None:
                raise ValueError(f"sheet 为空: {sheet_name}")
            indexes = _headers(header)
            for row_number, values in enumerate(iterator, start=2):
                sales_date = parse_date(_value(values, indexes, "sales_date"))
                sales_quantity = _integer(_value(values, indexes, "sales_quantity"))
                if sales_date is None or sales_quantity is None:
                    continue
                if sales_date.year not in HISTORICAL_YEARS:
                    continue
                if expected_year is not None and sales_date.year != expected_year:
                    raise ValueError(f"{sheet_name} 第 {row_number} 行日期不属于 {expected_year} 年")
                yield {
                    "brand": brand,
                    "sales_year": sales_date.year,
                    "sales_date": sales_date,
                    "channel": _text(_value(values, indexes, "channel")),
                    "style_code": _text(_value(values, indexes, "style_code")),
                    "product_code": _text(_value(values, indexes, "product_code")),
                    "original_sku": _text(_value(values, indexes, "original_sku")),
                    "size": _text(_value(values, indexes, "size")),
                    "color": _text(_value(values, indexes, "color")),
                    "sales_quantity": sales_quantity,
                    "sales_amount": _decimal(_value(values, indexes, "sales_amount")),
                    "source_workbook": path.name,
                    "source_sheet": sheet_name,
                    "source_row_number": row_number,
                }, sheet_name
    finally:
        workbook.close()


def import_rows(path: Path, *, brand: str) -> dict[str, object]:
    settings = load_settings(require_database=True)
    assert settings.database_url is not None
    engine = create_engine(settings.database_url, future=True)
    tables = {sales_year: ensure_product_goods_historical_sales_table(engine, sales_year) for sales_year in HISTORICAL_YEARS}
    inserted = 0
    read = 0
    sheet_counts: dict[str, int] = defaultdict(int)

    def insert_chunk(chunk: list[dict[str, object]]) -> int:
        if not chunk:
            return 0
        sales_year = int(chunk[0]["sales_year"])
        if any(int(row["sales_year"]) != sales_year for row in chunk):
            raise ValueError("历史销量批次不能混合不同年份")
        table = tables[sales_year]
        with engine.begin() as connection:
            statement = (
                pg_insert(table)
                .values(chunk)
                .on_conflict_do_nothing(index_elements=["source_workbook", "source_sheet", "source_row_number"])
                .returning(table.c.id)
            )
            return len(connection.execute(statement).all())

    chunks_by_year: dict[int, list[dict[str, object]]] = defaultdict(list)
    for row, sheet_name in iter_rows(path, brand=brand):
        read += 1
        sheet_counts[sheet_name] += 1
        sales_year = int(row["sales_year"])
        chunk = chunks_by_year[sales_year]
        chunk.append(row)
        if len(chunk) < 1000:
            continue
        inserted += insert_chunk(chunk)
        if read % 50_000 == 0:
            print(f"[{brand}] processed={read} inserted={inserted}", flush=True)
        chunks_by_year[sales_year] = []
    for chunk in chunks_by_year.values():
        inserted += insert_chunk(chunk)
    return {"source": str(path), "read": read, "inserted": inserted, "skipped_existing": read - inserted, "sheets": dict(sheet_counts)}


def main() -> None:
    parser = argparse.ArgumentParser(description="Import immutable 2024/2025 historical sales for product goods")
    parser.add_argument("path", type=Path)
    parser.add_argument("--brand", required=True)
    args = parser.parse_args()
    print(import_rows(args.path, brand=args.brand))


if __name__ == "__main__":
    main()
