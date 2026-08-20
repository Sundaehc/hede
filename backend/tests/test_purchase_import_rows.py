from __future__ import annotations

import io
from types import SimpleNamespace

from openpyxl import Workbook

from api.routes.inventory import (
    _build_purchase_details_from_rows,
    _build_purchase_order_import_template,
    _group_purchase_import_rows_by_summary,
    _missing_purchase_order_import_fields,
    _purchase_order_import_has_size_columns,
    _purchase_import_brand_for_supplier,
    _purchase_import_brand_for_record,
    _read_purchase_import_rows,
    _split_purchase_product_code,
    _split_purchase_size_code,
    PURCHASE_SUMMARY_EXPORT_HEADERS,
    PURCHASE_SIZE_ROW_EXPORT_HEADERS,
)
from api.routes import inventory as inventory_routes


def test_purchase_size_export_uses_detail_color_barcode_when_archive_color_code_is_missing() -> None:
    product_code, size_barcode = inventory_routes._purchase_size_export_product_code(
        "EE563366D20",
        "20",
        "225",
        "eblan",
        {
            "sku": "EE563366D20",
            "original_sku": "EE563366D20",
            "color_code": None,
            "barcode_build_rule": "货号+颜色代码+尺码",
            "size_barcodes": {"225": "225"},
        },
    )

    assert product_code == "EE563366D2020225"
    assert size_barcode == "225"


def _sample_purchase_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["供应商", "商品编码", "数量", "采购单备注", "采购日期", "协议到货日期", "收货仓库", "经办人", "附加说明"])
    worksheet.append([
        "友宝保罗（千百度）",
        "C5563406D8080240",
        6,
        "26.06.29友宝保罗（千百度）新款下单160双 未打",
        "2026/6/29",
        "2026/7/15",
        "赫德仙岩仓",
        "陈希华",
        "工厂需确认交期",
    ])
    worksheet.append([
        "友宝保罗（千百度）",
        "C5563406D8080245",
        10,
        "26.06.29友宝保罗（千百度）新款下单160双 未打",
        "2026/6/29",
        "2026/7/15",
        "赫德仙岩仓",
        "陈希华",
        "",
    ])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _legacy_size_column_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["商品编码", "220", "225"])
    worksheet.append(["C5563406D8080", 6, 10])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _combined_size_column_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["商品编码", "235-240"])
    worksheet.append(["C5563406D80", 12])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class _StubConnection:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        return False


class _StubEngine:
    def connect(self):
        return _StubConnection()


class _StubRepository:
    engine = _StubEngine()


class _WarehouseBrandRepository:
    def get_warehouse_by_name(self, name: str):
        return {"brand": "NI仓库"} if name == "NI仙岩仓库" else None


def _legacy_single_document_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["商品编码", "数量"])
    worksheet.append(["C5563406D8080240", 6])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_purchase_import_rows_include_document_fields_and_group_by_summary() -> None:
    rows, sheet_name = _read_purchase_import_rows(_sample_purchase_workbook())
    groups = _group_purchase_import_rows_by_summary(rows, "")

    assert sheet_name == "Sheet"
    assert len(rows) == 2
    assert rows[0]["product_code"] == "C5563406D8080240"
    assert rows[0]["quantity"] == "6"
    assert rows[0]["date"] == "2026-06-29"
    assert rows[0]["delivery_date"] == "2026-07-15"
    assert rows[0]["supplier"] == "友宝保罗（千百度）"
    assert rows[0]["warehouse"] == "赫德仙岩仓"
    assert rows[0]["handler"] == "陈希华"
    assert rows[0]["summary"] == "26.06.29友宝保罗（千百度）新款下单160双 未打"
    assert rows[0]["remark"] == ""
    assert rows[0]["additional_note"] == "工厂需确认交期"

    assert len(groups) == 1
    assert groups[0]["fields"]["summary"] == "26.06.29友宝保罗（千百度）新款下单160双 未打"
    assert groups[0]["fields"]["additional_note"] == "工厂需确认交期"
    assert len(groups[0]["rows"]) == 2


def test_purchase_import_keeps_same_summary_separate_when_warehouse_differs() -> None:
    rows, _ = _read_purchase_import_rows(_sample_purchase_workbook())
    rows[1]["warehouse"] = "赫德公司仓"

    groups = _group_purchase_import_rows_by_summary(rows, "")

    assert len(groups) == 2


def test_purchase_import_keeps_same_summary_separate_when_date_differs() -> None:
    rows, _ = _read_purchase_import_rows(_sample_purchase_workbook())
    rows[1]["date"] = "2026-06-30"

    groups = _group_purchase_import_rows_by_summary(rows, "")

    assert len(groups) == 2


def test_purchase_order_import_rejects_legacy_single_document_template() -> None:
    rows, _ = _read_purchase_import_rows(_legacy_single_document_workbook())

    assert _missing_purchase_order_import_fields(rows) == [
        "supplier",
        "summary",
        "date",
        "delivery_date",
        "handler",
    ]


def test_purchase_order_import_allows_an_empty_receiving_warehouse() -> None:
    rows, _ = _read_purchase_import_rows(_sample_purchase_workbook())
    for row in rows:
        row["warehouse"] = ""

    assert _missing_purchase_order_import_fields(rows) == []


def test_inventory_detail_list_refreshes_legacy_size_labels_from_product_size_group(monkeypatch) -> None:
    class _Connection:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback):
            return False

    class _Engine:
        def connect(self):
            return _Connection()

    class _Repository:
        engine = _Engine()

        def get_record(self, record_id):
            return {"id": record_id, "document_type": "报溢单", "supplier": "", "raw_payload": {}}

        def list_details(self, record_id):
            return [{
                "id": 1,
                "document_id": record_id,
                "product_code": "NI24Q3A030108",
                "size_quantities": {"230": "1", "240": "3"},
                "extra_fields": {"size_labels": "230|240|250", "size_range": "旧尺码段"},
            }]

        def get_supplier_by_name(self, name):
            return None

    monkeypatch.setattr(
        inventory_routes,
        "_load_inventory_detail_size_ranges",
        lambda connection, product_codes, preferred_brand: {"NI24Q3A030108": "NI尺码段35-47"},
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_size_group_items",
        lambda connection, size_ranges: {
            "NI尺码段35-47": (("35", "35"), ("36", "36"), ("37", "37"))
        },
    )
    request = SimpleNamespace(app=SimpleNamespace(state=SimpleNamespace(inventory_repository=_Repository())))

    result = inventory_routes.list_inventory_details(request, 1)

    assert result["items"][0]["extra_fields"]["size_range"] == "NI尺码段35-47"
    assert result["items"][0]["extra_fields"]["size_labels"] == "35|36|37"


def test_inventory_detail_list_maps_legacy_millimeter_size_to_combined_group(monkeypatch) -> None:
    class _Connection:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback):
            return False

    class _Engine:
        def connect(self):
            return _Connection()

    class _Repository:
        engine = _Engine()

        def get_record(self, record_id):
            return {"id": record_id, "document_type": "报溢单", "supplier": "", "raw_payload": {}}

        def list_details(self, record_id):
            return [{
                "id": 1,
                "document_id": record_id,
                "product_code": "NI24Q1A02030143-",
                "size_quantities": {"270": "3"},
                "extra_fields": {},
            }]

        def get_supplier_by_name(self, name):
            return None

    monkeypatch.setattr(
        inventory_routes,
        "_load_inventory_detail_size_ranges",
        lambda connection, product_codes, preferred_brand: {"NI24Q1A02030143-": "NI合码35-46"},
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_legacy_ni_combined_detail_profiles",
        lambda connection, base_codes_by_legacy_code: {
            "NI24Q1A02030143-": {
                "product_code": "NI24Q1A020301",
                "product_name": "NI24Q1A020301暗夜黑",
                "color_barcode": "01",
                "color_name": "暗夜黑",
            }
        },
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_size_group_items",
        lambda connection, size_ranges: {
            "NI合码35-46": (
                ("35-36", "35-36"),
                ("37-38", "37-38"),
                ("39-40", "39-40"),
                ("41-42", "41-42"),
                ("43-44", "43-44"),
                ("45-46", "45-46"),
            )
        },
    )
    request = SimpleNamespace(app=SimpleNamespace(state=SimpleNamespace(inventory_repository=_Repository())))

    result = inventory_routes.list_inventory_details(request, 1)

    assert result["items"][0]["size_quantities"] == {"43-44": "3"}
    assert result["items"][0]["product_code"] == "NI24Q1A020301"
    assert result["items"][0]["color_barcode"] == "01"
    assert result["items"][0]["color_name"] == "暗夜黑"


def test_purchase_order_import_rejects_legacy_size_column_template() -> None:
    rows, _ = _read_purchase_import_rows(_legacy_size_column_workbook())

    assert _purchase_order_import_has_size_columns(rows) is True


def test_purchase_import_reads_combined_size_column_as_one_size() -> None:
    rows, _ = _read_purchase_import_rows(_combined_size_column_workbook())

    assert rows[0]["quantity"] == "12"
    assert rows[0]["size_quantities"] == {"235-240": "12"}


def test_purchase_import_keeps_combined_size_quantity_and_full_base_code(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(inventory_routes, "_load_purchase_product_lookup", lambda connection, brand, product_codes: {})

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{"product_code": "C5563406D80235-240", "quantity": "12"}],
        brand="cbanner_womens",
        fallback_unit_price=0,
    )

    assert details[0]["product_code"] == "C5563406D80"
    assert details[0]["size_quantities"] == {"235-240": "12"}


def test_purchase_import_parses_combined_size_before_single_size_suffix() -> None:
    assert _split_purchase_size_code("C5563406D80235-240", "cbanner_womens") == ("C5563406D80", "235-240")
    assert _split_purchase_product_code("C5563406D80235-240", [], "cbanner_womens")[0] == "C5563406D80"
    assert _split_purchase_size_code("C5563406D8080240", "cbanner_womens") == ("C5563406D8080", "240")
    assert _split_purchase_size_code("NI24Q1A02030345-46", "ni") == ("NI24Q1A020303", "45-46")


def test_purchase_import_parses_truncated_ni_combined_size_suffix() -> None:
    assert _split_purchase_size_code("NI24Q1A02030143-", "ni") == ("NI24Q1A020301", "43")
    assert _split_purchase_product_code("NI24Q1A02030143-", [], "ni") == (
        "NI24Q1A020301",
        "NI24Q1A020301",
        "01",
        "",
        "43",
    )


def test_inventory_detail_list_recovers_color_for_full_ni_combined_size(monkeypatch) -> None:
    class _Connection:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback):
            return False

    class _Engine:
        def connect(self):
            return _Connection()

    class _Repository:
        engine = _Engine()

        def get_record(self, record_id):
            return {"id": record_id, "document_type": "报溢单", "supplier": "", "raw_payload": {}}

        def list_details(self, record_id):
            return [{
                "id": 1,
                "document_id": record_id,
                "product_code": "NI24Q1A02030345-46",
                "size_quantities": {"45-46": "2"},
                "extra_fields": {},
            }]

        def get_supplier_by_name(self, name):
            return None

    monkeypatch.setattr(
        inventory_routes,
        "_load_inventory_detail_size_ranges",
        lambda connection, product_codes, preferred_brand: {"NI24Q1A02030345-46": "NI合码35-46"},
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_legacy_ni_combined_detail_profiles",
        lambda connection, base_codes_by_legacy_code: {
            "NI24Q1A02030345-46": {
                "product_code": "NI24Q1A020303",
                "product_name": "NI24Q1A020303皓月白",
                "color_barcode": "03",
                "color_name": "皓月白",
            }
        },
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_size_group_items",
        lambda connection, size_ranges: {
            "NI合码35-46": (("45-46", "45-46"),)
        },
    )
    request = SimpleNamespace(app=SimpleNamespace(state=SimpleNamespace(inventory_repository=_Repository())))

    result = inventory_routes.list_inventory_details(request, 1)

    assert result["items"][0]["product_code"] == "NI24Q1A020303"
    assert result["items"][0]["color_barcode"] == "03"
    assert result["items"][0]["color_name"] == "皓月白"


def test_purchase_import_uses_product_size_group_labels(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_product_lookup",
        lambda connection, brand, product_codes: {
            "RCT63957D06": {
                "original_goods_code": "RCT63957D06",
                "color_name": "咖色",
                "color_code": "06",
                "size_range": "女鞋定制尺码",
                "barcode_build_rule": "货号+颜色代码+尺码",
            },
        },
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_size_group_items",
        lambda connection, size_ranges: {
            "女鞋定制尺码": (("34", "220"), ("35", "225")),
        },
    )

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{"product_code": "RCT63957D06", "quantity": "3", "size_quantities": {"220": "3"}}],
        brand="cbanner_womens",
        fallback_unit_price=0,
    )

    assert details[0]["extra_fields"]["size_range"] == "女鞋定制尺码"
    assert details[0]["extra_fields"]["size_labels"] == "34|35"
    assert details[0]["size_quantities"] == {"34": "3"}


def test_internal_sales_import_preserves_explicit_zero_price(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_product_lookup",
        lambda connection, brand, product_codes: {
            "RCT63957D06": {
                "original_goods_code": "RCT63957D06",
                "unit_price": "199",
            },
        },
    )
    monkeypatch.setattr(inventory_routes, "_load_purchase_size_group_items", lambda connection, size_ranges: {})

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{"product_code": "RCT63957D06", "quantity": "3", "unit_price": "0"}],
        brand="cbanner_womens",
        fallback_unit_price=0,
        prefer_lookup_unit_price=True,
        preserve_explicit_zero_price=True,
    )

    assert details[0]["unit_price"] == "0"
    assert details[0]["amount"] == "0"


def test_internal_sales_manual_detail_preserves_explicit_zero_price() -> None:
    class _InternalSalesRepository:
        def is_internal_sales_customer(self, name):
            return name == "千百度-内销客户"

    payload = inventory_routes._apply_product_archive_cost(
        _InternalSalesRepository(),
        {
            "document_type": "批发销售单",
            "supplier": "千百度-内销客户",
        },
        {
            "product_code": "RCT63957D06",
            "quantity": "3",
            "unit_price": "0.00",
            "amount": "597",
        },
    )

    assert payload["unit_price"] == "0"
    assert payload["amount"] == "0"


def test_internal_sales_import_still_fills_blank_price(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_product_lookup",
        lambda connection, brand, product_codes: {
            "RCT63957D06": {
                "original_goods_code": "RCT63957D06",
                "unit_price": "199",
            },
        },
    )
    monkeypatch.setattr(inventory_routes, "_load_purchase_size_group_items", lambda connection, size_ranges: {})

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{"product_code": "RCT63957D06", "quantity": "3", "unit_price": ""}],
        brand="cbanner_womens",
        fallback_unit_price=0,
        prefer_lookup_unit_price=True,
        preserve_explicit_zero_price=True,
    )

    assert details[0]["unit_price"] == "199"
    assert details[0]["amount"] == "597"


def test_blank_price_without_archive_cost_remains_empty(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(inventory_routes, "_load_purchase_product_lookup", lambda connection, brand, product_codes: {})

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{"product_code": "UNKNOWN", "quantity": "3", "unit_price": ""}],
        brand="cbanner_womens",
        fallback_unit_price=0,
        preserve_explicit_zero_price=True,
    )

    assert details[0]["unit_price"] is None
    assert details[0]["amount"] is None


class _WholesalePriceRepository(_StubRepository):
    def __init__(self, prices: dict[str, str]):
        self.prices = prices
        self.calls: list[dict[str, object]] = []

    def latest_wholesale_sales_prices(self, **kwargs):
        self.calls.append(kwargs)
        return self.prices


def _mock_wholesale_product(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_product_lookup",
        lambda connection, brand, product_codes: {
            "RCT63957D06": {
                "original_goods_code": "RCT63957D06",
                "unit_price": "199",
            },
        },
    )
    monkeypatch.setattr(inventory_routes, "_load_purchase_size_group_items", lambda connection, size_ranges: {})


def test_wholesale_imported_price_overrides_history_and_archive(monkeypatch) -> None:
    _mock_wholesale_product(monkeypatch)
    repository = _WholesalePriceRepository({"RCT63957D06": "120"})

    details = _build_purchase_details_from_rows(
        repository,
        [{"product_code": "RCT63957D06", "quantity": "3", "unit_price": "150"}],
        brand="cbanner_womens",
        fallback_unit_price=0,
        wholesale_customer="客户A",
        wholesale_price_date="2026-08-19",
    )

    assert details[0]["unit_price"] == "150"
    assert details[0]["amount"] == "450"


def test_wholesale_blank_price_uses_latest_same_customer_sales_price(monkeypatch) -> None:
    _mock_wholesale_product(monkeypatch)
    repository = _WholesalePriceRepository({"RCT63957D06": "120"})

    details = _build_purchase_details_from_rows(
        repository,
        [{"product_code": "RCT63957D06", "quantity": "3", "unit_price": ""}],
        brand="cbanner_womens",
        fallback_unit_price=0,
        wholesale_customer="客户A",
        wholesale_price_date="2026-08-19",
    )

    assert details[0]["unit_price"] == "120"
    assert details[0]["amount"] == "360"
    assert repository.calls[0]["customer"] == "客户A"
    assert repository.calls[0]["as_of_date"] == "2026-08-19"


def test_wholesale_blank_price_without_history_does_not_use_archive_cost(monkeypatch) -> None:
    _mock_wholesale_product(monkeypatch)
    repository = _WholesalePriceRepository({})

    details = _build_purchase_details_from_rows(
        repository,
        [{"product_code": "RCT63957D06", "quantity": "3", "unit_price": ""}],
        brand="cbanner_womens",
        fallback_unit_price=0,
        wholesale_customer="客户A",
        wholesale_price_date="2026-08-19",
    )

    assert details[0]["unit_price"] is None
    assert details[0]["amount"] is None


def test_purchase_import_splits_ni_mixed_gender_sizes_for_costs(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_product_lookup",
        lambda connection, brand, product_codes: {
            "NIA2253A020115": {
                "original_goods_code": "NIA2253A020115",
                "color_name": "黑黄",
                "color_code": "A2",
                "gender_costs": {"女": "178", "男": "190"},
            },
        },
    )
    monkeypatch.setattr(inventory_routes, "_load_purchase_size_group_items", lambda connection, size_ranges: {})

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{
            "product_code": "NIA2253A020115",
            "quantity": "3",
            "size_quantities": {"36": "2", "42": "1"},
        }],
        brand="ni",
        fallback_unit_price=0,
    )

    assert [(item["unit_price"], item["quantity"], item["size_quantities"]) for item in details] == [
        ("178", "2", {"36": "2"}),
        ("190", "1", {"42": "1"}),
    ]


def test_purchase_import_uses_matched_product_code_when_color_barcode_is_repeated(monkeypatch) -> None:
    monkeypatch.setattr(inventory_routes, "_load_color_barcodes", lambda connection: [])
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_product_lookup",
        lambda connection, brand, product_codes: {
            "RCW62308S28": {
                "original_goods_code": "RCW62308S28",
                "color_name": "米白",
                "color_code": "28",
                "size_range": "合码225-250",
            },
        },
    )
    monkeypatch.setattr(
        inventory_routes,
        "_load_purchase_size_group_items",
        lambda connection, size_ranges: {
            "合码225-250": (("225-230", "225-230"), ("235-240", "235-240"), ("245-250", "245-250")),
        },
    )

    details = _build_purchase_details_from_rows(
        _StubRepository(),
        [{"product_code": "RCW62308S2828", "quantity": "3", "size_quantities": {"235-240": "3"}}],
        brand="cbanner_womens",
        fallback_unit_price=0,
    )

    assert details[0]["product_code"] == "RCW62308S28"
    assert details[0]["color_barcode"] == "28"
    assert details[0]["extra_fields"]["size_range"] == "合码225-250"
    assert details[0]["size_quantities"] == {"235-240": "3"}


def test_purchase_order_import_template_does_not_require_unit_price() -> None:
    workbook = _build_purchase_order_import_template()
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]

    assert headers == [
        "供应商",
        "商品编码",
        "数量",
        "采购单备注",
        "采购日期",
        "协议到货日期",
        "收货仓库",
        "经办人",
        "附加说明",
    ]
    assert "单价" not in headers


def test_surplus_import_uses_warehouse_brand_when_supplier_is_empty() -> None:
    assert _purchase_import_brand_for_supplier(
        _WarehouseBrandRepository(),
        "",
        "报溢单",
        "cbanner_mens",
        "NI仙岩仓库",
    ) == "ni"


def test_warehouse_brand_overrides_legacy_default_brand_for_saved_record() -> None:
    assert _purchase_import_brand_for_record(
        _WarehouseBrandRepository(),
        {
            "supplier": "",
            "warehouse": "NI仙岩仓库",
            "document_type": "报溢单",
            "raw_payload": {"brand": "cbanner_mens"},
        },
    ) == "ni"


def test_purchase_size_row_export_does_not_include_duplicate_note_columns() -> None:
    assert "行号" not in PURCHASE_SIZE_ROW_EXPORT_HEADERS
    assert "摘要" not in PURCHASE_SIZE_ROW_EXPORT_HEADERS
    assert "采购单备注" in PURCHASE_SIZE_ROW_EXPORT_HEADERS


def test_purchase_summary_export_does_not_include_row_number() -> None:
    assert "行号" not in PURCHASE_SUMMARY_EXPORT_HEADERS
    assert "附加说明" in PURCHASE_SUMMARY_EXPORT_HEADERS
    assert "工厂货号" in PURCHASE_SUMMARY_EXPORT_HEADERS
