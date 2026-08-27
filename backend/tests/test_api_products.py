from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from api.routes.import_export import _build_product_import_template
from transform.rows import build_admin_record


def test_product_import_template_workbook_contains_headers_and_guidance():
    buffer = _build_product_import_template()
    workbook = load_workbook(buffer)
    worksheet = workbook["商品导入模板"]
    headers = [cell.value for cell in worksheet[1]]

    assert headers[:4] == ["货号", "原始货号", "品名", "组别"]
    assert "供应商名" in headers
    assert "条码构成逻辑" in headers
    assert "跟高" in headers
    assert "后跟高" in headers
    assert headers.index("后跟高") == headers.index("跟高") + 1
    assert "跟底款式" in headers
    assert "流行元素" in headers
    assert "开口深度" in headers
    assert "靴筒" in headers
    assert "鞋网面类型" in headers
    assert "图片" not in headers
    assert "实际导入品牌由页面当前选中的Tab决定" in workbook["填写说明"]["C2"].value
    assert workbook["填写说明"]["C3"].value.startswith("黄色表头")
    assert len(worksheet.data_validations.dataValidation) == 1


def test_import_products_uses_brand_specific_heel_height_label(
    test_app_client: TestClient,
    repository,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "跟高", "后跟高"])
    worksheet.append(["HEEL-WOMENS-001", "4cm", "6cm"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    womens_response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_womens"},
        files={
            "file": (
                "womens-heel-height.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert womens_response.status_code == 200
    womens_product = repository.find_by_sku("cbanner_womens", "HEEL-WOMENS-001")
    assert womens_product is not None
    assert womens_product["heel_height"] == "6cm"

    mens_workbook = Workbook()
    mens_worksheet = mens_workbook.active
    mens_worksheet.append(["货号", "跟高", "后跟高"])
    mens_worksheet.append(["HEEL-MENS-001", "4cm", "6cm"])
    mens_buffer = io.BytesIO()
    mens_workbook.save(mens_buffer)
    mens_buffer.seek(0)

    mens_response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "mens-heel-height.xlsx",
                mens_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert mens_response.status_code == 200
    mens_product = repository.find_by_sku("cbanner_mens", "HEEL-MENS-001")
    assert mens_product is not None
    assert mens_product["heel_height"] == "4cm"


def test_import_cbanner_womens_preserves_style_detail_fields(
    test_app_client: TestClient,
    repository,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([
        "货号",
        "跟底款式",
        "流行元素",
        "鞋帮高度",
        "开口深度",
        "靴筒",
        "闭合方式",
        "鞋网面类型",
    ])
    worksheet.append([
        "WOMENS-STYLE-DETAIL-001",
        "平底",
        "钉珠",
        "低帮",
        "深口",
        "短筒",
        "系带",
        "网面",
    ])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_womens"},
        files={
            "file": (
                "womens-style-detail.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200
    product = repository.find_by_sku("cbanner_womens", "WOMENS-STYLE-DETAIL-001")
    assert product is not None
    assert product["sole_style"] == "平底"
    assert product["fashion_elements"] == "钉珠"
    assert product["upper_height"] == "低帮"
    assert product["opening_depth"] == "深口"
    assert product["boot_shaft"] == "短筒"
    assert product["closure_type"] == "系带"
    assert product["mesh_upper_type"] == "网面"


def test_get_products_returns_paginated_rows(test_app_client: TestClient, repository):
    repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "A1001",
                "original_sku": "OA1001",
            },
        ),
    )

    response = test_app_client.get(
        "/products",
        params={"brand": "cbanner_mens", "page": 1, "page_size": 20},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["brand"] == "cbanner_mens"
    assert body["items"][0]["original_sku"] == "OA1001"


def test_product_auxiliary_options_use_womens_and_other_brand_scopes(test_app_client: TestClient):
    from sqlalchemy import delete, insert

    from domain.product_auxiliary_attribute_schema import PRODUCT_AUXILIARY_ATTRIBUTE_TABLE

    repository = test_app_client.app.state.repository
    with repository.engine.begin() as connection:
        connection.execute(delete(PRODUCT_AUXILIARY_ATTRIBUTE_TABLE))
        connection.execute(insert(PRODUCT_AUXILIARY_ATTRIBUTE_TABLE), [
            {
                "brand_scope": "cbanner_womens",
                "attribute_type": "鞋面材质",
                "attribute_name": "女鞋专属材质",
                "source_workbook": "womens.xls",
                "source_sheet": "千百度女鞋",
                "source_row_number": "5",
            },
            {
                "brand_scope": "other",
                "attribute_type": "品名",
                "attribute_name": "其他品牌品名",
                "source_workbook": "other.xls",
                "source_sheet": "导出数据",
                "source_row_number": "5",
            },
            {
                "brand_scope": "other",
                "attribute_type": "品牌",
                "attribute_name": "不支持的类型",
                "source_workbook": "other.xls",
                "source_sheet": "导出数据",
                "source_row_number": "6",
            },
        ])

    womens_response = test_app_client.get("/products/auxiliary-options", params={"brand": "cbanner_womens"})
    mens_response = test_app_client.get("/products/auxiliary-options", params={"brand": "cbanner_mens"})

    assert womens_response.status_code == 200
    assert womens_response.json()["brand_scope"] == "cbanner_womens"
    assert womens_response.json()["items"] == [
        {"field": "upper_material", "type_name": "鞋面材质", "options": ["女鞋专属材质"]}
    ]
    assert mens_response.status_code == 200
    assert mens_response.json()["brand_scope"] == "other"
    assert mens_response.json()["items"] == [
        {"field": "product_name", "type_name": "品名", "options": ["其他品牌品名"]}
    ]


def test_get_product_returns_404_when_missing(test_app_client: TestClient):
    response = test_app_client.get("/products/cbanner_mens/99999")

    assert response.status_code == 404
    assert response.json()["detail"] == "Product not found"


def test_post_products_creates_product_via_build_admin_record(test_app_client: TestClient):
    response = test_app_client.post(
        "/products",
        json={
            "brand": "cbanner_mens",
            "payload": {
                "sku": "A1001",
                "original_sku": "OA1001",
                "product_name": "女士皮鞋",
                "category": "女鞋",
                "color": "黑色",
                "barcode_build_rule": "货号+颜色代码+尺码",
            },
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["message"] == "Product created"
    assert body["item"]["brand"] == "cbanner_mens"
    assert body["item"]["source_workbook"] == "manual_admin"
    assert body["item"]["raw_payload"]["sku"] == "A1001"
    assert body["item"]["product_name"] == "女士皮鞋"
    assert body["item"]["category"] == "女鞋"
    assert body["item"]["barcode_build_rule"] == "货号+颜色代码+尺码"


def test_cbanner_womens_product_style_fields_can_be_created(test_app_client: TestClient):
    response = test_app_client.post(
        "/products",
        json={
            "brand": "cbanner_womens",
            "payload": {
                "sku": "WOMENS-STYLE-001",
                "original_sku": "WOMENS-STYLE-001",
                "sole_style": "粗跟",
                "fashion_elements": "金属装饰",
                "heel_height": "5cm",
                "upper_height": "低帮",
                "opening_depth": "浅口",
                "boot_shaft": "短筒",
                "closure_type": "套脚",
                "mesh_upper_type": "单网面",
                "barcode_build_rule": "货号+颜色代码+尺码",
            },
        },
    )

    assert response.status_code == 200
    item = response.json()["item"]
    assert item["sole_style"] == "粗跟"
    assert item["fashion_elements"] == "金属装饰"
    assert item["heel_height"] == "5cm"
    assert item["upper_height"] == "低帮"
    assert item["opening_depth"] == "浅口"
    assert item["boot_shaft"] == "短筒"
    assert item["closure_type"] == "套脚"
    assert item["mesh_upper_type"] == "单网面"


def test_put_products_preserves_existing_metadata(test_app_client: TestClient, repository):
    created = repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "A1001",
                "original_sku": "OA1001",
                "color": "黑色",
                "extra_fields": {"数据源列": "原始值"},
            },
        ),
    )

    response = test_app_client.put(
        f"/products/cbanner_mens/{created['id']}",
        json={
            "brand": "cbanner_mens",
            "payload": {
                "sku": "A1001",
                "original_sku": "OA1001",
                "color": "白金",
            },
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["message"] == "Product updated"
    assert body["item"]["color"] == "白金"
    assert body["item"]["source_workbook"] == created["source_workbook"]
    assert body["item"]["source_sheet"] == created["source_sheet"]
    assert body["item"]["source_row_number"] == created["source_row_number"]
    assert body["item"]["extra_fields"] == {"数据源列": "原始值"}


def test_put_products_rejects_brand_mismatch(test_app_client: TestClient, repository):
    created = repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "A1001",
                "original_sku": "OA1001",
            },
        ),
    )

    response = test_app_client.put(
        f"/products/cbanner_mens/{created['id']}",
        json={
            "brand": "yandou",
            "payload": {
                "sku": "A1001",
                "original_sku": "OA1001",
            },
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Brand mismatch"



def test_put_products_returns_404_for_missing_row(test_app_client: TestClient):
    response = test_app_client.put(
        "/products/cbanner_mens/99999",
        json={
            "brand": "cbanner_mens",
            "payload": {
                "sku": "A1001",
                "original_sku": "OA1001",
            },
        },
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "Product not found"



def test_delete_products_returns_message_and_removes_row(test_app_client: TestClient, repository):
    created = repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "A1001",
                "original_sku": "OA1001",
            },
        ),
    )

    response = test_app_client.delete(f"/products/cbanner_mens/{created['id']}")

    assert response.status_code == 200
    assert response.json() == {"message": "Product deleted"}
    assert repository.get_product("cbanner_mens", created["id"]) is None


def test_delete_products_returns_404_when_missing(test_app_client: TestClient):
    response = test_app_client.delete("/products/cbanner_mens/99999")

    assert response.status_code == 404
    assert response.json()["detail"] == "Product not found"


def test_post_products_rejects_empty_payload(test_app_client: TestClient):
    response = test_app_client.post(
        "/products",
        json={
            "brand": "cbanner_mens",
            "payload": {
                "sku": "   ",
                "original_sku": None,
                "color": "",
            },
        },
    )

    assert response.status_code == 422



def test_product_write_request_forbids_extra_top_level_fields(test_app_client: TestClient):
    response = test_app_client.post(
        "/products",
        json={
            "brand": "cbanner_mens",
            "payload": {
                "sku": "A1001",
                "original_sku": "OA1001",
            },
            "unexpected": "nope",
        },
    )

    assert response.status_code == 422



def test_product_payload_forbids_extra_fields(test_app_client: TestClient):
    response = test_app_client.post(
        "/products",
        json={
            "brand": "cbanner_mens",
            "payload": {
                "sku": "A1001",
                "original_sku": "OA1001",
                "unexpected": "nope",
            },
        },
    )

    assert response.status_code == 422


def test_import_products_updates_by_original_sku_without_clearing_blank_cells(
    test_app_client: TestClient,
    repository,
):
    existing = repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "SKU-OLD",
                "original_sku": "ORIG-001",
                "color": "黑色",
                "upper_material": "牛皮",
                "execution_standard": "QB/T 1002",
                "season_category": "春秋",
            },
        ),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "原始货号", "颜色", "鞋面材质", "执行标准", "季节分类", "条码构成逻辑"])
    worksheet.append(["", "ORIG-001", "白色", "", "", "", "货号+颜色代码+尺码"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "partial-products.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 0
    assert body["updated"] == 1

    updated = repository.get_product("cbanner_mens", existing["id"])
    assert updated is not None
    assert updated["color"] == "白色"
    assert updated["upper_material"] == "牛皮"
    assert updated["execution_standard"] == "QB/T 1002"
    assert updated["season_category"] == "春秋"

    listing = repository.list_products("cbanner_mens", query="ORIG-001", page=1, page_size=10)
    assert listing["total"] == 1


def test_import_products_restores_recycled_product_with_the_same_sku(
    test_app_client: TestClient,
    repository,
):
    existing = repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "RECYCLED-IMPORT-001",
                "original_sku": "RECYCLED-IMPORT-001",
                "color": "黑色",
            },
        ),
    )
    assert repository.delete_product("cbanner_mens", existing["id"]) is True

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "原始货号", "颜色"])
    worksheet.append(["RECYCLED-IMPORT-001", "RECYCLED-IMPORT-001", "白色"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "restore-recycled-product.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200
    assert response.json()["created"] == 0
    assert response.json()["updated"] == 1
    restored = repository.get_product("cbanner_mens", existing["id"])
    assert restored is not None
    assert restored["color"] == "白色"
    assert restored["deleted_at"] is None
    assert repository.list_recycled_products(brand="cbanner_mens", page=1, page_size=10)["total"] == 0


def test_import_products_keeps_distinct_skus_with_shared_original_sku(
    test_app_client: TestClient,
    repository,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "原始货号", "颜色", "条码构成逻辑"])
    worksheet.append(["SHARED-ORIG-01", "ORIG-SHARED", "黑色", "货号+颜色代码+尺码"])
    worksheet.append(["SHARED-ORIG-02", "ORIG-SHARED", "白色", "货号+颜色代码+尺码"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "shared-original-sku.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200
    assert response.json()["created"] == 2
    assert response.json()["updated"] == 0
    first = repository.find_by_sku("cbanner_mens", "SHARED-ORIG-01")
    second = repository.find_by_sku("cbanner_mens", "SHARED-ORIG-02")
    assert first is not None
    assert second is not None
    assert first["original_sku"] == second["original_sku"] == "ORIG-SHARED"
    assert first["color"] == "黑色"
    assert second["color"] == "白色"

    logs = test_app_client.app.state.operation_log_repository.list_logs(
        module="product",
        query="shared-original-sku.xlsx",
        page=1,
        page_size=10,
    )
    assert logs["total"] == 1
    after_data = logs["items"][0]["after_data"]
    assert after_data["created_item_count"] == 2
    assert after_data["updated_item_count"] == 0
    assert [item["sku"] for item in after_data["created_items"]] == [
        "SHARED-ORIG-01",
        "SHARED-ORIG-02",
    ]


def test_import_products_rejects_supplier_mismatch_for_existing_product(
    test_app_client: TestClient,
    repository,
):
    existing = repository.create_product(
        "cbanner_mens",
        build_admin_record(
            "cbanner_mens",
            {
                "sku": "SUPPLIER-MISMATCH-001",
                "supplier_name": "原供应商",
                "barcode_build_rule": "货号+尺码",
            },
        ),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "供应商名", "条码构成逻辑"])
    worksheet.append(["SUPPLIER-MISMATCH-001", "导入供应商", "货号+尺码"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "supplier-mismatch.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 400
    assert "供应商不一致" in response.json()["detail"]
    assert "原供应商" in response.json()["detail"]
    assert "导入供应商" in response.json()["detail"]

    unchanged = repository.get_product("cbanner_mens", existing["id"])
    assert unchanged is not None
    assert unchanged["supplier_name"] == "原供应商"


def test_import_products_rolls_back_all_rows_when_a_row_is_invalid(
    test_app_client: TestClient,
    repository,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "尺码段", "条码构成逻辑"])
    worksheet.append(["TXN-OK", "", "货号+颜色代码+尺码"])
    worksheet.append(["TXN-BAD", "不存在的尺码组", "货号+尺码"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "invalid-size-group.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 400
    assert "第 3 行导入失败" in response.json()["detail"]
    assert repository.find_by_sku("cbanner_mens", "TXN-OK") is None
    assert repository.find_by_sku("cbanner_mens", "TXN-BAD") is None


def test_import_products_uses_fixed_barcode_build_rule_when_omitted(
    test_app_client: TestClient,
    repository,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "品名"])
    worksheet.append(["BARCODE-RULE-DEFAULT", "测试鞋"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "cbanner_mens"},
        files={
            "file": (
                "default-barcode-rule.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200
    imported = repository.find_by_sku("cbanner_mens", "BARCODE-RULE-DEFAULT")
    assert imported is not None
    assert imported["barcode_build_rule"] == "货号+颜色代码+尺码"


def test_import_products_requires_barcode_build_rule_when_no_fixed_rule(
    test_app_client: TestClient,
    repository,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["货号", "品名"])
    worksheet.append(["BARCODE-RULE-REQUIRED", "测试鞋"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = test_app_client.post(
        "/import",
        params={"brand": "yandou"},
        files={
            "file": (
                "missing-barcode-rule.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 400
    assert "第 2 行导入失败" in response.json()["detail"]
    assert "BARCODE-RULE-REQUIRED" in response.json()["detail"]
    assert "未填写条码构成逻辑" in response.json()["detail"]
    assert repository.find_by_sku("yandou", "BARCODE-RULE-REQUIRED") is None


def test_download_product_import_template_matches_supported_import_fields(
    test_app_client: TestClient,
):
    response = test_app_client.get(
        "/import/template",
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    worksheet = workbook["商品导入模板"]
    headers = [cell.value for cell in worksheet[1]]
    assert headers[:4] == ["货号", "原始货号", "品名", "组别"]
    assert "供应商名" in headers
    assert "条码构成逻辑" in headers
    assert "图片" not in headers
    assert workbook["填写说明"]["C3"].value.startswith("黄色表头")
    assert len(worksheet.data_validations.dataValidation) == 1
