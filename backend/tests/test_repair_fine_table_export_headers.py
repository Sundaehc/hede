from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from scripts.repair_fine_table_export_headers import inspect_workbook, repair_workbook


def test_repair_workbook_changes_only_known_english_headers(tmp_path: Path) -> None:
    path = tmp_path / "fine-table.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["sku", "platform", "selling_points", "category", "future_field"])
    worksheet.append(["ABC", "天猫", "卖点内容", "女鞋", "保留值"])
    workbook.save(path)

    replacements, unknown = inspect_workbook(path)
    assert replacements == {
        "sku": "货号",
        "platform": "所属平台",
        "selling_points": "卖点",
        "category": "分类",
    }
    assert unknown == ["future_field"]

    repair_workbook(path, replacements)

    repaired, unknown = inspect_workbook(path)
    assert repaired == {}
    assert unknown == ["future_field"]
    saved_workbook = load_workbook(path, read_only=True)
    try:
        rows = list(saved_workbook.active.iter_rows(values_only=True))
    finally:
        saved_workbook.close()
    assert rows == [
        ("货号", "所属平台", "卖点", "分类", "future_field"),
        ("ABC", "天猫", "卖点内容", "女鞋", "保留值"),
    ]
