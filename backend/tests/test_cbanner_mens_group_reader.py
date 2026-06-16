from __future__ import annotations

from openpyxl import Workbook

from fileio.cbanner_mens_group_reader import (
    read_cbanner_mens_group_map,
    read_cbanner_mens_product_level_map,
    read_cbanner_womens_product_level_map,
    read_eblan_product_level_map,
)


def test_read_cbanner_mens_group_map_reads_product_detail_sheet(tmp_path):
    source_root = tmp_path / "千百度男鞋"
    workbook_dir = source_root / "2026年" / "6月份"
    workbook_dir.mkdir(parents=True)
    workbook_path = workbook_dir / "赫德货品表（千百度男鞋）6.4.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "商品明细表"
    worksheet.append(["标题行"])
    worksheet.append(["货号", "品名", "组别", "商品等级"])
    worksheet.append([" A1001 ", "男鞋", "一组", "A"])
    worksheet.append(["A1002", "男鞋", "二组", "B"])
    workbook.save(workbook_path)
    workbook.close()

    group_map = read_cbanner_mens_group_map(source_root)
    product_level_map = read_cbanner_mens_product_level_map(source_root)

    assert group_map == {"A1001": "一组", "A1002": "二组"}
    assert product_level_map == {"A1001": "A", "A1002": "B"}


def test_read_cbanner_womens_product_level_map_reads_latest_product_detail_workbook(tmp_path):
    source_root = tmp_path / "千百度女鞋"
    source_root.mkdir()
    old_path = source_root / "赫德货品表（千百度）6.15.xlsx"
    new_path = source_root / "赫德货品表（千百度）6.16.xlsx"

    for workbook_path, level in ((old_path, "旧等级"), (new_path, "新等级")):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "明细"
        worksheet.append(["货号", "商品等级"])
        worksheet.append(["W1001", level])
        workbook.save(workbook_path)
        workbook.close()

    assert read_cbanner_womens_product_level_map(source_root) == {"W1001": "新等级"}


def test_read_eblan_product_level_map_reads_xlsm_product_detail_workbook(tmp_path):
    source_root = tmp_path / "伊伴" / "2026" / "2026-06"
    source_root.mkdir(parents=True)
    workbook_path = source_root / "伊伴货品表（06.16）.xlsm"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "商品"
    worksheet.append(["说明"])
    worksheet.append(["原始货号", "货品等级"])
    worksheet.append(["E1001", "重点"])
    workbook.save(workbook_path)
    workbook.close()

    assert read_eblan_product_level_map(source_root) == {"E1001": "重点"}
