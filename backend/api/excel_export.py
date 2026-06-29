from __future__ import annotations

import unicodedata
from collections.abc import Mapping

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TEXT_HEADER_KEYWORDS = (
    "ID",
    "id",
    "编号",
    "编码",
    "货号",
    "条码",
    "款号",
    "代码",
    "单据编号",
)
DEFAULT_NUMERIC_HEADER_KEYWORDS = (
    "数量",
    "金额",
    "成本",
    "单价",
    "价格",
    "完成率",
    "库存",
    "销量",
    "销售额",
    "退货",
)
DEFAULT_DATE_HEADER_KEYWORDS = ("日期", "时间")

DEFAULT_WIDTH_BY_HEADER = {
    "行号": 8,
    "序号": 8,
    "ID": 14,
    "id": 14,
    "商品ID": 18,
    "商品编码": 20,
    "货品编码": 20,
    "商品编号": 20,
    "货号": 20,
    "原始货号": 20,
    "款式编码": 20,
    "款号": 20,
    "工厂货号": 18,
    "颜色条码": 12,
    "尺码条码": 12,
    "条码": 28,
    "单据编号": 24,
    "商品名": 28,
    "商品名称": 28,
    "商品全名": 30,
    "颜色及规格": 18,
    "颜色名称": 14,
    "仓库全名": 18,
    "单位全名": 18,
    "供应商名": 20,
    "供应商商品款号": 20,
    "摘要": 24,
    "鞋面材质": 16,
    "内里材质": 16,
    "大底材质": 16,
    "鞋垫材质": 16,
    "执行标准": 18,
    "产品型号": 18,
    "录单日期": 13,
    "到货日期": 13,
    "订货日期": 13,
    "交货日期": 13,
    "上市时间": 13,
    "首单时间": 13,
}


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _display_width(value: object) -> int:
    text = _cell_text(value)
    width = 0
    for char in text:
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W", "A"} else 1
    return width


def _header_matches(header: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in header for keyword in keywords)


def style_excel_worksheet(
    worksheet,
    *,
    width_by_header: Mapping[str, int] | None = None,
    text_headers: set[str] | None = None,
    numeric_headers: set[str] | None = None,
    min_width: int = 10,
    max_width: int = 42,
    freeze_panes: str | None = "A2",
    auto_filter: bool = True,
) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    widths = {**DEFAULT_WIDTH_BY_HEADER, **(dict(width_by_header or {}))}
    explicit_text_headers = text_headers or set()
    explicit_numeric_headers = numeric_headers or set()
    header_font = Font(name="宋体", size=10, bold=True)
    body_font = Font(name="宋体", size=10)
    fill = PatternFill("solid", fgColor="F2F2F2")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    number_alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)

    worksheet.row_dimensions[1].height = 24
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = header_alignment

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 20

    for col_index in range(1, worksheet.max_column + 1):
        header = _cell_text(worksheet.cell(row=1, column=col_index).value)
        content_width = max(
            (_display_width(worksheet.cell(row=row_index, column=col_index).value) for row_index in range(1, worksheet.max_row + 1)),
            default=0,
        ) + 2
        if header in widths:
            width = max(widths[header], min(content_width, max_width))
        elif header.isdigit() and len(header) <= 3:
            width = 7
        else:
            width = max(min_width, min(content_width, max_width))
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

        is_text_column = header in explicit_text_headers or _header_matches(header, DEFAULT_TEXT_HEADER_KEYWORDS)
        is_numeric_column = header in explicit_numeric_headers or _header_matches(header, DEFAULT_NUMERIC_HEADER_KEYWORDS)
        is_date_column = _header_matches(header, DEFAULT_DATE_HEADER_KEYWORDS)
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.font = body_font
            if is_text_column:
                cell.alignment = text_alignment
                cell.number_format = "@"
            elif is_numeric_column:
                cell.alignment = number_alignment
            elif is_date_column:
                cell.alignment = center_alignment
            else:
                cell.alignment = text_alignment

    if freeze_panes:
        worksheet.freeze_panes = freeze_panes
    if auto_filter:
        worksheet.auto_filter.ref = worksheet.dimensions


def style_excel_workbook(workbook, **kwargs) -> None:
    for worksheet in workbook.worksheets:
        style_excel_worksheet(worksheet, **kwargs)
