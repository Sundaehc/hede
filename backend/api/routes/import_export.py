from __future__ import annotations

import io
import logging
import urllib.parse
import unicodedata
from collections.abc import Iterator
from datetime import date as date_type
from datetime import datetime, time, timedelta
from zoneinfo import ZoneInfo

from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, desc, or_, select

from api.excel_export import DEFAULT_WIDTH_BY_HEADER, style_excel_worksheet
from api.fine_table_cache import clear_fine_table_cache
from api.product_goods_cache import clear_product_goods_cache
from api.operation_log_utils import write_operation_log
from api.routes.images import get_image_matcher, image_url_for
from domain.excluded_skus import is_excluded_sku, not_excluded_sku_condition
from domain.fields import PRODUCT_FIELDS
from domain.gj_schema import GJ_MERGED_PRODUCT_INFO_TABLE
from domain.product_defaults import apply_product_defaults
from domain.product_size_code import build_product_size_code
from domain.sources import CANONICAL_COLUMNS, COLUMN_ALIASES
from domain.color_barcode_schema import COLOR_BARCODE_TABLE
from domain.size_group_schema import SIZE_GROUP_ITEMS_TABLE, SIZE_GROUPS_TABLE
from domain.vip_schema import JST_PRODUCT_PROFILE_TABLE
from storage.product_repository import (
    PRODUCT_COLOR_BARCODE_SOURCE_BRANDS,
    _color_name_variants,
    apply_jst_product_costs,
)
from transform.rows import EXCLUDED_EXTRA_FIELD_KEYS, build_admin_record, filter_extra_fields, normalize_admin_field

router = APIRouter()
logger = logging.getLogger(__name__)

EXPORT_LABELS = {field.name: field.label for field in PRODUCT_FIELDS}
# Keep the existing export wording for this field while using the canonical field labels elsewhere.
EXPORT_LABELS["toe_shape"] = "鞋头款式"

EXPORT_COLUMNS = [c for c in CANONICAL_COLUMNS if c != "image_path"]
CN_TO_FIELD = {cn: en for cn, en in COLUMN_ALIASES.items() if en in EXPORT_COLUMNS}
SIZE_EXPORT_MODE = "with_sizes"
SIZE_EXPORT_HEADERS = [
    "供应商名",
    "商品编码",
    "款式编码",
    "商品名",
    "颜色名称",
    "尺码条码",
    "鞋面材质",
    "品名",
    "执行标准",
    "产品型号",
    "内里材质",
    "大底材质",
    "鞋垫材质",
    "原始货号",
    "供应商商品款号",
    "品牌",
    "颜色及规格",
    "分类",
    "成本价",
    "LOGO",
]
LOOKUP_CHUNK_SIZE = 2000
SHANGHAI_TIME_ZONE = ZoneInfo("Asia/Shanghai")
SIZE_EXPORT_MAX_WIDTH = 42
SIZE_EXPORT_MIN_WIDTH = 10
SIZE_EXPORT_WIDTH_BY_HEADER = {
    "供应商名": 20,
    "商品编码": 24,
    "款式编码": 20,
    "商品名": 32,
    "颜色名称": 16,
    "尺码条码": 12,
    "鞋面材质": 16,
    "品名": 18,
    "执行标准": 18,
    "产品型号": 18,
    "内里材质": 16,
    "大底材质": 16,
    "鞋垫材质": 16,
    "原始货号": 20,
    "供应商商品款号": 20,
    "品牌": 14,
    "颜色及规格": 18,
    "分类": 18,
    "成本价": 12,
    "LOGO": 14,
}
GJ_SIZE_EXPORT_COLUMNS = (
    "goods_code",
    "original_goods_code",
    "goods_full_name",
    "factory_code",
    "product_name",
    "execution_standard",
    "insole_material",
    "outsole_material",
    "lining_material",
    "upper_material",
    "brand",
    "source_date_value",
    "updated_at",
    "id",
)

BRAND_LABELS = {
    "cbanner_mens": "千百度男鞋",
    "cbanner_womens": "千百度女鞋",
    "yandou": "烟斗",
    "eblan": "伊伴",
    "smiley": "笑脸",
    "ni": "NI",
    "all": "总览",
}


def _activity_date_export_condition(table, activity_date: date_type):
    start = datetime.combine(activity_date, time.min, tzinfo=SHANGHAI_TIME_ZONE)
    end = start + timedelta(days=1)
    return or_(
        and_(table.c.created_at >= start, table.c.created_at < end),
        and_(table.c.last_imported_at >= start, table.c.last_imported_at < end),
    )


def _year_export_condition(table, year: str):
    normalized_year = year.strip()
    if not normalized_year:
        return None
    return or_(
        table.c.year.startswith(normalized_year),
        table.c.year.startswith(normalized_year[-2:]),
    )


def _activity_export_label(activity_date: date_type | None) -> str:
    return f"{activity_date.isoformat()}导入新增" if activity_date else "总览"


def _iter_all_export_rows(
    repository,
    *,
    activity_date: date_type | None = None,
    year: str | None = None,
) -> Iterator[tuple[str, list[object]]]:
    for brand in repository.product_archive_brands():
        table = repository._table_for_brand(brand)
        conditions = [not_excluded_sku_condition(table.c.sku, table.c.original_sku)]
        if activity_date:
            conditions.append(_activity_date_export_condition(table, activity_date))
        if year:
            year_condition = _year_export_condition(table, year)
            if year_condition is not None:
                conditions.append(year_condition)
        statement = select(*(table.c[column] for column in EXPORT_COLUMNS)).where(*conditions).order_by(desc(table.c.id))
        with repository.engine.connect() as connection:
            items = [dict(row) for row in connection.execute(statement).mappings()]

        # The archive already stores the displayed cost. Only backfill blank
        # costs during export instead of rescanning the large price history for
        # every product in a full overview export.
        items_missing_cost = [item for item in items if item.get("cost") in (None, "")]
        if items_missing_cost:
            apply_jst_product_costs(repository.engine, items_missing_cost)
        for item in items:
            yield brand, [item.get(column) for column in EXPORT_COLUMNS]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _display_width(value: object) -> int:
    width = 0
    for char in _cell_text(value):
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W", "A"} else 1
    return width


def _excel_cell_value(value: object) -> object:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _excel_streaming_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    quoted_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}",
            "Content-Length": str(buf.getbuffer().nbytes),
        },
    )


def _export_all_products(
    request: Request,
    repository,
    *,
    activity_date: date_type | None = None,
    year: str | None = None,
) -> StreamingResponse:
    headers = ["品牌"] + [EXPORT_LABELS.get(c, c) for c in EXPORT_COLUMNS]
    wb = Workbook(write_only=True)
    export_label = _activity_export_label(activity_date)
    ws = wb.create_sheet(title=export_label)
    header_font = Font(name="宋体", size=10, bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    max_width = 42
    min_width = 10
    column_widths = [max(DEFAULT_WIDTH_BY_HEADER.get(header, min_width), _display_width(header) + 2) for header in headers]

    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_cells.append(cell)
    ws.append(header_cells)

    row_count = 1
    for brand, values in _iter_all_export_rows(repository, activity_date=activity_date, year=year):
        row = [BRAND_LABELS.get(brand, brand)] + [_excel_cell_value(value) for value in values]
        # A small sample is enough to keep widths readable without scanning
        # every cell in a 60k+ row overview export.
        if row_count <= 1000:
            for index, value in enumerate(row):
                column_widths[index] = max(column_widths[index], min(_display_width(value) + 2, max_width))
        ws.append(row)
        row_count += 1

    for index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(min_width, min(width, max_width))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    exported_rows = max(0, row_count - 1)
    write_operation_log(
        request,
        module="product",
        action="export",
        entity_type="product_export",
        entity_label=export_label,
        summary=f"导出商品信息档案{export_label}：{exported_rows} 条",
        after_data={
            "brand": "all",
            "brand_label": export_label,
            "activity_date": activity_date.isoformat() if activity_date else None,
            "exported_rows": exported_rows,
            "filename": f"{export_label}商品信息档案.xlsx",
        },
    )
    return _excel_streaming_response(buf, f"{export_label}商品信息档案.xlsx")


def _dict_or_empty(value: object) -> dict[str, object]:
    return value if isinstance(value, dict) else {}


def _first_text(*values: object) -> str:
    for value in values:
        text = _cell_text(value)
        if text:
            return text
    return ""


def _chunk_values(values: set[str]) -> list[list[str]]:
    ordered_values = sorted(values)
    return [
        ordered_values[index:index + LOOKUP_CHUNK_SIZE]
        for index in range(0, len(ordered_values), LOOKUP_CHUNK_SIZE)
    ]


def _parse_id_list(ids: str | None) -> list[int]:
    if not ids:
        return []
    try:
        return [int(i.strip()) for i in ids.split(",") if i.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="无效的商品 ID")


def _validate_product_export_request(repository, brand: str, mode: str | None = None) -> None:
    if mode == SIZE_EXPORT_MODE and brand == "all":
        raise HTTPException(status_code=400, detail="带尺码导出请选择具体品牌")
    if brand != "all" and not repository.is_product_archive_brand(brand):
        raise HTTPException(status_code=400, detail="无效品牌")


def _validate_import_size_group(repository, size_range: object) -> None:
    normalized_size_group = str(size_range or "").strip()
    if not normalized_size_group:
        return
    with repository.engine.connect() as connection:
        exists = connection.execute(
            select(SIZE_GROUPS_TABLE.c.id)
            .where(SIZE_GROUPS_TABLE.c.name == normalized_size_group)
            .limit(1)
        ).scalar()
    if exists is None:
        raise HTTPException(status_code=400, detail=f"尺码组 {normalized_size_group} 不存在或已被删除")


def _load_size_export_source_items(
    repository,
    brand: str,
    ids: str | None,
    *,
    activity_date: date_type | None = None,
    year: str | None = None,
) -> list[dict[str, object]]:
    table = repository._table_for_brand(brand)
    statement = (
        select(table)
        .where(not_excluded_sku_condition(table.c.sku, table.c.original_sku))
    )
    id_list = _parse_id_list(ids)
    if id_list:
        statement = statement.where(table.c.id.in_(id_list))
    if activity_date:
        statement = statement.where(_activity_date_export_condition(table, activity_date))
    if year:
        year_condition = _year_export_condition(table, year)
        if year_condition is not None:
            statement = statement.where(year_condition)

    with repository.engine.connect() as connection:
        return [
            {**dict(row), "_archive_brand": brand}
            for row in connection.execute(statement).mappings()
        ]


def _size_export_source_codes(items: list[dict[str, object]]) -> set[str]:
    selected_codes: set[str] = set()
    for item in items:
        for key in ("sku", "original_sku"):
            code = _cell_text(item.get(key))
            if code:
                selected_codes.add(code)
    return selected_codes


def _load_product_archive_rows(repository, connection, brand: str, codes: set[str]) -> dict[str, dict[str, object]]:
    if not codes:
        return {}
    table = repository._table_for_brand(brand)

    rows_by_code: dict[str, dict[str, object]] = {}
    for chunk in _chunk_values(codes):
        for row in connection.execute(
            select(table)
            .where(or_(table.c.sku.in_(chunk), table.c.original_sku.in_(chunk)))
            .order_by(desc(table.c.updated_at), desc(table.c.id))
        ).mappings():
            item = dict(row)
            for key in ("sku", "original_sku"):
                code = _cell_text(item.get(key))
                if code and code in codes and code not in rows_by_code:
                    rows_by_code[code] = item
    return rows_by_code


def _load_gj_rows_for_codes(
    connection,
    codes: set[str],
    rows_by_code: dict[str, dict[str, object]],
    *,
    source_date_value: object | None = None,
) -> None:
    remaining_codes = set(codes)
    if not remaining_codes:
        return

    columns = [GJ_MERGED_PRODUCT_INFO_TABLE.c[name] for name in GJ_SIZE_EXPORT_COLUMNS]
    for chunk in _chunk_values(remaining_codes):
        statement = (
            select(*columns)
            .where(or_(
                GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_code.in_(chunk),
                GJ_MERGED_PRODUCT_INFO_TABLE.c.original_goods_code.in_(chunk),
            ))
        )
        if source_date_value is not None:
            statement = statement.where(GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value == source_date_value)
        statement = statement.order_by(
            GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value.desc().nulls_last(),
            desc(GJ_MERGED_PRODUCT_INFO_TABLE.c.updated_at),
            desc(GJ_MERGED_PRODUCT_INFO_TABLE.c.id),
        )
        for row in connection.execute(statement).mappings():
            item = dict(row)
            for key in ("goods_code", "original_goods_code"):
                code = _cell_text(item.get(key))
                if code and code in codes and code not in rows_by_code:
                    rows_by_code[code] = item
                    remaining_codes.discard(code)


def _load_gj_rows(connection, codes: set[str]) -> dict[str, dict[str, object]]:
    if not codes:
        return {}

    rows_by_code: dict[str, dict[str, object]] = {}
    latest_source_date = connection.execute(
        select(GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value)
        .where(GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value.is_not(None))
        .order_by(GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value.desc())
        .limit(1)
    ).scalar()
    if latest_source_date is not None:
        _load_gj_rows_for_codes(connection, codes, rows_by_code, source_date_value=latest_source_date)

    missing_codes = {code for code in codes if code not in rows_by_code}
    if missing_codes:
        _load_gj_rows_for_codes(connection, missing_codes, rows_by_code)
    return rows_by_code


def _load_product_profile_rows(connection, codes: set[str]) -> list[dict[str, object]]:
    if not codes:
        return []

    profiles: list[dict[str, object]] = []
    seen_ids: set[object] = set()
    for chunk in _chunk_values(codes):
        statement = (
            select(JST_PRODUCT_PROFILE_TABLE)
            .where(or_(
                JST_PRODUCT_PROFILE_TABLE.c.product_code.in_(chunk),
                JST_PRODUCT_PROFILE_TABLE.c.style_code.in_(chunk),
            ))
            .order_by(JST_PRODUCT_PROFILE_TABLE.c.style_code, JST_PRODUCT_PROFILE_TABLE.c.product_code)
        )
        for row in connection.execute(statement).mappings():
            item = dict(row)
            row_id = item.get("id") or item.get("product_code")
            if row_id in seen_ids:
                continue
            seen_ids.add(row_id)
            profiles.append(item)

    profiles.sort(key=lambda item: (
        _cell_text(item.get("style_code")),
        _cell_text(item.get("product_code")),
    ))
    return profiles


def _parse_size_range_labels(value: object) -> list[dict[str, str]]:
    import re

    text = _cell_text(value)
    if not text:
        return []
    range_match = re.fullmatch(r"(\d+)\s*[-~至]\s*(\d+)", text)
    if range_match:
        start, end = (int(part) for part in range_match.groups())
        if start <= end:
            step = 5 if end >= 100 else 1
            return [{"size_name": str(size), "barcode": str(size)} for size in range(start, end + 1, step)]
    return [
        {"size_name": size, "barcode": size}
        for size in re.split(r"[,，、/／|\s]+", text)
        if size
    ]


def _load_size_group_items(connection, source_items: list[dict[str, object]]) -> dict[str, list[dict[str, str]]]:
    size_group_names = {
        size_range
        for item in source_items
        if (size_range := _cell_text(item.get("size_range")))
    }
    if not size_group_names:
        return {}
    rows = connection.execute(
        select(
            SIZE_GROUPS_TABLE.c.name,
            SIZE_GROUP_ITEMS_TABLE.c.size_name,
            SIZE_GROUP_ITEMS_TABLE.c.barcode,
        )
        .select_from(
            SIZE_GROUPS_TABLE.join(
                SIZE_GROUP_ITEMS_TABLE,
                SIZE_GROUP_ITEMS_TABLE.c.size_group_id == SIZE_GROUPS_TABLE.c.id,
            )
        )
        .where(SIZE_GROUPS_TABLE.c.name.in_(size_group_names))
        .order_by(SIZE_GROUPS_TABLE.c.name, SIZE_GROUP_ITEMS_TABLE.c.sort_order, SIZE_GROUP_ITEMS_TABLE.c.id)
    ).mappings()
    items_by_group: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        size_group = _cell_text(row.get("name"))
        size_name = _cell_text(row.get("size_name"))
        size_barcode = _first_text(row.get("barcode"), size_name)
        if size_group and size_name and size_barcode:
            items_by_group.setdefault(size_group, []).append({
                "size_name": size_name,
                "barcode": size_barcode,
            })
    for size_group in size_group_names:
        if size_group not in items_by_group:
            parsed_items = _parse_size_range_labels(size_group)
            if parsed_items:
                items_by_group[size_group] = parsed_items
    return items_by_group


def _resolve_size_export_color_codes(
    connection,
    source_items: list[dict[str, object]],
) -> dict[object, str]:
    """Resolve only blank archive color codes from the maintained color mapping."""
    color_names_by_source_brand: dict[str, set[str]] = {}
    item_keys_by_brand_and_color: dict[tuple[str, str], list[object]] = {}
    for item in source_items:
        if _cell_text(item.get("color_code")):
            continue
        archive_brand = _cell_text(item.get("_archive_brand"))
        source_brand = PRODUCT_COLOR_BARCODE_SOURCE_BRANDS.get(archive_brand, archive_brand)
        color_name = _cell_text(item.get("color"))
        item_id = item.get("id")
        if not source_brand or not color_name or item_id is None:
            continue
        color_names_by_source_brand.setdefault(source_brand, set()).add(color_name)
        item_keys_by_brand_and_color.setdefault((source_brand, color_name), []).append(item_id)

    if not item_keys_by_brand_and_color:
        return {}

    color_codes_by_brand_and_name: dict[tuple[str, str], set[str]] = {}
    for source_brand, color_names in color_names_by_source_brand.items():
        for chunk in _chunk_values(color_names):
            rows = connection.execute(
                select(COLOR_BARCODE_TABLE.c.color_name, COLOR_BARCODE_TABLE.c.color_barcode)
                .where(COLOR_BARCODE_TABLE.c.brand == source_brand)
                .where(COLOR_BARCODE_TABLE.c.color_name.in_(chunk))
            ).mappings()
            for row in rows:
                color_code = _cell_text(row.get("color_barcode"))
                if not color_code:
                    continue
                for color_name in _color_name_variants(row.get("color_name")):
                    color_codes_by_brand_and_name.setdefault((source_brand, color_name), set()).add(color_code)

    resolved_by_item_id: dict[object, str] = {}
    for key, item_ids in item_keys_by_brand_and_color.items():
        color_codes = color_codes_by_brand_and_name.get(key, set())
        # A color name with multiple codes is intentionally not guessed.
        if len(color_codes) == 1:
            color_code = next(iter(color_codes))
            resolved_by_item_id.update({item_id: color_code for item_id in item_ids})
    return resolved_by_item_id


def _build_size_export_product_code(
    item: dict[str, object],
    size_barcode: str,
    *,
    resolved_color_code: str | None = None,
) -> str:
    return build_product_size_code(
        _first_text(item.get("sku"), item.get("original_sku")),
        _first_text(item.get("color_code"), resolved_color_code),
        size_barcode,
        item.get("barcode_build_rule"),
    )


def _size_export_product_name(style_code: str, color_name: str, product_code: str) -> str:
    name = f"{style_code}{color_name}"
    return name or product_code


def _size_export_profiles_from_size_groups(
    source_items: list[dict[str, object]],
    size_group_items: dict[str, list[dict[str, str]]],
    *,
    resolved_color_codes: dict[object, str] | None = None,
) -> tuple[list[dict[str, object]], set[str]]:
    profiles: list[dict[str, object]] = []
    source_codes_with_size_groups: set[str] = set()
    for item in source_items:
        size_range = _cell_text(item.get("size_range"))
        size_items = size_group_items.get(size_range, [])
        if not size_items:
            continue
        source_codes_with_size_groups.update(
            code
            for code in (_cell_text(item.get("sku")), _cell_text(item.get("original_sku")))
            if code
        )
        style_code = _first_text(item.get("original_sku"), item.get("sku"))
        for size_item in size_items:
            size_barcode = _first_text(size_item.get("barcode"), size_item.get("size_name"))
            product_code = _build_size_export_product_code(
                item,
                size_barcode,
                resolved_color_code=(resolved_color_codes or {}).get(item.get("id")),
            )
            if not product_code:
                continue
            profiles.append({
                "id": f"archive-{item.get('id')}-{size_item['size_name']}",
                "product_code": product_code,
                "style_code": style_code,
                "color_name": _cell_text(item.get("color")),
                "size_barcode": size_barcode,
                "raw_payload": _dict_or_empty(item.get("raw_payload")),
            })
    return profiles, source_codes_with_size_groups


def _size_export_fallback_profiles(
    source_items: list[dict[str, object]],
    profiles: list[dict[str, object]],
) -> list[dict[str, object]]:
    profile_codes = {
        code
        for profile in profiles
        for code in (_cell_text(profile.get("product_code")), _cell_text(profile.get("style_code")))
        if code
    }
    fallback_profiles: list[dict[str, object]] = []
    for item in source_items:
        source_codes = {
            code
            for code in (_cell_text(item.get("sku")), _cell_text(item.get("original_sku")))
            if code
        }
        if source_codes & profile_codes:
            continue
        product_code = _first_text(item.get("sku"), item.get("original_sku"))
        style_code = _first_text(item.get("original_sku"), item.get("sku"))
        if not product_code and not style_code:
            continue
        fallback_profiles.append({
            "id": f"archive-{item.get('id')}",
            "product_code": product_code,
            "style_code": style_code,
            "color_name": _cell_text(item.get("color")),
            "size_barcode": "",
            "raw_payload": _dict_or_empty(item.get("raw_payload")),
        })
    return fallback_profiles


def _write_only_header_cells(worksheet, headers: list[str]) -> list[WriteOnlyCell]:
    header_font = Font(name="宋体", size=10, bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cells = []
    for header in headers:
        cell = WriteOnlyCell(worksheet, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cells.append(cell)
    return cells


def _size_export_style_context(
    style_code: str,
    product_code: str,
    archive_rows: dict[str, dict[str, object]],
    gj_rows: dict[str, dict[str, object]],
) -> dict[str, str]:
    archive = archive_rows.get(style_code) or archive_rows.get(product_code) or {}
    gj = gj_rows.get(style_code) or gj_rows.get(product_code) or {}
    archive_extra = _dict_or_empty(archive.get("extra_fields"))
    return {
        "product_name": _first_text(gj.get("goods_full_name"), archive.get("product_name")),
        "category": _first_text(archive.get("group_name")),
        "logo": _first_text(gj.get("brand")),
        "upper_material": _first_text(gj.get("upper_material"), archive.get("upper_material")),
        "product_item_name": _first_text(archive.get("product_name"), archive_extra.get("品名"), gj.get("product_name")),
        "execution_standard": _first_text(gj.get("execution_standard"), archive.get("execution_standard")),
        "product_model": _first_text(archive.get("product_model")),
        "lining_material": _first_text(gj.get("lining_material"), archive.get("lining_material")),
        "outsole_material": _first_text(gj.get("outsole_material"), archive.get("outsole_material")),
        "insole_material": _first_text(gj.get("insole_material"), archive.get("insole_material")),
        "original_sku": _first_text(gj.get("original_goods_code"), archive.get("original_sku")),
        "factory_code": _first_text(archive.get("factory_sku"), gj.get("factory_code")),
        "supplier_name": _first_text(archive.get("supplier_name")),
        "brand": _first_text(gj.get("brand")),
        "cost": _first_text(archive.get("cost")),
    }


def _export_products_with_sizes(
    request: Request,
    repository,
    brand: str,
    ids: str | None,
    *,
    activity_date: date_type | None = None,
    year: str | None = None,
) -> StreamingResponse:
    _validate_product_export_request(repository, brand, SIZE_EXPORT_MODE)
    source_items = _load_size_export_source_items(
        repository,
        brand,
        None if activity_date else ids,
        activity_date=activity_date,
        year=year,
    )
    if ids and not source_items:
        raise HTTPException(status_code=404, detail="未找到可导出的选中商品")
    selected_codes = _size_export_source_codes(source_items)

    wb = Workbook(write_only=True)
    brand_label = BRAND_LABELS.get(brand, brand)
    ws = wb.create_sheet(title=f"{brand_label}带尺码")
    column_widths = [
        max(
            SIZE_EXPORT_WIDTH_BY_HEADER.get(header, DEFAULT_WIDTH_BY_HEADER.get(header, SIZE_EXPORT_MIN_WIDTH)),
            _display_width(header) + 2,
        )
        for header in SIZE_EXPORT_HEADERS
    ]
    for index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(SIZE_EXPORT_MIN_WIDTH, min(width, SIZE_EXPORT_MAX_WIDTH))
    ws.freeze_panes = "A2"
    ws.append(_write_only_header_cells(ws, SIZE_EXPORT_HEADERS))

    with repository.engine.connect() as connection:
        profiles = _load_product_profile_rows(connection, selected_codes)
        size_group_items = _load_size_group_items(connection, source_items)
        resolved_color_codes = _resolve_size_export_color_codes(connection, source_items)
        size_group_profiles, source_codes_with_size_groups = _size_export_profiles_from_size_groups(
            source_items,
            size_group_items,
            resolved_color_codes=resolved_color_codes,
        )
        profiles = [
            profile
            for profile in profiles
            if not {
                _cell_text(profile.get("product_code")),
                _cell_text(profile.get("style_code")),
            } & source_codes_with_size_groups
        ]
        source_items_without_size_groups = [
            item
            for item in source_items
            if not {
                _cell_text(item.get("sku")),
                _cell_text(item.get("original_sku")),
            } & source_codes_with_size_groups
        ]
        fallback_profiles = _size_export_fallback_profiles(source_items_without_size_groups, profiles)
        export_profiles = [*size_group_profiles, *profiles, *fallback_profiles]

        profile_style_codes = {
            code
            for profile in export_profiles
            for code in (_cell_text(profile.get("style_code")),)
            if code
        }
        lookup_codes = set(profile_style_codes)
        lookup_codes.update(selected_codes)
        loaded_archive_rows = _load_product_archive_rows(repository, connection, brand, lookup_codes)
        apply_jst_product_costs(repository.engine, list(loaded_archive_rows.values()))
        archive_rows = dict(loaded_archive_rows)
        gj_rows = _load_gj_rows(connection, lookup_codes)

    row_count = 1
    style_contexts: dict[str, dict[str, str]] = {}
    for profile in export_profiles:
        raw_payload = _dict_or_empty(profile.get("raw_payload"))
        product_code = _cell_text(profile.get("product_code"))
        style_code = _cell_text(profile.get("style_code"))
        color_name = _cell_text(profile.get("color_name"))
        size_barcode = _cell_text(profile.get("size_barcode"))
        context_key = style_code or product_code
        context = style_contexts.get(context_key)
        if context is None:
            context = _size_export_style_context(style_code, product_code, archive_rows, gj_rows)
            style_contexts[context_key] = context

        product_name = _size_export_product_name(style_code, color_name, product_code)
        category = "男鞋" if brand == "cbanner_mens" else _first_text(context["category"], raw_payload.get("分类"))
        logo = _first_text(context["logo"], raw_payload.get("LOGO"), raw_payload.get("品牌"))
        row = [
            _first_text(context["supplier_name"], raw_payload.get("供应商名"), raw_payload.get("供应商")),
            product_code,
            style_code,
            product_name,
            color_name,
            size_barcode,
            _first_text(context["upper_material"], raw_payload.get("鞋面材质")),
            _first_text(context["product_item_name"], raw_payload.get("品名")),
            _first_text(context["execution_standard"], raw_payload.get("执行标准")),
            _first_text(raw_payload.get("产品型号"), context["product_model"]),
            _first_text(context["lining_material"], raw_payload.get("内里材质")),
            _first_text(context["outsole_material"], raw_payload.get("大底材质")),
            _first_text(context["insole_material"], raw_payload.get("鞋垫材质")),
            _first_text(context["original_sku"], raw_payload.get("原始货号")),
            _first_text(raw_payload.get("供应商商品款号"), context["factory_code"]),
            _first_text(raw_payload.get("品牌"), context["brand"]),
            f"{color_name};{size_barcode}" if color_name or size_barcode else "",
            category,
            _first_text(context["cost"], raw_payload.get("成本价"), raw_payload.get("成本")),
            logo,
        ]
        row = [_excel_cell_value(value) for value in row]
        ws.append(row)
        row_count += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(SIZE_EXPORT_HEADERS))}{row_count}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    export_label = f"{brand_label}{_activity_export_label(activity_date) if activity_date else ''}带尺码"
    raw_filename = f"{export_label}商品档案.xlsx"
    write_operation_log(
        request,
        module="product",
        action="export",
        entity_type="product_export",
        entity_label=export_label,
        summary=f"导出商品信息档案带尺码：{export_label}，{len(export_profiles)} 条",
        after_data={
            "brand": brand,
            "brand_label": brand_label,
            "mode": SIZE_EXPORT_MODE,
            "ids": None if activity_date else ids,
            "activity_date": activity_date.isoformat() if activity_date else None,
            "exported_rows": len(export_profiles),
            "size_profile_rows": len(size_group_profiles),
            "fallback_rows": len(fallback_profiles),
            "filename": raw_filename,
        },
    )
    return _excel_streaming_response(buf, raw_filename)


@router.get("/export")
def export_products(
    request: Request,
    brand: str = Query(...),
    ids: str | None = Query(None),
    mode: str | None = Query(None),
    activity_date: date_type | None = Query(None),
    year: str | None = Query(None),
    today_only: bool = Query(False),
):
    repository = request.app.state.repository
    _validate_product_export_request(repository, brand, mode)
    if request.method == "HEAD":
        return Response(status_code=200)

    export_date = activity_date or (datetime.now(SHANGHAI_TIME_ZONE).date() if today_only else None)
    export_year = year.strip() if year and not export_date else None

    if mode == SIZE_EXPORT_MODE:
        return _export_products_with_sizes(
            request,
            repository,
            brand,
            ids,
            activity_date=export_date,
            year=export_year,
        )

    if brand == "all":
        return _export_all_products(request, repository, activity_date=export_date, year=export_year)

    if export_date:
        table = repository._table_for_brand(brand)
        with repository.engine.connect() as connection:
            items = [
                dict(row)
                for row in connection.execute(
                    select(table)
                    .where(not_excluded_sku_condition(table.c.sku, table.c.original_sku))
                    .where(_activity_date_export_condition(table, export_date))
                    .order_by(desc(table.c.id))
                ).mappings()
            ]
        apply_jst_product_costs(repository.engine, items)
    elif ids:
        id_list = [int(i.strip()) for i in ids.split(",") if i.strip()]
        items = repository.get_products_by_ids(brand, id_list)
    else:
        table = repository.list_products(brand, query=None, page=1, page_size=1_000_000, year=export_year)
        items = table["items"]

    wb = Workbook()
    ws = wb.active
    ws.title = BRAND_LABELS.get(brand, brand)

    headers = [EXPORT_LABELS.get(c, c) for c in EXPORT_COLUMNS]
    ws.append(headers)

    for item in items:
        row = [_excel_cell_value(item.get(c)) for c in EXPORT_COLUMNS]
        ws.append(row)

    style_excel_worksheet(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    brand_label = BRAND_LABELS.get(brand, brand)
    export_label = f"{brand_label}{_activity_export_label(export_date) if export_date else ''}"
    raw_filename = f"{export_label}.xlsx"
    write_operation_log(
        request,
        module="product",
        action="export",
        entity_type="product_export",
        entity_label=export_label,
        summary=f"导出商品信息档案：{export_label}，{len(items)} 条",
        after_data={
            "brand": brand,
            "brand_label": brand_label,
            "ids": None if export_date else ids,
            "activity_date": export_date.isoformat() if export_date else None,
            "exported_rows": len(items),
            "filename": raw_filename,
        },
    )
    return _excel_streaming_response(buf, raw_filename)


@router.head("/export")
def check_export_products(
    request: Request,
    brand: str = Query(...),
    mode: str | None = Query(None),
):
    repository = request.app.state.repository
    _validate_product_export_request(repository, brand, mode)
    return Response(status_code=200)


def _finish_product_import(
    request: Request,
    *,
    brand: str,
    filename: str | None,
    created: int,
    updated: int,
    imported_skus: list[str],
) -> None:
    try:
        clear_fine_table_cache()
        clear_product_goods_cache()
    except Exception:
        logger.exception("Failed to clear product caches after import")

    try:
        write_operation_log(
            request,
            module="product",
            action="import",
            entity_type="product_import",
            entity_label=filename or "商品档案导入",
            summary=f"导入商品档案：新增 {created} 条，更新 {updated} 条",
            after_data={
                "brand": brand,
                "filename": filename,
                "created": created,
                "updated": updated,
                "skus": imported_skus[:500],
                "sku_count": len(imported_skus),
            },
        )
    except Exception:
        logger.exception("Failed to write product import operation log")


@router.post("/import")
async def import_products(
    request: Request,
    background_tasks: BackgroundTasks,
    brand: str = Query(...),
    file: UploadFile = None,
):
    repository = request.app.state.repository
    if not repository.is_product_archive_brand(brand):
        raise HTTPException(status_code=400, detail="无效品牌")
    if file is None:
        raise HTTPException(status_code=400, detail="No file uploaded")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        wb.close()
        raise HTTPException(status_code=400, detail="Empty file")

    from transform.rows import normalize_header

    headers = [normalize_header(h) for h in header_row]

    reverse_aliases = {}
    for cn_label, en_field in CN_TO_FIELD.items():
        reverse_aliases[cn_label] = en_field
        reverse_aliases[en_field] = en_field

    repository = request.app.state.repository
    image_matcher = get_image_matcher(request, brand)
    created = 0
    updated = 0
    imported_skus: list[str] = []
    imported_product_ids: list[int] = []

    with repository.engine.begin() as connection:
        for row_number, row in enumerate(iterator, start=2):
            try:
                row_dict = {}
                for idx, cell_value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = cell_value

                payload = {}
                extra_fields = {}
                known_fields = set(CN_TO_FIELD.values()) | set(CN_TO_FIELD.keys())
                for key, value in row_dict.items():
                    field = reverse_aliases.get(key)
                    if field:
                        payload[field] = value
                    elif key and key not in known_fields:
                        if key in EXCLUDED_EXTRA_FIELD_KEYS:
                            continue
                        normalized = normalize_admin_field(key, value)
                        if normalized is not None and str(normalized).strip():
                            extra_fields[key] = normalized

                raw_sku = payload.get("sku")
                if raw_sku is not None:
                    payload["sku"] = str(int(raw_sku)) if isinstance(raw_sku, float) and raw_sku.is_integer() else str(raw_sku).strip()

                raw_orig = payload.get("original_sku")
                if raw_orig is not None:
                    payload["original_sku"] = str(int(raw_orig)) if isinstance(raw_orig, float) and raw_orig.is_integer() else str(raw_orig).strip()

                if not payload.get("original_sku") and not payload.get("sku"):
                    continue
                if payload.get("original_sku") and not payload.get("sku"):
                    payload["sku"] = payload["original_sku"]
                if is_excluded_sku(payload.get("sku"), payload.get("original_sku")):
                    continue
                if extra_fields:
                    payload["extra_fields"] = extra_fields

                sku_val = str(payload.get("sku", "") or "").strip()
                original_sku_val = str(payload.get("original_sku", "") or "").strip()
                payload = dict(apply_product_defaults(brand, payload))
                barcode_build_rule = normalize_admin_field(
                    "barcode_build_rule",
                    payload.get("barcode_build_rule"),
                )
                if barcode_build_rule is None or not str(barcode_build_rule).strip():
                    display_sku = sku_val or original_sku_val or "未填写货号"
                    raise HTTPException(
                        status_code=400,
                        detail=f"第 {row_number} 行导入失败：货号 {display_sku} 未填写条码构成逻辑",
                    )
                existing = repository.find_by_sku(brand, sku_val, connection=connection) if sku_val else None
                if existing is None and original_sku_val:
                    existing = repository.find_by_original_sku(brand, original_sku_val, connection=connection)
                if sku_val:
                    imported_skus.append(sku_val)

                incoming_supplier = normalize_admin_field(
                    "supplier_name",
                    payload.get("supplier_name"),
                )
                if existing is not None and incoming_supplier:
                    existing_supplier = normalize_admin_field(
                        "supplier_name",
                        existing.get("supplier_name"),
                    )
                    if existing_supplier and existing_supplier != incoming_supplier:
                        display_sku = sku_val or original_sku_val or "未填写货号"
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"第 {row_number} 行导入失败：货号 {display_sku} 的供应商不一致，"
                                f"现有供应商为“{existing_supplier}”，导入供应商为“{incoming_supplier}”，"
                                "请修正后再导入"
                            ),
                        )

                import_fields = {}
                for key, value in payload.items():
                    if key in ("sku", "extra_fields"):
                        continue
                    normalized = normalize_admin_field(key, value)
                    if normalized is not None and str(normalized).strip():
                        import_fields[key] = normalized

                if image_matcher and not import_fields.get("image_path"):
                    found_path = image_matcher.find(original_sku_val) if original_sku_val else None
                    if not found_path and sku_val:
                        found_path = image_matcher.find(sku_val)
                    if found_path:
                        import_fields["image_path"] = found_path

                if existing is not None:
                    merged = {key: value for key, value in existing.items() if value is not None}
                    merged.update(import_fields)
                    existing_extra = filter_extra_fields(existing.get("extra_fields")) or {}
                    new_extra = filter_extra_fields(payload.get("extra_fields")) or {}
                    if existing_extra or new_extra:
                        merged["extra_fields"] = {**existing_extra, **new_extra}
                    record = build_admin_record(brand, merged, existing_metadata={
                        "source_workbook": existing["source_workbook"],
                        "source_sheet": existing["source_sheet"],
                        "source_row_number": existing["source_row_number"],
                    })
                    _validate_import_size_group(repository, record.get("size_range"))
                    saved_item = repository.update_product(brand, existing["id"], record, connection=connection)
                    if saved_item is not None:
                        imported_product_ids.append(int(saved_item["id"]))
                    updated += 1
                else:
                    record = build_admin_record(brand, payload)
                    _validate_import_size_group(repository, record.get("size_range"))
                    saved_item = repository.create_product(brand, record, connection=connection)
                    imported_product_ids.append(int(saved_item["id"]))
                    created += 1
            except HTTPException as error:
                raise HTTPException(status_code=error.status_code, detail=f"第 {row_number} 行导入失败：{error.detail}") from error
            except Exception as error:
                raise HTTPException(status_code=400, detail=f"第 {row_number} 行导入失败：{error}") from error

        repository.mark_products_imported(brand, imported_product_ids, connection=connection)

    try:
        wb.close()
    except Exception:
        logger.exception("Failed to close product import workbook")

    background_tasks.add_task(
        _finish_product_import,
        request,
        brand=brand,
        filename=file.filename,
        created=created,
        updated=updated,
        imported_skus=imported_skus,
    )
    return {
        "created": created,
        "updated": updated,
        "skus": imported_skus,
        "message": f"导入完成：新增 {created} 条，更新 {updated} 条",
    }
