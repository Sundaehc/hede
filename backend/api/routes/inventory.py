from __future__ import annotations

import io
import urllib.parse
from collections import defaultdict
from decimal import Decimal
from datetime import date, datetime, timedelta

import xlrd
from fastapi import APIRouter, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import desc, or_, select as sa_select

from api.excel_export import style_excel_workbook, style_excel_worksheet
from api.operation_log_utils import (
    DETAIL_FIELD_LABELS,
    GENERAL_CUSTOMER_BRAND_FIELD_LABELS,
    GENERAL_CUSTOMER_SHOP_FIELD_LABELS,
    INVENTORY_FIELD_LABELS,
    build_changed_fields,
    detail_entity_label,
    inventory_entity_label,
    inventory_module_for_record,
    summarize_changes,
    write_operation_log,
)
from domain.color_barcode_schema import COLOR_BARCODE_TABLE
from domain.gj_brand import CBANNER_MENS_BRAND, CBANNER_WOMENS_BRAND, EBLAN_BRAND, NI_BRAND, SMILEY_BRAND, SUPPLIER_BRANDS, YANDOU_BRAND, infer_supplier_brand_from_name
from domain.gj_schema import GJ_MERGED_PRODUCT_INFO_TABLE
from domain.inventory_schema import SUPPLIER_TABLE
from domain.smiley_schema import SMILEY_FINE_TABLE
from domain.inventory_sources import (
    ACCOUNTING_DOCUMENT_TYPES,
    DOCUMENT_TYPES,
    normalize_document_type,
    INVENTORY_CANONICAL_COLUMNS,
    INVENTORY_COLUMN_ALIASES,
    INVENTORY_DETAIL_ALIASES,
    INVENTORY_DETAIL_COLUMNS,
    INVENTORY_EXPORT_LABELS,
)
from domain.schema import PRODUCT_TABLES
from domain.vip_schema import JST_PRICE_TABLE

router = APIRouter()

CN_TO_FIELD = {cn: en for cn, en in INVENTORY_COLUMN_ALIASES.items() if en in INVENTORY_CANONICAL_COLUMNS}
DETAIL_CN_TO_FIELD = {cn: en for cn, en in INVENTORY_DETAIL_ALIASES.items() if en in INVENTORY_DETAIL_COLUMNS}

EXCEL_EPOCH = datetime(1899, 12, 30)
PURCHASE_IMPORT_TYPES = {"进货订单", "进货单", "进货退货单", "报溢单", "报损单", "批发销售单", "批发销售退货单", "同价调拨单"}
EU_PURCHASE_SIZE_LABELS = ("35", "36", "37", "38", "39", "40", "41", "42", "43", "44")
MILLIMETER_PURCHASE_SIZE_LABELS = ("220", "225", "230", "235", "240", "245", "250", "255", "260", "265", "270", "275", "280", "285")
PURCHASE_EXPORT_SIZE_LABELS = (*MILLIMETER_PURCHASE_SIZE_LABELS, *EU_PURCHASE_SIZE_LABELS)
EU_SIZE_BRANDS = {"smiley", "ni", "nike"}
PURCHASE_SIZE_CODE_MAPS = {
    "cbanner_mens": {
        "01": "38",
        "02": "39",
        "03": "40",
        "04": "41",
        "05": "42",
        "06": "43",
        "07": "44",
    },
    "cbanner_womens": {
        "01": "35",
        "02": "36",
        "03": "37",
        "04": "38",
        "05": "39",
        "06": "40",
        "07": "41",
        "08": "42",
    },
}
PURCHASE_MILLIMETER_SIZE_MAP = {
    "220": "34",
    "225": "35",
    "230": "36",
    "235": "37",
    "240": "38",
    "245": "39",
    "250": "40",
    "255": "41",
    "260": "42",
    "265": "43",
    "270": "44",
    "275": "45",
    "280": "46",
    "285": "47",
}
PURCHASE_EU_TO_MILLIMETER_SIZE_MAP = {
    eu_size: millimeter_size
    for millimeter_size, eu_size in PURCHASE_MILLIMETER_SIZE_MAP.items()
}
PURCHASE_DETAIL_EXTRA_FIELDS = {
    "image_code": "图片",
    "factory_code": "工厂货号",
    "style_code": "款号（鞋内丝印）",
    "inner_color_code": "色号（鞋内丝印）",
    "upper_material": "鞋面材质",
    "lining_material": "内里材质",
    "outsole_material": "大底材质",
    "insole_material": "鞋垫材质",
    "shoe_box_spec": "鞋盒规格",
}
PURCHASE_IMPORT_DOC_FIELD_ALIASES = {
    "supplier": {"供货单位", "供应商", "单位全名", "单位名称"},
    "summary": {"摘要"},
    "date": {"采购日期", "订货日期", "单据日期", "日期"},
    "delivery_date": {"协议到货日期", "交货日期", "到货日期", "要求交货日期"},
    "warehouse": {"收货仓库", "仓库", "入货仓库"},
    "handler": {"经办人", "经手人"},
}
PURCHASE_ORDER_IMPORT_REQUIRED_DOC_FIELDS = ("supplier", "summary", "date", "delivery_date", "warehouse", "handler")
PURCHASE_ORDER_IMPORT_FIELD_LABELS = {
    "supplier": "供货单位",
    "summary": "摘要",
    "date": "采购日期",
    "delivery_date": "协议到货日期",
    "warehouse": "收货仓库",
    "handler": "经办人",
}
PURCHASE_ORDER_IMPORT_CODE_HEADERS = {"商品编码", "货品编码", "商品编号", "商品货号", "货号"}
PURCHASE_DETAIL_REMARK_HEADERS = {"商品备注", "采购单备注", "备注"}
PURCHASE_DETAIL_REMARK_LIMIT = 20
PURCHASE_PRODUCTION_ORDER_EXPORT_MODE = "production_order"
PURCHASE_EXPORT_MODES = {"summary", "size_rows", PURCHASE_PRODUCTION_ORDER_EXPORT_MODE}
PURCHASE_SUMMARY_EXPORT_HEADERS = [
    "行号",
    "单据类型",
    "货号",
    "原始货号",
    "商品全名",
    "单据编号",
    "摘要",
    "录单日期",
    "到货日期",
    "仓库全名",
    "单位编号",
    "单位全名",
    "职员全名",
    "订货数量",
    "订货金额",
    "到货数量",
    "未到货数量",
    "完成率%",
]
PURCHASE_SIZE_ROW_EXPORT_HEADERS = [
    "供应商",
    "商品编码",
    "数量",
    "采购单备注",
    "采购日期",
    "协议到货日期",
    "仓库全名",
    "商品全名",
    "颜色条码",
    "颜色名称",
    "尺码条码",
    "尺码名称",
    "条码",
    "单据类型",
    "单据编号",
    "商品备注",
    "单位全名",
    "单位编号",
    "职员全名",
    "订货金额",
    "到货数量",
    "未到货数量",
    "完成率%",
]
PURCHASE_PRODUCTION_FIXED_HEADERS = [
    "商品编号",
    "图片",
    "工厂货号",
    "款号\n（鞋内丝印）",
    "色号\n（鞋内丝印）",
    "颜色名称",
    "鞋面材质",
    "内里材质",
    "大底材质",
    "鞋垫材质",
    "鞋盒规格",
]
PURCHASE_PRODUCTION_TAIL_HEADERS = [
    "数量",
    "单价",
    "金额",
    "要求交货日期",
    "工厂回复交货日期",
]
CBANNER_PURCHASE_ORDER_REQUIREMENTS = "\n".join([
    "1、用皮标准、检验标准、鞋垫做法及颜色搭配必须按我司要求及确认样品的做法；",
    "2、鞋面LOGO压标：分左右脚，按规格压印（另附文件）；大底LOGO激光按我公司要求；",
    "3、所有须用到中底的产品必须使用千百度品牌专用中底，鞋撑必须用千百度品牌专用纸包裹；",
    "4、鞋内里打印货号、尺码、鞋型号（男鞋：2.5、女鞋：1.5）；",
    "5、装箱要求：鞋盒条码贴纸统一由我司提供，贴纸位置按照我司要求；外箱箱唛统一贴外箱左上角的贴标签处；",
    "送货清单打印标签货号、颜色、尺码、数量等明细；",
    "6、工厂接单后，48小时内必须向我司回复准确交货日期进度表，最后交货期不得延误，如有特殊原因不能按时出货请及时和我司采购部联系；",
    "7、鞋盒由浙江芙蓉有限公司提供，手机:13587732200 微信同号",
    "8、中底，防霉片、合格证、包装纸、鞋撑包纸由工厂直接向永嘉润森辅料经营部购买   电话：13857738780，18858880661  林建灯",
    "9、电商总部电话：0577-88267616",
    "10、千百度专用鞋垫（吉利鞋材 传真：67385799.胡邦平 联系电话：13777708608）",
    "11、仓库收货地址：浙江省温州市瓯海区仙岩街道凤康路1515号中胤儿童时尚智造产业园4幢2楼    联系人：邹锡明 15858528246",
    "12、条码领取地址：浙江省 温州市 瓯海区 娄桥街道 商汇路888号广川大厦5楼502",
    "13、商品售卖中因质量问题被平台禁售，此商品我司将有权全部退还给供应商，鞋底,帮面不能有侵权出现,否则后果自负！",
    "14、鞋内发现有钉子，罚款2000元/双；如因此被平台从重处罚的，需贵司额外承担费用！",
    "15、商标不符，罚款1000元/双；",
])
CBANNER_WOMENS_PURCHASE_ORDER_REQUIREMENTS = "\n".join([
    "1、供货单位保证我司采购的产品外观、功能等不得有侵权行为，如有侵权则由供货单位承担侵权行为产生的全部损失以及法律责任。",
    "2、用皮标准、检验标准、鞋垫做法及颜色搭配必须按我司要求及确认样品的做法；",
    "3、鞋面LOGO压标：分左右脚，按规格压印（另附文件）；大底LOGO激光按我公司要求；",
    "4、鞋内里打印款号、色号、尺码、鞋型号（女鞋：1.5）；",
    "5、装箱要求：鞋盒条码贴纸统一由我司提供，贴纸位置按照我司要求；外箱箱唛统一贴外箱左上角的贴标签处；送货清单打印标签货号、颜色、尺码、数量等明细；",
    "6、工厂接单后，48小时内必须向我司回复准确交货日期进度表，最后交货期不得延误，如有特殊原因不能按时出货请及时和我司商品计划部联系；",
    "7、千百度确认样收件地址：浙江省温州市瓯海区娄桥街道商汇路888号广川大厦5楼502 钱月辉18166487530 （收到确认样后发出鞋盒条码）",
    "8、鞋盒由浙江芙蓉有限公司提供，联系人：陈小姐   电话：15167781102",
    "9、专用鞋撑、包纸、中底布、合格证、防霉片：永嘉润森辅料经营部：地址：五星工业区裕隆产业园9幢1楼109号 联系人：林建灯   电话：  13857738780    18858880661",
    "10、仓库收货地址：浙江省温州市瓯海区仙岩街道凤康路1515号中胤儿童时尚智造产业园4幢2楼   联系人：  邹锡明 15858528246",
    "11、千百度烫底联系方式：千百度指定烫底恒硕鞋材，联系方式：程学兵15858514249 程雪妹13736739865",
])
PURCHASE_ORDER_REQUIREMENTS_BY_BRAND = {
    CBANNER_MENS_BRAND: CBANNER_PURCHASE_ORDER_REQUIREMENTS,
    EBLAN_BRAND: CBANNER_PURCHASE_ORDER_REQUIREMENTS,
    YANDOU_BRAND: CBANNER_PURCHASE_ORDER_REQUIREMENTS,
    CBANNER_WOMENS_BRAND: CBANNER_WOMENS_PURCHASE_ORDER_REQUIREMENTS,
}
PURCHASE_ORDER_REQUIREMENT_BRANDS = (
    CBANNER_MENS_BRAND,
    CBANNER_WOMENS_BRAND,
    YANDOU_BRAND,
    EBLAN_BRAND,
    SMILEY_BRAND,
    NI_BRAND,
)
PURCHASE_ORDER_REQUIREMENT_BRAND_LABELS = {
    CBANNER_MENS_BRAND: "千百度男鞋",
    CBANNER_WOMENS_BRAND: "千百度女鞋",
    YANDOU_BRAND: "烟斗",
    EBLAN_BRAND: "伊伴",
    SMILEY_BRAND: "笑脸",
    NI_BRAND: "NI",
}


def _normalize_date(value: object) -> str | None:
    """Convert Excel serial date number to YYYY-MM-DD string. Passes through non-numeric values unchanged."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if not value:
        return value
    try:
        serial = float(value)
        # Only treat as Excel serial if it looks like a date number (between 1 and ~100000 days from epoch)
        if 1 <= serial <= 100000:
            return (EXCEL_EPOCH + timedelta(days=int(serial))).strftime("%Y-%m-%d")
    except (ValueError, OverflowError):
        pass
    text = _cell_text(value)
    date_part = text.split()[0] if text else ""
    for separator in ("-", "/", "."):
        parts = date_part.split(separator)
        if len(parts) != 3 or not all(part.isdigit() for part in parts):
            continue
        year, month, day = parts
        year_number = int(year)
        if year_number < 100:
            year_number += 2000
        try:
            return f"{year_number:04d}-{int(month):02d}-{int(day):02d}"
        except ValueError:
            break
    return text


def _today_text() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _to_decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).strip())
    except Exception:
        return Decimal("0")


def _fmt_decimal(value: Decimal) -> str:
    normalized = value.normalize()
    return str(normalized) if normalized.as_tuple().exponent < 0 else str(int(normalized))


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _first_text(*values: object) -> str:
    for value in values:
        text = _cell_text(value)
        if text:
            return text
    return ""


def _dict_or_empty(value: object) -> dict[str, object]:
    return value if isinstance(value, dict) else {}


def _fmt_export_decimal(value: Decimal, *, blank_zero: bool = False) -> str:
    if blank_zero and value == 0:
        return ""
    return _fmt_decimal(value)


def _fmt_completion_rate(arrived_quantity: Decimal, ordered_quantity: Decimal) -> str:
    if ordered_quantity == 0:
        return ""
    try:
        return _fmt_decimal((arrived_quantity * Decimal("100") / ordered_quantity).quantize(Decimal("0.01")))
    except Exception:
        return ""


def _purchase_export_size_labels(details: list[dict[str, object]]) -> tuple[str, ...]:
    extra_sizes: set[str] = set()
    for detail in details:
        size_quantities = _dict_or_empty(detail.get("size_quantities"))
        for size in size_quantities:
            size_text = _cell_text(size)
            if size_text and size_text not in PURCHASE_EXPORT_SIZE_LABELS:
                extra_sizes.add(size_text)
    return (*PURCHASE_EXPORT_SIZE_LABELS, *sorted(extra_sizes))


def _purchase_detail_quantity(detail: dict[str, object], size_quantities: dict[str, object]) -> Decimal:
    quantity = _to_decimal(detail.get("quantity"))
    if quantity == 0 and size_quantities:
        quantity = sum((_to_decimal(value) for value in size_quantities.values()), Decimal("0"))
    return quantity


def _purchase_detail_amount(detail: dict[str, object], quantity: Decimal) -> Decimal:
    amount = _to_decimal(detail.get("amount"))
    if amount == 0 and quantity != 0:
        unit_price = _to_decimal(detail.get("unit_price"))
        if unit_price != 0:
            amount = quantity * unit_price
    return amount


def _purchase_prorated_decimal(total: Decimal, part: Decimal, base: Decimal) -> Decimal:
    if total == 0 or part == 0 or base == 0:
        return Decimal("0")
    try:
        return (total * part / base).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def _purchase_color_barcode(product_code: str, brand: str | None = None) -> str:
    if not product_code:
        return ""
    if str(brand or "").strip().lower() == "smiley":
        return product_code[-4:] if len(product_code) >= 4 else product_code
    return product_code[-2:] if len(product_code) >= 2 else product_code


def _purchase_export_barcode(product_code: str, color_barcode: str, size: str, brand: str) -> str:
    if not product_code:
        return ""
    if str(brand or "").strip().lower() in EU_SIZE_BRANDS:
        return f"{product_code}{size}"
    return f"{product_code}{color_barcode}{size}"


def _purchase_record_brand(record: dict[str, object]) -> str:
    raw_payload = _dict_or_empty(record.get("raw_payload"))
    brand = _cell_text(raw_payload.get("brand")).lower()
    if brand in EU_SIZE_BRANDS:
        return brand

    supplier_name = _cell_text(record.get("supplier"))
    upper_name = supplier_name.upper()
    if "NI" in upper_name:
        return "ni"
    if "笑脸" in supplier_name:
        return "smiley"
    return brand


def _purchase_detail_arrival_text(extra_fields: dict[str, object]) -> str:
    return _first_text(
        extra_fields.get("arrival_quantity"),
        extra_fields.get("arrived_quantity"),
        extra_fields.get("received_quantity"),
        extra_fields.get("到货数量"),
    )


def _purchase_detail_unreceived_text(extra_fields: dict[str, object]) -> str:
    return _first_text(
        extra_fields.get("unreceived_quantity"),
        extra_fields.get("pending_quantity"),
        extra_fields.get("undelivered_quantity"),
        extra_fields.get("未到货数量"),
    )


def _purchase_detail_completion_text(extra_fields: dict[str, object]) -> str:
    return _first_text(
        extra_fields.get("completion_rate"),
        extra_fields.get("complete_rate"),
        extra_fields.get("完成率"),
        extra_fields.get("完成率%"),
    )


def _validate_purchase_detail_remark(record: dict[str, object], payload: dict[str, object]) -> None:
    if _cell_text(record.get("document_type")) != "进货订单":
        return
    remark = _cell_text(payload.get("remark"))
    if len(remark) > PURCHASE_DETAIL_REMARK_LIMIT:
        raise HTTPException(status_code=400, detail=f"商品备注最多 {PURCHASE_DETAIL_REMARK_LIMIT} 个字")
    payload["remark"] = remark


def _log_record_operation(
    request: Request,
    *,
    action: str,
    prefix: str,
    before: dict[str, object] | None = None,
    after: dict[str, object] | None = None,
) -> None:
    record = after or before or {}
    label = inventory_entity_label(record)
    changes = build_changed_fields(before, after, INVENTORY_FIELD_LABELS) if before and after else []
    write_operation_log(
        request,
        module=inventory_module_for_record(record),
        action=action,
        entity_type="inventory_record",
        entity_id=record.get("id"),
        entity_label=label,
        summary=summarize_changes(prefix, label, changes) if changes else f"{prefix} {label}".strip(),
        changed_fields=changes,
        before_data=before,
        after_data=after,
    )


def _log_detail_operation(
    request: Request,
    *,
    action: str,
    prefix: str,
    record: dict[str, object] | None,
    before: dict[str, object] | None = None,
    after: dict[str, object] | None = None,
) -> None:
    detail = after or before or {}
    label = detail_entity_label(detail)
    changes = build_changed_fields(before, after, DETAIL_FIELD_LABELS) if before and after else []
    record_label = inventory_entity_label(record)
    summary_prefix = f"{prefix} {record_label} 的明细" if record_label else f"{prefix}明细"
    write_operation_log(
        request,
        module=inventory_module_for_record(record),
        action=action,
        entity_type="inventory_detail",
        entity_id=detail.get("id"),
        entity_label=label,
        summary=summarize_changes(summary_prefix, label, changes) if changes else f"{summary_prefix} {label}".strip(),
        changed_fields=changes,
        before_data=before,
        after_data=after,
    )


def _load_supplier_export_lookup(repository, items: list[dict[str, object]]) -> dict[str, list[dict[str, str]]]:
    supplier_names = sorted({
        supplier_name
        for supplier_name in (_cell_text(item.get("supplier")) for item in items)
        if supplier_name
    })
    if not supplier_names:
        return {}

    supplier_lookup: dict[str, list[dict[str, str]]] = defaultdict(list)
    with repository.engine.connect() as connection:
        rows = connection.execute(
            sa_select(SUPPLIER_TABLE.c.name, SUPPLIER_TABLE.c.factory_code, SUPPLIER_TABLE.c.brand)
            .where(SUPPLIER_TABLE.c.name.in_(supplier_names))
        ).mappings()
        for row in rows:
            name = _cell_text(row.get("name"))
            if not name:
                continue
            supplier_lookup[name].append({
                "factory_code": _cell_text(row.get("factory_code")),
                "brand": _cell_text(row.get("brand")).lower(),
            })
    return supplier_lookup


def _supplier_export_context(
    record: dict[str, object],
    supplier_lookup: dict[str, list[dict[str, str]]],
) -> dict[str, str]:
    supplier_name = _cell_text(record.get("supplier"))
    candidates = supplier_lookup.get(supplier_name, [])
    if not candidates:
        return {}

    record_brand = _purchase_record_brand(record)
    if record_brand:
        brand_candidates = [candidate for candidate in candidates if candidate.get("brand") == record_brand]
        if brand_candidates:
            for candidate in brand_candidates:
                if candidate.get("factory_code"):
                    return candidate
            return brand_candidates[0]
    for candidate in candidates:
        if candidate.get("factory_code"):
            return candidate
    for candidate in candidates:
        if candidate.get("brand"):
            return candidate
    return candidates[0]


def _purchase_record_export_context(
    record: dict[str, object],
    supplier_lookup: dict[str, list[dict[str, str]]],
) -> dict[str, str]:
    extra_fields = _dict_or_empty(record.get("extra_fields"))
    supplier_name = _cell_text(record.get("supplier"))
    supplier_context = _supplier_export_context(record, supplier_lookup)
    brand = supplier_context.get("brand") or _purchase_record_brand(record)
    return {
        "document_type": _cell_text(record.get("document_type")),
        "document_number": _first_text(record.get("document_number"), record.get("id")),
        "summary": _cell_text(record.get("summary")),
        "brand": brand,
        "date": _cell_text(record.get("date")),
        "delivery_date": _first_text(extra_fields.get("delivery_date"), extra_fields.get("到货日期")),
        "warehouse_code": _first_text(extra_fields.get("warehouse_code"), extra_fields.get("仓库编号")),
        "warehouse_name": _cell_text(record.get("warehouse")),
        "unit_code": supplier_context.get("factory_code", ""),
        "unit_name": supplier_name,
        "handler_code": _first_text(extra_fields.get("handler_code"), extra_fields.get("职员编号")),
        "handler_name": _cell_text(record.get("handler")),
    }


def _append_purchase_summary_export(
    worksheet,
    details: list[dict[str, object]],
    records_by_id: dict[object, dict[str, object]],
    supplier_lookup: dict[str, list[dict[str, str]]],
) -> None:
    worksheet.append(PURCHASE_SUMMARY_EXPORT_HEADERS)

    groups: dict[tuple[object, str], dict[str, object]] = {}
    ordered_keys: list[tuple[object, str]] = []
    for detail in details:
        document_id = detail.get("document_id")
        record = records_by_id.get(document_id)
        if not record:
            continue
        product_code = _cell_text(detail.get("product_code"))
        if not product_code:
            continue
        key = (document_id, product_code)
        if key not in groups:
            detail_extra_fields = _dict_or_empty(detail.get("extra_fields"))
            groups[key] = {
                "record_context": _purchase_record_export_context(record, supplier_lookup),
                "product_code": product_code,
                "original_code": _first_text(detail_extra_fields.get("image_code"), product_code),
                "product_name": _first_text(detail.get("product_name"), product_code),
                "size_quantities": defaultdict(Decimal),
                "quantity": Decimal("0"),
                "amount": Decimal("0"),
                "arrival_quantity": Decimal("0"),
                "has_arrival_quantity": False,
                "unreceived_quantity": Decimal("0"),
                "has_unreceived_quantity": False,
                "completion_rate": "",
            }
            ordered_keys.append(key)

        group = groups[key]
        detail_extra_fields = _dict_or_empty(detail.get("extra_fields"))
        size_quantities = _dict_or_empty(detail.get("size_quantities"))
        quantity = _purchase_detail_quantity(detail, size_quantities)
        amount = _purchase_detail_amount(detail, quantity)
        group["quantity"] = group["quantity"] + quantity
        group["amount"] = group["amount"] + amount

        for size, value in size_quantities.items():
            size_text = _cell_text(size)
            size_quantity = _to_decimal(value)
            if size_text and size_quantity != 0:
                group["size_quantities"][size_text] += size_quantity

        arrival_text = _purchase_detail_arrival_text(detail_extra_fields)
        if arrival_text:
            group["arrival_quantity"] = group["arrival_quantity"] + _to_decimal(arrival_text)
            group["has_arrival_quantity"] = True

        unreceived_text = _purchase_detail_unreceived_text(detail_extra_fields)
        if unreceived_text:
            group["unreceived_quantity"] = group["unreceived_quantity"] + _to_decimal(unreceived_text)
            group["has_unreceived_quantity"] = True

        completion_text = _purchase_detail_completion_text(detail_extra_fields)
        if completion_text and not group["completion_rate"]:
            group["completion_rate"] = completion_text

    for index, key in enumerate(ordered_keys, start=1):
        group = groups[key]
        record_context = group["record_context"]
        quantity = group["quantity"]
        amount = group["amount"]
        has_arrival = bool(group["has_arrival_quantity"])
        has_unreceived = bool(group["has_unreceived_quantity"])
        arrival_quantity = group["arrival_quantity"]
        unreceived_quantity = (
            group["unreceived_quantity"]
            if has_unreceived
            else max(quantity - arrival_quantity, Decimal("0")) if has_arrival
            else quantity
        )
        completion_rate = group["completion_rate"] or (
            _fmt_completion_rate(arrival_quantity, quantity) if has_arrival else ""
        )

        worksheet.append([
            index,
            record_context["document_type"],
            group["product_code"],
            group["original_code"],
            group["product_name"],
            record_context["document_number"],
            record_context["summary"],
            record_context["date"],
            record_context["delivery_date"],
            record_context["warehouse_name"],
            record_context["unit_code"],
            record_context["unit_name"],
            record_context["handler_name"],
            _fmt_export_decimal(quantity, blank_zero=True),
            _fmt_export_decimal(amount, blank_zero=True),
            _fmt_export_decimal(arrival_quantity, blank_zero=True) if has_arrival else "",
            _fmt_export_decimal(unreceived_quantity, blank_zero=True),
            completion_rate,
        ])
    _style_purchase_detail_export_sheet(worksheet)


def _append_purchase_size_rows_export(
    worksheet,
    details: list[dict[str, object]],
    records_by_id: dict[object, dict[str, object]],
    supplier_lookup: dict[str, list[dict[str, str]]],
) -> None:
    worksheet.append(PURCHASE_SIZE_ROW_EXPORT_HEADERS)
    size_labels = _purchase_export_size_labels(details)

    groups: dict[tuple[object, str, str, str], dict[str, object]] = {}
    ordered_keys: list[tuple[object, str, str, str]] = []
    for detail in details:
        document_id = detail.get("document_id")
        record = records_by_id.get(document_id)
        if not record:
            continue
        record_context = _purchase_record_export_context(record, supplier_lookup)
        detail_extra_fields = _dict_or_empty(detail.get("extra_fields"))
        detail_remark = _cell_text(detail.get("remark"))
        product_code = _cell_text(detail.get("product_code"))
        if not product_code:
            continue
        product_name = _first_text(detail.get("product_name"), product_code)
        color_barcode = _purchase_color_barcode(product_code, record_context["brand"])
        color_name = _first_text(detail.get("color_name"), detail.get("color_spec"))
        size_quantities = _dict_or_empty(detail.get("size_quantities"))
        detail_quantity = _purchase_detail_quantity(detail, size_quantities)
        detail_amount = _purchase_detail_amount(detail, detail_quantity)
        unit_price = _to_decimal(detail.get("unit_price"))

        size_entries = [
            (size, _to_decimal(size_quantities.get(size)))
            for size in size_labels
            if _to_decimal(size_quantities.get(size)) != 0
        ]
        if not size_entries and detail_quantity != 0:
            size_entries = [("", detail_quantity)]

        arrival_text = _purchase_detail_arrival_text(detail_extra_fields)
        unreceived_text = _purchase_detail_unreceived_text(detail_extra_fields)
        completion_text = _purchase_detail_completion_text(detail_extra_fields)
        detail_arrival_quantity = _to_decimal(arrival_text) if arrival_text else Decimal("0")
        detail_unreceived_quantity = _to_decimal(unreceived_text) if unreceived_text else Decimal("0")

        for size, quantity in size_entries:
            size_text = _cell_text(size)
            key = (document_id, product_code, color_barcode, size_text)
            if key not in groups:
                groups[key] = {
                    "record_context": record_context,
                    "product_code": product_code,
                    "product_name": product_name,
                    "detail_remark": detail_remark,
                    "color_barcode": color_barcode,
                    "color_name": color_name,
                    "size": size_text,
                    "quantity": Decimal("0"),
                    "amount": Decimal("0"),
                    "arrival_quantity": Decimal("0"),
                    "has_arrival_quantity": False,
                    "unreceived_quantity": Decimal("0"),
                    "has_unreceived_quantity": False,
                    "completion_rate": "",
                }
                ordered_keys.append(key)

            group = groups[key]
            if detail_remark and not group.get("detail_remark"):
                group["detail_remark"] = detail_remark
            size_amount = quantity * unit_price if unit_price != 0 else _purchase_prorated_decimal(detail_amount, quantity, detail_quantity)
            group["quantity"] = group["quantity"] + quantity
            group["amount"] = group["amount"] + size_amount
            if arrival_text:
                group["arrival_quantity"] = group["arrival_quantity"] + _purchase_prorated_decimal(detail_arrival_quantity, quantity, detail_quantity)
                group["has_arrival_quantity"] = True
            if unreceived_text:
                group["unreceived_quantity"] = group["unreceived_quantity"] + _purchase_prorated_decimal(detail_unreceived_quantity, quantity, detail_quantity)
                group["has_unreceived_quantity"] = True
            if completion_text and not group["completion_rate"]:
                group["completion_rate"] = completion_text

    for key in ordered_keys:
        group = groups[key]
        record_context = group["record_context"]
        quantity = group["quantity"]
        has_arrival = bool(group["has_arrival_quantity"])
        has_unreceived = bool(group["has_unreceived_quantity"])
        arrival_quantity = group["arrival_quantity"]
        unreceived_quantity = (
            group["unreceived_quantity"]
            if has_unreceived
            else max(quantity - arrival_quantity, Decimal("0")) if has_arrival
            else quantity
        )
        completion_rate = group["completion_rate"] or (
            _fmt_completion_rate(arrival_quantity, quantity) if has_arrival else ""
        )
        size = group["size"]

        worksheet.append([
            record_context["unit_name"],
            group["product_code"],
            _fmt_export_decimal(quantity, blank_zero=True),
            record_context["summary"],
            record_context["date"],
            record_context["delivery_date"],
            record_context["warehouse_name"],
            group["product_name"],
            group["color_barcode"],
            group["color_name"],
            size,
            size,
            _purchase_export_barcode(group["product_code"], group["color_barcode"], size, record_context["brand"]),
            record_context["document_type"],
            record_context["document_number"],
            group["detail_remark"],
            record_context["unit_name"],
            record_context["unit_code"],
            record_context["handler_name"],
            _fmt_export_decimal(group["amount"], blank_zero=True),
            _fmt_export_decimal(arrival_quantity, blank_zero=True) if has_arrival else "",
            _fmt_export_decimal(unreceived_quantity, blank_zero=True),
            completion_rate,
        ])
    _style_purchase_detail_export_sheet(worksheet)


def _style_purchase_detail_export_sheet(worksheet) -> None:
    style_excel_worksheet(worksheet)


def _short_sheet_title(value: str, fallback: str) -> str:
    title = (value or fallback).strip() or fallback
    for ch in "[]:*?/\\":
        title = title.replace(ch, "-")
    return title[:31] or fallback[:31]


def _purchase_sheet_date(value: object) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    return text.replace("-", ".")


def _purchase_hyphen_date(value: object) -> str:
    text = _normalize_date(_cell_text(value))
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text).strftime("%Y-%m-%d")
    except ValueError:
        pass
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-").replace(".", "-")
    parts = [part for part in normalized.split("-") if part]
    if len(parts) == 3 and all(part.isdigit() for part in parts):
        year, month, day = parts
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    if len(parts) == 3 and parts[0].isdigit() and parts[1].isdigit() and parts[2].split()[0].isdigit():
        year, month, day = parts
        return f"{int(year):04d}-{int(month):02d}-{int(day.split()[0]):02d}"
    if len(normalized) == 8 and normalized.isdigit():
        return f"{normalized[:4]}-{normalized[4:6]}-{normalized[6:]}"
    return normalized


def _purchase_production_size_labels(details: list[dict[str, object]]) -> tuple[str, ...]:
    used_sizes: set[str] = set()
    for detail in details:
        size_quantities = _dict_or_empty(detail.get("size_quantities"))
        for size, quantity in size_quantities.items():
            size_text = _cell_text(size)
            if size_text and _to_decimal(quantity) != 0:
                used_sizes.add(size_text)
    ordered = [size for size in PURCHASE_EXPORT_SIZE_LABELS if size in used_sizes]
    extra = sorted(size for size in used_sizes if size not in PURCHASE_EXPORT_SIZE_LABELS)
    return tuple(ordered + extra) or ("220", "225", "230", "235", "240", "245", "250")


def _purchase_production_detail_rows(
    details: list[dict[str, object]],
    size_labels: tuple[str, ...],
    delivery_date: str,
) -> list[list[object]]:
    rows: list[list[object]] = []
    for detail in details:
        extra_fields = _dict_or_empty(detail.get("extra_fields"))
        size_quantities = _dict_or_empty(detail.get("size_quantities"))
        quantity = _purchase_detail_quantity(detail, size_quantities)
        unit_price = _to_decimal(detail.get("unit_price"))
        amount = _purchase_detail_amount(detail, quantity)
        product_code = _cell_text(detail.get("product_code"))
        if not product_code and quantity == 0:
            continue
        rows.append([
            product_code,
            _cell_text(extra_fields.get("image_code") or product_code),
            _cell_text(extra_fields.get("factory_code")),
            _cell_text(extra_fields.get("image_code") or extra_fields.get("style_code") or product_code),
            _cell_text(extra_fields.get("inner_color_code")),
            _first_text(detail.get("color_name"), detail.get("color_spec")),
            _cell_text(extra_fields.get("upper_material")),
            _cell_text(extra_fields.get("lining_material")),
            _cell_text(extra_fields.get("outsole_material")),
            _cell_text(extra_fields.get("insole_material")),
            _cell_text(extra_fields.get("shoe_box_spec")),
            *[
                _fmt_export_decimal(_to_decimal(size_quantities.get(size)), blank_zero=True)
                for size in size_labels
            ],
            _fmt_export_decimal(quantity, blank_zero=True),
            _fmt_export_decimal(unit_price, blank_zero=True),
            _fmt_export_decimal(amount, blank_zero=True),
            _purchase_hyphen_date(_first_text(extra_fields.get("required_delivery_date"), delivery_date)),
            _purchase_sheet_date(_first_text(extra_fields.get("factory_reply_delivery_date"), extra_fields.get("reply_delivery_date"))),
        ])
    return rows


def _purchase_production_total_row(
    detail_rows: list[list[object]],
    size_labels: tuple[str, ...],
) -> list[object]:
    fixed_count = len(PURCHASE_PRODUCTION_FIXED_HEADERS)
    quantity_index = fixed_count + len(size_labels)
    amount_index = quantity_index + 2
    row_length = fixed_count + len(size_labels) + len(PURCHASE_PRODUCTION_TAIL_HEADERS)
    total_row: list[object] = [""] * row_length

    for size_offset in range(len(size_labels)):
        col_index = fixed_count + size_offset
        total = sum((_to_decimal(row[col_index]) for row in detail_rows), Decimal("0"))
        total_row[col_index] = _fmt_export_decimal(total, blank_zero=True)

    quantity_total = sum((_to_decimal(row[quantity_index]) for row in detail_rows), Decimal("0"))
    amount_total = sum((_to_decimal(row[amount_index]) for row in detail_rows), Decimal("0"))
    total_row[quantity_index] = _fmt_export_decimal(quantity_total, blank_zero=True)
    total_row[amount_index] = _fmt_export_decimal(amount_total, blank_zero=True)
    return total_row


def _purchase_order_requirements_for_brand(
    brand: str,
    requirement_templates: dict[str, str] | None = None,
) -> str:
    normalized_brand = _cell_text(brand).lower()
    if requirement_templates is not None and normalized_brand in requirement_templates:
        return requirement_templates[normalized_brand]
    return PURCHASE_ORDER_REQUIREMENTS_BY_BRAND.get(normalized_brand, "")


def _append_purchase_production_order_requirements(
    worksheet,
    row_index: int,
    max_col: int,
    requirements: str,
) -> None:
    if not requirements:
        return

    if max_col > 1:
        worksheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=max_col)

    line_count = requirements.count("\n") + 1
    worksheet.row_dimensions[row_index].height = min(max(240, line_count * 24), 480)
    worksheet.cell(row=row_index, column=1, value="订单要求")
    worksheet.cell(row=row_index, column=2 if max_col > 1 else 1, value=requirements)

    black_side = Side(style="thin", color="000000")
    black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    label_font = Font(name="宋体", size=12, bold=True, color="000000")
    requirement_font = Font(name="宋体", size=12, bold=True, color="000000")
    label_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_index in range(1, max_col + 1):
        cell = worksheet.cell(row=row_index, column=col_index)
        cell.border = black_border
        cell.font = requirement_font
        cell.alignment = text_alignment
    worksheet.cell(row=row_index, column=1).font = label_font
    worksheet.cell(row=row_index, column=1).alignment = label_alignment


def _style_purchase_production_spacer_row(worksheet, row_index: int, max_col: int) -> None:
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    fill = PatternFill("solid", fgColor="D9EAD3")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font = Font(name="宋体", size=10)

    worksheet.row_dimensions[row_index].height = 42
    for col_index in range(1, max_col + 1):
        cell = worksheet.cell(row=row_index, column=col_index, value="")
        cell.border = border
        cell.fill = fill
        cell.alignment = alignment
        cell.font = font


def _style_purchase_production_total_row(worksheet, row_index: int, max_col: int) -> None:
    font = Font(name="宋体", size=10, bold=True)
    fill = PatternFill("solid", fgColor="F7F7F7")
    for col_index in range(1, max_col + 1):
        cell = worksheet.cell(row=row_index, column=col_index)
        cell.font = font
        cell.fill = fill


def _style_purchase_production_sheet(worksheet, max_row: int, max_col: int, size_start_col: int, tail_start_col: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    no_wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    title_font = Font(name="宋体", size=18, bold=True)
    header_font = Font(name="宋体", size=10, bold=True)
    body_font = Font(name="宋体", size=10)
    red_font = Font(name="宋体", size=10, bold=True, color="FF0000")
    fill = PatternFill("solid", fgColor="F7F7F7")

    worksheet.row_dimensions[1].height = 32
    worksheet.row_dimensions[2].height = 34
    worksheet.row_dimensions[3].height = 34
    worksheet.row_dimensions[4].height = 82
    for row_index in range(5, max_row + 1):
        worksheet.row_dimensions[row_index].height = 42

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = center
            cell.border = border
            cell.font = body_font

    for cell in worksheet[1]:
        cell.font = title_font

    for row_index in (2, 3, 4):
        for cell in worksheet[row_index]:
            cell.font = header_font
            cell.fill = fill if row_index == 4 else PatternFill(fill_type=None)

    for col_index in range(tail_start_col + 3, tail_start_col + 5):
        worksheet.cell(row=4, column=col_index).font = red_font
        worksheet.cell(row=4, column=col_index).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=255)

    for row_index in range(5, max_row + 1):
        for col_index in (1, 2, 3, 4):
            worksheet.cell(row=row_index, column=col_index).alignment = no_wrap_center

    width_map = {
        1: 16,
        2: 16,
        3: 12,
        4: 16,
        5: 8,
        6: 9,
        7: 9,
        8: 8,
        9: 8,
        10: 8,
        11: 8,
    }
    for col_index in range(1, max_col + 1):
        if col_index in width_map:
            width = width_map[col_index]
        elif size_start_col <= col_index < tail_start_col:
            width = 5
        elif col_index >= tail_start_col:
            width = 7 if col_index < tail_start_col + 3 else 8
        else:
            width = 9
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    worksheet.page_margins.top = 0.3
    worksheet.page_margins.bottom = 0.3


def _append_purchase_production_order_sheet(
    worksheet,
    record: dict[str, object],
    details: list[dict[str, object]],
    supplier_lookup: dict[str, list[dict[str, str]]],
    requirement_templates: dict[str, str] | None = None,
) -> None:
    context = _purchase_record_export_context(record, supplier_lookup)
    size_labels = _purchase_production_size_labels(details)
    headers = [*PURCHASE_PRODUCTION_FIXED_HEADERS, *size_labels, *PURCHASE_PRODUCTION_TAIL_HEADERS]
    max_col = len(headers)
    size_start_col = len(PURCHASE_PRODUCTION_FIXED_HEADERS) + 1
    tail_start_col = size_start_col + len(size_labels)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    worksheet.cell(row=1, column=1, value="赫德电商（千百度）生产采购单")

    worksheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
    worksheet.merge_cells(start_row=2, start_column=6, end_row=2, end_column=8)
    worksheet.merge_cells(start_row=2, start_column=10, end_row=2, end_column=min(max_col, 14))
    worksheet.cell(row=2, column=1, value="供货单位")
    worksheet.cell(row=2, column=2, value=context["unit_name"])
    worksheet.cell(row=2, column=5, value="收货仓库")
    worksheet.cell(row=2, column=6, value=context["warehouse_name"])
    worksheet.cell(row=2, column=9, value="订货日期")
    worksheet.cell(row=2, column=10, value=context["date"])

    worksheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=8)
    worksheet.merge_cells(start_row=3, start_column=10, end_row=3, end_column=min(max_col, 14))
    worksheet.cell(row=3, column=1, value="摘要")
    worksheet.cell(row=3, column=2, value=context["summary"])
    worksheet.cell(row=3, column=9, value="单据编号")
    worksheet.cell(row=3, column=10, value=context["document_number"])
    if max_col >= 15:
        worksheet.cell(row=2, column=15, value="交货日期")
        worksheet.cell(row=2, column=16, value=context["delivery_date"])
        worksheet.cell(row=3, column=15, value="经手人")
        worksheet.cell(row=3, column=16, value=context["handler_name"])
        if max_col >= 16:
            worksheet.merge_cells(start_row=2, start_column=16, end_row=2, end_column=max_col)
            worksheet.merge_cells(start_row=3, start_column=16, end_row=3, end_column=max_col)
    else:
        worksheet.cell(row=2, column=15, value="交货日期")
        worksheet.cell(row=3, column=max_col, value=context["handler_name"])

    for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=4, column=col_index, value=header)

    detail_rows = _purchase_production_detail_rows(details, size_labels, context["delivery_date"])
    data_start_row = 5
    for row_offset, row_values in enumerate(detail_rows):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=data_start_row + row_offset, column=col_index, value=value)

    content_last_row = data_start_row + len(detail_rows) - 1
    total_row_index: int | None = None
    if detail_rows:
        total_row_index = data_start_row + len(detail_rows)
        total_row_values = _purchase_production_total_row(detail_rows, size_labels)
        for col_index, value in enumerate(total_row_values, start=1):
            worksheet.cell(row=total_row_index, column=col_index, value=value)
        content_last_row = total_row_index
    else:
        content_last_row = data_start_row
        worksheet.cell(row=content_last_row, column=1, value="暂无明细")
        worksheet.merge_cells(start_row=content_last_row, start_column=1, end_row=content_last_row, end_column=max_col)

    _style_purchase_production_sheet(worksheet, content_last_row, max_col, size_start_col, tail_start_col)
    if total_row_index is not None:
        _style_purchase_production_total_row(worksheet, total_row_index, max_col)
    requirements = _purchase_order_requirements_for_brand(context["brand"], requirement_templates)
    if requirements:
        _style_purchase_production_spacer_row(worksheet, content_last_row + 1, max_col)
        _append_purchase_production_order_requirements(worksheet, content_last_row + 2, max_col, requirements)


def _append_purchase_production_order_export(
    workbook: Workbook,
    details: list[dict[str, object]],
    records_by_id: dict[object, dict[str, object]],
    supplier_lookup: dict[str, list[dict[str, str]]],
    requirement_templates: dict[str, str] | None = None,
) -> None:
    workbook.remove(workbook.active)
    details_by_document: dict[object, list[dict[str, object]]] = defaultdict(list)
    for detail in details:
        details_by_document[detail.get("document_id")].append(detail)

    for index, record in enumerate(records_by_id.values(), start=1):
        title = _short_sheet_title(_first_text(record.get("document_number"), record.get("summary")), f"采购单{index}")
        worksheet = workbook.create_sheet(title)
        _append_purchase_production_order_sheet(
            worksheet,
            record,
            details_by_document.get(record.get("id"), []),
            supplier_lookup,
            requirement_templates,
        )

    if not records_by_id:
        worksheet = workbook.create_sheet("生产采购单")
        _append_purchase_production_order_sheet(worksheet, {}, [], supplier_lookup, requirement_templates)


def _stream_excel_workbook(workbook: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    quoted_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}",
            "Content-Length": str(buf.getbuffer().nbytes),
            "X-Content-Type-Options": "nosniff",
            "Cache-Control": "no-store",
        },
    )


def _build_purchase_order_import_template() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "采购单导入模板"
    headers = [
        "供货单位",
        "摘要",
        "采购日期",
        "协议到货日期",
        "收货仓库",
        "经办人",
        "商品货号",
        "商品备注",
        "数量",
    ]
    worksheet.append(headers)
    worksheet.append([
        "说明：同一个摘要会导入为同一张采购单；不同摘要会自动分成多张采购单。商品货号请填写带颜色和尺码的完整商品编码，系统导入时会自动拆解并匹配单价。",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ])
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    style_excel_worksheet(
        worksheet,
        width_by_header={
            "供货单位": 24,
            "摘要": 28,
            "采购日期": 14,
            "协议到货日期": 14,
            "收货仓库": 18,
            "经办人": 12,
            "商品货号": 24,
            "商品备注": 18,
            "数量": 10,
        },
        freeze_panes="A2",
        auto_filter=False,
    )
    worksheet.row_dimensions[2].height = 36
    worksheet.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return workbook


def _purchase_import_brand(document_type: str) -> str:
    return "cbanner_womens" if "女鞋" in document_type else "cbanner_mens"


def _normalize_purchase_brand(brand: str | None) -> str:
    return str(brand or "").strip().lower()


def _uses_millimeter_purchase_sizes(brand: str | None) -> bool:
    normalized = _normalize_purchase_brand(brand)
    return normalized not in EU_SIZE_BRANDS


def _purchase_size_labels(brand: str | None) -> tuple[str, ...]:
    return MILLIMETER_PURCHASE_SIZE_LABELS if _uses_millimeter_purchase_sizes(brand) else EU_PURCHASE_SIZE_LABELS


def _purchase_size_from_eu(size: str, brand: str | None) -> str:
    if _uses_millimeter_purchase_sizes(brand):
        return PURCHASE_EU_TO_MILLIMETER_SIZE_MAP.get(size, size)
    return size


def _purchase_size_from_millimeter(size_code: str, brand: str | None) -> str:
    if _uses_millimeter_purchase_sizes(brand):
        return size_code
    return PURCHASE_MILLIMETER_SIZE_MAP.get(size_code, size_code)


def _purchase_import_doc_field_indexes(headers: list[str]) -> dict[int, str]:
    return {
        index: field
        for index, value in enumerate(headers)
        for field, aliases in PURCHASE_IMPORT_DOC_FIELD_ALIASES.items()
        if value in aliases
    }


def _purchase_import_doc_fields_from_values(values: object, indexes: dict[int, str]) -> dict[str, str]:
    if not isinstance(values, (list, tuple)):
        return {}
    fields: dict[str, str] = {}
    for index, field in indexes.items():
        raw_value = values[index] if index < len(values) else None
        if field in {"date", "delivery_date"}:
            text = _normalize_date(raw_value) or ""
        else:
            text = _cell_text(raw_value)
        if text:
            fields[field] = text
    return fields


def _read_purchase_import_rows(content: bytes) -> tuple[list[dict[str, object]], str]:
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return _read_purchase_import_rows_xls(content)
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Excel 文件格式无法读取，请使用 .xlsx/.xlsm，或确认 .xls 文件未损坏") from error
    try:
        worksheet = workbook.active
        if worksheet is None:
            return [], ""
        header_index = None
        code_index = None
        qty_index = None
        unit_price_index = None
        remark_index = None
        size_indexes: dict[int, str] = {}
        extra_indexes: dict[int, str] = {}
        doc_field_indexes: dict[int, str] = {}
        iterator = worksheet.iter_rows(values_only=True)
        for row_number, row in enumerate(iterator):
            headers = [_cell_text(value).replace("\n", "").replace("\r", "") for value in row]
            code_index = next((index for index, value in enumerate(headers) if value in PURCHASE_ORDER_IMPORT_CODE_HEADERS), None)
            qty_index = next((index for index, value in enumerate(headers) if value == "数量"), None)
            unit_price_index = next((index for index, value in enumerate(headers) if value == "单价"), None)
            remark_index = next((index for index, value in enumerate(headers) if value in PURCHASE_DETAIL_REMARK_HEADERS), None)
            doc_field_indexes = _purchase_import_doc_field_indexes(headers)
            extra_indexes = {
                index: key
                for index, value in enumerate(headers)
                for key, label in PURCHASE_DETAIL_EXTRA_FIELDS.items()
                if value == label.replace("\n", "")
            }
            size_indexes = {
                index: value
                for index, value in enumerate(headers)
                if value in PURCHASE_EXPORT_SIZE_LABELS
            }
            if code_index is not None and (qty_index is not None or size_indexes):
                header_index = row_number
                break
            if row_number >= 29:
                break
        if header_index is None or code_index is None or (qty_index is None and not size_indexes):
            raise HTTPException(status_code=400, detail="Excel 中需要包含 商品货号/商品编码/货品编码/商品编号/货号，以及 数量 或尺码数量列")

        parsed_rows: list[dict[str, object]] = []
        empty_streak = 0
        for row in iterator:
            product_code = _cell_text(row[code_index] if code_index < len(row) else None)
            quantity = _cell_text(row[qty_index] if qty_index is not None and qty_index < len(row) else None)
            unit_price = _cell_text(row[unit_price_index] if unit_price_index is not None and unit_price_index < len(row) else None)
            remark = _cell_text(row[remark_index] if remark_index is not None and remark_index < len(row) else None)
            if len(remark) > PURCHASE_DETAIL_REMARK_LIMIT:
                raise HTTPException(status_code=400, detail=f"商品备注最多 {PURCHASE_DETAIL_REMARK_LIMIT} 个字")
            size_quantities = {
                size: _cell_text(row[index] if index < len(row) else None)
                for index, size in size_indexes.items()
                if _cell_text(row[index] if index < len(row) else None)
            }
            extra_fields = {
                key: _cell_text(row[index] if index < len(row) else None)
                for index, key in extra_indexes.items()
                if _cell_text(row[index] if index < len(row) else None) or key == "inner_color_code"
            }
            if product_code and (quantity or size_quantities):
                if not quantity and size_quantities:
                    quantity = _fmt_decimal(sum((_to_decimal(value) for value in size_quantities.values()), Decimal("0")))
                parsed_rows.append({
                    "product_code": product_code,
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "remark": remark,
                    "size_quantities": size_quantities,
                    "extra_fields": extra_fields,
                    **_purchase_import_doc_fields_from_values(row, doc_field_indexes),
                })
                empty_streak = 0
            elif not any(_cell_text(value) for value in row):
                empty_streak += 1
                if empty_streak >= 50:
                    break
            else:
                empty_streak = 0
        return parsed_rows, worksheet.title
    finally:
        workbook.close()


def _read_purchase_import_rows_xls(content: bytes) -> tuple[list[dict[str, object]], str]:
    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Excel .xls 文件无法读取，请另存为 .xlsx 后重试") from error
    if workbook.nsheets == 0:
        return [], ""
    worksheet = workbook.sheet_by_index(0)
    code_index = None
    qty_index = None
    unit_price_index = None
    remark_index = None
    size_indexes: dict[int, str] = {}
    extra_indexes: dict[int, str] = {}
    doc_field_indexes: dict[int, str] = {}
    data_start_row = 0
    for row_index in range(min(worksheet.nrows, 30)):
        headers = [_cell_text(value).replace("\n", "").replace("\r", "") for value in worksheet.row_values(row_index)]
        code_index = next((index for index, value in enumerate(headers) if value in PURCHASE_ORDER_IMPORT_CODE_HEADERS), None)
        qty_index = next((index for index, value in enumerate(headers) if value == "数量"), None)
        unit_price_index = next((index for index, value in enumerate(headers) if value == "单价"), None)
        remark_index = next((index for index, value in enumerate(headers) if value in PURCHASE_DETAIL_REMARK_HEADERS), None)
        doc_field_indexes = _purchase_import_doc_field_indexes(headers)
        extra_indexes = {
            index: key
            for index, value in enumerate(headers)
            for key, label in PURCHASE_DETAIL_EXTRA_FIELDS.items()
            if value == label.replace("\n", "")
        }
        size_indexes = {
            index: value
            for index, value in enumerate(headers)
            if value in PURCHASE_EXPORT_SIZE_LABELS
        }
        if code_index is not None and (qty_index is not None or size_indexes):
            data_start_row = row_index + 1
            break
    if code_index is None or (qty_index is None and not size_indexes):
        raise HTTPException(status_code=400, detail="Excel 中需要包含 商品货号/商品编码/货品编码/商品编号/货号，以及 数量 或尺码数量列")

    parsed_rows: list[dict[str, object]] = []
    empty_streak = 0
    for row_index in range(data_start_row, worksheet.nrows):
        values = worksheet.row_values(row_index)
        product_code = _cell_text(values[code_index] if code_index < len(values) else None)
        quantity = _cell_text(values[qty_index] if qty_index is not None and qty_index < len(values) else None)
        unit_price = _cell_text(values[unit_price_index] if unit_price_index is not None and unit_price_index < len(values) else None)
        remark = _cell_text(values[remark_index] if remark_index is not None and remark_index < len(values) else None)
        if len(remark) > PURCHASE_DETAIL_REMARK_LIMIT:
            raise HTTPException(status_code=400, detail=f"商品备注最多 {PURCHASE_DETAIL_REMARK_LIMIT} 个字")
        size_quantities = {
            size: _cell_text(values[index] if index < len(values) else None)
            for index, size in size_indexes.items()
            if _cell_text(values[index] if index < len(values) else None)
        }
        extra_fields = {
            key: _cell_text(values[index] if index < len(values) else None)
            for index, key in extra_indexes.items()
            if _cell_text(values[index] if index < len(values) else None) or key == "inner_color_code"
        }
        if product_code and (quantity or size_quantities):
            if not quantity and size_quantities:
                quantity = _fmt_decimal(sum((_to_decimal(value) for value in size_quantities.values()), Decimal("0")))
            parsed_rows.append({
                "product_code": product_code,
                "quantity": quantity,
                "unit_price": unit_price,
                "remark": remark,
                "size_quantities": size_quantities,
                "extra_fields": extra_fields,
                **_purchase_import_doc_fields_from_values(values, doc_field_indexes),
            })
            empty_streak = 0
        elif not any(_cell_text(value) for value in values):
            empty_streak += 1
            if empty_streak >= 50:
                break
        else:
            empty_streak = 0
    return parsed_rows, worksheet.name


def _load_color_barcodes(connection, brand: str | None = None) -> list[tuple[str, str]]:
    statement = sa_select(COLOR_BARCODE_TABLE.c.color_barcode, COLOR_BARCODE_TABLE.c.color_name)
    if brand:
        statement = statement.where(COLOR_BARCODE_TABLE.c.brand == brand)
    rows = connection.execute(
        statement
    ).all()
    return sorted(
        [(str(row[0]), str(row[1])) for row in rows if row[0] and row[1]],
        key=lambda item: len(item[0]),
        reverse=True,
    )


def _lookup_color_name_by_tail(connection, product_code: str, brand: str | None = None) -> tuple[str, str]:
    color_barcode = _purchase_color_barcode(product_code, brand)
    if not color_barcode:
        return "", ""
    row = connection.execute(
        sa_select(COLOR_BARCODE_TABLE.c.color_name)
        .where(COLOR_BARCODE_TABLE.c.color_barcode == color_barcode)
        .order_by(COLOR_BARCODE_TABLE.c.brand)
        .limit(1)
    ).first()
    return color_barcode, _cell_text(row[0] if row else None)


def _purchase_detail_color_values(
    product_info: dict[str, object],
    fallback_color_barcode: object = "",
    fallback_color_name: object = "",
) -> tuple[str, str]:
    return (
        _first_text(product_info.get("color_barcode"), product_info.get("color_code"), fallback_color_barcode),
        _first_text(product_info.get("color_name"), product_info.get("color"), fallback_color_name),
    )


def _split_purchase_size_code(product_code: str, brand: str) -> tuple[str, str]:
    if len(product_code) >= 3:
        size_code = product_code[-3:]
        if size_code in PURCHASE_MILLIMETER_SIZE_MAP:
            return product_code[:-3], _purchase_size_from_millimeter(size_code, brand)
    if len(product_code) >= 2:
        size_code = product_code[-2:]
        if size_code in EU_PURCHASE_SIZE_LABELS:
            return product_code[:-2], _purchase_size_from_eu(size_code, brand)
        size = PURCHASE_SIZE_CODE_MAPS.get(brand, {}).get(size_code)
        if size:
            return product_code[:-2], _purchase_size_from_eu(size, brand)
    return product_code, ""


def _split_purchase_product_code(product_code: str, color_barcodes: list[tuple[str, str]], brand: str) -> tuple[str, str, str, str, str]:
    if len(product_code) < 3:
        return product_code, product_code, "", "", ""
    if len(product_code) >= 5:
        raw_size_code = product_code[-3:]
        if raw_size_code in PURCHASE_MILLIMETER_SIZE_MAP:
            if str(brand or "").strip().lower() == "smiley":
                style_color_code = product_code[:-3]
                color_barcode = _purchase_color_barcode(style_color_code, brand)
                color_name = next((name for barcode, name in color_barcodes if barcode == color_barcode), "")
                return style_color_code, style_color_code, color_barcode, color_name, _purchase_size_from_millimeter(raw_size_code, brand)
            original_sku = product_code[:-5]
            color_barcode = product_code[-5:-3]
            color_name = next((name for barcode, name in color_barcodes if barcode == color_barcode), "")
            return original_sku, original_sku, color_barcode, color_name, _purchase_size_from_millimeter(raw_size_code, brand)
    style_color_code, size = _split_purchase_size_code(product_code, brand)
    color_barcode = _purchase_color_barcode(style_color_code, brand)
    color_name = next((name for barcode, name in color_barcodes if barcode == color_barcode), "")
    return style_color_code, style_color_code, color_barcode, color_name, size


def _split_color_from_style_code(style_color_code: str, color_barcodes: list[tuple[str, str]]) -> tuple[str, str]:
    for color_barcode, color_name in color_barcodes:
        if style_color_code.endswith(color_barcode):
            return color_barcode, color_name
    return "", ""


def _merge_product_info(target: dict[str, object], values: dict[str, object]) -> None:
    for key, value in values.items():
        if value not in (None, "") and not target.get(key):
            target[key] = value


def _extract_color_name_from_text(*values: object) -> str:
    for value in values:
        text = _cell_text(value)
        if not text:
            continue
        for separator in ("/", "／", "-", " ", "　"):
            if separator in text:
                candidate = text.rsplit(separator, 1)[-1].strip()
                if candidate:
                    return candidate
        return text
    return ""


def _load_purchase_product_lookup(connection, brand: str, product_codes: set[str]) -> dict[str, dict[str, object]]:
    lookup: dict[str, dict[str, object]] = {code: {} for code in product_codes if code}
    if not lookup:
        return {}

    codes = set(lookup)
    gj_statement = (
        sa_select(
            GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_code,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.original_goods_code,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_full_name,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.product_name,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.barcode,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.factory_code,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.upper_material,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.lining_material,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.outsole_material,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.insole_material,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.shoe_box_spec,
            GJ_MERGED_PRODUCT_INFO_TABLE.c.extra_fields,
        )
        .where(or_(
            GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_code.in_(codes),
            GJ_MERGED_PRODUCT_INFO_TABLE.c.original_goods_code.in_(codes),
        ))
        .order_by(GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value.desc().nulls_last(), desc(GJ_MERGED_PRODUCT_INFO_TABLE.c.updated_at), desc(GJ_MERGED_PRODUCT_INFO_TABLE.c.id))
    )
    for row in connection.execute(gj_statement).mappings():
        product_name = (
            str(row.get("goods_full_name") or "").strip()
            or str(row.get("product_name") or "").strip()
            or None
        )
        extra_fields = _dict_or_empty(row.get("extra_fields"))
        color_name = _first_text(
            extra_fields.get("颜色名称"),
            extra_fields.get("颜色"),
            extra_fields.get("新色"),
            extra_fields.get("颜色及规格"),
            extra_fields.get("色号"),
            _extract_color_name_from_text(row.get("product_name"), row.get("goods_full_name")) if brand in EU_SIZE_BRANDS else "",
        )
        product_info = {
            "goods_code": _cell_text(row.get("goods_code")),
            "original_goods_code": _cell_text(row.get("original_goods_code")),
            "product_name": product_name,
            "color_name": color_name,
            "color_barcode": _first_text(extra_fields.get("颜色条码"), extra_fields.get("色号")),
            "barcode": _cell_text(row.get("barcode")),
            "factory_code": _cell_text(row.get("factory_code")),
            "upper_material": _cell_text(row.get("upper_material")),
            "lining_material": _cell_text(row.get("lining_material")),
            "outsole_material": _cell_text(row.get("outsole_material")),
            "insole_material": _cell_text(row.get("insole_material")),
            "shoe_box_spec": _cell_text(row.get("shoe_box_spec")),
        }
        for code_key in ("goods_code", "original_goods_code"):
            code = str(row.get(code_key) or "").strip()
            if not code or code not in lookup:
                continue
            _merge_product_info(lookup[code], product_info)

    price_statement = (
        sa_select(
            JST_PRICE_TABLE.c.goods_code,
            JST_PRICE_TABLE.c.goods_full_name,
            JST_PRICE_TABLE.c.latest_purchase_price,
            JST_PRICE_TABLE.c.preset_price,
            JST_PRICE_TABLE.c.cost_unit_price,
        )
        .where(JST_PRICE_TABLE.c.goods_code.in_(codes))
        .order_by(JST_PRICE_TABLE.c.source_date_value.desc().nulls_last(), desc(JST_PRICE_TABLE.c.updated_at), desc(JST_PRICE_TABLE.c.id))
    )
    for row in connection.execute(price_statement).mappings():
        code = str(row.get("goods_code") or "").strip()
        if not code or code not in lookup:
            continue
        price = (
            row.get("cost_unit_price")
            or row.get("latest_purchase_price")
            or row.get("preset_price")
        )
        product_name = str(row.get("goods_full_name") or "").strip()
        if product_name and not lookup[code].get("product_name"):
            lookup[code]["product_name"] = product_name
        if price not in (None, "") and not lookup[code].get("unit_price"):
            lookup[code]["unit_price"] = price

    product_table = PRODUCT_TABLES.get(brand)
    if product_table is not None:
        for row in connection.execute(
            sa_select(
                product_table.c.sku,
                product_table.c.original_sku,
                product_table.c.cost,
                product_table.c.color,
                product_table.c.color_code,
                product_table.c.factory_sku,
                product_table.c.upper_material,
                product_table.c.lining_material,
                product_table.c.outsole_material,
                product_table.c.insole_material,
                product_table.c.shoe_box_spec,
            )
            .where(or_(
                product_table.c.sku.in_(codes),
                product_table.c.original_sku.in_(codes),
            ))
            .order_by(desc(product_table.c.updated_at), desc(product_table.c.id))
        ).mappings():
            for code_key in ("sku", "original_sku"):
                code = str(row.get(code_key) or "").strip()
                if not code or code not in lookup:
                    continue
                if not lookup[code].get("unit_price") and row.get("cost") not in (None, ""):
                    lookup[code]["unit_price"] = row.get("cost")
                product_color = _cell_text(row.get("color"))
                product_color_code = _cell_text(row.get("color_code"))
                if product_color:
                    lookup[code]["color_name"] = product_color
                    lookup[code]["color"] = product_color
                if product_color_code:
                    lookup[code]["color_barcode"] = product_color_code
                    lookup[code]["color_code"] = product_color_code
                fallback_fields = {
                    "original_goods_code": _cell_text(row.get("original_sku")),
                    "factory_code": _cell_text(row.get("factory_sku")),
                    "upper_material": _cell_text(row.get("upper_material")),
                    "lining_material": _cell_text(row.get("lining_material")),
                    "outsole_material": _cell_text(row.get("outsole_material")),
                    "insole_material": _cell_text(row.get("insole_material")),
                    "shoe_box_spec": _cell_text(row.get("shoe_box_spec")),
                }
                _merge_product_info(lookup[code], fallback_fields)
    if brand == "smiley":
        for row in connection.execute(
            sa_select(
                SMILEY_FINE_TABLE.c.sku,
                SMILEY_FINE_TABLE.c.original_sku,
                SMILEY_FINE_TABLE.c.factory_code,
                SMILEY_FINE_TABLE.c.factory_sku,
                SMILEY_FINE_TABLE.c.cost,
                SMILEY_FINE_TABLE.c.product_name,
                SMILEY_FINE_TABLE.c.barcode,
                SMILEY_FINE_TABLE.c.upper_material,
                SMILEY_FINE_TABLE.c.lining_material,
                SMILEY_FINE_TABLE.c.outsole_material,
                SMILEY_FINE_TABLE.c.insole_material,
                SMILEY_FINE_TABLE.c.shoe_box_spec,
                SMILEY_FINE_TABLE.c.raw_payload,
            )
            .where(or_(
                SMILEY_FINE_TABLE.c.sku.in_(codes),
                SMILEY_FINE_TABLE.c.original_sku.in_(codes),
            ))
            .order_by(SMILEY_FINE_TABLE.c.snapshot_date.desc().nulls_last(), desc(SMILEY_FINE_TABLE.c.updated_at), desc(SMILEY_FINE_TABLE.c.id))
        ).mappings():
            raw_payload = _dict_or_empty(row.get("raw_payload"))
            smiley_info = {
                "original_goods_code": _first_text(row.get("original_sku"), row.get("sku")),
                "factory_code": _first_text(row.get("factory_sku"), row.get("factory_code")),
                "unit_price": row.get("cost"),
                "product_name": _cell_text(row.get("product_name")),
                "color_name": _first_text(
                    raw_payload.get("颜色名称"),
                    raw_payload.get("颜色"),
                    raw_payload.get("新色"),
                    raw_payload.get("颜色及规格"),
                    _extract_color_name_from_text(row.get("product_name")),
                ),
                "barcode": _cell_text(row.get("barcode")),
                "upper_material": _cell_text(row.get("upper_material")),
                "lining_material": _cell_text(row.get("lining_material")),
                "outsole_material": _cell_text(row.get("outsole_material")),
                "insole_material": _cell_text(row.get("insole_material")),
                "shoe_box_spec": _cell_text(row.get("shoe_box_spec")),
            }
            for code_key in ("sku", "original_sku"):
                code = str(row.get(code_key) or "").strip()
                if code and code in lookup:
                    _merge_product_info(lookup[code], smiley_info)
    return lookup


def _build_purchase_detail_candidates(connection, query: str, brand: str | None = None, limit: int = 20) -> list[dict[str, str]]:
    keyword = _cell_text(query)
    if len(keyword) < 2:
        return []

    candidate_brands = [brand] if brand else ["cbanner_mens", "cbanner_womens"]
    seen: set[tuple[str, str]] = set()
    items: list[dict[str, str]] = []

    for candidate_brand in candidate_brands:
        normalized_brand = _cell_text(candidate_brand).lower()
        if normalized_brand == "nike":
            normalized_brand = "ni"
        if not normalized_brand:
            continue
        if normalized_brand == "smiley":
            statement = (
                sa_select(
                    SMILEY_FINE_TABLE.c.sku,
                    SMILEY_FINE_TABLE.c.original_sku,
                    SMILEY_FINE_TABLE.c.product_name,
                    SMILEY_FINE_TABLE.c.factory_sku,
                    SMILEY_FINE_TABLE.c.raw_payload,
                )
                .where(or_(
                    SMILEY_FINE_TABLE.c.sku.ilike(f"%{keyword}%"),
                    SMILEY_FINE_TABLE.c.original_sku.ilike(f"%{keyword}%"),
                ))
                .order_by(SMILEY_FINE_TABLE.c.snapshot_date.desc().nulls_last(), desc(SMILEY_FINE_TABLE.c.updated_at), desc(SMILEY_FINE_TABLE.c.id))
                .limit(limit)
            )
            rows = connection.execute(statement).mappings()
            for row in rows:
                sku = _cell_text(row.get("sku"))
                original_sku = _cell_text(row.get("original_sku")) or sku
                key = (normalized_brand, sku)
                if not sku or key in seen:
                    continue
                raw_payload = _dict_or_empty(row.get("raw_payload"))
                items.append({
                    "product_code": sku,
                    "sku": sku,
                    "original_sku": original_sku,
                    "product_name": _cell_text(row.get("product_name")),
                    "color_name": _first_text(
                        raw_payload.get("颜色名称"),
                        raw_payload.get("颜色"),
                        raw_payload.get("新色"),
                        raw_payload.get("颜色及规格"),
                    ),
                    "factory_code": _cell_text(row.get("factory_sku")),
                    "brand": normalized_brand,
                })
                seen.add(key)
                if len(items) >= limit:
                    return items

        product_table = PRODUCT_TABLES.get(normalized_brand)
        if product_table is not None:
            rows = connection.execute(
                sa_select(
                    product_table.c.sku,
                    product_table.c.original_sku,
                    product_table.c.color,
                    product_table.c.factory_sku,
                    product_table.c.product_model,
                )
                .where(or_(
                    product_table.c.sku.ilike(f"%{keyword}%"),
                    product_table.c.original_sku.ilike(f"%{keyword}%"),
                ))
                .order_by(desc(product_table.c.updated_at), desc(product_table.c.id))
                .limit(limit)
            ).mappings()
            for row in rows:
                sku = _cell_text(row.get("sku"))
                original_sku = _cell_text(row.get("original_sku")) or sku
                key = (normalized_brand, sku)
                if not sku or key in seen:
                    continue
                items.append({
                    "product_code": sku,
                    "sku": sku,
                    "original_sku": original_sku,
                    "product_name": _cell_text(row.get("product_model")),
                    "color_name": _cell_text(row.get("color")),
                    "factory_code": _cell_text(row.get("factory_sku")),
                    "brand": normalized_brand,
                })
                seen.add(key)
                if len(items) >= limit:
                    return items

        gj_statement = (
            sa_select(
                GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_code,
                GJ_MERGED_PRODUCT_INFO_TABLE.c.original_goods_code,
                GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_full_name,
                GJ_MERGED_PRODUCT_INFO_TABLE.c.product_name,
                GJ_MERGED_PRODUCT_INFO_TABLE.c.factory_code,
                GJ_MERGED_PRODUCT_INFO_TABLE.c.fine_table_brand,
            )
            .where(or_(
                GJ_MERGED_PRODUCT_INFO_TABLE.c.goods_code.ilike(f"%{keyword}%"),
                GJ_MERGED_PRODUCT_INFO_TABLE.c.original_goods_code.ilike(f"%{keyword}%"),
            ))
            .where(GJ_MERGED_PRODUCT_INFO_TABLE.c.fine_table_brand == normalized_brand)
            .order_by(
                GJ_MERGED_PRODUCT_INFO_TABLE.c.source_date_value.desc().nulls_last(),
                desc(GJ_MERGED_PRODUCT_INFO_TABLE.c.updated_at),
                desc(GJ_MERGED_PRODUCT_INFO_TABLE.c.id),
            )
            .limit(limit)
        )
        for row in connection.execute(gj_statement).mappings():
            sku = _cell_text(row.get("goods_code"))
            original_sku = _cell_text(row.get("original_goods_code")) or sku
            key = (normalized_brand, sku)
            if not sku or key in seen:
                continue
            items.append({
                "product_code": sku,
                "sku": sku,
                "original_sku": original_sku,
                "product_name": _first_text(row.get("goods_full_name"), row.get("product_name")),
                "color_name": "",
                "factory_code": _cell_text(row.get("factory_code")),
                "brand": _cell_text(row.get("fine_table_brand")) or normalized_brand,
            })
            seen.add(key)
            if len(items) >= limit:
                return items

    return items


def _purchase_detail_extra_fields(product_info: dict[str, object], detail_code: object, imported_extra_fields: object | None = None) -> dict[str, str]:
    code = _cell_text(detail_code)
    original_code = _first_text(product_info.get("original_goods_code"), code)
    extra_fields = {
        "image_code": original_code,
        "factory_code": _cell_text(product_info.get("factory_code")),
        "style_code": original_code,
        "inner_color_code": "",
        "upper_material": _cell_text(product_info.get("upper_material")),
        "lining_material": _cell_text(product_info.get("lining_material")),
        "outsole_material": _cell_text(product_info.get("outsole_material")),
        "insole_material": _cell_text(product_info.get("insole_material")),
        "shoe_box_spec": _cell_text(product_info.get("shoe_box_spec")),
    }
    if isinstance(imported_extra_fields, dict):
        for key, value in imported_extra_fields.items():
            if key not in PURCHASE_DETAIL_EXTRA_FIELDS:
                continue
            text = _cell_text(value)
            if text or key == "inner_color_code":
                extra_fields[key] = text
    extra_fields["style_code"] = _first_text(extra_fields.get("image_code"), original_code)
    return extra_fields


def _purchase_size_label(size_code: str, brand: str) -> str:
    if size_code in EU_PURCHASE_SIZE_LABELS:
        return size_code
    return PURCHASE_SIZE_CODE_MAPS.get(brand, {}).get(size_code, size_code)


def _lookup_has_product_data(item: dict[str, object]) -> bool:
    return bool(item.get("_matched_product"))


def _build_purchase_detail_lookup_for_brand(connection, product_code: str, quantity: Decimal, brand: str) -> dict[str, object]:
    raw_code = product_code.strip()
    if not raw_code:
        raise HTTPException(status_code=400, detail="货号不能为空")

    stripped_code, size = _split_purchase_size_code(raw_code, brand)
    lookup_codes = {raw_code}
    if stripped_code:
        lookup_codes.add(stripped_code)
    product_lookup = _load_purchase_product_lookup(connection, brand, lookup_codes)

    product_info = product_lookup.get(raw_code) or {}
    detail_code = raw_code
    if not product_info and stripped_code and stripped_code != raw_code:
        product_info = product_lookup.get(stripped_code) or {}
        if product_info:
            detail_code = stripped_code

    if not product_info and stripped_code != raw_code:
        detail_code = stripped_code

    fallback_color_barcode, fallback_color_name = _lookup_color_name_by_tail(connection, detail_code, brand)
    color_barcode, color_name = _purchase_detail_color_values(product_info, fallback_color_barcode, fallback_color_name)
    unit_price = _to_decimal(product_info.get("unit_price"))
    amount = quantity * unit_price if quantity and unit_price else Decimal("0")
    product_name = str(product_info.get("product_name") or "").strip()
    if not product_name:
        product_name = f"{detail_code}{color_name}" if color_name else detail_code

    size_quantities = {size: _fmt_decimal(quantity)} if size and quantity else {}
    extra_fields = _purchase_detail_extra_fields(product_info, detail_code)
    return {
        "product_code": detail_code,
        "product_name": product_name,
        "color_spec": color_name,
        "color_barcode": color_barcode,
        "color_name": color_name,
        "extra_fields": extra_fields,
        "quantity": _fmt_decimal(quantity) if quantity else None,
        "unit_price": _fmt_decimal(unit_price) if unit_price else None,
        "amount": _fmt_decimal(amount) if amount else None,
        "size_quantities": size_quantities,
        "_matched_product": bool(product_info),
    }


def _build_purchase_detail_lookup(connection, product_code: str, quantity: Decimal, brand: str | None = None) -> dict[str, object]:
    brands = [brand] if brand else ["cbanner_mens", "cbanner_womens"]
    fallback: dict[str, object] | None = None
    for candidate in brands:
        if not candidate:
            continue
        item = _build_purchase_detail_lookup_for_brand(connection, product_code, quantity, candidate)
        if _lookup_has_product_data(item):
            item.pop("_matched_product", None)
            return item
        if fallback is None:
            fallback = item
    item = fallback or _build_purchase_detail_lookup_for_brand(connection, product_code, quantity, "cbanner_mens")
    item.pop("_matched_product", None)
    return item


def _build_purchase_details_from_rows(
    repository,
    rows: list[dict[str, object]],
    *,
    brand: str,
    fallback_unit_price: Decimal,
    prefer_lookup_unit_price: bool = False,
) -> list[dict[str, object]]:
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的明细")

    with repository.engine.connect() as connection:
        color_barcodes = _load_color_barcodes(connection)

    parsed_rows = []
    product_codes = set()
    for row in rows:
        raw_code = str(row["product_code"])
        quantity = _to_decimal(row["quantity"])
        if quantity == 0:
            continue
        style_color_code, _ = _split_purchase_size_code(raw_code, brand)
        imported_size_quantities = row.get("size_quantities")
        if isinstance(imported_size_quantities, dict) and imported_size_quantities:
            sku = raw_code
            original_sku = raw_code
            color_barcode = _purchase_color_barcode(raw_code, brand)
            color_name = next((name for barcode, name in color_barcodes if barcode == color_barcode), "")
            size = ""
            normalized_sizes = {
                _purchase_size_from_millimeter(str(size_key), brand): _to_decimal(size_quantity)
                for size_key, size_quantity in imported_size_quantities.items()
                if _to_decimal(size_quantity) != 0
            }
        else:
            sku, original_sku, color_barcode, color_name, size = _split_purchase_product_code(raw_code, color_barcodes, brand)
            normalized_sizes = {}
        parsed_rows.append({
            "raw_code": raw_code,
            "quantity": quantity,
            "sku": sku,
            "original_sku": original_sku,
            "style_color_code": style_color_code,
            "color_barcode": color_barcode,
            "color_name": color_name,
            "size": size,
            "unit_price": row.get("unit_price") or "",
            "remark": _cell_text(row.get("remark")),
            "size_quantities": normalized_sizes,
            "extra_fields": row.get("extra_fields") if isinstance(row.get("extra_fields"), dict) else {},
        })
        if original_sku:
            product_codes.add(original_sku)
        if sku:
            product_codes.add(sku)
        if raw_code:
            product_codes.add(raw_code)
        if style_color_code:
            product_codes.add(style_color_code)

    with repository.engine.connect() as connection:
        product_lookup = _load_purchase_product_lookup(connection, brand, product_codes)

    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for row in parsed_rows:
        raw_code = row["raw_code"]
        quantity = row["quantity"]
        original_sku = row["original_sku"]
        color_barcode = row["color_barcode"]
        color_name = row["color_name"] or ""
        product_info = (
            product_lookup.get(original_sku)
            or product_lookup.get(row.get("style_color_code"))
            or product_lookup.get(row.get("sku"))
            or product_lookup.get(raw_code)
            or {}
        )
        color_barcode, color_name = _purchase_detail_color_values(product_info, color_barcode, color_name)
        imported_unit_price = _to_decimal(row.get("unit_price"))
        lookup_unit_price = _to_decimal(product_info.get("unit_price"))
        if prefer_lookup_unit_price:
            unit_price = lookup_unit_price or fallback_unit_price
        else:
            unit_price = imported_unit_price or lookup_unit_price or fallback_unit_price
        product_name = str(product_info.get("product_name") or "").strip()
        if not product_name:
            product_name = f"{original_sku}{color_name}" if color_name else original_sku
        extra_fields = _purchase_detail_extra_fields(product_info, original_sku, row.get("extra_fields"))
        size = row["size"]
        key = (original_sku, color_barcode)
        item = grouped.setdefault(
            key,
            {
                "product_code": original_sku,
                "product_name": product_name,
                "color_spec": color_name,
                "color_barcode": color_barcode,
                "color_name": color_name,
                "size_quantities": defaultdict(Decimal),
                "quantity": Decimal("0"),
                "unit_price": unit_price,
                "remark": _cell_text(row.get("remark")),
                "extra_fields": extra_fields,
                "raw_codes": [],
            },
        )
        if row.get("remark") and not item.get("remark"):
            item["remark"] = _cell_text(row.get("remark"))
        if isinstance(item.get("extra_fields"), dict):
            for field_key, field_value in extra_fields.items():
                if (field_value or field_key == "inner_color_code") and not item["extra_fields"].get(field_key):
                    item["extra_fields"][field_key] = field_value
        item["quantity"] = item["quantity"] + quantity
        item["raw_codes"].append(raw_code)
        row_size_quantities = row.get("size_quantities")
        if isinstance(row_size_quantities, dict) and row_size_quantities:
            for size_key, size_quantity in row_size_quantities.items():
                item["size_quantities"][size_key] += size_quantity
        elif size:
            item["size_quantities"][size] += quantity

    if not grouped:
        raise HTTPException(status_code=400, detail="Excel 中没有有效数量")

    details = []
    size_labels = _purchase_size_labels(brand)
    for item in grouped.values():
        quantity = item["quantity"]
        item_unit_price = _to_decimal(item.get("unit_price"))
        amount = quantity * item_unit_price if item_unit_price else Decimal("0")
        size_quantities = {
            size: _fmt_decimal(item["size_quantities"].get(size, Decimal("0")))
            for size in size_labels
            if item["size_quantities"].get(size, Decimal("0")) != 0
        }
        details.append({
            "product_code": item["product_code"],
            "product_name": item["product_name"],
            "color_spec": item["color_spec"],
            "color_barcode": item["color_barcode"],
            "color_name": item["color_name"],
            "quantity": _fmt_decimal(quantity),
            "unit_price": _fmt_decimal(item_unit_price) if item_unit_price else None,
            "amount": _fmt_decimal(amount) if amount else None,
            "remark": item.get("remark") or "",
            "size_quantities": size_quantities,
            "extra_fields": item.get("extra_fields") or {},
        })
    return details


def _build_purchase_details_from_excel(
    repository,
    content: bytes,
    *,
    brand: str,
    fallback_unit_price: Decimal,
    prefer_lookup_unit_price: bool = False,
) -> tuple[list[dict[str, object]], str]:
    rows, sheet_name = _read_purchase_import_rows(content)
    details = _build_purchase_details_from_rows(
        repository,
        rows,
        brand=brand,
        fallback_unit_price=fallback_unit_price,
        prefer_lookup_unit_price=prefer_lookup_unit_price,
    )
    return details, sheet_name


def _purchase_import_brand_for_supplier(
    repository,
    supplier: str,
    document_type: str,
    fallback_brand: str,
) -> str:
    supplier_name = _cell_text(supplier)
    if supplier_name:
        supplier_record = repository.get_supplier_by_name(supplier_name)
        supplier_brand = _cell_text(supplier_record.get("brand") if supplier_record else "").lower()
        if supplier_brand:
            return supplier_brand
        inferred_brand = infer_supplier_brand_from_name(supplier_name)
        if inferred_brand:
            return inferred_brand
    return fallback_brand or _purchase_import_brand(document_type)


def _group_purchase_import_rows_by_summary(
    rows: list[dict[str, object]],
    fallback_summary: str,
) -> list[dict[str, object]]:
    groups: dict[str, dict[str, object]] = {}
    ordered_keys: list[str] = []
    carried_fields: dict[str, str] = {}
    doc_fields = tuple(PURCHASE_IMPORT_DOC_FIELD_ALIASES.keys())

    for row_index, row in enumerate(rows, start=1):
        row_fields: dict[str, str] = {}
        for field in doc_fields:
            value = _cell_text(row.get(field))
            if value:
                carried_fields[field] = value
                row_fields[field] = value

        effective_fields = {
            field: row_fields.get(field) or carried_fields.get(field, "")
            for field in doc_fields
        }
        summary = effective_fields.get("summary") or fallback_summary
        group_key = summary or f"__missing_summary_{row_index}"
        if group_key not in groups:
            groups[group_key] = {"summary": summary, "fields": {}, "rows": []}
            ordered_keys.append(group_key)

        group = groups[group_key]
        fields = group["fields"]
        if isinstance(fields, dict):
            for field, value in effective_fields.items():
                if value and not fields.get(field):
                    fields[field] = value
        group_rows = group["rows"]
        if isinstance(group_rows, list):
            group_rows.append(row)

    return [groups[key] for key in ordered_keys]


def _missing_purchase_order_import_fields(rows: list[dict[str, object]]) -> list[str]:
    return [
        field
        for field in PURCHASE_ORDER_IMPORT_REQUIRED_DOC_FIELDS
        if not any(_cell_text(row.get(field)) for row in rows)
    ]


def _purchase_order_import_has_size_columns(rows: list[dict[str, object]]) -> bool:
    return any(bool(_dict_or_empty(row.get("size_quantities"))) for row in rows)


@router.get("/inventory")
def list_inventory(
    request: Request,
    date_start: str | None = None,
    date_end: str | None = None,
    supplier: str | None = None,
    warehouse: str | None = None,
    document_type: str | None = None,
    exclude_document_type: str | None = None,
    summary: str | None = None,
    original_sku: str | None = None,
    product_code: str | None = None,
    handler: str | None = None,
    completion_status: str | None = None,
    page: int = 1,
    page_size: int = 20,
):
    repository = request.app.state.inventory_repository
    document_type = normalize_document_type(document_type) if document_type else None
    exclude_document_type = normalize_document_type(exclude_document_type) if exclude_document_type else None
    return repository.list_records(
        date_start=date_start,
        date_end=date_end,
        supplier=supplier,
        warehouse=warehouse,
        document_type=document_type,
        exclude_document_type=exclude_document_type,
        summary=summary,
        original_sku=original_sku,
        product_code=product_code,
        handler=handler,
        completion_status=completion_status,
        page=page,
        page_size=page_size,
    )


@router.get("/inventory-reports/purchase-inbound-details")
def list_purchase_inbound_details(
    request: Request,
    date_start: str | None = None,
    date_end: str | None = None,
    document_type: str | None = None,
    supplier: str | None = None,
    warehouse: str | None = None,
    product_code: str | None = None,
    product_name: str | None = None,
    color_name: str | None = None,
    size_name: str | None = None,
    page: int = 1,
    page_size: int = 50,
):
    repository = request.app.state.inventory_repository
    normalized_document_type = normalize_document_type(document_type) if document_type else None
    if normalized_document_type and normalized_document_type not in {"进货单", "进货退货单"}:
        raise HTTPException(status_code=400, detail="单据类型仅支持进货单或进货退货单")
    return repository.list_purchase_inbound_details(
        date_start=date_start,
        date_end=date_end,
        document_type=normalized_document_type,
        supplier=supplier,
        warehouse=warehouse,
        product_code=product_code,
        product_name=product_name,
        color_name=color_name,
        size_name=size_name,
        page=page,
        page_size=page_size,
    )


@router.get("/inventory/recycle-bin")
def list_inventory_recycle_bin(
    request: Request,
    document_type: str | None = None,
    exclude_document_type: str | None = None,
    page: int = 1,
    page_size: int = 20,
):
    repository = request.app.state.inventory_repository
    document_type = normalize_document_type(document_type) if document_type else None
    exclude_document_type = normalize_document_type(exclude_document_type) if exclude_document_type else None
    return repository.list_deleted_records(
        page=page,
        page_size=page_size,
        document_type=document_type,
        exclude_document_type=exclude_document_type,
    )


@router.get("/inventory/export")
def export_inventory(
    request: Request,
    ids: str | None = None,
    date_start: str | None = None,
    date_end: str | None = None,
    supplier: str | None = None,
    warehouse: str | None = None,
    document_type: str | None = None,
    exclude_document_type: str | None = None,
    summary: str | None = None,
    original_sku: str | None = None,
    product_code: str | None = None,
    handler: str | None = None,
    completion_status: str | None = None,
    purchase_export_mode: str | None = None,
):
    repository = request.app.state.inventory_repository
    document_type = normalize_document_type(document_type) if document_type else None
    exclude_document_type = normalize_document_type(exclude_document_type) if exclude_document_type else None
    selected_ids: set[int] | None = None
    if ids:
        selected_ids = set()
        for raw_id in ids.split(","):
            raw_id = raw_id.strip()
            if not raw_id:
                continue
            try:
                selected_ids.add(int(raw_id))
            except ValueError:
                raise HTTPException(status_code=400, detail="导出单据 ID 格式错误") from None
        if not selected_ids:
            selected_ids = None
    result = repository.list_records(
        date_start=date_start,
        date_end=date_end,
        supplier=supplier,
        warehouse=warehouse,
        document_type=document_type,
        exclude_document_type=exclude_document_type,
        summary=summary,
        original_sku=original_sku,
        product_code=product_code,
        handler=handler,
        completion_status=completion_status,
        page=1,
        page_size=100_000,
    )
    items = [
        item for item in result["items"]
        if item.get("document_type") not in ACCOUNTING_DOCUMENT_TYPES
    ]
    if selected_ids is not None:
        items = [item for item in items if int(item.get("id") or 0) in selected_ids]
    details = repository.list_details_for_documents([int(item["id"]) for item in items])
    records_by_id = {item["id"]: item for item in items}
    is_purchase_detail_export = document_type == "进货订单"

    if is_purchase_detail_export:
        normalized_purchase_export_mode = str(purchase_export_mode or "summary").strip() or "summary"
        if normalized_purchase_export_mode not in PURCHASE_EXPORT_MODES:
            raise HTTPException(status_code=400, detail="采购单导出类型仅支持 summary、size_rows 或 production_order")

        wb = Workbook()
        ws = wb.active
        supplier_lookup = _load_supplier_export_lookup(repository, items)
        if normalized_purchase_export_mode == PURCHASE_PRODUCTION_ORDER_EXPORT_MODE:
            requirement_templates = repository.get_purchase_order_requirement_template_map()
            _append_purchase_production_order_export(wb, details, records_by_id, supplier_lookup, requirement_templates)
            return _stream_excel_workbook(wb, "生产采购单.xlsx")
        if normalized_purchase_export_mode == "size_rows":
            ws.title = "尺码明细"
            _append_purchase_size_rows_export(ws, details, records_by_id, supplier_lookup)
            return _stream_excel_workbook(wb, "进货订单尺码明细.xlsx")

        ws.title = "汇总明细"
        _append_purchase_summary_export(ws, details, records_by_id, supplier_lookup)
        return _stream_excel_workbook(wb, "进货订单汇总明细.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "经营历程"

    headers = [INVENTORY_EXPORT_LABELS.get(c, c) for c in INVENTORY_CANONICAL_COLUMNS]
    ws.append(headers)

    for item in items:
        row = [item.get(c) for c in INVENTORY_CANONICAL_COLUMNS]
        ws.append(row)

    detail_ws = wb.create_sheet("单据明细")
    detail_headers = [
        "单据编号",
        "订货日期",
        "交货日期",
        "单据类型",
        "供应商/客户/出货仓库",
        "仓库",
        "经手人",
        "摘要",
        "货号",
        "商品全名",
        "颜色条码",
        "颜色名称",
        *PURCHASE_EXPORT_SIZE_LABELS,
        "数量",
        "单价",
        "金额",
    ]
    detail_ws.append(detail_headers)
    for detail in details:
        record = records_by_id.get(detail.get("document_id"), {})
        extra_fields = record.get("extra_fields") or {}
        if not isinstance(extra_fields, dict):
            extra_fields = {}
        size_quantities = detail.get("size_quantities") or {}
        if not isinstance(size_quantities, dict):
            size_quantities = {}
        detail_extra_fields = detail.get("extra_fields") or {}
        if not isinstance(detail_extra_fields, dict):
            detail_extra_fields = {}
        common_values = [
            record.get("document_number") or record.get("id") or "",
            record.get("date") or "",
            extra_fields.get("delivery_date") or "",
            record.get("document_type") or "",
            record.get("supplier") or "",
            record.get("warehouse") or "",
            record.get("handler") or "",
            record.get("summary") or "",
        ]
        detail_ws.append([
            *common_values,
            detail.get("product_code") or "",
            detail.get("product_name") or "",
            detail.get("color_barcode") or "",
            detail.get("color_name") or detail.get("color_spec") or "",
            *[size_quantities.get(size, "") for size in PURCHASE_EXPORT_SIZE_LABELS],
            detail.get("quantity") or "",
            detail.get("unit_price") or "",
            detail.get("amount") or "",
        ])

    style_excel_workbook(wb)
    return _stream_excel_workbook(wb, "进销存数据.xlsx")


@router.get("/inventory/ending-balance")
def get_ending_inventory(
    request: Request,
    stock_date: str,
    date_start: str | None = None,
    date_end: str | None = None,
    product_code: str | None = None,
    page: int = 1,
    page_size: int = 20,
):
    repository = request.app.state.inventory_repository
    settings = request.app.state.settings
    return repository.get_ending_inventory(
        jst_stock_root=settings.jst_stock_root,
        stock_date=stock_date,
        date_start=date_start,
        date_end=date_end,
        product_code=product_code,
        page=page,
        page_size=page_size,
    )


@router.post("/inventory/import-jst-stock")
def import_jst_stock(request: Request, stock_date: str | None = None):
    if stock_date is None:
        from datetime import datetime
        now = datetime.now()
        stock_date = f"{now.month:02d}.{now.day:02d}"
    repository = request.app.state.inventory_repository
    settings = request.app.state.settings
    return repository.import_jst_stock(
        jst_stock_root=settings.jst_stock_root,
        stock_date=stock_date,
    )


@router.get("/inventory/import-purchase/template")
def download_purchase_import_template():
    return _stream_excel_workbook(_build_purchase_order_import_template(), "采购单导入模板.xlsx")


@router.post("/inventory/import-purchase")
async def import_purchase_inventory(request: Request, file: UploadFile = None):
    if file is None:
        raise HTTPException(status_code=400, detail="No file uploaded")

    form = await request.form()
    document_type = normalize_document_type(form.get("document_type")) or "进货订单"
    if document_type not in PURCHASE_IMPORT_TYPES:
        raise HTTPException(status_code=400, detail="只支持进货订单、进货单、进货退货单、报溢单、报损单、批发销售单、批发销售退货单、同价调拨单导入")
    supplier = str(form.get("supplier") or "").strip()
    warehouse = str(form.get("warehouse") or "").strip()
    handler = str(form.get("handler") or "").strip()
    summary = str(form.get("summary") or "").strip()
    date_value = _normalize_date(str(form.get("date") or "")) or _today_text()
    delivery_date = _normalize_date(str(form.get("delivery_date") or "")) or None
    brand = str(form.get("brand") or "").strip() or _purchase_import_brand(document_type)
    fallback_unit_price = _to_decimal(form.get("unit_price"))
    overwrite_existing = str(form.get("overwrite_existing") or "").strip().lower() in {"1", "true", "yes", "on"}

    content = await file.read()
    repository = request.app.state.inventory_repository
    rows, sheet_name = _read_purchase_import_rows(content)
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的明细")

    is_purchase_order_import = document_type == "进货订单"
    if is_purchase_order_import:
        if _purchase_order_import_has_size_columns(rows):
            raise HTTPException(status_code=400, detail="采购单导入只支持新模板：商品货号（完整商品编码）+ 数量，不再支持尺码列旧模板")
        missing_fields = _missing_purchase_order_import_fields(rows)
        if missing_fields:
            labels = "、".join(PURCHASE_ORDER_IMPORT_FIELD_LABELS[field] for field in missing_fields)
            raise HTTPException(status_code=400, detail=f"采购单导入模板必须包含并填写：{labels}")

    groups = _group_purchase_import_rows_by_summary(rows, "" if is_purchase_order_import else summary)
    if not groups:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的明细")

    plans: list[dict[str, object]] = []
    duplicate_summaries: list[str] = []
    for index, group in enumerate(groups, start=1):
        fields = group.get("fields") if isinstance(group.get("fields"), dict) else {}
        group_summary = _cell_text(fields.get("summary")) if is_purchase_order_import else _first_text(fields.get("summary"), summary)
        group_supplier = _cell_text(fields.get("supplier")) if is_purchase_order_import else _first_text(fields.get("supplier"), supplier)
        group_warehouse = _cell_text(fields.get("warehouse")) if is_purchase_order_import else _first_text(fields.get("warehouse"), warehouse)
        group_handler = _cell_text(fields.get("handler")) if is_purchase_order_import else _first_text(fields.get("handler"), handler)
        group_date = (
            _normalize_date(fields.get("date"))
            if is_purchase_order_import
            else _normalize_date(_first_text(fields.get("date"), date_value)) or _today_text()
        )
        group_delivery_date = (
            _normalize_date(fields.get("delivery_date"))
            if is_purchase_order_import
            else _normalize_date(_first_text(fields.get("delivery_date"), delivery_date)) or None
        )

        label = group_summary or f"第 {index} 组"
        if document_type not in {"报溢单", "报损单"} and not group_supplier:
            if document_type == "同价调拨单":
                raise HTTPException(status_code=400, detail=f"{label}：出货仓库不能为空")
            field_name = "收货客户" if document_type.startswith("批发销售") else "供货单位"
            raise HTTPException(status_code=400, detail=f"{label}：{field_name}不能为空")
        if not group_warehouse:
            if document_type == "同价调拨单":
                raise HTTPException(status_code=400, detail=f"{label}：入货仓库不能为空")
            field_name = "发货仓库" if document_type.startswith("批发销售") else "收货仓库"
            raise HTTPException(status_code=400, detail=f"{label}：{field_name}不能为空")
        if not group_handler:
            raise HTTPException(status_code=400, detail=f"{label}：经手人不能为空")
        if not group_summary:
            raise HTTPException(status_code=400, detail=f"第 {index} 组：摘要不能为空")
        if is_purchase_order_import and not group_date:
            raise HTTPException(status_code=400, detail=f"{label}：采购日期不能为空")
        if document_type == "进货订单" and not group_delivery_date:
            raise HTTPException(status_code=400, detail=f"{label}：交货日期不能为空")
        existing_record = repository.get_record_by_summary(group_summary)
        if existing_record:
            duplicate_summaries.append(group_summary)

        group_rows = group.get("rows") if isinstance(group.get("rows"), list) else []
        group_brand = _purchase_import_brand_for_supplier(repository, group_supplier, document_type, brand)
        plans.append({
            "summary": group_summary,
            "supplier": group_supplier,
            "warehouse": group_warehouse,
            "handler": group_handler,
            "date": group_date,
            "delivery_date": group_delivery_date,
            "brand": group_brand,
            "rows": group_rows,
            "existing_record": existing_record,
        })

    if duplicate_summaries and not overwrite_existing:
        preview = "、".join(duplicate_summaries[:5])
        suffix = "等" if len(duplicate_summaries) > 5 else ""
        return {
            "created": 0,
            "details": 0,
            "requires_confirmation": True,
            "duplicate_summaries": duplicate_summaries,
            "message": f"摘要 {preview}{suffix} 已存在，确认后将覆盖这些单据的主信息和全部明细，是否继续？",
        }

    for plan in plans:
        plan["details"] = _build_purchase_details_from_rows(
            repository,
            plan["rows"],
            brand=str(plan["brand"]),
            fallback_unit_price=fallback_unit_price,
            prefer_lookup_unit_price=is_purchase_order_import,
        )

    created_docs = 0
    created_details = 0
    overwritten_docs = 0
    overwritten_details = 0
    first_doc: dict[str, object] | None = None
    created_records: list[dict[str, object]] = []
    overwritten_records: list[dict[str, object]] = []
    overwritten_before: list[dict[str, object]] = []
    for plan in plans:
        details = plan["details"] if isinstance(plan.get("details"), list) else []
        total_count = sum((_to_decimal(detail.get("quantity")) for detail in details), Decimal("0"))
        total_amount = sum((_to_decimal(detail.get("amount")) for detail in details), Decimal("0"))
        plan_delivery_date = _cell_text(plan.get("delivery_date"))
        plan_brand = _cell_text(plan.get("brand"))
        doc_payload = {
            "date": plan["date"],
            "supplier": plan["supplier"],
            "warehouse": plan["warehouse"],
            "document_type": document_type,
            "handler": plan["handler"],
            "summary": plan["summary"],
            "total_count": _fmt_decimal(total_count),
            "amount": _fmt_decimal(total_amount) if total_amount else None,
            "source_workbook": file.filename or "",
            "source_sheet": sheet_name,
            "source_row_number": "import_purchase",
            "extra_fields": {"delivery_date": plan_delivery_date} if plan_delivery_date else None,
            "raw_payload": {"import_type": "purchase_detail", "brand": plan_brand, "delivery_date": plan_delivery_date},
        }
        detail_payloads = []
        existing_record = plan.get("existing_record") if isinstance(plan.get("existing_record"), dict) else None
        if existing_record and overwrite_existing:
            doc_id = existing_record.get("id")
            if doc_id is None:
                raise HTTPException(status_code=400, detail=f"摘要 {plan['summary']} 对应的旧单据编号异常，无法覆盖")
            before_details = repository.list_details(int(doc_id))
            doc_payload["deleted_at"] = None
            doc = repository.update_record(int(doc_id), doc_payload)
            if doc is None:
                raise HTTPException(status_code=404, detail=f"摘要 {plan['summary']} 对应的旧单据不存在，无法覆盖")
            for detail in details:
                item = dict(detail)
                item["document_id"] = doc["id"]
                detail_payloads.append(item)
            repository.replace_details(doc["id"], detail_payloads)
            doc = repository.get_record(int(doc["id"])) or doc
            overwritten_docs += 1
            overwritten_details += len(detail_payloads)
            overwritten_records.append(doc)
            overwritten_before.append({
                "record": existing_record,
                "details": before_details[:200],
                "detail_count": len(before_details),
            })
        else:
            doc = repository.create_record(doc_payload)
            created_records.append(doc)
            created_docs += 1
            for detail in details:
                item = dict(detail)
                item["document_id"] = doc["id"]
                detail_payloads.append(item)
            repository.create_details(detail_payloads, doc["id"])
            created_details += len(detail_payloads)
        if first_doc is None:
            first_doc = doc
    total_details = created_details + overwritten_details
    summary_parts = [f"新增 {created_docs} 条单据"]
    if overwritten_docs:
        summary_parts.append(f"覆盖 {overwritten_docs} 条单据")
    summary_parts.append(f"{total_details} 条明细")
    result_message = f"导入完成：{'，'.join(summary_parts)}"

    write_operation_log(
        request,
        module="purchase" if document_type == "进货订单" else "inventory",
        action="import_purchase",
        entity_type="inventory_record",
        entity_label=file.filename or "采购/进销存导入",
        summary=f"导入{document_type}：{'，'.join(summary_parts)}",
        before_data={
            "overwritten": overwritten_before[:200],
            "overwritten_count": overwritten_docs,
        } if overwritten_docs else None,
        after_data={
            "filename": file.filename,
            "document_type": document_type,
            "documents": created_records[:200],
            "overwritten_documents": overwritten_records[:200],
            "document_count": created_docs,
            "overwritten_count": overwritten_docs,
            "detail_count": total_details,
            "created_detail_count": created_details,
            "overwritten_detail_count": overwritten_details,
        },
    )
    return {
        "created": created_docs,
        "overwritten": overwritten_docs,
        "details": total_details,
        "message": result_message,
        "item": first_doc,
    }


@router.get("/inventory/purchase-order-requirements")
def list_purchase_order_requirements(request: Request):
    repository = request.app.state.inventory_repository
    saved_rows = {
        _cell_text(row.get("brand")).lower(): row
        for row in repository.list_purchase_order_requirement_templates()
        if _cell_text(row.get("brand"))
    }
    items: list[dict[str, object]] = []
    seen_brands: set[str] = set()
    for brand in PURCHASE_ORDER_REQUIREMENT_BRANDS:
        row = saved_rows.get(brand)
        seen_brands.add(brand)
        items.append({
            "brand": brand,
            "label": PURCHASE_ORDER_REQUIREMENT_BRAND_LABELS.get(brand, brand),
            "content": str(row.get("content") if row and row.get("content") is not None else PURCHASE_ORDER_REQUIREMENTS_BY_BRAND.get(brand, "")),
            "default_content": PURCHASE_ORDER_REQUIREMENTS_BY_BRAND.get(brand, ""),
            "updated_at": row.get("updated_at") if row else None,
            "is_custom": row is not None,
        })

    for brand, row in sorted(saved_rows.items()):
        if brand in seen_brands:
            continue
        items.append({
            "brand": brand,
            "label": PURCHASE_ORDER_REQUIREMENT_BRAND_LABELS.get(brand, brand),
            "content": str(row.get("content") if row.get("content") is not None else ""),
            "default_content": PURCHASE_ORDER_REQUIREMENTS_BY_BRAND.get(brand, ""),
            "updated_at": row.get("updated_at"),
            "is_custom": True,
        })
    return {"items": items}


@router.put("/inventory/purchase-order-requirements/{brand}")
def update_purchase_order_requirement(request: Request, brand: str, payload: dict):
    normalized_brand = str(brand or "").strip().lower()
    if normalized_brand not in SUPPLIER_BRANDS:
        raise HTTPException(status_code=400, detail="品牌无效")
    repository = request.app.state.inventory_repository
    content = payload.get("content")
    item = repository.upsert_purchase_order_requirement_template(
        normalized_brand,
        "" if content is None else str(content),
    )
    write_operation_log(
        request,
        module="purchase",
        action="update_requirement",
        entity_type="purchase_order_requirement",
        entity_id=normalized_brand,
        entity_label=PURCHASE_ORDER_REQUIREMENT_BRAND_LABELS.get(normalized_brand, normalized_brand),
        summary=f"修改订单要求 {PURCHASE_ORDER_REQUIREMENT_BRAND_LABELS.get(normalized_brand, normalized_brand)}",
        after_data=item,
    )
    return {
        "item": {
            **item,
            "label": PURCHASE_ORDER_REQUIREMENT_BRAND_LABELS.get(normalized_brand, normalized_brand),
            "default_content": PURCHASE_ORDER_REQUIREMENTS_BY_BRAND.get(normalized_brand, ""),
            "is_custom": True,
        },
        "message": "保存成功",
    }


@router.get("/inventory/general-customer-shops")
def list_general_customer_shops(request: Request):
    repository = request.app.state.inventory_repository
    return {"items": repository.list_general_customer_shops()}


@router.get("/inventory/account-subjects")
def list_inventory_account_subjects(request: Request):
    repository = request.app.state.inventory_repository
    return {"items": repository.list_account_subjects()}


@router.post("/inventory/account-subjects")
def create_inventory_account_subject(request: Request, payload: dict):
    repository = request.app.state.inventory_repository
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="科目名称不能为空")
    try:
        item = repository.create_account_subject({
            "code": payload.get("code"),
            "name": name,
        })
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"科目 '{name}' 已存在或无法创建") from error
    label = str(item.get("name") or item.get("id") or "").strip()
    write_operation_log(
        request,
        module="account_subject",
        action="create",
        entity_type="inventory_account_subject",
        entity_id=item.get("id"),
        entity_label=label,
        summary=f"新增科目 {label}".strip(),
        before_data=None,
        after_data=item,
    )
    return {
        "item": item,
        "message": "创建成功",
    }


@router.delete("/inventory/account-subjects/{subject_id}")
def delete_inventory_account_subject(request: Request, subject_id: int):
    repository = request.app.state.inventory_repository
    before = repository.get_account_subject(subject_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Subject not found")
    if not repository.delete_account_subject(subject_id):
        raise HTTPException(status_code=404, detail="Subject not found")
    label = str(before.get("name") or subject_id).strip()
    write_operation_log(
        request,
        module="account_subject",
        action="delete",
        entity_type="inventory_account_subject",
        entity_id=subject_id,
        entity_label=label,
        summary=f"删除科目 {label}".strip(),
        before_data=before,
        after_data=None,
    )
    return {"message": "删除成功"}


@router.get("/inventory/general-customer-brands")
def list_general_customer_brands(request: Request):
    repository = request.app.state.inventory_repository
    return {"items": repository.list_general_customer_brands()}


@router.post("/inventory/general-customer-brands")
def create_general_customer_brand(request: Request, payload: dict):
    repository = request.app.state.inventory_repository
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="品牌名称不能为空")
    if repository.get_general_customer_brand_by_name(name):
        raise HTTPException(status_code=400, detail=f"品牌 '{name}' 已存在")
    item = repository.create_general_customer_brand({
        "name": name,
    })
    label = str(item.get("name") or item.get("id") or "").strip()
    write_operation_log(
        request,
        module="general_customer",
        action="create",
        entity_type="general_customer_brand",
        entity_id=item.get("id"),
        entity_label=label,
        summary=f"新增一般客户品牌 {label}".strip(),
        before_data=None,
        after_data=item,
    )
    return {
        "item": item,
        "message": "创建成功",
    }


@router.put("/inventory/general-customer-brands/{brand_id}")
def update_general_customer_brand(request: Request, brand_id: int, payload: dict):
    repository = request.app.state.inventory_repository
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="品牌名称不能为空")
    existing = repository.get_general_customer_brand_by_name(name)
    if existing and existing.get("id") != brand_id:
        raise HTTPException(status_code=400, detail=f"品牌 '{name}' 已存在")
    before = repository.get_general_customer_brand(brand_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Brand not found")
    record = repository.update_general_customer_brand(brand_id, {
        "name": name,
    })
    if record is None:
        raise HTTPException(status_code=404, detail="Brand not found")
    label = str(record.get("name") or before.get("name") or brand_id).strip()
    changes = build_changed_fields(before, record, GENERAL_CUSTOMER_BRAND_FIELD_LABELS)
    write_operation_log(
        request,
        module="general_customer",
        action="update",
        entity_type="general_customer_brand",
        entity_id=brand_id,
        entity_label=label,
        summary=summarize_changes("编辑一般客户品牌", label, changes),
        changed_fields=changes,
        before_data=before,
        after_data=record,
    )
    return {"item": record, "message": "更新成功"}


@router.delete("/inventory/general-customer-brands/{brand_id}")
def delete_general_customer_brand(request: Request, brand_id: int):
    repository = request.app.state.inventory_repository
    before = repository.get_general_customer_brand(brand_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Brand not found")
    result = repository.delete_general_customer_brand(brand_id)
    if result == "not_found":
        raise HTTPException(status_code=404, detail="Brand not found")
    label = str(before.get("name") or brand_id).strip()
    write_operation_log(
        request,
        module="general_customer",
        action="delete",
        entity_type="general_customer_brand",
        entity_id=brand_id,
        entity_label=label,
        summary=f"删除一般客户品牌 {label}".strip(),
        before_data=before,
        after_data=None,
    )
    return {"message": "删除成功"}


@router.post("/inventory/general-customer-shops")
def create_general_customer_shop(request: Request, payload: dict):
    repository = request.app.state.inventory_repository
    customer_name = str(payload.get("customer_name") or "").strip()
    shop_name = str(payload.get("shop_name") or "").strip()
    if not customer_name:
        raise HTTPException(status_code=400, detail="品牌名称不能为空")
    if not shop_name:
        raise HTTPException(status_code=400, detail="店铺名称不能为空")
    existing = repository.get_general_customer_shop_by_name(customer_name, shop_name)
    if existing:
        raise HTTPException(status_code=400, detail=f"店铺 '{customer_name} / {shop_name}' 已存在")
    payload["customer_name"] = customer_name
    payload["shop_name"] = shop_name
    item = repository.create_general_customer_shop(payload)
    label = f"{customer_name} / {shop_name}"
    write_operation_log(
        request,
        module="general_customer",
        action="create",
        entity_type="general_customer_shop",
        entity_id=item.get("id"),
        entity_label=label,
        summary=f"新增一般客户店铺 {label}".strip(),
        before_data=None,
        after_data=item,
    )
    return {"item": item, "message": "创建成功"}


@router.put("/inventory/general-customer-shops/{shop_id}")
def update_general_customer_shop(request: Request, shop_id: int, payload: dict):
    repository = request.app.state.inventory_repository
    payload["customer_name"] = str(payload.get("customer_name") or "").strip()
    payload["shop_name"] = str(payload.get("shop_name") or "").strip()
    if not payload["customer_name"]:
        raise HTTPException(status_code=400, detail="品牌名称不能为空")
    if not payload["shop_name"]:
        raise HTTPException(status_code=400, detail="店铺名称不能为空")
    before = repository.get_general_customer_shop(shop_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Shop not found")
    record = repository.update_general_customer_shop(shop_id, payload)
    if record is None:
        raise HTTPException(status_code=404, detail="Shop not found")
    label = f"{record.get('customer_name') or ''} / {record.get('shop_name') or ''}".strip()
    changes = build_changed_fields(before, record, GENERAL_CUSTOMER_SHOP_FIELD_LABELS)
    write_operation_log(
        request,
        module="general_customer",
        action="update",
        entity_type="general_customer_shop",
        entity_id=shop_id,
        entity_label=label,
        summary=summarize_changes("编辑一般客户店铺", label, changes),
        changed_fields=changes,
        before_data=before,
        after_data=record,
    )
    return {"item": record, "message": "更新成功"}


@router.delete("/inventory/general-customer-shops/{shop_id}")
def delete_general_customer_shop(request: Request, shop_id: int):
    repository = request.app.state.inventory_repository
    before = repository.get_general_customer_shop(shop_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Shop not found")
    if not repository.delete_general_customer_shop(shop_id):
        raise HTTPException(status_code=404, detail="Shop not found")
    label = f"{before.get('customer_name') or ''} / {before.get('shop_name') or ''}".strip()
    write_operation_log(
        request,
        module="general_customer",
        action="delete",
        entity_type="general_customer_shop",
        entity_id=shop_id,
        entity_label=label,
        summary=f"删除一般客户店铺 {label}".strip(),
        before_data=before,
        after_data=None,
    )
    return {"message": "删除成功"}


@router.get("/inventory/counterparty-ledger")
def get_counterparty_ledger(
    request: Request,
    counterparty_type: str,
    name: str,
    date_start: str | None = None,
    date_end: str | None = None,
):
    counterparty_type = str(counterparty_type or "").strip()
    if counterparty_type not in {"supplier", "customer"}:
        raise HTTPException(status_code=400, detail="单位类型无效")
    name = str(name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="单位全名不能为空")
    repository = request.app.state.inventory_repository
    return repository.get_counterparty_ledger(
        counterparty_type=counterparty_type,
        name=name,
        date_start=_normalize_date(date_start),
        date_end=_normalize_date(date_end),
    )


@router.get("/inventory/detail-lookup")
def lookup_inventory_detail(
    request: Request,
    product_code: str,
    quantity: str | None = None,
    brand: str | None = None,
):
    repository = request.app.state.inventory_repository
    with repository.engine.connect() as connection:
        item = _build_purchase_detail_lookup(
            connection,
            str(product_code or ""),
            _to_decimal(quantity),
            brand,
        )
    return {"item": item}


@router.get("/inventory/detail-candidates")
def list_inventory_detail_candidates(
    request: Request,
    query: str,
    brand: str | None = None,
    limit: int = 20,
):
    normalized_limit = min(max(limit, 1), 50)
    repository = request.app.state.inventory_repository
    with repository.engine.connect() as connection:
        items = _build_purchase_detail_candidates(connection, query, brand, normalized_limit)
    return {"items": items}


@router.get("/inventory/{record_id}")
def get_inventory_record(request: Request, record_id: int):
    repository = request.app.state.inventory_repository
    record = repository.get_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    return record


@router.post("/inventory")
def create_inventory_record(request: Request, payload: dict):
    repository = request.app.state.inventory_repository
    if payload.get("date"):
        payload["date"] = _normalize_date(str(payload["date"]))
    else:
        payload["date"] = _today_text()
    if "document_type" in payload:
        payload["document_type"] = normalize_document_type(payload.get("document_type"))
    summary = (payload.get("summary") or "").strip()
    if summary:
        from domain.inventory_schema import INVENTORY_TABLE
        from sqlalchemy import select as sa_select
        with repository.engine.connect() as conn:
            existing = conn.execute(
                sa_select(INVENTORY_TABLE).where(INVENTORY_TABLE.c.summary == summary)
            ).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"摘要 '{summary}' 已存在")
    record = repository.create_record(payload)
    _log_record_operation(request, action="create", prefix="新增单据", after=record)
    return {"item": record, "message": "创建成功"}


@router.put("/inventory/{record_id}")
def update_inventory_record(request: Request, record_id: int, payload: dict):
    repository = request.app.state.inventory_repository
    before = repository.get_record(record_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Record not found")
    if payload.get("date"):
        payload["date"] = _normalize_date(str(payload["date"]))
    if "document_type" in payload:
        payload["document_type"] = normalize_document_type(payload.get("document_type"))
    record = repository.update_record(record_id, payload)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    _log_record_operation(request, action="update", prefix="编辑单据", before=before, after=record)
    return {"item": record, "message": "更新成功"}


@router.delete("/inventory/{record_id}")
def delete_inventory_record(request: Request, record_id: int):
    repository = request.app.state.inventory_repository
    before = repository.get_record(record_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Record not found")
    if not repository.delete_record(record_id):
        raise HTTPException(status_code=404, detail="Record not found")
    _log_record_operation(request, action="delete", prefix="移入回收站", before=before)
    return {"message": "已移入回收站，10 天内可恢复"}


@router.post("/inventory/{record_id}/restore")
def restore_inventory_record(request: Request, record_id: int):
    repository = request.app.state.inventory_repository
    before = repository.get_record_any_status(record_id)
    record = repository.restore_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    _log_record_operation(request, action="restore", prefix="恢复单据", before=before, after=record)
    return {"item": record, "message": "恢复成功"}


@router.post("/inventory/batch-restore")
def batch_restore_inventory(request: Request, payload: dict):
    ids = payload.get("ids", [])
    if not isinstance(ids, list):
        raise HTTPException(status_code=400, detail="ids 必须是数组")
    record_ids = []
    for value in ids:
        try:
            record_ids.append(int(value))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="ids 中包含非法编号") from None
    repository = request.app.state.inventory_repository
    before_records = [record for record_id in record_ids if (record := repository.get_record_any_status(record_id)) is not None]
    restored = repository.restore_records(record_ids)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in before_records:
        grouped[inventory_module_for_record(record)].append(record)
    for module, records in grouped.items():
        write_operation_log(
            request,
            module=module,
            action="batch_restore",
            entity_type="inventory_record",
            entity_label=f"{len(records)} 条单据",
            summary=f"批量恢复单据 {len(records)} 条",
            after_data={"ids": [record.get("id") for record in records], "records": records[:200]},
        )
    return {"restored": restored, "message": f"已恢复 {restored} 条记录"}


@router.post("/inventory/recycle-bin/batch-delete")
def batch_permanently_delete_inventory(request: Request, payload: dict):
    ids = payload.get("ids", [])
    if not isinstance(ids, list):
        raise HTTPException(status_code=400, detail="ids 必须是数组")
    record_ids = []
    for value in ids:
        try:
            record_ids.append(int(value))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="ids 中包含非法编号") from None
    repository = request.app.state.inventory_repository
    before_records = [record for record_id in record_ids if (record := repository.get_record_any_status(record_id)) is not None]
    deleted = repository.permanently_delete_records(record_ids)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in before_records:
        grouped[inventory_module_for_record(record)].append(record)
    for module, records in grouped.items():
        write_operation_log(
            request,
            module=module,
            action="batch_permanent_delete",
            entity_type="inventory_record",
            entity_label=f"{len(records)} 条单据",
            summary=f"彻底删除回收站单据 {len(records)} 条",
            before_data={"ids": [record.get("id") for record in records], "records": records[:200]},
        )
    return {"deleted": deleted, "message": f"已彻底删除 {deleted} 条单据"}


@router.post("/inventory/batch-delete")
def batch_delete_inventory(request: Request, payload: dict):
    ids = payload.get("ids", [])
    repository = request.app.state.inventory_repository
    record_ids = []
    for value in ids:
        try:
            record_ids.append(int(value))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="ids 中包含非法编号") from None
    before_records = [record for record_id in record_ids if (record := repository.get_record(record_id)) is not None]
    deleted = repository.delete_records(record_ids)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in before_records:
        grouped[inventory_module_for_record(record)].append(record)
    for module, records in grouped.items():
        write_operation_log(
            request,
            module=module,
            action="batch_delete",
            entity_type="inventory_record",
            entity_label=f"{len(records)} 条单据",
            summary=f"批量移入回收站 {len(records)} 条单据",
            before_data={"ids": [record.get("id") for record in records], "records": records[:200]},
        )
    return {"deleted": deleted, "message": f"已移入回收站 {deleted} 条记录，10 天内可恢复"}


@router.get("/inventory/{record_id}/details")
def list_inventory_details(request: Request, record_id: int):
    repository = request.app.state.inventory_repository
    return {"items": repository.list_details(record_id)}


@router.post("/inventory/{record_id}/details/import-replace")
async def replace_inventory_details_from_excel(request: Request, record_id: int, file: UploadFile = None):
    if file is None:
        raise HTTPException(status_code=400, detail="No file uploaded")
    repository = request.app.state.inventory_repository
    record = repository.get_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    before_details = repository.list_details(record_id)

    form = await request.form()
    document_type = normalize_document_type(record.get("document_type"))
    brand = str(form.get("brand") or "").strip() or _purchase_import_brand(document_type or "")
    fallback_unit_price = _to_decimal(form.get("unit_price"))
    content = await file.read()
    details, sheet_name = _build_purchase_details_from_excel(
        repository,
        content,
        brand=brand,
        fallback_unit_price=fallback_unit_price,
        prefer_lookup_unit_price=document_type == "进货订单",
    )
    detail_payloads = []
    for detail in details:
        item = dict(detail)
        item["document_id"] = record_id
        detail_payloads.append(item)
    repository.replace_details(record_id, detail_payloads)
    repository.update_record(record_id, {
        "source_workbook": file.filename or record.get("source_workbook") or "",
        "source_sheet": sheet_name or record.get("source_sheet") or "",
        "source_row_number": "replace_details",
    })
    write_operation_log(
        request,
        module=inventory_module_for_record(record),
        action="replace_details_import",
        entity_type="inventory_detail",
        entity_id=record_id,
        entity_label=inventory_entity_label(record),
        summary=f"重新导入并覆盖 {inventory_entity_label(record)} 的明细：{len(before_details)} 条 -> {len(details)} 条",
        before_data={"details": before_details[:200], "count": len(before_details)},
        after_data={"details": detail_payloads[:200], "count": len(details), "filename": file.filename},
    )
    return {
        "updated": 1,
        "details": len(details),
        "message": f"已重新导入并覆盖 {len(details)} 条明细",
    }


@router.post("/inventory/{record_id}/details")
def create_inventory_detail(request: Request, record_id: int, payload: dict):
    repository = request.app.state.inventory_repository
    record = repository.get_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    _validate_purchase_detail_remark(record, payload)
    payload["document_id"] = record_id
    detail = repository.create_detail(payload)
    _log_detail_operation(request, action="detail_create", prefix="新增", record=record, after=detail)
    return {"item": detail, "message": "明细添加成功"}


@router.put("/inventory/{record_id}/details/{detail_id}")
def update_inventory_detail(request: Request, record_id: int, detail_id: int, payload: dict):
    repository = request.app.state.inventory_repository
    record = repository.get_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    before = repository.get_detail(detail_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Detail not found")
    _validate_purchase_detail_remark(record, payload)
    payload["document_id"] = record_id
    detail = repository.update_detail(detail_id, payload)
    if detail is None:
        raise HTTPException(status_code=404, detail="Detail not found")
    _log_detail_operation(request, action="detail_update", prefix="编辑", record=record, before=before, after=detail)
    return {"item": detail, "message": "明细更新成功"}


@router.post("/inventory/{record_id}/details/batch-delete")
def batch_delete_inventory_details(request: Request, record_id: int, payload: dict):
    ids = payload.get("ids", [])
    if not isinstance(ids, list):
        raise HTTPException(status_code=400, detail="ids 必须是数组")
    detail_ids = []
    for value in ids:
        try:
            detail_ids.append(int(value))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="明细 ID 格式错误")
    repository = request.app.state.inventory_repository
    record = repository.get_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    before_details = [detail for detail_id in detail_ids if (detail := repository.get_detail(detail_id)) is not None]
    deleted = repository.delete_details(record_id, detail_ids)
    write_operation_log(
        request,
        module=inventory_module_for_record(record),
        action="detail_batch_delete",
        entity_type="inventory_detail",
        entity_id=record_id,
        entity_label=inventory_entity_label(record),
        summary=f"批量删除 {inventory_entity_label(record)} 的明细 {deleted} 条",
        before_data={"details": before_details[:200], "count": len(before_details)},
    )
    return {"deleted": deleted, "message": f"已删除 {deleted} 条明细"}


@router.delete("/inventory/{record_id}/details/{detail_id}")
def delete_inventory_detail(request: Request, record_id: int, detail_id: int):
    repository = request.app.state.inventory_repository
    record = repository.get_record(record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    before = repository.get_detail(detail_id)
    if before is None:
        raise HTTPException(status_code=404, detail="Detail not found")
    if not repository.delete_detail(detail_id):
        raise HTTPException(status_code=404, detail="Detail not found")
    _log_detail_operation(request, action="detail_delete", prefix="删除", record=record, before=before)
    return {"message": "明细删除成功"}


@router.post("/inventory/batch-update-costs")
def batch_update_inventory_costs(request: Request, payload: dict):
    repository = request.app.state.inventory_repository
    date_start = _normalize_date(str(payload.get("date_start") or "").strip()) or None
    date_end = _normalize_date(str(payload.get("date_end") or "").strip()) or None
    updates = payload.get("updates")
    if not isinstance(updates, dict) or not updates:
        raise HTTPException(status_code=400, detail="请填写货号和新单价")
    for product_code, unit_price in updates.items():
        if not str(product_code or "").strip():
            raise HTTPException(status_code=400, detail="货号不能为空")
        if _to_decimal(unit_price) <= 0:
            raise HTTPException(status_code=400, detail=f"货号 {product_code} 的新单价必须大于 0")
    result = repository.batch_update_purchase_costs(
        date_start=date_start,
        date_end=date_end,
        price_updates=updates,
    )
    write_operation_log(
        request,
        module="inventory",
        action="batch_update_costs",
        entity_type="inventory_detail",
        entity_label=f"{result['updated_details']} 条明细",
        summary=f"批量改成本价：更新 {result['updated_details']} 条明细，涉及 {result['updated_documents']} 张单据",
        after_data={
            "date_start": date_start,
            "date_end": date_end,
            "updates": updates,
            "result": result,
        },
    )
    return {
        **result,
        "message": f"已更新 {result['updated_details']} 条明细，涉及 {result['updated_documents']} 张单据",
    }


@router.post("/inventory/import")
async def import_inventory(request: Request, file: UploadFile = None):
    if file is None:
        raise HTTPException(status_code=400, detail="No file uploaded")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        wb.close()
        raise HTTPException(status_code=400, detail="Empty file")

    headers = [str(h).strip() if h else "" for h in header_row]

    # Build reverse alias maps
    reverse_aliases: dict[str, str] = {}
    for cn_label, en_field in CN_TO_FIELD.items():
        reverse_aliases[cn_label] = en_field
        reverse_aliases[en_field] = en_field

    detail_reverse_aliases: dict[str, str] = {}
    for cn_label, en_field in DETAIL_CN_TO_FIELD.items():
        detail_reverse_aliases[cn_label] = en_field
        detail_reverse_aliases[en_field] = en_field

    known_fields = set(CN_TO_FIELD.values()) | set(CN_TO_FIELD.keys())
    detail_known_fields = set(DETAIL_CN_TO_FIELD.values()) | set(DETAIL_CN_TO_FIELD.keys())
    repository = request.app.state.inventory_repository

    # Phase 1: Parse all rows and group by summary
    # Each row_entry = {doc: dict, detail: dict}
    groups: dict[str, list[dict]] = {}
    group_order: list[str] = []  # preserve insertion order

    for row in iterator:
        row_dict = {}
        for idx, cell_value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                row_dict[headers[idx]] = cell_value

        doc_payload: dict[str, object] = {}
        detail_payload: dict[str, object] = {}
        extra_fields: dict[str, str] = {}

        for key, value in row_dict.items():
            doc_field = reverse_aliases.get(key)
            detail_field = detail_reverse_aliases.get(key)
            str_value = str(value).strip() if value is not None else None

            if doc_field:
                if doc_field in ("total_count", "amount") and str_value:
                    try:
                        str_value = str(float(str_value))
                    except ValueError:
                        pass
                if doc_field == "date":
                    str_value = _normalize_date(str_value)
                doc_payload[doc_field] = str_value
            elif detail_field:
                if detail_field == "quantity" and str_value:
                    try:
                        str_value = str(int(float(str_value))) if float(str_value) == int(float(str_value)) else str(float(str_value))
                    except ValueError:
                        pass
                if detail_field in ("amount", "unit_price") and str_value:
                    try:
                        str_value = str(float(str_value))
                    except ValueError:
                        pass
                detail_payload[detail_field] = str_value
            elif key and key not in known_fields and key not in detail_known_fields:
                if value is not None and str(value).strip():
                    extra_fields[key] = str(value).strip()

        if not doc_payload.get("date"):
            continue

        # Validate document_type
        doc_type = normalize_document_type(doc_payload.get("document_type"))
        doc_payload["document_type"] = doc_type
        if doc_type and doc_type not in DOCUMENT_TYPES:
            extra_fields["原始单据类型"] = doc_type
            doc_payload["document_type"] = ""

        if extra_fields:
            doc_payload["extra_fields"] = extra_fields

        doc_payload.setdefault("source_workbook", file.filename or "")
        doc_payload.setdefault("source_sheet", ws.title or "")

        # raw_payload stores the original row data
        raw_payload = {}
        for k, v in row_dict.items():
            raw_payload[k] = str(v) if v is not None else ""
        for rp_key, rp_value in raw_payload.items():
            if reverse_aliases.get(rp_key) == "date":
                raw_payload[rp_key] = _normalize_date(rp_value) or rp_value
                break
        doc_payload["raw_payload"] = raw_payload

        summary = str(doc_payload.get("summary") or "").strip()
        if summary not in groups:
            groups[summary] = []
            group_order.append(summary)
        groups[summary].append({"doc": doc_payload, "detail": detail_payload})

    wb.close()

    # Phase 2: Create documents with details grouped by summary
    new_suppliers: set[str] = set()
    new_warehouses: set[str] = set()
    created_docs = 0
    created_details = 0
    skipped_docs = 0
    created_records: list[dict[str, object]] = []

    from domain.inventory_schema import INVENTORY_TABLE
    from sqlalchemy import select as sa_select

    for summary in group_order:
        group_rows = groups[summary]
        first = group_rows[0]
        doc_payload = first["doc"]

        supplier_name = str(doc_payload.get("supplier") or "").strip()
        warehouse_name = str(doc_payload.get("warehouse") or "").strip()
        if supplier_name:
            new_suppliers.add(supplier_name)
        if warehouse_name:
            new_warehouses.add(warehouse_name)

        # Skip if summary already exists in database
        if summary:
            with repository.engine.connect() as conn:
                existing = conn.execute(
                    sa_select(INVENTORY_TABLE).where(INVENTORY_TABLE.c.summary == summary)
                ).first()
                if existing:
                    skipped_docs += 1
                    continue

        try:
            doc = repository.create_record(doc_payload)
            created_records.append(doc)
            created_docs += 1
            doc_id = doc["id"]

            # Create detail rows that have a product_code
            for row in group_rows:
                detail = row["detail"]
                if detail.get("product_code"):
                    detail["document_id"] = doc_id
                    try:
                        repository.create_detail(detail)
                        created_details += 1
                    except Exception:
                        pass
        except Exception:
            skipped_docs += 1

    # Sync new suppliers and warehouses
    supplier_added = 0
    for name in new_suppliers:
        if not repository.get_supplier_by_name(name):
            repository.create_supplier({"name": name})
            supplier_added += 1

    warehouse_added = 0
    for name in new_warehouses:
        if not repository.get_warehouse_by_name(name):
            repository.create_warehouse({"name": name})
            warehouse_added += 1

    msg = f"导入完成：新增 {created_docs} 条单据，{created_details} 条明细"
    if skipped_docs > 0:
        msg += f"，跳过 {skipped_docs} 条已存在的单据"
    if supplier_added > 0:
        msg += f"，新增供应商 {supplier_added} 个"
    if warehouse_added > 0:
        msg += f"，新增仓库 {warehouse_added} 个"
    write_operation_log(
        request,
        module="inventory",
        action="import",
        entity_type="inventory_record",
        entity_label=file.filename or "进销存导入",
        summary=f"导入进销存：新增 {created_docs} 条单据，{created_details} 条明细，跳过 {skipped_docs} 条",
        after_data={
            "filename": file.filename,
            "documents": created_records[:200],
            "document_count": created_docs,
            "detail_count": created_details,
            "skipped": skipped_docs,
            "supplier_added": supplier_added,
            "warehouse_added": warehouse_added,
        },
    )
    return {"created": created_docs, "details": created_details, "skipped": skipped_docs, "message": msg}
