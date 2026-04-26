from __future__ import annotations

from pathlib import Path

import xlrd
from openpyxl import load_workbook

from domain.sources import SheetSpec, WorkbookSpec
from transform.rows import drop_empty_rows, normalize_header


class MissingRequiredSheetError(RuntimeError):
    pass



def _extract_rows_xlsx(workbook_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise KeyError(sheet_name)

    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    header = next(iterator, None)
    if header is None:
        workbook.close()
        return []

    headers = [normalize_header(value) for value in header]
    rows: list[dict[str, object]] = []
    for row in iterator:
        row_dict = {headers[index]: row[index] for index in range(len(headers)) if headers[index]}
        rows.append(row_dict)

    workbook.close()
    return drop_empty_rows(rows)



def _extract_rows_xls(workbook_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = xlrd.open_workbook(workbook_path)
    try:
        worksheet = workbook.sheet_by_name(sheet_name)
    except xlrd.biffh.XLRDError as error:
        raise KeyError(sheet_name) from error

    if worksheet.nrows == 0:
        return []

    headers = [normalize_header(value) for value in worksheet.row_values(0)]
    rows: list[dict[str, object]] = []
    for row_index in range(1, worksheet.nrows):
        values = worksheet.row_values(row_index)
        row_dict = {headers[index]: values[index] for index in range(len(headers)) if headers[index]}
        rows.append(row_dict)
    return drop_empty_rows(rows)



def read_sheet_rows(workbook_path: Path, sheet_spec: SheetSpec) -> list[dict[str, object]]:
    try:
        if workbook_path.suffix.lower() == ".xls":
            return _extract_rows_xls(workbook_path, sheet_spec.name)
        return _extract_rows_xlsx(workbook_path, sheet_spec.name)
    except KeyError:
        if sheet_spec.optional:
            return []
        raise MissingRequiredSheetError(f"Missing required sheet: {sheet_spec.name}")



def read_workbook_rows(spec: WorkbookSpec, root: Path) -> dict[str, list[dict[str, object]]]:
    workbook_path = spec.resolve_path(root)
    sheet_rows: dict[str, list[dict[str, object]]] = {}
    for sheet_spec in spec.sheets:
        rows = read_sheet_rows(workbook_path, sheet_spec)
        if rows or not sheet_spec.optional:
            sheet_rows[sheet_spec.name] = rows
    return sheet_rows
