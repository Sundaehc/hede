from __future__ import annotations

from pathlib import Path
from datetime import date
import re
from xml.etree import ElementTree
from zipfile import ZipFile

import xlrd
from openpyxl import load_workbook

from transform.rows import normalize_cell, normalize_header


DEFAULT_CBANNER_MENS_GROUP_SOURCE = Path(
    r"\\192.168.10.229\运营组资料\9商品组（卢嘉诚）\商品分析\商品运营货品表\千百度男鞋"
)
DEFAULT_CBANNER_WOMENS_PRODUCT_DETAIL_SOURCE = Path(
    r"\\Hede\运营组资料\9商品组（卢嘉诚）\商品分析\商品运营货品表\千百度女鞋"
)
DEFAULT_EBLAN_PRODUCT_DETAIL_SOURCE = Path(
    r"\\Hede\运营组资料\9商品组（卢嘉诚）\商品分析\商品运营货品表\伊伴\2026\2026-06"
)
CBANNER_MENS_GROUP_SHEET_NAME = "商品明细表"
SUPPORTED_WORKBOOK_SUFFIXES = (".xlsx", ".xlsm", ".xls")
CBANNER_MENS_WORKBOOK_NAME_PREFIX = "赫德货品表（千百度男鞋）"
CBANNER_WOMENS_WORKBOOK_NAME_PREFIX = "赫德货品表（千百度）"
EBLAN_WORKBOOK_NAME_PREFIX = "伊伴货品表"
SKU_HEADERS = ("货号", "商品货号", "商品编码", "货品编码", "货品货号", "原始货号", "款号")
GROUP_HEADERS = ("组别", "运营组别", "商品组别", "货品组别", "分组")
PRODUCT_LEVEL_HEADERS = ("商品等级", "等级", "货品等级")


def _safe_is_file(path: Path) -> bool:
    try:
        return path.is_file()
    except OSError:
        return False


def _safe_is_dir(path: Path) -> bool:
    try:
        return path.is_dir()
    except OSError:
        return False


def resolve_product_detail_workbook(source: Path | None, workbook_name_prefix: str) -> Path | None:
    if source is None:
        return None
    if _safe_is_file(source):
        return source

    if not _safe_is_dir(source):
        for suffix in SUPPORTED_WORKBOOK_SUFFIXES:
            candidate = Path(f"{source}{suffix}")
            if _safe_is_file(candidate):
                return candidate
        return None

    candidates: list[Path] = []
    try:
        iterator = source.rglob("*")
        for item in iterator:
            if (
                _safe_is_file(item)
                and item.suffix.lower() in SUPPORTED_WORKBOOK_SUFFIXES
                and item.stem.startswith(workbook_name_prefix)
            ):
                candidates.append(item)
    except OSError:
        return None
    if not candidates:
        return None

    def sort_key(path: Path) -> tuple[int, int, int, float, str]:
        path_text = str(path)
        year_matches = re.findall(r"20\d{2}", path_text)
        month_match = re.search(r"(\d{1,2})月份", path_text)
        day_match = re.search(r"(\d{1,2})\.(\d{1,2})", path.stem)
        try:
            modified_at = path.stat().st_mtime
        except OSError:
            modified_at = 0
        year = max((int(value) for value in year_matches), default=date.today().year)
        month = int(day_match.group(1)) if day_match else int(month_match.group(1)) if month_match else 0
        day = int(day_match.group(2)) if day_match else 0
        return year, month, day, modified_at, path.name

    return max(candidates, key=sort_key)


def resolve_cbanner_mens_group_workbook(source: Path | None) -> Path | None:
    return resolve_product_detail_workbook(source, CBANNER_MENS_WORKBOOK_NAME_PREFIX)


def _find_header_indexes(headers: list[str], value_headers: tuple[str, ...]) -> tuple[int, int] | None:
    sku_index = next((index for index, value in enumerate(headers) if value in SKU_HEADERS), None)
    value_index = next((index for index, value in enumerate(headers) if value in value_headers), None)
    if sku_index is None or value_index is None:
        return None
    return sku_index, value_index


def _get_value(row: tuple[object, ...], index: int) -> object:
    if index >= len(row):
        return None
    return row[index]


def _candidate_sheet_names(workbook_sheet_names: list[str], preferred_sheet_name: str | None) -> list[str]:
    if preferred_sheet_name and preferred_sheet_name in workbook_sheet_names:
        return [preferred_sheet_name]
    return workbook_sheet_names


def _read_shared_strings(archive: ZipFile) -> list[str]:
    try:
        source = archive.open("xl/sharedStrings.xml")
    except KeyError:
        return []

    strings: list[str] = []
    with source:
        for _, element in ElementTree.iterparse(source, events=("end",)):
            if element.tag.endswith("}si"):
                parts = [
                    text_element.text or ""
                    for text_element in element.iter()
                    if text_element.tag.endswith("}t")
                ]
                strings.append("".join(parts))
                element.clear()
    return strings


def _sheet_paths_by_name(archive: ZipFile) -> dict[str, str]:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root
        if rel.attrib.get("Id") and rel.attrib.get("Target")
    }
    sheets: dict[str, str] = {}
    for sheet in workbook_root.iter():
        if not sheet.tag.endswith("}sheet"):
            continue
        name = sheet.attrib.get("name")
        relationship_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = relationships.get(relationship_id or "")
        if not name or not target:
            continue
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        sheets[name] = target
    return sheets


def _column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    value = 0
    for char in letters:
        value = value * 26 + ord(char.upper()) - ord("A") + 1
    return max(value - 1, 0)


def _cell_text(cell: ElementTree.Element, shared_strings: list[str]) -> str | None:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        parts = [
            text_element.text or ""
            for text_element in cell.iter()
            if text_element.tag.endswith("}t")
        ]
        return "".join(parts)

    value_element = next((child for child in cell if child.tag.endswith("}v")), None)
    if value_element is None or value_element.text is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(value_element.text)]
        except (IndexError, ValueError):
            return None
    return value_element.text


def _row_values(row: ElementTree.Element, shared_strings: list[str]) -> tuple[object, ...]:
    values: list[object] = []
    for cell in row:
        if not cell.tag.endswith("}c"):
            continue
        cell_ref = cell.attrib.get("r", "")
        index = _column_index(cell_ref)
        while len(values) <= index:
            values.append(None)
        values[index] = _cell_text(cell, shared_strings)
    return tuple(values)


def _iter_xlsx_rows(archive: ZipFile, sheet_path: str, shared_strings: list[str]):
    with archive.open(sheet_path) as source:
        for _, element in ElementTree.iterparse(source, events=("end",)):
            if element.tag.endswith("}row"):
                yield _row_values(element, shared_strings)
                element.clear()


def _read_zipped_xlsx_value_map(
    workbook_path: Path,
    *,
    sheet_name: str | None,
    value_headers: tuple[str, ...],
) -> dict[str, str]:
    with ZipFile(workbook_path) as archive:
        shared_strings = _read_shared_strings(archive)
        sheet_paths = _sheet_paths_by_name(archive)
        for candidate_sheet in _candidate_sheet_names(list(sheet_paths), sheet_name):
            sheet_path = sheet_paths[candidate_sheet]
            rows = _iter_xlsx_rows(archive, sheet_path, shared_strings)
            header_indexes: tuple[int, int] | None = None
            for row_number, row in enumerate(rows, start=1):
                headers = [normalize_header(value) for value in row]
                header_indexes = _find_header_indexes(headers, value_headers)
                if header_indexes is not None:
                    break
                if row_number >= 30:
                    break

            if header_indexes is None:
                continue

            sku_index, value_index = header_indexes
            values_by_sku: dict[str, str] = {}
            for row in rows:
                sku = normalize_cell(_get_value(row, sku_index))
                value = normalize_cell(_get_value(row, value_index))
                if sku and value:
                    values_by_sku[str(sku)] = str(value)
            return values_by_sku
    return {}


def _read_xlsx_value_map(
    workbook_path: Path,
    *,
    sheet_name: str | None,
    value_headers: tuple[str, ...],
) -> dict[str, str]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        for candidate_sheet in _candidate_sheet_names(workbook.sheetnames, sheet_name):
            worksheet = workbook[candidate_sheet]
            header_indexes: tuple[int, int] | None = None
            rows = worksheet.iter_rows(values_only=True)
            for row_number, row in enumerate(rows, start=1):
                headers = [normalize_header(value) for value in row]
                header_indexes = _find_header_indexes(headers, value_headers)
                if header_indexes is not None:
                    break
                if row_number >= 30:
                    break

            if header_indexes is None:
                continue

            sku_index, value_index = header_indexes
            values_by_sku: dict[str, str] = {}
            for row in rows:
                sku = normalize_cell(_get_value(row, sku_index))
                value = normalize_cell(_get_value(row, value_index))
                if sku and value:
                    values_by_sku[str(sku)] = str(value)
            return values_by_sku
        return {}
    finally:
        workbook.close()


def _read_xls_value_map(
    workbook_path: Path,
    *,
    sheet_name: str | None,
    value_headers: tuple[str, ...],
) -> dict[str, str]:
    workbook = xlrd.open_workbook(workbook_path)
    candidate_sheet_names = _candidate_sheet_names(workbook.sheet_names(), sheet_name)
    for candidate_sheet in candidate_sheet_names:
        try:
            worksheet = workbook.sheet_by_name(candidate_sheet)
        except xlrd.biffh.XLRDError:
            continue

        header_indexes: tuple[int, int] | None = None
        data_start_row = 0
        for row_index in range(min(worksheet.nrows, 30)):
            headers = [normalize_header(value) for value in worksheet.row_values(row_index)]
            header_indexes = _find_header_indexes(headers, value_headers)
            if header_indexes is not None:
                data_start_row = row_index + 1
                break

        if header_indexes is None:
            continue

        sku_index, value_index = header_indexes
        values_by_sku: dict[str, str] = {}
        for row_index in range(data_start_row, worksheet.nrows):
            row = tuple(worksheet.row_values(row_index))
            sku = normalize_cell(_get_value(row, sku_index))
            value = normalize_cell(_get_value(row, value_index))
            if sku and value:
                values_by_sku[str(sku)] = str(value)
        return values_by_sku
    return {}


def read_product_detail_value_map(
    source: Path | None,
    *,
    workbook_name_prefix: str,
    value_headers: tuple[str, ...],
    sheet_name: str | None = None,
) -> dict[str, str]:
    workbook_path = resolve_product_detail_workbook(source, workbook_name_prefix)
    if workbook_path is None:
        return {}

    if workbook_path.suffix.lower() == ".xls":
        return _read_xls_value_map(workbook_path, sheet_name=sheet_name, value_headers=value_headers)
    try:
        return _read_zipped_xlsx_value_map(workbook_path, sheet_name=sheet_name, value_headers=value_headers)
    except Exception:
        return _read_xlsx_value_map(workbook_path, sheet_name=sheet_name, value_headers=value_headers)


def read_cbanner_mens_group_map(source: Path | None) -> dict[str, str]:
    return read_product_detail_value_map(
        source,
        workbook_name_prefix=CBANNER_MENS_WORKBOOK_NAME_PREFIX,
        value_headers=GROUP_HEADERS,
        sheet_name=CBANNER_MENS_GROUP_SHEET_NAME,
    )


def read_cbanner_mens_product_level_map(source: Path | None) -> dict[str, str]:
    return read_product_detail_value_map(
        source,
        workbook_name_prefix=CBANNER_MENS_WORKBOOK_NAME_PREFIX,
        value_headers=PRODUCT_LEVEL_HEADERS,
        sheet_name=CBANNER_MENS_GROUP_SHEET_NAME,
    )


def read_cbanner_womens_product_level_map(source: Path | None) -> dict[str, str]:
    return read_product_detail_value_map(
        source,
        workbook_name_prefix=CBANNER_WOMENS_WORKBOOK_NAME_PREFIX,
        value_headers=PRODUCT_LEVEL_HEADERS,
    )


def read_eblan_product_level_map(source: Path | None) -> dict[str, str]:
    return read_product_detail_value_map(
        source,
        workbook_name_prefix=EBLAN_WORKBOOK_NAME_PREFIX,
        value_headers=PRODUCT_LEVEL_HEADERS,
    )
