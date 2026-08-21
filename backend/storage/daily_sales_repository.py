from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import orjson
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.dialects.postgresql import insert as pg_insert

from domain.daily_sales_schema import ensure_jst_daily_sales_table, ensure_vip_daily_sales_table
from storage.date_normalization import parse_date
from storage.vip_repository import _xlsx_sheet_rows


JST_FILE_NAME = "聚水潭日销售报表.xlsx"
VIP_FILE_NAME = "唯品日销售报表.xlsx"


def _json_serializer(value: object) -> str:
    return orjson.dumps(value).decode("utf-8")


def _header(value: object) -> str:
    return str(value or "").strip().replace("\n", "")


def _text(value: object) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _key_text(value: object) -> str:
    return _text(value) or ""


def _integer(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(Decimal(str(value).replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def _decimal(value: object) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", ""))
    except InvalidOperation:
        return None


def _payload_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _value(row: dict[str, object], *names: str) -> object:
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip() != "":
            return value
    return None


class DailySalesRepository:
    def __init__(self, database_url: str):
        self.engine = create_engine(database_url, future=True, json_serializer=_json_serializer)

    def import_jst_daily_sales(self, file_path: Path) -> dict[str, object]:
        return self._import(file_path, source="jst")

    def import_vip_daily_sales(self, file_path: Path) -> dict[str, object]:
        return self._import(file_path, source="vip")

    def import_jst_daily_sales_range(
        self,
        file_path: Path,
        *,
        date_start: date,
        date_end: date,
        batch_size: int = 1_000,
    ) -> dict[str, object]:
        return self._import_range(
            file_path,
            source="jst",
            date_start=date_start,
            date_end=date_end,
            batch_size=batch_size,
        )

    def import_vip_daily_sales_range(
        self,
        file_path: Path,
        *,
        date_start: date,
        date_end: date,
        batch_size: int = 1_000,
    ) -> dict[str, object]:
        return self._import_range(
            file_path,
            source="vip",
            date_start=date_start,
            date_end=date_end,
            batch_size=batch_size,
        )

    def _import_range(
        self,
        file_path: Path,
        *,
        source: str,
        date_start: date,
        date_end: date,
        batch_size: int,
    ) -> dict[str, object]:
        if not file_path.exists():
            raise FileNotFoundError(file_path)
        if source not in {"jst", "vip"}:
            raise ValueError(f"Unsupported daily sales source: {source}")
        if date_start > date_end:
            raise ValueError("date_start cannot be later than date_end")
        if date_start.year != date_end.year:
            raise ValueError("Daily sales range import must stay within one calendar year")

        row_iterator = _xlsx_sheet_rows(file_path)
        try:
            _, header_values = next(row_iterator)
        except StopIteration:
            return {
                "source": source,
                "source_file": str(file_path),
                "read": 0,
                "matched": 0,
                "upserted": 0,
                "skipped": 0,
                "duplicate_keys_in_batch": 0,
                "sales_dates": [],
            }

        headers = {index: _header(value) for index, value in header_values.items() if _header(value)}
        date_column = next((index for index, header in headers.items() if header == "日期"), None)
        if date_column is None:
            raise ValueError("缺少必要列: 日期")

        if source == "jst":
            table = ensure_jst_daily_sales_table(self.engine, date_start.year)
            mapper = self._map_jst
            key_builder = self._jst_key
            key_columns = self._jst_key_columns
        else:
            table = ensure_vip_daily_sales_table(self.engine, date_start.year)
            mapper = self._map_vip
            key_builder = self._vip_key
            key_columns = self._vip_key_columns
        rows_read = matched_rows = skipped_rows = duplicate_rows = upserted_rows = 0
        sales_dates: set[date] = set()
        batch: dict[tuple[object, ...], dict[str, object]] = {}

        def flush() -> None:
            nonlocal upserted_rows
            if not batch:
                return
            rows = list(batch.values())
            self._upsert(table, rows, key_columns)
            upserted_rows += len(rows)
            batch.clear()

        for row_number, values in row_iterator:
            if not values:
                continue
            rows_read += 1
            sales_date = parse_date(values.get(date_column, ""))
            if sales_date is None or sales_date < date_start or sales_date > date_end:
                skipped_rows += 1
                continue

            raw = {
                header: _payload_value(values.get(index, ""))
                for index, header in headers.items()
            }
            mapped = mapper(raw)
            mapped.update({
                "sales_date": sales_date,
                "source_workbook": file_path.name,
                "source_sheet": "Sheet1",
                "source_row_number": row_number,
                "raw_payload": raw,
            })
            key = key_builder(mapped)
            if key in batch:
                duplicate_rows += 1
            batch[key] = mapped
            matched_rows += 1
            sales_dates.add(sales_date)
            if len(batch) >= batch_size:
                flush()
        flush()

        return {
            "source": source,
            "source_file": str(file_path),
            "read": rows_read,
            "matched": matched_rows,
            "upserted": upserted_rows,
            "skipped": skipped_rows,
            "duplicate_keys_in_batch": duplicate_rows,
            "sales_dates": sorted(value.isoformat() for value in sales_dates),
        }

    def _import(self, file_path: Path, *, source: str) -> dict[str, object]:
        if not file_path.exists():
            raise FileNotFoundError(file_path)
        rows_read = valid_rows = skipped_rows = duplicate_rows = 0
        rows_by_year: dict[int, dict[tuple[object, ...], dict[str, object]]] = defaultdict(dict)
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header_row = next(sheet.iter_rows(values_only=True), None)
            if header_row is None:
                return {"source": source, "source_file": str(file_path), "read": 0, "upserted": 0, "skipped": 0, "years": {}}
            headers = [_header(value) for value in header_row]
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=2):
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                rows_read += 1
                raw = {headers[index]: _payload_value(value) for index, value in enumerate(values) if headers[index]}
                sales_date = parse_date(_value(raw, "日期"))
                if sales_date is None:
                    skipped_rows += 1
                    continue
                mapped = self._map_jst(raw) if source == "jst" else self._map_vip(raw)
                mapped.update({
                    "sales_date": sales_date,
                    "source_workbook": file_path.name,
                    "source_sheet": sheet.title,
                    "source_row_number": row_number,
                    "raw_payload": raw,
                })
                key = self._jst_key(mapped) if source == "jst" else self._vip_key(mapped)
                year_rows = rows_by_year[sales_date.year]
                if key in year_rows:
                    duplicate_rows += 1
                year_rows[key] = mapped
                valid_rows += 1
        finally:
            workbook.close()

        result_years: dict[str, int] = {}
        for year, keyed_rows in rows_by_year.items():
            table = ensure_jst_daily_sales_table(self.engine, year) if source == "jst" else ensure_vip_daily_sales_table(self.engine, year)
            self._upsert(table, list(keyed_rows.values()), self._jst_key_columns if source == "jst" else self._vip_key_columns)
            result_years[str(year)] = len(keyed_rows)
        return {
            "source": source,
            "source_file": str(file_path),
            "read": rows_read,
            "valid": valid_rows,
            "upserted": sum(result_years.values()),
            "skipped": skipped_rows,
            "duplicate_keys_in_file": duplicate_rows,
            "years": result_years,
            "sales_dates": sorted({row["sales_date"].isoformat() for rows in rows_by_year.values() for row in rows.values()}),
        }

    _jst_key_columns = ["sales_date", "channel", "product_code", "style_code", "color_spec", "channel_code", "barcode"]
    _vip_key_columns = ["sales_date", "goods_id", "size_id"]

    @staticmethod
    def _jst_key(row: dict[str, object]) -> tuple[object, ...]:
        return tuple(row[column] for column in DailySalesRepository._jst_key_columns)

    @staticmethod
    def _vip_key(row: dict[str, object]) -> tuple[object, ...]:
        return tuple(row[column] for column in DailySalesRepository._vip_key_columns)

    @staticmethod
    def _map_jst(row: dict[str, object]) -> dict[str, object]:
        return {
            "channel": _key_text(_value(row, "渠道")),
            "product_code": _key_text(_value(row, "商品编码")),
            "style_code": _key_text(_value(row, "款式编码")),
            "color_spec": _key_text(_value(row, "颜色规格")),
            "channel_code": _key_text(_value(row, "渠道编码")),
            "barcode": _key_text(_value(row, "国标码", "条码")),
            "order_type": _text(_value(row, "类型")),
            "supplier": _text(_value(row, "供应商")),
            "supplier_style_code": _text(_value(row, "供应商款号")),
            "product_name": _text(_value(row, "商品名称")),
            "product_category": _text(_value(row, "产品分类")),
            "brand": _text(_value(row, "品牌")),
            "cost_price": _decimal(_value(row, "成本价")),
            "shipped_order_count": _integer(_value(row, "实发单数")),
            "sales_order_count": _integer(_value(row, "销售单数")),
            "return_order_count": _integer(_value(row, "退货单数")),
            "sales_quantity": _integer(_value(row, "销售数量")),
            "shipped_quantity": _integer(_value(row, "实发数量")),
            "return_quantity": _integer(_value(row, "退货数量")),
            "net_sales_quantity": _integer(_value(row, "净销量")),
            "sales_amount": _decimal(_value(row, "销售金额")),
            "net_sales_amount": _decimal(_value(row, "净销售额")),
            "cost_amount": _decimal(_value(row, "销售成本", "成本", "成本金额")),
            "gross_profit": _decimal(_value(row, "销售毛利", "毛利")),
        }

    @staticmethod
    def _map_vip(row: dict[str, object]) -> dict[str, object]:
        return {
            "barcode": _text(_value(row, "条码")),
            "size_id": _key_text(_value(row, "SIZE_ID")),
            "goods_id": _key_text(_value(row, "商品ID")),
            "product_name": _text(_value(row, "商品名称")),
            "goods_code": _text(_value(row, "货号")),
            "style_code": _text(_value(row, "款号")),
            "spu_id": _text(_value(row, "P_SPU_ID")),
            "size_name": _text(_value(row, "尺码名称")),
            "product_image": _text(_value(row, "商品图片")),
            "product_type": _text(_value(row, "货品类型")),
            "brand_sn": _text(_value(row, "品牌SN")),
            "brand_name": _text(_value(row, "品牌名称")),
            "sales_amount": _decimal(_value(row, "销售额")),
            "sales_quantity": _integer(_value(row, "销售量")),
            "customer_count": _integer(_value(row, "客户数")),
            "on_sale_stock": _integer(_value(row, "在售库存")),
            "product_link": _text(_value(row, "商品链接")),
        }

    def _upsert(self, table, rows: list[dict[str, object]], key_columns: list[str]) -> None:
        update_columns = [column.name for column in table.columns if column.name not in {"id", "created_at", *key_columns}]
        with self.engine.begin() as connection:
            for start in range(0, len(rows), 1000):
                chunk = rows[start:start + 1000]
                statement = pg_insert(table).values(chunk)
                statement = statement.on_conflict_do_update(
                    index_elements=key_columns,
                    set_={column: statement.excluded[column] for column in update_columns},
                )
                connection.execute(statement)
