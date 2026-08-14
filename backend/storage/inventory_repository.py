from __future__ import annotations

from collections.abc import Mapping
from datetime import date
from decimal import Decimal
from uuid import uuid4

from pathlib import Path

import orjson
from openpyxl import load_workbook
from sqlalchemy import Text, and_, case, create_engine, delete, desc, func, insert, inspect, or_, select, text, union_all, update
from sqlalchemy.dialects.postgresql import insert as pg_insert

from domain.gj_schema import GJ_MERGED_PRODUCT_INFO_TABLE
from domain.inventory_schema import GENERAL_CUSTOMER_BRAND_TABLE, GENERAL_CUSTOMER_SHOP_TABLE, GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE, GENERAL_CUSTOMER_UNIT_TABLE, INVENTORY_ACCOUNT_SUBJECT_TABLE, INVENTORY_DETAIL_TABLE, INVENTORY_TABLE, JST_STOCK_TABLE, PURCHASE_ORDER_REQUIREMENT_TABLE, SUPPLIER_BRAND_TABLE, SUPPLIER_TABLE, WAREHOUSE_BRAND_TABLE, WAREHOUSE_TABLE
from domain.inventory_sources import ACCOUNTING_DOCUMENT_TYPES
from domain.gj_brand import CBANNER_MENS_BRAND, GJ_FINE_TABLE_BRANDS, SUPPLIER_BRANDS, infer_supplier_brand_from_name
from domain import jst_stock_snapshot_schema  # noqa: F401 - register JST stock snapshot tables on METADATA
from domain import product_goods_schema  # noqa: F401 - register goods table overrides on METADATA
from domain import product_size_group_mapping_schema  # noqa: F401 - register product size group mappings on METADATA
from domain import master_data_schema  # noqa: F401 - register master-data tables on METADATA
from domain import data_governance_schema  # noqa: F401 - register data-governance tables on METADATA
from domain import size_group_schema  # noqa: F401 - register size group tables on METADATA
from domain.jst_stock_snapshot_schema import JST_SIZE_STOCK_SNAPSHOT_TABLE, JST_STOCK_SUMMARY_SNAPSHOT_TABLE
from domain.product_goods_schema import PRODUCT_GOODS_OVERRIDES_TABLE
from domain.product_goods_historical_sales_schema import HISTORICAL_SALES_YEARS, ensure_product_goods_historical_sales_table
from domain.product_size_group_mapping_schema import PRODUCT_SIZE_GROUP_MAPPINGS_TABLE
from domain.size_group_schema import SIZE_GROUP_ITEMS_TABLE, SIZE_GROUPS_TABLE
from storage.date_normalization import parse_date, parse_month_day


DOCUMENT_NUMBER_PREFIXES = {
    "进货订单": "JHDD",
    "进货单": "JHD",
    "进货退货单": "JHTHD",
    "报溢单": "BYD",
    "报损单": "BSD",
    "批发销售单": "PFXSD",
    "批发销售退货单": "PFXSTHD",
    "同价调拨单": "TJDBD",
    "应付款减少": "YFKJS",
    "应付款增加": "YFKZJ",
    "应收款减少": "YSKJS",
    "应收款增加": "YSKZJ",
}
DEFAULT_DOCUMENT_NUMBER_PREFIX = "DJ"
SUPPLIER_LEDGER_INCREASE_TYPES = ("进货单", "应付款增加")
SUPPLIER_LEDGER_DECREASE_TYPES = ("进货退货单", "应付款减少")
SUPPLIER_LEDGER_NEUTRAL_TYPES = ("同价调拨单",)
CUSTOMER_LEDGER_INCREASE_TYPES = ("批发销售单", "应收款增加")
CUSTOMER_LEDGER_DECREASE_TYPES = ("批发销售退货单", "应收款减少")
NEGATIVE_TOTAL_DOCUMENT_TYPES = {"进货退货单"}
PURCHASE_INBOUND_DETAIL_TYPES = ("进货单", "进货退货单")
GENERAL_CUSTOMER_SORT_SCOPE_BRAND = "brand"
GENERAL_CUSTOMER_SORT_SCOPE_SHOP = "shop"
GENERAL_CUSTOMER_SORT_SCOPE_UNIT = "unit"

GENERAL_CUSTOMER_ROOT_SORT_PARENT_ID = 0


def _json_serializer(value: object) -> str:
    return orjson.dumps(value).decode("utf-8")


class InventoryRepository:
    def __init__(self, database_url: str):
        self.engine = create_engine(
            database_url,
            future=True,
            json_serializer=_json_serializer,
        )
        self.create_tables()

    # ── Inventory Records ──────────────────────────────────────────

    def list_records(
        self,
        *,
        date_start: str | None = None,
        date_end: str | None = None,
        supplier: str | None = None,
        warehouse: str | None = None,
        document_type: str | None = None,
        exclude_document_type: str | None = None,
        summary: str | None = None,
        original_sku: str | None = None,
        product_code: str | None = None,
        handler: str | None = None,
        completion_status: str | None = None,
        sort_by: str | None = None,
        sort_direction: str = "desc",
        sort_rules: list[tuple[str, str]] | None = None,
        page: int,
        page_size: int,
    ) -> dict[str, object]:
        table = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        stock = JST_STOCK_TABLE
        count_statement = select(func.count()).select_from(table)
        items_statement = select(table)

        self.purge_expired_deleted_records()
        conditions = [table.c.deleted_at.is_(None)]
        if date_start:
            parsed = parse_date(date_start)
            conditions.append(table.c.date_value >= parsed if parsed else table.c.date >= date_start)
        if date_end:
            parsed = parse_date(date_end)
            conditions.append(table.c.date_value <= parsed if parsed else table.c.date <= date_end)
        if supplier:
            conditions.append(table.c.supplier.ilike(f"%{supplier.strip()}%"))
        if warehouse:
            conditions.append(table.c.warehouse == warehouse)
        if document_type:
            conditions.append(table.c.document_type == document_type)
        if exclude_document_type:
            conditions.append(or_(table.c.document_type.is_(None), table.c.document_type != exclude_document_type))
        if summary:
            conditions.append(table.c.summary.ilike(f"%{summary.strip()}%"))
        if handler:
            conditions.append(table.c.handler.ilike(f"%{handler.strip()}%"))
        if completion_status == "incomplete":
            is_accounting_document = table.c.document_type.in_(ACCOUNTING_DOCUMENT_TYPES)
            is_product_document = or_(table.c.document_type.is_(None), ~is_accounting_document)
            conditions.append(or_(
                ~select(detail.c.id)
                .where(detail.c.document_id == table.c.id)
                .exists(),
                and_(
                    is_accounting_document,
                    select(detail.c.id)
                    .where(
                        detail.c.document_id == table.c.id,
                        or_(
                            detail.c.amount.is_(None),
                            detail.c.amount == 0,
                        ),
                    )
                    .exists(),
                ),
                and_(
                    is_product_document,
                    select(detail.c.id)
                    .where(
                        detail.c.document_id == table.c.id,
                        or_(
                            detail.c.unit_price.is_(None),
                            detail.c.unit_price == 0,
                        ),
                    )
                    .exists(),
                ),
            ))
        elif completion_status == "completed":
            is_accounting_document = table.c.document_type.in_(ACCOUNTING_DOCUMENT_TYPES)
            is_product_document = or_(table.c.document_type.is_(None), ~is_accounting_document)
            conditions.append(
                select(detail.c.id)
                .where(detail.c.document_id == table.c.id)
                .exists()
            )
            conditions.append(or_(
                and_(
                    is_accounting_document,
                    ~select(detail.c.id)
                    .where(
                        detail.c.document_id == table.c.id,
                        or_(
                            detail.c.amount.is_(None),
                            detail.c.amount == 0,
                        ),
                    )
                    .exists(),
                ),
                and_(
                    is_product_document,
                    ~select(detail.c.id)
                    .where(
                        detail.c.document_id == table.c.id,
                        or_(
                            detail.c.unit_price.is_(None),
                            detail.c.unit_price == 0,
                        ),
                    )
                    .exists(),
                ),
            )
            )
        if original_sku:
            original_like = f"%{original_sku.strip()}%"
            conditions.append(
                select(detail.c.id)
                .where(
                    detail.c.document_id == table.c.id,
                    detail.c.product_code.ilike(original_like),
                )
                .exists()
            )
        if product_code:
            product_like = f"%{product_code.strip()}%"
            stock_code_matches = stock.c.product_code.ilike(product_like)
            stock_candidates = union_all(
                select(stock.c.product_code.label("candidate"))
                .where(stock_code_matches),
                select(func.left(stock.c.product_code, func.length(stock.c.product_code) - 5).label("candidate"))
                .where(stock_code_matches)
                .where(func.length(stock.c.product_code) > 5),
                select(func.left(stock.c.product_code, func.length(stock.c.product_code) - 3).label("candidate"))
                .where(stock_code_matches)
                .where(func.length(stock.c.product_code) > 3),
                select(func.left(stock.c.product_code, func.length(stock.c.product_code) - 2).label("candidate"))
                .where(stock_code_matches)
                .where(func.length(stock.c.product_code) > 2),
            ).subquery()
            conditions.append(
                select(detail.c.id)
                .where(
                    detail.c.document_id == table.c.id,
                    or_(
                        detail.c.product_code.ilike(product_like),
                        detail.c.product_code.in_(
                            select(stock_candidates.c.candidate)
                            .where(stock_candidates.c.candidate.isnot(None))
                            .where(stock_candidates.c.candidate != "")
                            .distinct()
                        ),
                    ),
                )
                .exists()
            )

        if conditions:
            criterion = conditions[0] if len(conditions) == 1 else and_(*conditions)
            items_statement = items_statement.where(criterion)
            count_statement = count_statement.where(criterion)

        sort_columns = {
            "document_number": table.c.document_number,
            "date": table.c.date_value,
            "delivery_date": table.c.extra_fields["delivery_date"].as_string(),
            "document_type": table.c.document_type,
            "supplier": table.c.supplier,
            "total_count": table.c.total_count,
            "amount": table.c.amount,
            "warehouse": table.c.warehouse,
            "handler": table.c.handler,
            "summary": table.c.summary,
            "additional_note": table.c.additional_note,
            "updated_at": table.c.updated_at,
        }
        normalized_sort_rules = list(sort_rules or [])
        if not normalized_sort_rules and sort_by:
            normalized_sort_rules = [(str(sort_by).strip(), str(sort_direction or "desc").lower())]
        order_by = []
        seen_sort_keys: set[str] = set()
        for sort_key, direction in normalized_sort_rules:
            normalized_key = str(sort_key or "").strip()
            sort_column = sort_columns.get(normalized_key)
            if sort_column is None or normalized_key in seen_sort_keys:
                continue
            seen_sort_keys.add(normalized_key)
            if str(direction or "").lower() == "asc":
                order_by.append(sort_column.asc().nulls_last())
            else:
                order_by.append(sort_column.desc().nulls_last())
        if not order_by:
            order_by = [desc(table.c.id)]
        else:
            order_by.append(desc(table.c.id))
        items_statement = items_statement.order_by(*order_by).offset((page - 1) * page_size).limit(page_size)

        with self.engine.connect() as connection:
            total = connection.execute(count_statement).scalar_one()
            items = [dict(row) for row in connection.execute(items_statement).mappings()]
        for item in items:
            self._clear_accounting_record_summary(item)

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    def list_purchase_inbound_details(
        self,
        *,
        date_start: str | None = None,
        date_end: str | None = None,
        document_type: str | None = None,
        supplier: str | None = None,
        warehouse: list[str] | None = None,
        product_code: str | None = None,
        product_name: str | None = None,
        color_name: str | None = None,
        size_name: str | None = None,
        page: int,
        page_size: int,
    ) -> dict[str, object]:
        record = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        joined = detail.join(record, detail.c.document_id == record.c.id)
        conditions = [
            record.c.deleted_at.is_(None),
            record.c.document_type.in_(PURCHASE_INBOUND_DETAIL_TYPES),
        ]
        if date_start:
            parsed = parse_date(date_start)
            conditions.append(record.c.date_value >= parsed if parsed else record.c.date >= date_start)
        if date_end:
            parsed = parse_date(date_end)
            conditions.append(record.c.date_value <= parsed if parsed else record.c.date <= date_end)
        if document_type:
            conditions.append(record.c.document_type == document_type)
        if supplier:
            conditions.append(record.c.supplier.ilike(f"%{supplier.strip()}%"))
        warehouse_values = [warehouse] if isinstance(warehouse, str) else (warehouse or [])
        warehouse_names = list(dict.fromkeys(
            warehouse_name.strip()
            for warehouse_name in warehouse_values
            if warehouse_name and warehouse_name.strip()
        ))
        if warehouse_names:
            conditions.append(record.c.warehouse.in_(warehouse_names))
        if product_code:
            conditions.append(detail.c.product_code.ilike(f"%{product_code.strip()}%"))
        if product_name:
            conditions.append(detail.c.product_name.ilike(f"%{product_name.strip()}%"))
        if color_name:
            color_like = f"%{color_name.strip()}%"
            conditions.append(or_(detail.c.color_name.ilike(color_like), detail.c.color_spec.ilike(color_like)))
        if size_name:
            conditions.append(detail.c.size_quantities.cast(Text).ilike(f"%\"{size_name.strip()}\"%"))

        criterion = and_(*conditions)
        signed_quantity = case(
            (record.c.document_type == "进货退货单", -func.coalesce(detail.c.quantity, 0)),
            else_=func.coalesce(detail.c.quantity, 0),
        )
        signed_amount = case(
            (record.c.document_type == "进货退货单", -func.coalesce(detail.c.amount, 0)),
            else_=func.coalesce(detail.c.amount, 0),
        )
        count_statement = select(func.count()).select_from(joined).where(criterion)
        totals_statement = select(
            func.coalesce(func.sum(signed_quantity), 0).label("quantity_total"),
            func.coalesce(func.sum(signed_amount), 0).label("purchase_amount_total"),
        ).select_from(joined).where(criterion)
        items_statement = (
            select(
                detail.c.id,
                detail.c.document_id,
                detail.c.product_code,
                detail.c.product_name,
                detail.c.color_name,
                detail.c.color_spec,
                detail.c.size_quantities,
                detail.c.quantity,
                detail.c.amount,
                detail.c.extra_fields,
                record.c.document_type,
                record.c.document_number,
                record.c.date,
                record.c.supplier,
                record.c.warehouse,
            )
            .select_from(joined)
            .where(criterion)
            .order_by(record.c.date_value.nulls_last(), record.c.date, record.c.document_number, detail.c.id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )

        with self.engine.connect() as connection:
            total = connection.execute(count_statement).scalar_one()
            totals = connection.execute(totals_statement).mappings().one()
            rows = [dict(row) for row in connection.execute(items_statement).mappings()]
            supplier_names = sorted({
                str(row.get("supplier") or "").strip()
                for row in rows
                if str(row.get("supplier") or "").strip()
            })
            supplier_codes: dict[str, str] = {}
            if supplier_names:
                supplier_rows = connection.execute(
                    select(SUPPLIER_TABLE.c.name, SUPPLIER_TABLE.c.factory_code)
                    .where(SUPPLIER_TABLE.c.name.in_(supplier_names))
                ).mappings()
                for supplier_row in supplier_rows:
                    name = str(supplier_row.get("name") or "").strip()
                    if not name:
                        continue
                    code = str(supplier_row.get("factory_code") or "").strip()
                    if code or name not in supplier_codes:
                        supplier_codes[name] = code

        items: list[dict[str, object]] = []
        offset = (page - 1) * page_size
        for index, row in enumerate(rows, start=1):
            document_type_text = str(row.get("document_type") or "")
            multiplier = Decimal("-1") if document_type_text == "进货退货单" else Decimal("1")
            quantity = Decimal(str(row.get("quantity") or "0")) * multiplier
            purchase_amount = Decimal(str(row.get("amount") or "0")) * multiplier
            extra_fields = row.get("extra_fields") if isinstance(row.get("extra_fields"), dict) else {}
            supplier_name = str(row.get("supplier") or "").strip()
            items.append({
                "row_number": offset + index,
                "detail_id": row.get("id"),
                "document_id": row.get("document_id"),
                "product_code": row.get("product_code"),
                "product_name": row.get("product_name"),
                "document_type": row.get("document_type"),
                "document_number": row.get("document_number"),
                "date": row.get("date"),
                "purchase_quantity": self._format_decimal(quantity),
                "purchase_amount": self._format_decimal(purchase_amount),
                "retail_amount": "",
                "factory_code": extra_fields.get("factory_code") or "",
                "unit_code": supplier_codes.get(supplier_name, ""),
                "unit_name": row.get("supplier"),
                "warehouse_name": row.get("warehouse"),
                "color_name": row.get("color_name") or row.get("color_spec") or "",
                "size_quantities": row.get("size_quantities") if isinstance(row.get("size_quantities"), dict) else {},
            })

        quantity_total = Decimal(str(totals.get("quantity_total") or "0"))
        purchase_amount_total = Decimal(str(totals.get("purchase_amount_total") or "0"))
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "totals": {
                "purchase_quantity": self._format_decimal(quantity_total),
                "purchase_amount": self._format_decimal(purchase_amount_total),
                "retail_amount": "",
            },
        }

    def get_record(self, record_id: int) -> dict[str, object] | None:
        table = INVENTORY_TABLE
        statement = select(table).where(table.c.id == record_id, table.c.deleted_at.is_(None))
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_record_any_status(self, record_id: int) -> dict[str, object] | None:
        table = INVENTORY_TABLE
        statement = select(table).where(table.c.id == record_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_record_for_append(
        self,
        *,
        date_value: object,
        warehouse: object,
        document_type: object,
        summary: object,
    ) -> dict[str, object] | None:
        normalized_summary = str(summary or "").strip()
        normalized_warehouse = str(warehouse or "").strip()
        normalized_document_type = str(document_type or "").strip()
        normalized_date = parse_date(date_value)
        if not (normalized_summary and normalized_warehouse and normalized_document_type and normalized_date):
            return None

        table = INVENTORY_TABLE
        statement = (
            select(table)
            .where(
                table.c.deleted_at.is_(None),
                or_(
                    table.c.date_value == normalized_date,
                    table.c.date == normalized_date.isoformat(),
                ),
                table.c.warehouse == normalized_warehouse,
                table.c.document_type == normalized_document_type,
                table.c.summary == normalized_summary,
            )
            .order_by(desc(table.c.id))
        )
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def create_record(self, record: Mapping[str, object]) -> dict[str, object]:
        table = INVENTORY_TABLE
        with self.engine.begin() as connection:
            payload = self._prepare_record(record)
            if not payload.get("document_number"):
                payload["document_number"] = self._generate_document_number(
                    connection,
                    payload.get("date_value") or payload.get("date"),
                    payload.get("document_type"),
                )
            statement = insert(table).values(**payload).returning(table)
            row = connection.execute(statement).mappings().one()
        return dict(row)

    def update_record(self, record_id: int, record: Mapping[str, object]) -> dict[str, object] | None:
        table = INVENTORY_TABLE
        payload = self._prepare_record(record)
        payload.pop("id", None)
        statement = update(table).where(table.c.id == record_id).values(**payload).returning(table)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().first()
        if row is None:
            return None
        item = dict(row)
        if "document_type" in payload:
            self.recalculate_totals(record_id)
            return self.get_record(record_id) or item
        return item

    def delete_record(self, record_id: int) -> bool:
        table = INVENTORY_TABLE
        statement = (
            update(table)
            .where(table.c.id == record_id, table.c.deleted_at.is_(None))
            .values(deleted_at=func.now())
        )
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount > 0

    def delete_records(self, ids: list[int]) -> int:
        if not ids:
            return 0
        table = INVENTORY_TABLE
        statement = (
            update(table)
            .where(table.c.id.in_(ids), table.c.deleted_at.is_(None))
            .values(deleted_at=func.now())
        )
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount

    def list_deleted_records(
        self,
        *,
        page: int,
        page_size: int,
        document_type: str | None = None,
        exclude_document_type: str | None = None,
    ) -> dict[str, object]:
        self.purge_expired_deleted_records()
        table = INVENTORY_TABLE
        conditions = [table.c.deleted_at.isnot(None)]
        if document_type:
            conditions.append(table.c.document_type == document_type)
        if exclude_document_type:
            conditions.append(or_(table.c.document_type.is_(None), table.c.document_type != exclude_document_type))
        criterion = and_(*conditions)
        count_statement = select(func.count()).select_from(table).where(criterion)
        items_statement = (
            select(table)
            .where(criterion)
            .order_by(desc(table.c.deleted_at), desc(table.c.id))
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        with self.engine.connect() as connection:
            total = connection.execute(count_statement).scalar_one()
            items = [dict(row) for row in connection.execute(items_statement).mappings()]
        for item in items:
            self._clear_accounting_record_summary(item)
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    def restore_record(self, record_id: int) -> dict[str, object] | None:
        self.purge_expired_deleted_records()
        table = INVENTORY_TABLE
        statement = (
            update(table)
            .where(table.c.id == record_id, table.c.deleted_at.isnot(None))
            .values(deleted_at=None)
            .returning(table)
        )
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def restore_records(self, ids: list[int]) -> int:
        self.purge_expired_deleted_records()
        if not ids:
            return 0
        table = INVENTORY_TABLE
        statement = (
            update(table)
            .where(table.c.id.in_(ids), table.c.deleted_at.isnot(None))
            .values(deleted_at=None)
        )
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount or 0

    def permanently_delete_records(self, ids: list[int]) -> int:
        self.purge_expired_deleted_records()
        if not ids:
            return 0
        table = INVENTORY_TABLE
        statement = delete(table).where(table.c.id.in_(ids), table.c.deleted_at.isnot(None))
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount or 0

    def purge_expired_deleted_records(self) -> int:
        table = INVENTORY_TABLE
        statement = delete(table).where(table.c.deleted_at < func.now() - text("interval '10 days'"))
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount or 0

    def get_counterparty_ledger(
        self,
        *,
        counterparty_type: str,
        name: str,
        date_start: str | None = None,
        date_end: str | None = None,
    ) -> dict[str, object]:
        table = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        normalized_name = str(name or "").strip()
        if counterparty_type == "customer":
            increase_types = CUSTOMER_LEDGER_INCREASE_TYPES
            decrease_types = CUSTOMER_LEDGER_DECREASE_TYPES
            neutral_types: tuple[str, ...] = ()
        else:
            increase_types = SUPPLIER_LEDGER_INCREASE_TYPES
            decrease_types = SUPPLIER_LEDGER_DECREASE_TYPES
            neutral_types = SUPPLIER_LEDGER_NEUTRAL_TYPES
        document_types = (*increase_types, *decrease_types, *neutral_types)

        detail_amount = (
            select(
                detail.c.document_id.label("document_id"),
                func.coalesce(func.sum(detail.c.amount), 0).label("detail_amount"),
            )
            .group_by(detail.c.document_id)
            .subquery()
        )
        effective_amount = func.coalesce(detail_amount.c.detail_amount, table.c.amount, 0)
        ledger_amount = func.abs(effective_amount)
        base_conditions = [
            table.c.deleted_at.is_(None),
            table.c.supplier == normalized_name,
            table.c.document_type.in_(document_types),
        ]

        start_date = parse_date(date_start) if date_start else None
        end_date = parse_date(date_end) if date_end else None
        range_conditions = list(base_conditions)
        if date_start:
            range_conditions.append(table.c.date_value >= start_date if start_date else table.c.date >= date_start)
        if date_end:
            range_conditions.append(table.c.date_value <= end_date if end_date else table.c.date <= date_end)

        increase_expr = case(
            (table.c.document_type.in_(increase_types), ledger_amount),
            else_=0,
        )
        decrease_expr = case(
            (table.c.document_type.in_(decrease_types), ledger_amount),
            else_=0,
        )

        items_statement = (
            select(
                table.c.id,
                table.c.document_number,
                table.c.date,
                table.c.document_type,
                table.c.summary,
                table.c.handler,
                table.c.warehouse,
                increase_expr.label("increase_amount"),
                decrease_expr.label("decrease_amount"),
            )
            .outerjoin(detail_amount, detail_amount.c.document_id == table.c.id)
            .where(and_(*range_conditions))
            .order_by(table.c.date_value.nulls_last(), table.c.date, table.c.id)
        )
        totals_statement = (
            select(
                func.coalesce(func.sum(increase_expr), 0).label("increase_total"),
                func.coalesce(func.sum(decrease_expr), 0).label("decrease_total"),
            )
            .select_from(table)
            .outerjoin(detail_amount, detail_amount.c.document_id == table.c.id)
            .where(and_(*range_conditions))
        )

        beginning_balance = Decimal("0")
        if date_start:
            beginning_conditions = list(base_conditions)
            beginning_conditions.append(table.c.date_value < start_date if start_date else table.c.date < date_start)
            beginning_statement = (
                select(func.coalesce(func.sum(increase_expr - decrease_expr), 0))
                .select_from(table)
                .outerjoin(detail_amount, detail_amount.c.document_id == table.c.id)
                .where(and_(*beginning_conditions))
            )
        else:
            beginning_statement = None

        with self.engine.connect() as connection:
            if beginning_statement is not None:
                beginning_balance = Decimal(str(connection.execute(beginning_statement).scalar_one() or "0"))
            totals = connection.execute(totals_statement).mappings().one()
            rows = [dict(row) for row in connection.execute(items_statement).mappings()]

        running_balance = beginning_balance
        items: list[dict[str, object]] = []
        for index, row in enumerate(rows, start=1):
            increase = Decimal(str(row.pop("increase_amount") or "0"))
            decrease = Decimal(str(row.pop("decrease_amount") or "0"))
            running_balance = running_balance + increase - decrease
            items.append({
                **row,
                "row_number": index,
                "increase_amount": self._format_decimal(increase) if increase else "",
                "decrease_amount": self._format_decimal(decrease) if decrease else "",
                "balance": self._format_decimal(running_balance),
            })

        increase_total = Decimal(str(totals.get("increase_total") or "0"))
        decrease_total = Decimal(str(totals.get("decrease_total") or "0"))
        ending_balance = beginning_balance + increase_total - decrease_total
        return {
            "items": items,
            "counterparty_type": counterparty_type,
            "name": normalized_name,
            "date_start": date_start,
            "date_end": date_end,
            "beginning_balance": self._format_decimal(beginning_balance),
            "increase_total": self._format_decimal(increase_total),
            "decrease_total": self._format_decimal(decrease_total),
            "ending_balance": self._format_decimal(ending_balance),
        }

    # ── Suppliers ──────────────────────────────────────────────────

    def list_supplier_brands(self) -> list[dict[str, object]]:
        statement = select(SUPPLIER_BRAND_TABLE).order_by(SUPPLIER_BRAND_TABLE.c.sort_order, SUPPLIER_BRAND_TABLE.c.id)
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def get_supplier_brand(self, brand_id: int) -> dict[str, object] | None:
        statement = select(SUPPLIER_BRAND_TABLE).where(SUPPLIER_BRAND_TABLE.c.id == brand_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_supplier_brand_by_code(self, code: str) -> dict[str, object] | None:
        statement = select(SUPPLIER_BRAND_TABLE).where(SUPPLIER_BRAND_TABLE.c.code == code)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def list_product_archive_brands(self) -> list[dict[str, object]]:
        statement = (
            select(SUPPLIER_BRAND_TABLE)
            .where(SUPPLIER_BRAND_TABLE.c.product_archive_enabled.is_(True))
            .order_by(SUPPLIER_BRAND_TABLE.c.sort_order, SUPPLIER_BRAND_TABLE.c.id)
        )
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def create_supplier_brand(self, data: Mapping[str, object]) -> dict[str, object]:
        name = str(data.get("name") or "").strip()
        code = str(data.get("code") or "").strip() or f"supplier_brand_{uuid4().hex[:12]}"
        with self.engine.begin() as connection:
            sort_order = connection.execute(select(func.coalesce(func.max(SUPPLIER_BRAND_TABLE.c.sort_order), 0))).scalar_one() + 1
            row = connection.execute(
                insert(SUPPLIER_BRAND_TABLE)
                .values(code=code, name=name, product_archive_enabled=True, sort_order=sort_order)
                .returning(SUPPLIER_BRAND_TABLE)
            ).mappings().one()
            item = dict(row)
            row = connection.execute(
                update(SUPPLIER_BRAND_TABLE)
                .where(SUPPLIER_BRAND_TABLE.c.id == item["id"])
                .values(product_table_name=f"manual_product_archive_{item['id']}")
                .returning(SUPPLIER_BRAND_TABLE)
            ).mappings().one()
        return dict(row)

    def update_supplier_brand(self, brand_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        name = str(data.get("name") or "").strip()
        with self.engine.begin() as connection:
            row = connection.execute(
                update(SUPPLIER_BRAND_TABLE)
                .where(SUPPLIER_BRAND_TABLE.c.id == brand_id)
                .values(name=name)
                .returning(SUPPLIER_BRAND_TABLE)
            ).mappings().first()
        return None if row is None else dict(row)

    def delete_supplier_brand(self, brand_id: int) -> dict[str, object] | None:
        with self.engine.begin() as connection:
            row = connection.execute(
                delete(SUPPLIER_BRAND_TABLE)
                .where(SUPPLIER_BRAND_TABLE.c.id == brand_id)
                .returning(SUPPLIER_BRAND_TABLE)
            ).mappings().first()
        return None if row is None else dict(row)

    def count_suppliers_by_brand(self, brand_code: str) -> int:
        statement = (
            select(func.count())
            .select_from(SUPPLIER_TABLE)
            .where(SUPPLIER_TABLE.c.brand == brand_code)
        )
        with self.engine.connect() as connection:
            return int(connection.execute(statement).scalar_one())

    def reorder_supplier_brands(self, ordered_ids: list[int]) -> bool:
        return self._replace_sort_order(SUPPLIER_BRAND_TABLE, ordered_ids)

    def list_suppliers(self, *, brand: str | None = None) -> list[dict[str, object]]:
        statement = select(SUPPLIER_TABLE).order_by(SUPPLIER_TABLE.c.brand, SUPPLIER_TABLE.c.id)
        if brand:
            statement = statement.where(SUPPLIER_TABLE.c.brand == brand)
        with self.engine.begin() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def list_suppliers_page(
        self,
        *,
        page: int,
        page_size: int,
        query: str | None = None,
        brand: str | None = None,
    ) -> dict[str, object]:
        count_statement = select(func.count()).select_from(SUPPLIER_TABLE)
        items_statement = (
            select(SUPPLIER_TABLE)
            .order_by(SUPPLIER_TABLE.c.id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        conditions = []
        if brand:
            conditions.append(SUPPLIER_TABLE.c.brand == brand)
        normalized_query = (query or "").strip()
        if normalized_query:
            like = f"%{normalized_query}%"
            conditions.append(or_(
                SUPPLIER_TABLE.c.name.ilike(like),
                SUPPLIER_TABLE.c.factory_code.ilike(like),
                SUPPLIER_TABLE.c.contact.ilike(like),
                SUPPLIER_TABLE.c.wechat.ilike(like),
            ))
        if conditions:
            criterion = conditions[0] if len(conditions) == 1 else and_(*conditions)
            count_statement = count_statement.where(criterion)
            items_statement = items_statement.where(criterion)
        with self.engine.begin() as connection:
            total = connection.execute(count_statement).scalar_one()
            items = [dict(row) for row in connection.execute(items_statement).mappings()]
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    def create_supplier(self, data: Mapping[str, object]) -> dict[str, object]:
        statement = insert(SUPPLIER_TABLE).values(**self._prepare_supplier(data)).returning(SUPPLIER_TABLE)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().one()
        return dict(row)

    def update_supplier(self, supplier_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        payload = self._prepare_supplier(data)
        payload.pop("id", None)
        statement = update(SUPPLIER_TABLE).where(SUPPLIER_TABLE.c.id == supplier_id).values(**payload).returning(SUPPLIER_TABLE)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_supplier(self, supplier_id: int) -> dict[str, object] | None:
        statement = select(SUPPLIER_TABLE).where(SUPPLIER_TABLE.c.id == supplier_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def delete_supplier(self, supplier_id: int) -> bool:
        statement = delete(SUPPLIER_TABLE).where(SUPPLIER_TABLE.c.id == supplier_id)
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount > 0

    def get_supplier_by_name(self, name: str, brand: str | None = None) -> dict[str, object] | None:
        statement = select(SUPPLIER_TABLE).where(SUPPLIER_TABLE.c.name == name)
        if brand:
            statement = statement.where(SUPPLIER_TABLE.c.brand == brand)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
            if row is None:
                normalized_name = str(name or "").strip().replace("（", "(").replace("）", ")").lower()
                if normalized_name:
                    normalized_supplier_name = func.lower(
                        func.replace(
                            func.replace(SUPPLIER_TABLE.c.name, "（", "("),
                            "）",
                            ")",
                        )
                    )
                    normalized_statement = select(SUPPLIER_TABLE).where(normalized_supplier_name == normalized_name)
                    if brand:
                        normalized_statement = normalized_statement.where(SUPPLIER_TABLE.c.brand == brand)
                    row = connection.execute(normalized_statement).mappings().first()
        return None if row is None else dict(row)

    @staticmethod
    def _next_sort_order(connection, table, where_clause=None) -> int:
        statement = select(func.coalesce(func.max(table.c.sort_order), 0))
        if where_clause is not None:
            statement = statement.where(where_clause)
        return int(connection.execute(statement).scalar_one() or 0) + 1

    def _replace_sort_order(self, table, ordered_ids: list[int], where_clause=None) -> bool:
        if len(ordered_ids) != len(set(ordered_ids)):
            return False
        statement = select(table.c.id)
        if where_clause is not None:
            statement = statement.where(where_clause)
        with self.engine.begin() as connection:
            existing_ids = [int(row[0]) for row in connection.execute(statement).all()]
            if set(existing_ids) != set(ordered_ids):
                return False
            for sort_order, item_id in enumerate(ordered_ids, start=1):
                connection.execute(
                    update(table)
                    .where(table.c.id == item_id)
                    .values(sort_order=sort_order)
                )
        return True

    # ── General Customer Brands & Shops ────────────────────────────

    def _general_customer_sort_orders(
        self,
        user_id: int | None,
        scope: str,
    ) -> dict[tuple[int, int], int]:
        if user_id is None:
            return {}
        statement = select(
            GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.parent_id,
            GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.item_id,
            GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.sort_order,
        ).where(
            GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.user_id == user_id,
            GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.scope == scope,
        )
        with self.engine.connect() as connection:
            rows = connection.execute(statement).mappings()
            return {
                (int(row["parent_id"]), int(row["item_id"])): int(row["sort_order"])
                for row in rows
            }

    @staticmethod
    def _general_customer_sort_key(
        preferences: Mapping[tuple[int, int], int],
        parent_id: int,
        item_id: int,
        default_sort_order: int,
    ) -> tuple[int, int, int]:
        preference_order = preferences.get((parent_id, item_id))
        if preference_order is not None:
            return (0, preference_order, item_id)
        return (1, default_sort_order, item_id)

    def _replace_general_customer_user_sort_order(
        self,
        *,
        user_id: int,
        scope: str,
        parent_id: int,
        table,
        ordered_ids: list[int],
        condition=None,
    ) -> bool:
        statement = select(table.c.id)
        if condition is not None:
            statement = statement.where(condition)
        with self.engine.begin() as connection:
            existing_ids = [int(value) for value in connection.execute(statement).scalars()]
            if set(existing_ids) != set(ordered_ids):
                return False
            connection.execute(
                delete(GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE).where(
                    GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.user_id == user_id,
                    GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.scope == scope,
                    GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.c.parent_id == parent_id,
                )
            )
            connection.execute(
                insert(GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE),
                [
                    {
                        "user_id": user_id,
                        "scope": scope,
                        "parent_id": parent_id,
                        "item_id": item_id,
                        "sort_order": sort_order,
                    }
                    for sort_order, item_id in enumerate(ordered_ids, start=1)
                ],
            )
        return True

    def list_general_customer_brands(self, user_id: int | None = None) -> list[dict[str, object]]:
        shop_count = (
            select(
                GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name.label("brand_name"),
                func.count(GENERAL_CUSTOMER_SHOP_TABLE.c.id).label("shop_count"),
            )
            .group_by(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name)
            .subquery()
        )
        statement = (
            select(
                GENERAL_CUSTOMER_BRAND_TABLE.c.id,
                GENERAL_CUSTOMER_BRAND_TABLE.c.name,
                GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order,
                GENERAL_CUSTOMER_BRAND_TABLE.c.created_at,
                GENERAL_CUSTOMER_BRAND_TABLE.c.updated_at,
                func.coalesce(shop_count.c.shop_count, 0).label("shop_count"),
            )
            .outerjoin(shop_count, shop_count.c.brand_name == GENERAL_CUSTOMER_BRAND_TABLE.c.name)
            .order_by(GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order, GENERAL_CUSTOMER_BRAND_TABLE.c.id)
        )
        with self.engine.connect() as connection:
            items = [dict(row) for row in connection.execute(statement).mappings()]
        preferences = self._general_customer_sort_orders(user_id, GENERAL_CUSTOMER_SORT_SCOPE_BRAND)
        return sorted(
            items,
            key=lambda item: self._general_customer_sort_key(
                preferences,
                GENERAL_CUSTOMER_ROOT_SORT_PARENT_ID,
                int(item["id"]),
                int(item["sort_order"]),
            ),
        )

    def create_general_customer_brand(self, data: Mapping[str, object]) -> dict[str, object]:
        payload = {
            "name": str(data.get("name") or "").strip(),
        }
        with self.engine.begin() as connection:
            payload["sort_order"] = self._next_sort_order(connection, GENERAL_CUSTOMER_BRAND_TABLE)
            statement = (
                insert(GENERAL_CUSTOMER_BRAND_TABLE)
                .values(**payload)
                .returning(
                    GENERAL_CUSTOMER_BRAND_TABLE.c.id,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.name,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.created_at,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.updated_at,
                )
            )
            row = connection.execute(statement).mappings().one()
        item = dict(row)
        item["shop_count"] = 0
        return item

    def get_general_customer_brand(self, brand_id: int) -> dict[str, object] | None:
        shop_count = (
            select(func.count())
            .select_from(GENERAL_CUSTOMER_SHOP_TABLE)
            .where(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == GENERAL_CUSTOMER_BRAND_TABLE.c.name)
            .scalar_subquery()
        )
        statement = select(
            GENERAL_CUSTOMER_BRAND_TABLE.c.id,
            GENERAL_CUSTOMER_BRAND_TABLE.c.name,
            GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order,
            GENERAL_CUSTOMER_BRAND_TABLE.c.created_at,
            GENERAL_CUSTOMER_BRAND_TABLE.c.updated_at,
            shop_count.label("shop_count"),
        ).where(GENERAL_CUSTOMER_BRAND_TABLE.c.id == brand_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def update_general_customer_brand(self, brand_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        payload = {
            "name": str(data.get("name") or "").strip(),
        }
        with self.engine.begin() as connection:
            existing = connection.execute(
                select(GENERAL_CUSTOMER_BRAND_TABLE.c.name).where(GENERAL_CUSTOMER_BRAND_TABLE.c.id == brand_id)
            ).mappings().first()
            if existing is None:
                return None
            old_name = existing["name"]
            row = connection.execute(
                update(GENERAL_CUSTOMER_BRAND_TABLE)
                .where(GENERAL_CUSTOMER_BRAND_TABLE.c.id == brand_id)
                .values(**payload)
                .returning(
                    GENERAL_CUSTOMER_BRAND_TABLE.c.id,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.name,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.created_at,
                    GENERAL_CUSTOMER_BRAND_TABLE.c.updated_at,
                )
            ).mappings().first()
            if row is None:
                return None
            new_name = row["name"]
            if old_name != new_name:
                connection.execute(
                    update(GENERAL_CUSTOMER_SHOP_TABLE)
                    .where(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == old_name)
                    .values(customer_name=new_name)
                )
            shop_count = connection.execute(
                select(func.count()).select_from(GENERAL_CUSTOMER_SHOP_TABLE).where(
                    GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == new_name
                )
            ).scalar_one()
        item = dict(row)
        item["shop_count"] = shop_count
        return item

    def delete_general_customer_brand(self, brand_id: int) -> str | None:
        with self.engine.begin() as connection:
            row = connection.execute(
                select(GENERAL_CUSTOMER_BRAND_TABLE.c.name).where(GENERAL_CUSTOMER_BRAND_TABLE.c.id == brand_id)
            ).first()
            if row is None:
                return "not_found"
            brand_name = row[0]
            connection.execute(delete(GENERAL_CUSTOMER_SHOP_TABLE).where(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == brand_name))
            result = connection.execute(delete(GENERAL_CUSTOMER_BRAND_TABLE).where(GENERAL_CUSTOMER_BRAND_TABLE.c.id == brand_id))
        return None if result.rowcount > 0 else "not_found"

    def get_general_customer_brand_by_name(self, name: str) -> dict[str, object] | None:
        statement = select(
            GENERAL_CUSTOMER_BRAND_TABLE.c.id,
            GENERAL_CUSTOMER_BRAND_TABLE.c.name,
            GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order,
            GENERAL_CUSTOMER_BRAND_TABLE.c.created_at,
            GENERAL_CUSTOMER_BRAND_TABLE.c.updated_at,
        ).where(GENERAL_CUSTOMER_BRAND_TABLE.c.name == name)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def list_general_customer_shops(self, user_id: int | None = None) -> list[dict[str, object]]:
        unit_count = (
            select(
                GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id,
                func.count(GENERAL_CUSTOMER_UNIT_TABLE.c.id).label("unit_count"),
            )
            .group_by(GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id)
            .subquery()
        )
        statement = (
            select(
                GENERAL_CUSTOMER_SHOP_TABLE.c.id,
                GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order,
                GENERAL_CUSTOMER_SHOP_TABLE.c.created_at,
                GENERAL_CUSTOMER_SHOP_TABLE.c.updated_at,
                GENERAL_CUSTOMER_BRAND_TABLE.c.id.label("_brand_id"),
                GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order.label("_brand_sort_order"),
                func.coalesce(unit_count.c.unit_count, 0).label("unit_count"),
            )
            .outerjoin(unit_count, unit_count.c.shop_id == GENERAL_CUSTOMER_SHOP_TABLE.c.id)
            .outerjoin(
                GENERAL_CUSTOMER_BRAND_TABLE,
                GENERAL_CUSTOMER_BRAND_TABLE.c.name == GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
            )
            .order_by(
                GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order,
                GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order,
                GENERAL_CUSTOMER_SHOP_TABLE.c.id,
            )
        )
        with self.engine.connect() as connection:
            items = [dict(row) for row in connection.execute(statement).mappings()]
        brand_preferences = self._general_customer_sort_orders(user_id, GENERAL_CUSTOMER_SORT_SCOPE_BRAND)
        shop_preferences = self._general_customer_sort_orders(user_id, GENERAL_CUSTOMER_SORT_SCOPE_SHOP)
        items.sort(
            key=lambda item: (
                self._general_customer_sort_key(
                    brand_preferences,
                    GENERAL_CUSTOMER_ROOT_SORT_PARENT_ID,
                    int(item["_brand_id"]),
                    int(item["_brand_sort_order"]),
                ),
                self._general_customer_sort_key(
                    shop_preferences,
                    int(item["_brand_id"]),
                    int(item["id"]),
                    int(item["sort_order"]),
                ),
            )
        )
        for item in items:
            item.pop("_brand_id", None)
            item.pop("_brand_sort_order", None)
        return items

    def create_general_customer_shop(self, data: Mapping[str, object]) -> dict[str, object]:
        payload = {
            "customer_name": str(data.get("customer_name") or "").strip(),
            "shop_name": str(data.get("shop_name") or "").strip(),
        }
        with self.engine.begin() as connection:
            self._ensure_general_customer_brand(connection, payload["customer_name"])
            payload["sort_order"] = self._next_sort_order(
                connection,
                GENERAL_CUSTOMER_SHOP_TABLE,
                GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == payload["customer_name"],
            )
            statement = (
                insert(GENERAL_CUSTOMER_SHOP_TABLE)
                .values(**payload)
                .returning(
                    GENERAL_CUSTOMER_SHOP_TABLE.c.id,
                    GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
                    GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
                    GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order,
                    GENERAL_CUSTOMER_SHOP_TABLE.c.created_at,
                    GENERAL_CUSTOMER_SHOP_TABLE.c.updated_at,
                )
            )
            row = connection.execute(statement).mappings().one()
        item = dict(row)
        item["unit_count"] = 0
        return item

    def get_general_customer_shop(self, shop_id: int) -> dict[str, object] | None:
        unit_count = (
            select(func.count())
            .select_from(GENERAL_CUSTOMER_UNIT_TABLE)
            .where(GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == GENERAL_CUSTOMER_SHOP_TABLE.c.id)
            .scalar_subquery()
        )
        statement = select(
            GENERAL_CUSTOMER_SHOP_TABLE.c.id,
            GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
            GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
            GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order,
            GENERAL_CUSTOMER_SHOP_TABLE.c.created_at,
            GENERAL_CUSTOMER_SHOP_TABLE.c.updated_at,
            unit_count.label("unit_count"),
        ).where(GENERAL_CUSTOMER_SHOP_TABLE.c.id == shop_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def reorder_general_customer_brands(self, user_id: int, ordered_ids: list[int]) -> bool:
        return self._replace_general_customer_user_sort_order(
            user_id=user_id,
            scope=GENERAL_CUSTOMER_SORT_SCOPE_BRAND,
            parent_id=GENERAL_CUSTOMER_ROOT_SORT_PARENT_ID,
            table=GENERAL_CUSTOMER_BRAND_TABLE,
            ordered_ids=ordered_ids,
        )

    def update_general_customer_shop(self, shop_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        payload = {
            "customer_name": str(data.get("customer_name") or "").strip(),
            "shop_name": str(data.get("shop_name") or "").strip(),
        }
        statement = (
            update(GENERAL_CUSTOMER_SHOP_TABLE)
            .where(GENERAL_CUSTOMER_SHOP_TABLE.c.id == shop_id)
            .values(**payload)
            .returning(
                GENERAL_CUSTOMER_SHOP_TABLE.c.id,
                GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.created_at,
                GENERAL_CUSTOMER_SHOP_TABLE.c.updated_at,
            )
        )
        with self.engine.begin() as connection:
            self._ensure_general_customer_brand(connection, payload["customer_name"])
            row = connection.execute(statement).mappings().first()
            unit_count = 0 if row is None else connection.execute(
                select(func.count()).select_from(GENERAL_CUSTOMER_UNIT_TABLE).where(
                    GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == shop_id
                )
            ).scalar_one()
        if row is None:
            return None
        item = dict(row)
        item["unit_count"] = unit_count
        return item

    def delete_general_customer_shop(self, shop_id: int) -> bool:
        statement = delete(GENERAL_CUSTOMER_SHOP_TABLE).where(GENERAL_CUSTOMER_SHOP_TABLE.c.id == shop_id)
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount > 0

    def get_general_customer_shop_by_name(self, customer_name: str, shop_name: str) -> dict[str, object] | None:
        statement = select(
            GENERAL_CUSTOMER_SHOP_TABLE.c.id,
            GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
            GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
            GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order,
            GENERAL_CUSTOMER_SHOP_TABLE.c.created_at,
            GENERAL_CUSTOMER_SHOP_TABLE.c.updated_at,
        ).where(
            GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == customer_name,
            GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name == shop_name,
        )
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def reorder_general_customer_shops(self, user_id: int, customer_name: str, ordered_ids: list[int]) -> bool:
        with self.engine.connect() as connection:
            brand_id = connection.execute(
                select(GENERAL_CUSTOMER_BRAND_TABLE.c.id).where(
                    GENERAL_CUSTOMER_BRAND_TABLE.c.name == customer_name
                )
            ).scalar_one_or_none()
        if brand_id is None:
            return False
        return self._replace_general_customer_user_sort_order(
            user_id=user_id,
            scope=GENERAL_CUSTOMER_SORT_SCOPE_SHOP,
            parent_id=int(brand_id),
            table=GENERAL_CUSTOMER_SHOP_TABLE,
            ordered_ids=ordered_ids,
            condition=GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == customer_name,
        )

    def list_general_customer_units(self, user_id: int | None = None) -> list[dict[str, object]]:
        statement = (
            select(
                GENERAL_CUSTOMER_UNIT_TABLE.c.id,
                GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id,
                GENERAL_CUSTOMER_UNIT_TABLE.c.unit_name,
                GENERAL_CUSTOMER_UNIT_TABLE.c.sort_order,
                GENERAL_CUSTOMER_UNIT_TABLE.c.created_at,
                GENERAL_CUSTOMER_UNIT_TABLE.c.updated_at,
                GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order.label("_shop_sort_order"),
                GENERAL_CUSTOMER_BRAND_TABLE.c.id.label("_brand_id"),
                GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order.label("_brand_sort_order"),
            )
            .join(GENERAL_CUSTOMER_SHOP_TABLE, GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == GENERAL_CUSTOMER_SHOP_TABLE.c.id)
            .join(GENERAL_CUSTOMER_BRAND_TABLE, GENERAL_CUSTOMER_BRAND_TABLE.c.name == GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name)
            .order_by(
                GENERAL_CUSTOMER_BRAND_TABLE.c.sort_order,
                GENERAL_CUSTOMER_SHOP_TABLE.c.sort_order,
                GENERAL_CUSTOMER_UNIT_TABLE.c.sort_order,
                GENERAL_CUSTOMER_UNIT_TABLE.c.id,
            )
        )
        with self.engine.connect() as connection:
            items = [dict(row) for row in connection.execute(statement).mappings()]
        brand_preferences = self._general_customer_sort_orders(user_id, GENERAL_CUSTOMER_SORT_SCOPE_BRAND)
        shop_preferences = self._general_customer_sort_orders(user_id, GENERAL_CUSTOMER_SORT_SCOPE_SHOP)
        unit_preferences = self._general_customer_sort_orders(user_id, GENERAL_CUSTOMER_SORT_SCOPE_UNIT)
        items.sort(
            key=lambda item: (
                self._general_customer_sort_key(
                    brand_preferences,
                    GENERAL_CUSTOMER_ROOT_SORT_PARENT_ID,
                    int(item["_brand_id"]),
                    int(item["_brand_sort_order"]),
                ),
                self._general_customer_sort_key(
                    shop_preferences,
                    int(item["_brand_id"]),
                    int(item["shop_id"]),
                    int(item["_shop_sort_order"]),
                ),
                self._general_customer_sort_key(
                    unit_preferences,
                    int(item["shop_id"]),
                    int(item["id"]),
                    int(item["sort_order"]),
                ),
            )
        )
        for item in items:
            item.pop("_brand_id", None)
            item.pop("_brand_sort_order", None)
            item.pop("_shop_sort_order", None)
        return items

    def create_general_customer_unit(self, data: Mapping[str, object]) -> dict[str, object]:
        payload = {
            "shop_id": int(data.get("shop_id") or 0),
            "unit_name": str(data.get("unit_name") or "").strip(),
        }
        with self.engine.begin() as connection:
            payload["sort_order"] = self._next_sort_order(
                connection,
                GENERAL_CUSTOMER_UNIT_TABLE,
                GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == payload["shop_id"],
            )
            statement = insert(GENERAL_CUSTOMER_UNIT_TABLE).values(**payload).returning(GENERAL_CUSTOMER_UNIT_TABLE)
            row = connection.execute(statement).mappings().one()
            shop = connection.execute(
                select(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name, GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name)
                .where(GENERAL_CUSTOMER_SHOP_TABLE.c.id == payload["shop_id"])
            ).mappings().one()
        item = dict(row)
        item.update(shop)
        return item

    def get_general_customer_unit(self, unit_id: int) -> dict[str, object] | None:
        statement = (
            select(
                GENERAL_CUSTOMER_UNIT_TABLE.c.id,
                GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id,
                GENERAL_CUSTOMER_UNIT_TABLE.c.unit_name,
                GENERAL_CUSTOMER_UNIT_TABLE.c.sort_order,
                GENERAL_CUSTOMER_UNIT_TABLE.c.created_at,
                GENERAL_CUSTOMER_UNIT_TABLE.c.updated_at,
                GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name,
                GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name,
            )
            .join(GENERAL_CUSTOMER_SHOP_TABLE, GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == GENERAL_CUSTOMER_SHOP_TABLE.c.id)
            .where(GENERAL_CUSTOMER_UNIT_TABLE.c.id == unit_id)
        )
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_general_customer_unit_by_name(self, shop_id: int, unit_name: str) -> dict[str, object] | None:
        statement = select(GENERAL_CUSTOMER_UNIT_TABLE).where(
            GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == shop_id,
            GENERAL_CUSTOMER_UNIT_TABLE.c.unit_name == unit_name,
        )
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def reorder_general_customer_units(self, user_id: int, shop_id: int, ordered_ids: list[int]) -> bool:
        return self._replace_general_customer_user_sort_order(
            user_id=user_id,
            scope=GENERAL_CUSTOMER_SORT_SCOPE_UNIT,
            parent_id=shop_id,
            table=GENERAL_CUSTOMER_UNIT_TABLE,
            ordered_ids=ordered_ids,
            condition=GENERAL_CUSTOMER_UNIT_TABLE.c.shop_id == shop_id,
        )

    def update_general_customer_unit(self, unit_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        payload = {
            "shop_id": int(data.get("shop_id") or 0),
            "unit_name": str(data.get("unit_name") or "").strip(),
        }
        statement = (
            update(GENERAL_CUSTOMER_UNIT_TABLE)
            .where(GENERAL_CUSTOMER_UNIT_TABLE.c.id == unit_id)
            .values(**payload)
            .returning(GENERAL_CUSTOMER_UNIT_TABLE)
        )
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().first()
            if row is None:
                return None
            shop = connection.execute(
                select(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name, GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name)
                .where(GENERAL_CUSTOMER_SHOP_TABLE.c.id == payload["shop_id"])
            ).mappings().one()
        item = dict(row)
        item.update(shop)
        return item

    def delete_general_customer_unit(self, unit_id: int) -> bool:
        statement = delete(GENERAL_CUSTOMER_UNIT_TABLE).where(GENERAL_CUSTOMER_UNIT_TABLE.c.id == unit_id)
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount > 0

    # ── Inventory Account Subjects ─────────────────────────────────

    def list_account_subjects(self) -> list[dict[str, object]]:
        statement = select(INVENTORY_ACCOUNT_SUBJECT_TABLE).order_by(
            INVENTORY_ACCOUNT_SUBJECT_TABLE.c.id,
            INVENTORY_ACCOUNT_SUBJECT_TABLE.c.name,
        )
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def create_account_subject(self, data: Mapping[str, object]) -> dict[str, object]:
        payload = {
            "code": str(data.get("code") or "").strip() or None,
            "name": str(data.get("name") or "").strip(),
        }
        statement = insert(INVENTORY_ACCOUNT_SUBJECT_TABLE).values(**payload).returning(INVENTORY_ACCOUNT_SUBJECT_TABLE)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().one()
        return dict(row)

    def get_account_subject(self, subject_id: int) -> dict[str, object] | None:
        statement = select(INVENTORY_ACCOUNT_SUBJECT_TABLE).where(INVENTORY_ACCOUNT_SUBJECT_TABLE.c.id == subject_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def delete_account_subject(self, subject_id: int) -> bool:
        statement = delete(INVENTORY_ACCOUNT_SUBJECT_TABLE).where(INVENTORY_ACCOUNT_SUBJECT_TABLE.c.id == subject_id)
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount > 0

    # ── Purchase Order Requirements ────────────────────────────────

    def list_purchase_order_requirement_templates(self) -> list[dict[str, object]]:
        statement = select(PURCHASE_ORDER_REQUIREMENT_TABLE).order_by(PURCHASE_ORDER_REQUIREMENT_TABLE.c.brand)
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def get_purchase_order_requirement_template_map(self) -> dict[str, str]:
        rows = self.list_purchase_order_requirement_templates()
        return {
            str(row.get("brand") or "").strip().lower(): str(row.get("content") or "")
            for row in rows
            if str(row.get("brand") or "").strip()
        }

    def upsert_purchase_order_requirement_template(self, brand: str, content: str) -> dict[str, object]:
        normalized_brand = str(brand or "").strip().lower()
        payload = {
            "brand": normalized_brand,
            "content": str(content or ""),
        }
        insert_statement = pg_insert(PURCHASE_ORDER_REQUIREMENT_TABLE).values(**payload)
        statement = insert_statement.on_conflict_do_update(
            index_elements=[PURCHASE_ORDER_REQUIREMENT_TABLE.c.brand],
            set_={
                "content": insert_statement.excluded.content,
                "updated_at": func.date_trunc('minute', func.now()),
            },
        ).returning(PURCHASE_ORDER_REQUIREMENT_TABLE)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().one()
        return dict(row)

    # ── Warehouses ─────────────────────────────────────────────────

    def list_warehouses(self) -> list[dict[str, object]]:
        statement = (
            select(WAREHOUSE_TABLE)
            .outerjoin(WAREHOUSE_BRAND_TABLE, WAREHOUSE_BRAND_TABLE.c.name == WAREHOUSE_TABLE.c.brand)
            .order_by(WAREHOUSE_BRAND_TABLE.c.sort_order, WAREHOUSE_TABLE.c.sort_order, WAREHOUSE_TABLE.c.id)
        )
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def list_warehouse_brands(self) -> list[dict[str, object]]:
        warehouse_count = (
            select(
                WAREHOUSE_TABLE.c.brand.label("brand_name"),
                func.count().label("warehouse_count"),
            )
            .group_by(WAREHOUSE_TABLE.c.brand)
            .subquery()
        )
        statement = (
            select(
                WAREHOUSE_BRAND_TABLE.c.id,
                WAREHOUSE_BRAND_TABLE.c.name,
                WAREHOUSE_BRAND_TABLE.c.sort_order,
                WAREHOUSE_BRAND_TABLE.c.created_at,
                WAREHOUSE_BRAND_TABLE.c.updated_at,
                func.coalesce(warehouse_count.c.warehouse_count, 0).label("warehouse_count"),
            )
            .outerjoin(warehouse_count, warehouse_count.c.brand_name == WAREHOUSE_BRAND_TABLE.c.name)
            .order_by(WAREHOUSE_BRAND_TABLE.c.sort_order, WAREHOUSE_BRAND_TABLE.c.id)
        )
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def create_warehouse_brand(self, data: Mapping[str, object]) -> dict[str, object]:
        payload = {"name": str(data.get("name") or "").strip()}
        with self.engine.begin() as connection:
            payload["sort_order"] = self._next_sort_order(connection, WAREHOUSE_BRAND_TABLE)
            statement = (
                insert(WAREHOUSE_BRAND_TABLE)
                .values(**payload)
                .returning(
                    WAREHOUSE_BRAND_TABLE.c.id,
                    WAREHOUSE_BRAND_TABLE.c.name,
                    WAREHOUSE_BRAND_TABLE.c.sort_order,
                    WAREHOUSE_BRAND_TABLE.c.created_at,
                    WAREHOUSE_BRAND_TABLE.c.updated_at,
                )
            )
            row = connection.execute(statement).mappings().one()
        record = dict(row)
        record["warehouse_count"] = 0
        return record

    def get_warehouse_brand(self, brand_id: int) -> dict[str, object] | None:
        warehouse_count = (
            select(func.count())
            .select_from(WAREHOUSE_TABLE)
            .where(WAREHOUSE_TABLE.c.brand == WAREHOUSE_BRAND_TABLE.c.name)
            .scalar_subquery()
        )
        statement = select(
            WAREHOUSE_BRAND_TABLE.c.id,
            WAREHOUSE_BRAND_TABLE.c.name,
            WAREHOUSE_BRAND_TABLE.c.sort_order,
            WAREHOUSE_BRAND_TABLE.c.created_at,
            WAREHOUSE_BRAND_TABLE.c.updated_at,
            warehouse_count.label("warehouse_count"),
        ).where(WAREHOUSE_BRAND_TABLE.c.id == brand_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_warehouse_brand_by_name(self, name: str) -> dict[str, object] | None:
        statement = select(WAREHOUSE_BRAND_TABLE).where(WAREHOUSE_BRAND_TABLE.c.name == name)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def reorder_warehouse_brands(self, ordered_ids: list[int]) -> bool:
        return self._replace_sort_order(WAREHOUSE_BRAND_TABLE, ordered_ids)

    def update_warehouse_brand(self, brand_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        payload = {"name": str(data.get("name") or "").strip()}
        with self.engine.begin() as connection:
            previous = connection.execute(
                select(WAREHOUSE_BRAND_TABLE.c.name).where(WAREHOUSE_BRAND_TABLE.c.id == brand_id)
            ).mappings().first()
            if previous is None:
                return None
            row = connection.execute(
                update(WAREHOUSE_BRAND_TABLE)
                .where(WAREHOUSE_BRAND_TABLE.c.id == brand_id)
                .values(**payload)
                .returning(
                    WAREHOUSE_BRAND_TABLE.c.id,
                    WAREHOUSE_BRAND_TABLE.c.name,
                    WAREHOUSE_BRAND_TABLE.c.created_at,
                    WAREHOUSE_BRAND_TABLE.c.updated_at,
                )
            ).mappings().first()
            if row is None:
                return None
            old_name = str(previous["name"] or "")
            if old_name != payload["name"]:
                connection.execute(
                    update(WAREHOUSE_TABLE)
                    .where(WAREHOUSE_TABLE.c.brand == old_name)
                    .values(brand=payload["name"])
                )
            warehouse_count = connection.execute(
                select(func.count()).select_from(WAREHOUSE_TABLE).where(WAREHOUSE_TABLE.c.brand == payload["name"])
            ).scalar_one()
        record = dict(row)
        record["warehouse_count"] = warehouse_count
        return record

    def delete_warehouse_brand(self, brand_id: int) -> str | None:
        with self.engine.begin() as connection:
            row = connection.execute(
                select(WAREHOUSE_BRAND_TABLE.c.name).where(WAREHOUSE_BRAND_TABLE.c.id == brand_id)
            ).first()
            if row is None:
                return "not_found"
            brand_name = str(row[0] or "")
            has_warehouses = connection.execute(
                select(WAREHOUSE_TABLE.c.id).where(WAREHOUSE_TABLE.c.brand == brand_name).limit(1)
            ).first()
            if has_warehouses is not None:
                return "in_use"
            result = connection.execute(delete(WAREHOUSE_BRAND_TABLE).where(WAREHOUSE_BRAND_TABLE.c.id == brand_id))
        return None if result.rowcount > 0 else "not_found"

    def create_warehouse(self, data: Mapping[str, object]) -> dict[str, object]:
        payload = dict(data)
        payload["brand"] = str(payload.get("brand") or "通用").strip() or "通用"
        with self.engine.begin() as connection:
            payload["sort_order"] = self._next_sort_order(
                connection,
                WAREHOUSE_TABLE,
                WAREHOUSE_TABLE.c.brand == payload["brand"],
            )
            statement = insert(WAREHOUSE_TABLE).values(**payload).returning(WAREHOUSE_TABLE)
            row = connection.execute(statement).mappings().one()
        return dict(row)

    def reorder_warehouses(self, brand: str, ordered_ids: list[int]) -> bool:
        return self._replace_sort_order(
            WAREHOUSE_TABLE,
            ordered_ids,
            WAREHOUSE_TABLE.c.brand == brand,
        )

    def update_warehouse(self, warehouse_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        payload = dict(data)
        payload.pop("id", None)
        with self.engine.begin() as connection:
            before = connection.execute(
                select(WAREHOUSE_TABLE.c.name).where(WAREHOUSE_TABLE.c.id == warehouse_id)
            ).scalar_one_or_none()
            if before is None:
                return None
            previous_name = str(before).strip()
            next_name = str(payload.get("name") or previous_name).strip()
            if next_name and next_name != previous_name:
                connection.execute(
                    update(INVENTORY_TABLE)
                    .where(INVENTORY_TABLE.c.warehouse == previous_name)
                    .values(warehouse=next_name)
                )
                connection.execute(
                    update(INVENTORY_TABLE)
                    .where(
                        INVENTORY_TABLE.c.document_type == "同价调拨单",
                        INVENTORY_TABLE.c.supplier == previous_name,
                    )
                    .values(supplier=next_name)
                )
            statement = update(WAREHOUSE_TABLE).where(WAREHOUSE_TABLE.c.id == warehouse_id).values(**payload).returning(WAREHOUSE_TABLE)
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def get_warehouse(self, warehouse_id: int) -> dict[str, object] | None:
        statement = select(WAREHOUSE_TABLE).where(WAREHOUSE_TABLE.c.id == warehouse_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def delete_warehouse(self, warehouse_id: int) -> bool:
        statement = delete(WAREHOUSE_TABLE).where(WAREHOUSE_TABLE.c.id == warehouse_id)
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        return result.rowcount > 0

    def get_warehouse_by_name(self, name: str) -> dict[str, object] | None:
        statement = select(WAREHOUSE_TABLE).where(WAREHOUSE_TABLE.c.name == name)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    # ── Warehouse inventory ───────────────────────────────────────

    @staticmethod
    def _warehouse_inventory_expressions(warehouse_name: str):
        """Return inbound, outbound and scope expressions for one warehouse."""
        record = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        quantity = func.coalesce(detail.c.quantity, 0)
        inbound_types = ("进货单", "报溢单", "批发销售退货单")
        outbound_types = ("进货退货单", "报损单", "批发销售单")

        inbound = case(
            (
                or_(
                    and_(record.c.document_type.in_(inbound_types), record.c.warehouse == warehouse_name),
                    and_(record.c.document_type == "同价调拨单", record.c.warehouse == warehouse_name),
                ),
                quantity,
            ),
            else_=0,
        )
        outbound = case(
            (
                or_(
                    and_(record.c.document_type.in_(outbound_types), record.c.warehouse == warehouse_name),
                    and_(record.c.document_type == "同价调拨单", record.c.supplier == warehouse_name),
                ),
                quantity,
            ),
            else_=0,
        )
        affected = or_(
            and_(record.c.document_type.in_(inbound_types + outbound_types), record.c.warehouse == warehouse_name),
            and_(
                record.c.document_type == "同价调拨单",
                or_(record.c.warehouse == warehouse_name, record.c.supplier == warehouse_name),
            ),
        )
        return inbound, outbound, affected

    def get_warehouse_inventory(
        self,
        *,
        warehouse_name: str,
        date_start: str | None = None,
        date_end: str | None = None,
        product_code: str | None = None,
        page: int,
        page_size: int,
    ) -> dict[str, object]:
        record = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        joined = detail.join(record, detail.c.document_id == record.c.id)
        inbound, outbound, affected = self._warehouse_inventory_expressions(warehouse_name)
        start_date = parse_date(date_start) if date_start else None
        end_date = parse_date(date_end) if date_end else None

        conditions = [record.c.deleted_at.is_(None), affected]
        if end_date:
            conditions.append(or_(
                record.c.date_value <= end_date,
                and_(record.c.date_value.is_(None), record.c.date <= date_end),
            ))
        elif date_end:
            conditions.append(record.c.date <= date_end)
        if product_code:
            conditions.append(detail.c.product_code.ilike(f"%{product_code.strip()}%"))

        if start_date:
            before_period = or_(
                record.c.date_value < start_date,
                and_(record.c.date_value.is_(None), record.c.date < date_start),
            )
            in_period = or_(
                record.c.date_value >= start_date,
                and_(record.c.date_value.is_(None), record.c.date >= date_start),
            )
        elif date_start:
            before_period = record.c.date < date_start
            in_period = record.c.date >= date_start
        else:
            before_period = False
            in_period = True

        beginning_qty = func.coalesce(func.sum(case((before_period, inbound - outbound), else_=0)), 0).label("beginning_qty")
        inbound_qty = func.coalesce(func.sum(case((in_period, inbound), else_=0)), 0).label("inbound_qty")
        outbound_qty = func.coalesce(func.sum(case((in_period, outbound), else_=0)), 0).label("outbound_qty")
        grouped = (
            select(
                detail.c.product_code,
                detail.c.product_name,
                detail.c.color_name,
                detail.c.color_spec,
                beginning_qty,
                inbound_qty,
                outbound_qty,
            )
            .select_from(joined)
            .where(and_(*conditions))
            .group_by(
                detail.c.product_code,
                detail.c.product_name,
                detail.c.color_name,
                detail.c.color_spec,
            )
            .subquery()
        )
        ending_qty = (grouped.c.beginning_qty + grouped.c.inbound_qty - grouped.c.outbound_qty).label("ending_qty")
        count_statement = select(func.count()).select_from(grouped)
        totals_statement = select(
            func.coalesce(func.sum(grouped.c.beginning_qty), 0).label("beginning_qty"),
            func.coalesce(func.sum(grouped.c.inbound_qty), 0).label("inbound_qty"),
            func.coalesce(func.sum(grouped.c.outbound_qty), 0).label("outbound_qty"),
            func.coalesce(func.sum(ending_qty), 0).label("ending_qty"),
        )
        items_statement = (
            select(grouped, ending_qty)
            .order_by(grouped.c.product_code, grouped.c.color_name, grouped.c.color_spec)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )

        with self.engine.connect() as connection:
            total = connection.execute(count_statement).scalar_one()
            totals = connection.execute(totals_statement).mappings().one()
            rows = [dict(row) for row in connection.execute(items_statement).mappings()]

        def format_quantity(value: object) -> str:
            return self._format_decimal(Decimal(str(value or "0")))

        items = [{
            "product_code": row.get("product_code"),
            "product_name": row.get("product_name"),
            "color_name": row.get("color_name"),
            "color_spec": row.get("color_spec"),
            "beginning_qty": format_quantity(row.get("beginning_qty")),
            "inbound_qty": format_quantity(row.get("inbound_qty")),
            "outbound_qty": format_quantity(row.get("outbound_qty")),
            "ending_qty": format_quantity(row.get("ending_qty")),
        } for row in rows]
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "totals": {
                "beginning_qty": format_quantity(totals.get("beginning_qty")),
                "inbound_qty": format_quantity(totals.get("inbound_qty")),
                "outbound_qty": format_quantity(totals.get("outbound_qty")),
                "ending_qty": format_quantity(totals.get("ending_qty")),
            },
        }

    def list_warehouse_inventory_movements(
        self,
        *,
        warehouse_name: str,
        date_start: str | None = None,
        date_end: str | None = None,
        product_code: str | None = None,
        color_name: str | None = None,
        color_spec: str | None = None,
        page: int,
        page_size: int,
    ) -> dict[str, object]:
        record = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        joined = detail.join(record, detail.c.document_id == record.c.id)
        inbound, outbound, affected = self._warehouse_inventory_expressions(warehouse_name)
        conditions = [record.c.deleted_at.is_(None), affected]
        start_date = parse_date(date_start) if date_start else None
        end_date = parse_date(date_end) if date_end else None
        if start_date:
            conditions.append(or_(
                record.c.date_value >= start_date,
                and_(record.c.date_value.is_(None), record.c.date >= date_start),
            ))
        elif date_start:
            conditions.append(record.c.date >= date_start)
        if end_date:
            conditions.append(or_(
                record.c.date_value <= end_date,
                and_(record.c.date_value.is_(None), record.c.date <= date_end),
            ))
        elif date_end:
            conditions.append(record.c.date <= date_end)
        if product_code:
            conditions.append(detail.c.product_code.ilike(f"%{product_code.strip()}%"))
        if color_name:
            conditions.append(detail.c.color_name == color_name)
        if color_spec:
            conditions.append(detail.c.color_spec == color_spec)

        criterion = and_(*conditions)
        count_statement = select(func.count()).select_from(joined).where(criterion)
        items_statement = (
            select(
                detail.c.id.label("detail_id"),
                record.c.id.label("document_id"),
                record.c.date,
                record.c.date_value,
                record.c.document_type,
                record.c.document_number,
                record.c.supplier,
                record.c.warehouse,
                record.c.summary,
                record.c.handler,
                detail.c.product_code,
                detail.c.product_name,
                detail.c.color_name,
                detail.c.color_spec,
                inbound.label("inbound_qty"),
                outbound.label("outbound_qty"),
                (inbound - outbound).label("change_qty"),
            )
            .select_from(joined)
            .where(criterion)
            .order_by(record.c.date_value.nulls_last(), record.c.date, record.c.document_number, detail.c.id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        with self.engine.connect() as connection:
            total = connection.execute(count_statement).scalar_one()
            rows = [dict(row) for row in connection.execute(items_statement).mappings()]

        def format_quantity(value: object) -> str:
            return self._format_decimal(Decimal(str(value or "0")))

        return {
            "items": [{
                **row,
                "inbound_qty": format_quantity(row.get("inbound_qty")),
                "outbound_qty": format_quantity(row.get("outbound_qty")),
                "change_qty": format_quantity(row.get("change_qty")),
            } for row in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    # ── Inventory Details ───────────────────────────────────────────

    def list_details(self, document_id: int) -> list[dict[str, object]]:
        table = INVENTORY_DETAIL_TABLE
        statement = select(table).where(table.c.document_id == document_id).order_by(table.c.id)
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def list_details_page(self, document_id: int, *, page: int, page_size: int) -> dict[str, object]:
        table = INVENTORY_DETAIL_TABLE
        normalized_page = max(1, page)
        normalized_page_size = min(max(1, page_size), 500)
        offset = (normalized_page - 1) * normalized_page_size
        statement = (
            select(table)
            .where(table.c.document_id == document_id)
            .order_by(table.c.id)
            .limit(normalized_page_size)
            .offset(offset)
        )
        with self.engine.connect() as connection:
            total = int(connection.execute(
                select(func.count()).select_from(table).where(table.c.document_id == document_id)
            ).scalar_one())
            items = [dict(row) for row in connection.execute(statement).mappings()]
        return {
            "items": items,
            "total": total,
            "page": normalized_page,
            "page_size": normalized_page_size,
        }

    def get_detail(self, detail_id: int) -> dict[str, object] | None:
        table = INVENTORY_DETAIL_TABLE
        statement = select(table).where(table.c.id == detail_id)
        with self.engine.connect() as connection:
            row = connection.execute(statement).mappings().first()
        return None if row is None else dict(row)

    def list_details_for_documents(self, document_ids: list[int]) -> list[dict[str, object]]:
        if not document_ids:
            return []
        table = INVENTORY_DETAIL_TABLE
        statement = (
            select(table)
            .where(table.c.document_id.in_(document_ids))
            .order_by(table.c.document_id, table.c.id)
        )
        with self.engine.connect() as connection:
            return [dict(row) for row in connection.execute(statement).mappings()]

    def create_detail(self, data: Mapping[str, object]) -> dict[str, object]:
        table = INVENTORY_DETAIL_TABLE
        payload = self._filter_table_payload(table, self._coerce_empty(data))
        statement = insert(table).values(**payload).returning(table)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().one()
        self.recalculate_totals(payload.get("document_id"))
        return dict(row)

    def create_details(self, rows: list[Mapping[str, object]], document_id: object) -> int:
        if not rows:
            return 0
        table = INVENTORY_DETAIL_TABLE
        payload = [self._filter_table_payload(table, self._coerce_empty(row)) for row in rows]
        with self.engine.begin() as connection:
            result = connection.execute(insert(table), payload)
        self.recalculate_totals(document_id)
        return result.rowcount if result.rowcount and result.rowcount > 0 else len(payload)

    def replace_details(self, document_id: object, rows: list[Mapping[str, object]]) -> int:
        table = INVENTORY_DETAIL_TABLE
        payload = []
        for row in rows:
            item = self._filter_table_payload(table, self._coerce_empty(row))
            item["document_id"] = document_id
            payload.append(item)
        with self.engine.begin() as connection:
            connection.execute(delete(table).where(table.c.document_id == document_id))
            result = connection.execute(insert(table), payload) if payload else None
        self.recalculate_totals(document_id)
        if result is None:
            return 0
        return result.rowcount if result.rowcount and result.rowcount > 0 else len(payload)

    def merge_imported_details(self, document_id: object, rows: list[Mapping[str, object]]) -> dict[str, int]:
        """Merge Excel-imported details without removing existing detail rows.

        The import parser has already grouped rows by product and color.  Matching
        rows refresh the corresponding detail, while rows absent from the workbook
        remain untouched so re-importing a partial workbook cannot remove them.
        """
        table = INVENTORY_DETAIL_TABLE
        payloads = []
        for row in rows:
            payload = self._filter_table_payload(table, self._coerce_empty(row))
            payload["document_id"] = document_id
            payloads.append(payload)

        def match_key(row: Mapping[str, object]) -> tuple[str, str]:
            product_code = str(row.get("product_code") or "").strip()
            color_key = str(
                row.get("color_barcode")
                or row.get("color_name")
                or row.get("color_spec")
                or ""
            ).strip()
            return product_code, color_key

        added = 0
        updated = 0
        with self.engine.begin() as connection:
            existing_rows = connection.execute(
                select(table).where(table.c.document_id == document_id).order_by(table.c.id)
            ).mappings()
            existing_by_key: dict[tuple[str, str], int] = {}
            existing_by_product: dict[str, list[int]] = {}
            for existing in existing_rows:
                existing_id = int(existing["id"])
                key = match_key(existing)
                existing_by_key.setdefault(key, existing_id)
                if key[0]:
                    existing_by_product.setdefault(key[0], []).append(existing_id)

            for payload in payloads:
                key = match_key(payload)
                existing_id = existing_by_key.get(key)
                # A blank color can safely match only when the document has one
                # existing row for that product code.
                if existing_id is None and not key[1]:
                    candidates = existing_by_product.get(key[0], [])
                    if len(candidates) == 1:
                        existing_id = candidates[0]

                if existing_id is None:
                    existing_id = int(
                        connection.execute(
                            insert(table).values(**payload).returning(table.c.id)
                        ).scalar_one()
                    )
                    added += 1
                    existing_by_key.setdefault(key, existing_id)
                    if key[0]:
                        existing_by_product.setdefault(key[0], []).append(existing_id)
                    continue

                update_payload = dict(payload)
                update_payload.pop("document_id", None)
                connection.execute(
                    update(table).where(table.c.id == existing_id).values(**update_payload)
                )
                updated += 1

        self.recalculate_totals(document_id)
        return {"added": added, "updated": updated, "total": len(payloads)}

    def update_detail(self, detail_id: int, data: Mapping[str, object]) -> dict[str, object] | None:
        table = INVENTORY_DETAIL_TABLE
        payload = self._filter_table_payload(table, self._coerce_empty(data))
        document_id = payload.pop("document_id", None)
        statement = update(table).where(table.c.id == detail_id).values(**payload).returning(table)
        with self.engine.begin() as connection:
            row = connection.execute(statement).mappings().first()
        if row is None:
            return None
        if document_id is not None:
            self.recalculate_totals(document_id)
        else:
            detail = dict(row)
            self.recalculate_totals(detail.get("document_id"))
        return dict(row)

    def delete_detail(self, detail_id: int) -> bool:
        table = INVENTORY_DETAIL_TABLE
        # Get document_id before deleting for recalculation
        with self.engine.connect() as connection:
            detail = connection.execute(select(table.c.document_id).where(table.c.id == detail_id)).first()
        document_id = detail[0] if detail else None
        statement = delete(table).where(table.c.id == detail_id)
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        if result.rowcount > 0 and document_id is not None:
            self.recalculate_totals(document_id)
        return result.rowcount > 0

    def delete_details(self, document_id: object, detail_ids: list[int]) -> int:
        if not detail_ids:
            return 0
        table = INVENTORY_DETAIL_TABLE
        statement = delete(table).where(
            table.c.document_id == document_id,
            table.c.id.in_(detail_ids),
        )
        with self.engine.begin() as connection:
            result = connection.execute(statement)
        deleted = result.rowcount or 0
        if deleted > 0:
            self.recalculate_totals(document_id)
        return deleted

    def recalculate_totals(self, document_id: object) -> None:
        detail = INVENTORY_DETAIL_TABLE
        record = INVENTORY_TABLE
        with self.engine.connect() as connection:
            document_type = connection.execute(
                select(record.c.document_type).where(record.c.id == document_id)
            ).scalar_one_or_none()
            if document_type in ACCOUNTING_DOCUMENT_TYPES:
                total_count = None
                amount = connection.execute(
                    select(func.coalesce(func.sum(detail.c.amount), 0))
                    .where(detail.c.document_id == document_id)
                ).scalar_one()
            else:
                stmt = select(
                    func.coalesce(func.sum(detail.c.quantity), 0),
                    func.coalesce(func.sum(detail.c.amount), 0),
                ).where(detail.c.document_id == document_id)
                total_count, amount = connection.execute(stmt).one()
                total_count = self._apply_document_total_sign(document_type, total_count)
                amount = self._apply_document_total_sign(document_type, amount)
        update_stmt = (
            update(INVENTORY_TABLE)
            .where(INVENTORY_TABLE.c.id == document_id)
            .values(total_count=total_count, amount=amount)
        )
        with self.engine.begin() as connection:
            connection.execute(update_stmt)

    def batch_update_purchase_costs(
        self,
        *,
        date_start: str | None,
        date_end: str | None,
        price_updates: Mapping[str, object],
    ) -> dict[str, object]:
        normalized_updates = {
            str(product_code or "").strip(): Decimal(str(unit_price).strip())
            for product_code, unit_price in price_updates.items()
            if str(product_code or "").strip() and str(unit_price or "").strip()
        }
        if not normalized_updates:
            return {"updated_details": 0, "updated_documents": 0, "items": []}

        record = INVENTORY_TABLE
        detail = INVENTORY_DETAIL_TABLE
        conditions = [
            record.c.document_type.in_(("进货单", "进货退货单")),
            detail.c.product_code.in_(normalized_updates.keys()),
        ]
        if date_start:
            parsed = parse_date(date_start)
            conditions.append(record.c.date_value >= parsed if parsed else record.c.date >= date_start)
        if date_end:
            parsed = parse_date(date_end)
            conditions.append(record.c.date_value <= parsed if parsed else record.c.date <= date_end)

        joined = detail.join(record, detail.c.document_id == record.c.id)
        select_stmt = (
            select(
                detail.c.id,
                detail.c.document_id,
                detail.c.product_code,
                detail.c.quantity,
                detail.c.unit_price,
                detail.c.amount,
                record.c.document_number,
                record.c.date,
                record.c.document_type,
            )
            .select_from(joined)
            .where(and_(*conditions))
            .order_by(record.c.date_value, record.c.id, detail.c.id)
        )

        changed_documents: dict[int, str] = {}
        updated_items: list[dict[str, object]] = []
        with self.engine.begin() as connection:
            rows = [dict(row) for row in connection.execute(select_stmt).mappings()]
            for row in rows:
                product_code = str(row.get("product_code") or "").strip()
                new_price = normalized_updates.get(product_code)
                if new_price is None:
                    continue
                quantity = Decimal(str(row.get("quantity") or "0"))
                old_price = row.get("unit_price")
                new_amount = quantity * new_price
                connection.execute(
                    update(detail)
                    .where(detail.c.id == row["id"])
                    .values(unit_price=new_price, amount=new_amount)
                )
                changed_documents[int(row["document_id"])] = str(row.get("document_type") or "")
                updated_items.append({
                    "detail_id": row["id"],
                    "document_id": row["document_id"],
                    "document_number": row.get("document_number"),
                    "date": row.get("date"),
                    "document_type": row.get("document_type"),
                    "product_code": product_code,
                    "quantity": str(quantity.normalize()) if quantity.as_tuple().exponent < 0 else str(int(quantity)),
                    "old_unit_price": None if old_price is None else str(old_price),
                    "new_unit_price": str(new_price),
                    "new_amount": str(new_amount),
                })

            for document_id, document_type in changed_documents.items():
                totals = connection.execute(
                    select(
                        func.coalesce(func.sum(detail.c.quantity), 0),
                        func.coalesce(func.sum(detail.c.amount), 0),
                    ).where(detail.c.document_id == document_id)
                ).one()
                connection.execute(
                    update(record)
                    .where(record.c.id == document_id)
                    .values(
                        total_count=self._apply_document_total_sign(document_type, totals[0]),
                        amount=self._apply_document_total_sign(document_type, totals[1]),
                    )
                )

        return {
            "updated_details": len(updated_items),
            "updated_documents": len(changed_documents),
            "items": updated_items,
        }

    # ── Ending Inventory ─────────────────────────────────────────────

    def get_ending_inventory(
        self,
        *,
        jst_stock_root: Path | None,
        stock_date: str,
        date_start: str | None = None,
        date_end: str | None = None,
        product_code: str | None = None,
        page: int,
        page_size: int,
    ) -> dict[str, object]:
        detail = INVENTORY_DETAIL_TABLE
        record = INVENTORY_TABLE

        # Build the aggregated inventory changes query
        inbound_types = ("进货单", "报溢单", "批发销售退货单")
        outbound_types = ("进货退货单", "报损单", "批发销售单")
        inbound = func.sum(case(
            (record.c.document_type.in_(inbound_types), detail.c.quantity),
            else_=0,
        ))
        return_qty = func.sum(case(
            (record.c.document_type.in_(outbound_types), detail.c.quantity),
            else_=0,
        ))

        joined = detail.join(record, detail.c.document_id == record.c.id)
        conditions = []
        if date_start:
            parsed = parse_date(date_start)
            conditions.append(record.c.date_value >= parsed if parsed else record.c.date >= date_start)
        if date_end:
            parsed = parse_date(date_end)
            conditions.append(record.c.date_value <= parsed if parsed else record.c.date <= date_end)
        if product_code:
            conditions.append(detail.c.product_code.like(f"{product_code}%"))
        criterion = and_(*conditions) if conditions else None

        base = (
            select(
                detail.c.product_code,
                detail.c.product_name,
                detail.c.color_spec,
                func.min(record.c.date).label("first_doc_date"),
                inbound.label("inbound_qty"),
                return_qty.label("return_qty"),
            )
            .select_from(joined)
            .group_by(detail.c.product_code, detail.c.product_name, detail.c.color_spec)
        )
        if criterion is not None:
            base = base.where(criterion)

        # Subquery for counting total groups
        sub = base.subquery()
        count_stmt = select(func.count()).select_from(sub)
        with self.engine.connect() as connection:
            total = connection.execute(count_stmt).scalar_one()

        # Paginated query
        data_stmt = (
            select(sub)
            .order_by(sub.c.product_code)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        with self.engine.connect() as connection:
            rows = [dict(row) for row in connection.execute(data_stmt).mappings()]

        # Determine beginning stock for each product
        # If date_start is given, all products share the same beginning date.
        # Otherwise, each product uses its own earliest document date.

        def _fmt(v: int | Decimal) -> str:
            if isinstance(v, int):
                return str(v)
            d = v.normalize()
            return str(d) if d.as_tuple().exponent < 0 else str(int(d))

        def _to_mmdd(date_str: str) -> str:
            """Convert YYYY-MM-DD to MM.DD."""
            try:
                parts = date_str.split("-")
                return f"{parts[1]}.{parts[2]}"
            except (IndexError, ValueError):
                return stock_date

        beginning_by_date: dict[str, dict[str, int]] = {}

        def _stock_for_date(mmdd: str) -> dict[str, int]:
            if mmdd not in beginning_by_date:
                beginning_by_date[mmdd] = self._read_jst_stock(jst_stock_root, mmdd) if jst_stock_root else {}
            return beginning_by_date[mmdd]

        if date_start:
            # Global beginning stock date
            begin_date = _to_mmdd(date_start)
            global_stock = _stock_for_date(begin_date)
        else:
            global_stock = None

        items = []
        for row in rows:
            code = row.get("product_code") or ""
            if global_stock is not None:
                beginning = global_stock.get(str(code), 0)
            else:
                first_date = row.get("first_doc_date")
                if first_date:
                    per_stock = _stock_for_date(_to_mmdd(str(first_date)))
                    beginning = per_stock.get(str(code), 0)
                else:
                    beginning = 0

            inbound_val = row.get("inbound_qty") or Decimal("0")
            return_val = row.get("return_qty") or Decimal("0")
            ending = beginning + inbound_val - return_val

            items.append({
                "product_code": row.get("product_code"),
                "product_name": row.get("product_name"),
                "color_spec": row.get("color_spec"),
                "beginning_qty": _fmt(beginning),
                "inbound_qty": _fmt(inbound_val),
                "return_qty": _fmt(return_val),
                "ending_qty": _fmt(ending),
            })

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    def _read_jst_stock(self, jst_stock_root: Path | None, stock_date: str) -> dict:
        """Read beginning stock from DB, fallback to Excel if no data."""
        if jst_stock_root is None:
            return {}

        # Try DB first
        table = JST_STOCK_TABLE
        stmt = select(table.c.product_code, table.c.available_qty).where(table.c.stock_date == stock_date)
        with self.engine.connect() as connection:
            rows = connection.execute(stmt).all()

        if rows:
            return {str(row.product_code): int(row.available_qty) for row in rows}

        # Fallback to Excel
        return self._read_jst_stock_from_excel(jst_stock_root, stock_date)

    def import_jst_stock(self, jst_stock_root: Path | None, stock_date: str) -> dict[str, object]:
        """Import daily stock from 聚水潭 Excel into jst_daily_stock table."""
        if jst_stock_root is None:
            return {"imported": 0, "message": "JST_STOCK_ROOT 未配置"}

        data = self._read_jst_stock_from_excel(jst_stock_root, stock_date)
        if not data:
            return {"imported": 0, "message": f"未找到 {stock_date} 的库存数据"}

        table = JST_STOCK_TABLE
        stock_date_value = parse_month_day(stock_date)
        rows = [
            {
                "stock_date": stock_date,
                "stock_date_value": stock_date_value,
                "product_code": product_code,
                "available_qty": qty,
            }
            for product_code, qty in data.items()
        ]
        with self.engine.begin() as connection:
            for i in range(0, len(rows), 1000):
                stmt = pg_insert(table).values(rows[i:i + 1000]).on_conflict_do_update(
                    index_elements=["stock_date", "product_code"],
                    set_={
                        "stock_date_value": pg_insert(table).excluded.stock_date_value,
                        "available_qty": pg_insert(table).excluded.available_qty,
                    },
                )
                connection.execute(stmt)

        return {"imported": len(rows), "message": f"已导入 {len(rows)} 条 {stock_date} 库存数据"}

    @staticmethod
    def _read_jst_stock_from_excel(jst_stock_root: Path, stock_date: str) -> dict:
        """Read product available stock from 聚水潭 daily stock Excel."""

        stock_dir = jst_stock_root / stock_date
        if not stock_dir.exists():
            return {}

        # Try common extensions
        candidates = []
        for ext in (".xlsx", ".xls", ".xlsm"):
            p = stock_dir / f"商品库存{ext}"
            if p.exists():
                candidates.append(p)

        stock_file = candidates[0] if candidates else None
        if stock_file is None:
            return {}

        wb = load_workbook(stock_file, data_only=True, read_only=True)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        if ws is None:
            wb.close()
            return {}
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            wb.close()
            return {}

        headers = [str(h).strip() if h else "" for h in header_row]

        # Find column indices for "商品编码" and "可用数"
        code_idx = None
        avail_idx = None
        for i, h in enumerate(headers):
            if "商品编码" in h or h == "商品编码":
                code_idx = i
            elif "可用数" in h or "可用库存" in h:
                avail_idx = i

        if code_idx is None or avail_idx is None:
            wb.close()
            return {}

        result = {}
        for row in iterator:
            code = str(row[code_idx]).strip() if code_idx < len(row) and row[code_idx] is not None else ""
            avail = row[avail_idx] if avail_idx < len(row) else None
            if code and avail is not None:
                try:
                    result[code] = int(float(str(avail)))
                except Exception:
                    result[code] = 0

        wb.close()
        return result

    def create_tables(self) -> None:
        with self.engine.begin() as connection:
            INVENTORY_TABLE.create(connection, checkfirst=True)
            INVENTORY_DETAIL_TABLE.create(connection, checkfirst=True)
            JST_STOCK_TABLE.create(connection, checkfirst=True)
            JST_SIZE_STOCK_SNAPSHOT_TABLE.create(connection, checkfirst=True)
            JST_STOCK_SUMMARY_SNAPSHOT_TABLE.create(connection, checkfirst=True)
            PRODUCT_GOODS_OVERRIDES_TABLE.create(connection, checkfirst=True)
            PRODUCT_SIZE_GROUP_MAPPINGS_TABLE.create(connection, checkfirst=True)
            SIZE_GROUPS_TABLE.create(connection, checkfirst=True)
            SIZE_GROUP_ITEMS_TABLE.create(connection, checkfirst=True)
            for sales_year in HISTORICAL_SALES_YEARS:
                ensure_product_goods_historical_sales_table(connection, sales_year)
            connection.execute(
                text(
                    """
                    DO $$
                    BEGIN
                        IF EXISTS (
                            SELECT 1
                            FROM information_schema.columns
                            WHERE table_schema = current_schema()
                              AND table_name = 'product_goods_overrides'
                              AND column_name = 'douyin_hot'
                              AND data_type = 'boolean'
                        ) THEN
                            ALTER TABLE product_goods_overrides
                            ALTER COLUMN douyin_hot TYPE TEXT
                            USING CASE WHEN douyin_hot THEN '是' ELSE NULL END;
                        END IF;
                        IF EXISTS (
                            SELECT 1
                            FROM information_schema.columns
                            WHERE table_schema = current_schema()
                              AND table_name = 'product_goods_overrides'
                              AND column_name = 'clearance'
                              AND data_type = 'boolean'
                        ) THEN
                            ALTER TABLE product_goods_overrides
                            ALTER COLUMN clearance TYPE TEXT
                            USING CASE WHEN clearance THEN '是' ELSE NULL END;
                        END IF;
                    END $$;
                    """
                )
            )
            connection.execute(text("ALTER TABLE IF EXISTS inventory_records ADD COLUMN IF NOT EXISTS document_number TEXT"))
            self._backfill_document_numbers(connection)
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_inventory_records_document_number ON inventory_records (document_number)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_inventory_records_warehouse_date ON inventory_records (warehouse, date_value)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_inventory_records_supplier_date ON inventory_records (supplier, date_value)"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_records ADD COLUMN IF NOT EXISTS handler TEXT"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_records ADD COLUMN IF NOT EXISTS additional_note TEXT"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_records ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMPTZ"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_inventory_records_deleted_at ON inventory_records (deleted_at)"))
            connection.execute(
                text(
                    """
                    DO $$
                    BEGIN
                        IF EXISTS (
                            SELECT 1
                            FROM inventory_records
                            WHERE deleted_at IS NULL
                              AND NULLIF(BTRIM(document_number), '') IS NOT NULL
                            GROUP BY document_number
                            HAVING COUNT(*) > 1
                        ) THEN
                            RAISE EXCEPTION 'duplicate active inventory document_number exists';
                        END IF;
                    END $$;
                    """
                )
            )
            connection.execute(
                text(
                    """
                    CREATE UNIQUE INDEX IF NOT EXISTS uq_inventory_records_active_document_number
                    ON inventory_records (document_number)
                    WHERE deleted_at IS NULL
                      AND NULLIF(BTRIM(document_number), '') IS NOT NULL
                    """
                )
            )
            connection.execute(text("ALTER TABLE IF EXISTS inventory_details ADD COLUMN IF NOT EXISTS color_barcode TEXT"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_details ADD COLUMN IF NOT EXISTS color_name TEXT"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_details ADD COLUMN IF NOT EXISTS size_quantities JSON"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_details ADD COLUMN IF NOT EXISTS remark TEXT"))
            connection.execute(text("ALTER TABLE IF EXISTS inventory_details ADD COLUMN IF NOT EXISTS extra_fields JSON"))
            connection.execute(text("CREATE EXTENSION IF NOT EXISTS pg_trgm"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_inventory_details_product_code ON inventory_details (product_code)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_jst_stock_product_code ON jst_daily_stock (product_code)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_inventory_details_product_code_trgm ON inventory_details USING GIN (product_code gin_trgm_ops)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS idx_jst_stock_product_code_trgm ON jst_daily_stock USING GIN (product_code gin_trgm_ops)"))
            INVENTORY_ACCOUNT_SUBJECT_TABLE.create(connection, checkfirst=True)
            self._seed_account_subjects(connection)
            PURCHASE_ORDER_REQUIREMENT_TABLE.create(connection, checkfirst=True)
            SUPPLIER_TABLE.create(connection, checkfirst=True)
            SUPPLIER_BRAND_TABLE.create(connection, checkfirst=True)
            self._ensure_supplier_brand_schema(connection)
            WAREHOUSE_TABLE.create(connection, checkfirst=True)
            WAREHOUSE_BRAND_TABLE.create(connection, checkfirst=True)
            self._ensure_warehouse_schema(connection)
            self._ensure_supplier_schema(connection)
            self._sync_suppliers_from_gj(connection)
            GENERAL_CUSTOMER_BRAND_TABLE.create(connection, checkfirst=True)
            GENERAL_CUSTOMER_SHOP_TABLE.create(connection, checkfirst=True)
            GENERAL_CUSTOMER_UNIT_TABLE.create(connection, checkfirst=True)
            GENERAL_CUSTOMER_SORT_PREFERENCE_TABLE.create(connection, checkfirst=True)
            self._ensure_general_customer_schema(connection)
            self._seed_general_customer_shops(connection)
            connection.execute(text("UPDATE inventory_records SET total_count = NULL, warehouse = NULL WHERE document_type IN ('应付款减少', '应付款增加', '应收款减少', '应收款增加') AND (total_count IS NOT NULL OR warehouse IS NOT NULL)"))
            connection.execute(text("""
                UPDATE inventory_records AS record
                SET amount = accounting_amounts.amount
                FROM (
                    SELECT details.document_id, SUM(details.amount) AS amount
                    FROM inventory_details AS details
                    JOIN inventory_records AS source_record ON source_record.id = details.document_id
                    WHERE source_record.document_type IN ('应付款减少', '应付款增加', '应收款减少', '应收款增加')
                      AND details.amount IS NOT NULL
                    GROUP BY details.document_id
                ) AS accounting_amounts
                WHERE record.id = accounting_amounts.document_id
                  AND record.amount IS NULL
                  AND record.deleted_at IS NULL
            """))
            connection.execute(text("UPDATE inventory_records SET total_count = -abs(total_count) WHERE document_type = '进货退货单' AND total_count IS NOT NULL AND total_count > 0"))
            connection.execute(text("UPDATE inventory_records SET amount = -abs(amount) WHERE document_type = '进货退货单' AND amount IS NOT NULL AND amount > 0"))
            connection.execute(text("DELETE FROM inventory_records WHERE deleted_at < now() - interval '10 days'"))

    @staticmethod
    def _seed_account_subjects(connection) -> None:
        defaults = [
            {"code": "0337", "name": "罚款收入"},
            {"code": None, "name": "付货款"},
        ]
        for row in defaults:
            exists = connection.execute(
                select(INVENTORY_ACCOUNT_SUBJECT_TABLE.c.id).where(
                    INVENTORY_ACCOUNT_SUBJECT_TABLE.c.name == row["name"]
                )
            ).first()
            if exists is None:
                connection.execute(insert(INVENTORY_ACCOUNT_SUBJECT_TABLE).values(**row))

    @staticmethod
    def _ensure_supplier_brand_schema(connection) -> None:
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ADD COLUMN IF NOT EXISTS code TEXT"))
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ADD COLUMN IF NOT EXISTS name TEXT"))
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ADD COLUMN IF NOT EXISTS product_archive_enabled BOOLEAN NOT NULL DEFAULT TRUE"))
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ADD COLUMN IF NOT EXISTS product_table_name TEXT"))
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0"))
        default_brands = (
            ("cbanner_mens", "千百度男鞋"),
            ("cbanner_womens", "千百度女鞋"),
            ("yandou", "烟斗"),
            ("eblan", "伊伴"),
            ("smiley", "笑脸"),
            ("ni", "NI"),
        )
        for index, (code, name) in enumerate(default_brands, start=1):
            connection.execute(
                text(
                    """
                    INSERT INTO supplier_brands (code, name, sort_order)
                    VALUES (:code, :name, :sort_order)
                    ON CONFLICT (code) DO NOTHING
                    """
                ),
                {"code": code, "name": name, "sort_order": index},
            )
        manual_brand_ids = connection.execute(text("""
            SELECT id
            FROM supplier_brands
            WHERE product_archive_enabled = TRUE
              AND product_table_name IS NULL
              AND code NOT IN ('cbanner_mens', 'cbanner_womens', 'yandou', 'eblan', 'smiley', 'ni')
        """)).scalars()
        for brand_id in manual_brand_ids:
            connection.execute(
                text("UPDATE supplier_brands SET product_table_name = :table_name WHERE id = :id"),
                {"id": brand_id, "table_name": f"manual_product_archive_{brand_id}"},
            )
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ALTER COLUMN code SET NOT NULL"))
        connection.execute(text("ALTER TABLE IF EXISTS supplier_brands ALTER COLUMN name SET NOT NULL"))
        connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_supplier_brands_code ON supplier_brands (code)"))
        connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_supplier_brands_name ON supplier_brands (name)"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_supplier_brands_sort ON supplier_brands (sort_order)"))

    @staticmethod
    def _ensure_warehouse_schema(connection) -> None:
        connection.execute(text("ALTER TABLE IF EXISTS warehouses ADD COLUMN IF NOT EXISTS brand TEXT"))
        connection.execute(text("ALTER TABLE IF EXISTS warehouses ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0"))
        connection.execute(text("ALTER TABLE IF EXISTS warehouse_brands ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_warehouses_brand_sort ON warehouses (brand, sort_order)"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_warehouse_brands_sort ON warehouse_brands (sort_order)"))
        connection.execute(
            text(
                """
                INSERT INTO warehouse_brands (name)
                VALUES ('通用')
                ON CONFLICT (name) DO NOTHING
                """
            )
        )
        connection.execute(
            text(
                """
                UPDATE warehouses
                SET brand = '通用'
                WHERE brand IS NULL OR btrim(brand) = ''
                """
            )
        )
        connection.execute(
            text(
                """
                UPDATE warehouse_brands AS target
                SET sort_order = ranked.sort_order
                FROM (
                    SELECT id, row_number() OVER (ORDER BY id) AS sort_order
                    FROM warehouse_brands
                    WHERE sort_order = 0
                ) AS ranked
                WHERE target.id = ranked.id
                """
            )
        )
        connection.execute(
            text(
                """
                UPDATE warehouses AS target
                SET sort_order = ranked.sort_order
                FROM (
                    SELECT id, row_number() OVER (PARTITION BY brand ORDER BY id) AS sort_order
                    FROM warehouses
                    WHERE sort_order = 0
                ) AS ranked
                WHERE target.id = ranked.id
                """
            )
        )

    @staticmethod
    def _ensure_supplier_schema(connection) -> None:
        connection.execute(text("ALTER TABLE IF EXISTS suppliers ADD COLUMN IF NOT EXISTS brand TEXT"))
        connection.execute(text("ALTER TABLE IF EXISTS suppliers ADD COLUMN IF NOT EXISTS wechat TEXT"))
        connection.execute(text("ALTER TABLE IF EXISTS suppliers ADD COLUMN IF NOT EXISTS cooperation_status TEXT"))
        connection.execute(
            text(
                """
                DELETE FROM suppliers AS bad
                USING suppliers AS good
                WHERE bad.id <> good.id
                  AND bad.name = good.name
                  AND good.brand = CASE
                      WHEN upper(coalesce(bad.name, '')) LIKE '%TRUMPPIPE%'
                        OR coalesce(bad.name, '') LIKE '%烟斗%' THEN 'yandou'
                      WHEN upper(coalesce(bad.name, '')) LIKE '%EBLAN%'
                        OR coalesce(bad.name, '') LIKE '%伊伴%' THEN 'eblan'
                      WHEN upper(coalesce(bad.name, '')) LIKE '%SMILEY%'
                        OR coalesce(bad.name, '') LIKE '%笑脸%'
                        OR coalesce(bad.name, '') LIKE '%小莲%' THEN 'smiley'
                      WHEN upper(btrim(coalesce(bad.name, ''))) ~ '(^|[（([:space:]])NI($|[）)[:space:]])' THEN 'ni'
                      WHEN coalesce(bad.name, '') LIKE '%千百度女鞋%' THEN 'cbanner_womens'
                      ELSE bad.brand
                  END
                  AND bad.brand IS DISTINCT FROM good.brand
                  AND (
                      upper(coalesce(bad.name, '')) LIKE '%TRUMPPIPE%'
                      OR coalesce(bad.name, '') LIKE '%烟斗%'
                      OR upper(coalesce(bad.name, '')) LIKE '%EBLAN%'
                      OR coalesce(bad.name, '') LIKE '%伊伴%'
                      OR upper(coalesce(bad.name, '')) LIKE '%SMILEY%'
                      OR coalesce(bad.name, '') LIKE '%笑脸%'
                      OR coalesce(bad.name, '') LIKE '%小莲%'
                      OR upper(btrim(coalesce(bad.name, ''))) ~ '(^|[（([:space:]])NI($|[）)[:space:]])'
                      OR coalesce(bad.name, '') LIKE '%千百度女鞋%'
                  )
                """
            )
        )
        connection.execute(
            text(
                """
                UPDATE suppliers
                SET brand = CASE
                    WHEN upper(coalesce(name, '')) LIKE '%TRUMPPIPE%'
                      OR coalesce(name, '') LIKE '%烟斗%' THEN 'yandou'
                    WHEN upper(coalesce(name, '')) LIKE '%EBLAN%'
                      OR coalesce(name, '') LIKE '%伊伴%' THEN 'eblan'
                    WHEN upper(coalesce(name, '')) LIKE '%SMILEY%'
                      OR coalesce(name, '') LIKE '%笑脸%'
                      OR coalesce(name, '') LIKE '%小莲%' THEN 'smiley'
                    WHEN upper(btrim(coalesce(name, ''))) ~ '(^|[（([:space:]])NI($|[）)[:space:]])' THEN 'ni'
                    WHEN coalesce(name, '') LIKE '%千百度品牌方%' THEN :default_brand
                    WHEN coalesce(name, '') LIKE '%千百度女鞋%' THEN 'cbanner_womens'
                    WHEN coalesce(name, '') LIKE '%千百度%' THEN 'cbanner_mens'
                    ELSE :default_brand
                END
                WHERE brand IS NULL
                   OR brand = ''
                   OR (
                        upper(coalesce(name, '')) LIKE '%TRUMPPIPE%'
                        OR coalesce(name, '') LIKE '%烟斗%'
                        OR upper(coalesce(name, '')) LIKE '%EBLAN%'
                        OR coalesce(name, '') LIKE '%伊伴%'
                        OR upper(coalesce(name, '')) LIKE '%SMILEY%'
                        OR coalesce(name, '') LIKE '%笑脸%'
                        OR coalesce(name, '') LIKE '%小莲%'
                        OR upper(btrim(coalesce(name, ''))) ~ '(^|[（([:space:]])NI($|[）)[:space:]])'
                        OR coalesce(name, '') LIKE '%千百度女鞋%'
                   )
                """
            ),
            {"default_brand": CBANNER_MENS_BRAND},
        )
        connection.execute(text("ALTER TABLE IF EXISTS suppliers ALTER COLUMN brand SET NOT NULL"))
        connection.execute(text("DROP INDEX IF EXISTS idx_suppliers_brand"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_suppliers_brand ON suppliers (brand)"))
        connection.execute(
            text(
                """
                do $$
                begin
                    if exists (
                        select 1
                        from pg_constraint
                        where conname = 'uq_supplier_name'
                    ) then
                        alter table suppliers drop constraint uq_supplier_name;
                    end if;
                end $$;
                """
            )
        )
        connection.execute(
            text(
                """
                do $$
                begin
                    if not exists (
                        select 1
                        from pg_constraint
                        where conname = 'uq_supplier_brand_name'
                    ) then
                        alter table suppliers add constraint uq_supplier_brand_name unique (brand, name);
                    end if;
                end $$;
                """
            )
        )

    @staticmethod
    def _sync_suppliers_from_gj(connection) -> int:
        if not inspect(connection).has_table(GJ_MERGED_PRODUCT_INFO_TABLE.name):
            return 0
        synced = 0
        rows = connection.execute(
            select(
                GJ_MERGED_PRODUCT_INFO_TABLE.c.fine_table_brand,
                GJ_MERGED_PRODUCT_INFO_TABLE.c.primary_supplier,
            )
            .where(GJ_MERGED_PRODUCT_INFO_TABLE.c.fine_table_brand.in_(GJ_FINE_TABLE_BRANDS))
            .where(GJ_MERGED_PRODUCT_INFO_TABLE.c.primary_supplier.isnot(None))
            .where(GJ_MERGED_PRODUCT_INFO_TABLE.c.primary_supplier != "")
            .distinct()
        ).mappings()
        for row in rows:
            name = str(row["primary_supplier"] or "").strip()
            brand = infer_supplier_brand_from_name(name) or str(row["fine_table_brand"] or "").strip()
            if not brand or not name:
                continue
            exists = connection.execute(
                select(SUPPLIER_TABLE.c.id).where(
                    SUPPLIER_TABLE.c.brand == brand,
                    SUPPLIER_TABLE.c.name == name,
                )
            ).first()
            if exists is None:
                connection.execute(insert(SUPPLIER_TABLE).values(brand=brand, name=name))
                synced += 1
        return synced

    def _seed_general_customer_shops(self, connection) -> None:
        defaults = [
            {"customer_name": "烟斗", "shop_name": "烟斗唯品会店铺"},
        ]
        for row in defaults:
            self._ensure_general_customer_brand(connection, row["customer_name"])
            exists = connection.execute(
                select(GENERAL_CUSTOMER_SHOP_TABLE.c.id).where(
                    GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == row["customer_name"],
                    GENERAL_CUSTOMER_SHOP_TABLE.c.shop_name == row["shop_name"],
                )
            ).first()
            if exists is None:
                connection.execute(
                    insert(GENERAL_CUSTOMER_SHOP_TABLE).values(
                        **row,
                        sort_order=self._next_sort_order(
                            connection,
                            GENERAL_CUSTOMER_SHOP_TABLE,
                            GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name == row["customer_name"],
                        ),
                    )
                )

        existing_brands = connection.execute(select(GENERAL_CUSTOMER_SHOP_TABLE.c.customer_name).distinct()).all()
        for row in existing_brands:
            self._ensure_general_customer_brand(connection, row[0])

    @staticmethod
    def _ensure_general_customer_schema(connection) -> None:
        connection.execute(text("ALTER TABLE IF EXISTS general_customer_brands ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0"))
        connection.execute(text("ALTER TABLE IF EXISTS general_customer_shops ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0"))
        connection.execute(text("ALTER TABLE IF EXISTS general_customer_units ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_general_customer_brands_sort ON general_customer_brands (sort_order)"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_general_customer_shops_customer_sort ON general_customer_shops (customer_name, sort_order)"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS idx_general_customer_units_shop_sort ON general_customer_units (shop_id, sort_order)"))
        connection.execute(
            text(
                """
                UPDATE general_customer_brands AS target
                SET sort_order = ranked.sort_order
                FROM (
                    SELECT id, row_number() OVER (ORDER BY id) AS sort_order
                    FROM general_customer_brands
                    WHERE sort_order = 0
                ) AS ranked
                WHERE target.id = ranked.id
                """
            )
        )
        connection.execute(
            text(
                """
                UPDATE general_customer_shops AS target
                SET sort_order = ranked.sort_order
                FROM (
                    SELECT id, row_number() OVER (PARTITION BY customer_name ORDER BY id) AS sort_order
                    FROM general_customer_shops
                    WHERE sort_order = 0
                ) AS ranked
                WHERE target.id = ranked.id
                """
            )
        )
        connection.execute(
            text(
                """
                UPDATE general_customer_units AS target
                SET sort_order = ranked.sort_order
                FROM (
                    SELECT id, row_number() OVER (PARTITION BY shop_id ORDER BY id) AS sort_order
                    FROM general_customer_units
                    WHERE sort_order = 0
                ) AS ranked
                WHERE target.id = ranked.id
                """
            )
        )

    def _ensure_general_customer_brand(self, connection, name: str) -> None:
        if not name:
            return
        exists = connection.execute(
            select(GENERAL_CUSTOMER_BRAND_TABLE.c.id).where(GENERAL_CUSTOMER_BRAND_TABLE.c.name == name)
        ).first()
        if exists is None:
            connection.execute(
                insert(GENERAL_CUSTOMER_BRAND_TABLE).values(
                    name=name,
                    sort_order=self._next_sort_order(connection, GENERAL_CUSTOMER_BRAND_TABLE),
                )
            )

    # ── Helpers ────────────────────────────────────────────────────

    @staticmethod
    def _coerce_empty(data: Mapping[str, object]) -> dict[str, object]:
        return {k: (None if v == "" else v) for k, v in data.items()}

    @staticmethod
    def _filter_table_payload(table, data: Mapping[str, object]) -> dict[str, object]:
        allowed = set(table.c.keys())
        return {
            key: value
            for key, value in data.items()
            if key in allowed
        }

    @staticmethod
    def _format_decimal(value: Decimal) -> str:
        normalized = value.normalize()
        return str(normalized) if normalized.as_tuple().exponent < 0 else str(int(normalized))

    @staticmethod
    def _clear_accounting_record_summary(record: dict[str, object]) -> dict[str, object]:
        if record.get("document_type") in ACCOUNTING_DOCUMENT_TYPES:
            record["total_count"] = None
            record["warehouse"] = None
        return record

    @staticmethod
    def _apply_document_total_sign(document_type: object, value: object) -> object:
        if value is None or str(document_type or "").strip() not in NEGATIVE_TOTAL_DOCUMENT_TYPES:
            return value
        decimal_value = Decimal(str(value or "0"))
        return -abs(decimal_value)

    @staticmethod
    def _prepare_record(record: Mapping[str, object]) -> dict[str, object]:
        payload = {}
        for key, value in record.items():
            if value == "":
                payload[key] = None
            else:
                payload[key] = value
        if "date" in payload and "date_value" not in payload:
            payload["date_value"] = parse_date(payload.get("date"))
        if payload.get("document_type") in ACCOUNTING_DOCUMENT_TYPES:
            payload["total_count"] = None
            payload["amount"] = None
            payload["warehouse"] = None
        else:
            for field in ("total_count", "amount"):
                if field in payload:
                    payload[field] = InventoryRepository._apply_document_total_sign(payload.get("document_type"), payload[field])
        raw_payload = payload.get("raw_payload")
        if isinstance(raw_payload, Mapping):
            payload["raw_payload"] = {
                k: str(v) if isinstance(v, Decimal) else v
                for k, v in raw_payload.items()
            }
        return payload

    @staticmethod
    def _document_number_prefix(document_type: object) -> str:
        return DOCUMENT_NUMBER_PREFIXES.get(str(document_type or "").strip(), DEFAULT_DOCUMENT_NUMBER_PREFIX)

    @staticmethod
    def _document_number_date_text(date_value: object) -> str:
        parsed_date = date_value if isinstance(date_value, date) else parse_date(date_value)
        return (parsed_date or date.today()).strftime("%Y-%m-%d")

    @staticmethod
    def _format_document_number(document_type: object, date_value: object, sequence: int) -> str:
        prefix = InventoryRepository._document_number_prefix(document_type)
        date_text = InventoryRepository._document_number_date_text(date_value)
        return f"{prefix}-{date_text}-{sequence:04d}"

    @staticmethod
    def _generate_document_number(connection, date_value: object, document_type: object) -> str:
        prefix = InventoryRepository._document_number_prefix(document_type)
        date_text = InventoryRepository._document_number_date_text(date_value)
        pattern = f"{prefix}-{date_text}-%"
        rows = connection.execute(
            select(INVENTORY_TABLE.c.document_number)
            .where(INVENTORY_TABLE.c.document_number.like(pattern))
        ).all()
        max_sequence = 0
        for row in rows:
            suffix = str(row[0] or "").rsplit("-", 1)[-1]
            if suffix.isdigit():
                max_sequence = max(max_sequence, int(suffix))
        return f"{prefix}-{date_text}-{max_sequence + 1:04d}"

    @staticmethod
    def _backfill_document_numbers(connection) -> None:
        rows = list(connection.execute(
            select(
                INVENTORY_TABLE.c.id,
                INVENTORY_TABLE.c.date,
                INVENTORY_TABLE.c.date_value,
                INVENTORY_TABLE.c.document_type,
                INVENTORY_TABLE.c.document_number,
            )
            .order_by(
                INVENTORY_TABLE.c.date_value.nulls_last(),
                INVENTORY_TABLE.c.date,
                INVENTORY_TABLE.c.document_type,
                INVENTORY_TABLE.c.id,
            )
        ).mappings())
        occupied_numbers = {
            str(row.get("document_number") or "").strip()
            for row in rows
            if str(row.get("document_number") or "").strip()
        }
        counters: dict[tuple[str, str], int] = {}
        for row in rows:
            existing_number = str(row.get("document_number") or "").strip()
            if existing_number:
                continue
            date_value = row.get("date_value") or row.get("date")
            document_type = row.get("document_type")
            prefix = InventoryRepository._document_number_prefix(document_type)
            date_text = InventoryRepository._document_number_date_text(date_value)
            key = (prefix, date_text)
            sequence = counters.get(key, 0)
            while True:
                sequence += 1
                document_number = InventoryRepository._format_document_number(document_type, date_value, sequence)
                if document_number not in occupied_numbers:
                    break
            counters[key] = sequence
            connection.execute(
                update(INVENTORY_TABLE)
                .where(INVENTORY_TABLE.c.id == row["id"])
                .values(document_number=document_number)
            )
            occupied_numbers.add(document_number)

    @staticmethod
    def _prepare_supplier(data: Mapping[str, object]) -> dict[str, object]:
        name = str(data.get("name") or "").strip()
        brand = str(data.get("brand") or "").strip() or infer_supplier_brand_from_name(name) or CBANNER_MENS_BRAND
        payload = {
            "brand": brand,
            "name": name,
            "factory_code": str(data.get("factory_code") or "").strip() or None,
            "contact": str(data.get("contact") or "").strip() or None,
            "wechat": str(data.get("wechat") or "").strip() or None,
            "cooperation_status": str(data.get("cooperation_status") or "").strip() or None,
            "address": str(data.get("address") or "").strip() or None,
            "notes": str(data.get("notes") or "").strip() or None,
        }
        return payload
