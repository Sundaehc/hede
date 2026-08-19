from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import orjson
from openpyxl import load_workbook
from sqlalchemy import create_engine, delete, insert

from domain.dewu_order_schema import DEWU_ORDERS_TABLE


@dataclass(frozen=True)
class DewuOrderSource:
    brand_group: str
    brand_label: str
    filename: str


DEWU_ORDER_SOURCES = (
    DewuOrderSource("cbanner", "千百度", "千百度得物订单.xlsx"),
    DewuOrderSource("eblan", "伊伴", "伊伴得物订单.xlsx"),
    DewuOrderSource("yandou", "烟斗", "烟斗得物订单.xlsx"),
    DewuOrderSource("smiley", "笑脸", "笑脸得物订单.xlsx"),
)

DEWU_ORDER_HEADER_MAP = {
    "订单号": "order_number",
    "订单类型": "order_type",
    "spuID": "spu_id",
    "skuID": "sku_id",
    "商品名称": "product_name",
    "货号": "goods_code",
    "SKU货号": "sku_goods_code",
    "品牌": "brand_name",
    "规格": "specification",
    "数量": "quantity",
    "出价金额（元）": "bid_amount",
    "卖家承担优惠金额（元）": "seller_discount_amount",
    "消费者邮费补贴金额（元）": "consumer_postage_subsidy_amount",
    "预计收入金额（元）": "estimated_income_amount",
    "剩余发货时效": "remaining_shipping_time",
    "关闭原因": "close_reason",
    "配送方": "delivery_party",
    "物流公司（卖家>平台or买家）": "logistics_company",
    "物流单号（卖家>平台or买家）": "logistics_tracking_number",
    "预约单号": "appointment_number",
    "订单状态": "order_status",
    "订单来源": "order_source",
    "得物收货地址": "dewu_receiving_address",
    "收件人姓名": "recipient_name",
    "收件人手机号": "recipient_phone",
    "收件人省": "recipient_province",
    "收件人市": "recipient_city",
    "收件人区": "recipient_district",
    "收件人街道": "recipient_street",
    "收件人详细地址": "recipient_detail_address",
    "服务保障": "service_guarantee",
    "直发预约上门取件时间": "direct_shipping_pickup_time",
    "直发发货仓": "direct_shipping_warehouse",
    "用户标识": "user_identifier",
    "订单备注": "order_remark",
    "订单标记": "order_tag",
    "买家下单时间": "buyer_order_time",
    "买家支付时间": "buyer_payment_time",
    "承诺送达时间": "promised_delivery_time",
    "卖家发货时间": "seller_shipping_time",
    "平台收货时间（现货）": "platform_receiving_time",
    "平台发货时间（现货）": "platform_shipping_time",
    "订单关闭时间": "order_closed_time",
    "意向金（元）": "intent_deposit_amount",
    "全款金额（元）": "full_payment_amount",
    "子运单号": "sub_waybill_number",
    "第三方订单号": "third_party_order_number",
    "现场取票地址": "onsite_ticket_address",
    "是否有刻字备注": "has_engraving_remark",
    "买家备注": "buyer_remark",
    "定制商品明细": "custom_product_detail",
    "预订单号": "preorder_number",
    "预约门店": "appointment_store",
    "核销门店": "verification_store",
    "核销时间": "verification_time",
    "订单标识": "order_identifier",
    "预约上门取件承运商": "pickup_carrier",
    "预约上门取件运单号": "pickup_waybill_number",
    "预约上门取件取件码": "pickup_code",
    "SN码": "sn_code",
    "IMEI1": "imei1",
    "IMEI2": "imei2",
    "开放预约时间": "reservation_open_time",
}

INTEGER_FIELDS = {"quantity"}
DECIMAL_FIELDS = {
    "bid_amount",
    "seller_discount_amount",
    "consumer_postage_subsidy_amount",
    "estimated_income_amount",
    "intent_deposit_amount",
    "full_payment_amount",
}
INSERT_BATCH_SIZE = 1000


def _json_serializer(value: object) -> str:
    return orjson.dumps(value).decode("utf-8")


def _text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, "f").rstrip("0").rstrip(".")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    result = str(value).strip()
    return result or None


def _integer(value: object) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        return int(Decimal(text.replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def _decimal(value: object) -> Decimal | None:
    text = _text(value)
    if not text:
        return None
    try:
        return Decimal(text.replace(",", "")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _order_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("/", "-")).date()
    except ValueError:
        return None


def parse_dewu_order_workbook(source_file: Path, source: DewuOrderSource) -> list[dict[str, object]]:
    if not source_file.exists():
        raise FileNotFoundError(f"得物订单文件不存在: {source_file}")

    workbook = load_workbook(source_file, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        # The robot export writes <dimension ref="A1"> even when thousands of
        # rows exist. Read-only openpyxl trusts that value unless it is reset.
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if header_values is None:
            raise ValueError(f"得物订单文件没有表头: {source_file}")

        headers = [_text(value) or "" for value in header_values]
        missing = [header for header in DEWU_ORDER_HEADER_MAP if header not in headers]
        if missing:
            raise ValueError(f"得物订单文件缺少字段 {missing}: {source_file}")

        header_indexes = {header: index for index, header in enumerate(headers) if header}
        unknown_headers = [header for header in headers if header and header not in DEWU_ORDER_HEADER_MAP]
        source_modified_at = datetime.fromtimestamp(source_file.stat().st_mtime).astimezone()
        parsed: list[dict[str, object]] = []

        for row_number, values in enumerate(rows, start=2):
            def raw(header: str) -> object:
                index = header_indexes[header]
                return values[index] if index < len(values) else None

            order_number = _text(raw("订单号"))
            if not order_number:
                continue

            item: dict[str, object] = {
                "brand_group": source.brand_group,
                "brand_label": source.brand_label,
                "source_workbook": source_file.name,
                "source_sheet": worksheet.title,
                "source_row_number": row_number,
                "source_modified_at": source_modified_at,
                "order_number": order_number,
            }
            for header, field in DEWU_ORDER_HEADER_MAP.items():
                if field == "order_number":
                    continue
                value = raw(header)
                if field in INTEGER_FIELDS:
                    item[field] = _integer(value)
                elif field in DECIMAL_FIELDS:
                    item[field] = _decimal(value)
                else:
                    item[field] = _text(value)
            item["order_date"] = _order_date(item.get("buyer_order_time"))
            extra_fields = {
                header: _text(values[header_indexes[header]])
                for header in unknown_headers
                if header_indexes[header] < len(values) and _text(values[header_indexes[header]]) is not None
            }
            item["extra_fields"] = extra_fields or None
            parsed.append(item)
        return parsed
    finally:
        workbook.close()


class DewuOrderRepository:
    def __init__(self, database_url: str):
        self.engine = create_engine(database_url, future=True, json_serializer=_json_serializer)

    def ensure_table(self) -> None:
        DEWU_ORDERS_TABLE.create(self.engine, checkfirst=True)

    def import_all(self, source_root: Path) -> dict[str, object]:
        rows: list[dict[str, object]] = []
        counts: dict[str, int] = {}
        source_files: dict[str, str] = {}
        for source in DEWU_ORDER_SOURCES:
            source_file = source_root / source.filename
            source_rows = parse_dewu_order_workbook(source_file, source)
            rows.extend(source_rows)
            counts[source.brand_group] = len(source_rows)
            source_files[source.brand_group] = str(source_file)

        if not rows:
            raise ValueError("四份得物订单文件均无有效订单，已取消全量覆盖")

        self.ensure_table()
        with self.engine.begin() as connection:
            connection.execute(delete(DEWU_ORDERS_TABLE))
            for offset in range(0, len(rows), INSERT_BATCH_SIZE):
                connection.execute(insert(DEWU_ORDERS_TABLE), rows[offset:offset + INSERT_BATCH_SIZE])

        unique_orders = len({str(row["order_number"]) for row in rows})
        statuses = Counter(str(row.get("order_status") or "") for row in rows)
        return {
            "imported": len(rows),
            "unique_orders": unique_orders,
            "counts": counts,
            "source_files": source_files,
            "status_counts": dict(statuses),
            "message": f"得物订单全量覆盖完成，共导入 {len(rows)} 行、{unique_orders} 个订单",
        }
