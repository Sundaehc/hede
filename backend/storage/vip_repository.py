from __future__ import annotations

import re
from zipfile import BadZipFile
from datetime import date, datetime
from pathlib import Path

import orjson
import xlrd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import create_engine, func as sa_func, text
from sqlalchemy.dialects.postgresql import insert as pg_insert

from domain.vip_schema import VIP_DAILY_TABLE, VIP_DAILY_SNAPSHOT_TABLE, VIP_OPS_TABLE, VIP_OPS_SNAPSHOT_TABLE, JST_PRICE_TABLE, VIP_REALTIME_TABLE, JST_MONTHLY_ORDERS_TABLE, JST_SIZE_STOCK_TABLE, JST_STOCK_SUMMARY_TABLE, JST_PURCHASE_DIFF_TABLE, JST_PRODUCT_PROFILE_TABLE, JST_AFTERSALE_RETURN_TABLE
from domain.jst_stock_snapshot_schema import JST_SIZE_STOCK_SNAPSHOT_TABLE, JST_STOCK_SUMMARY_SNAPSHOT_TABLE
from domain.vip_sources import (
    VIP_DAILY_COLUMN_ALIASES,
    VIP_OPS_COLUMN_ALIASES,
    JST_PRICE_COLUMN_ALIASES,
    JST_PRODUCT_PROFILE_COLUMN_ALIASES,
    JST_AFTERSALE_RETURN_COLUMN_ALIASES,
    VIP_REALTIME_COLUMN_ALIASES,
    JST_MONTHLY_ORDERS_COLUMN_ALIASES,
)
from storage.date_normalization import parse_date, parse_date_range, parse_datetime


def _period_from_filename(filename: str) -> str | None:
    if "日罗" in filename:
        return "1d"
    if "月罗" in filename:
        return "30d"
    m = re.match(r"(\d+)", filename)
    if m:
        return f"{m.group(1)}d"
    return None


def _report_type_from_filename(filename: str) -> str | None:
    if "环比" in filename:
        return "环比"
    if "罗盘" in filename:
        return "罗盘"
    return None


def _snapshot_date_from_path(file_path: Path) -> date:
    parsed = parse_date(file_path.parent.name)
    if parsed is not None:
        return parsed
    return date.fromtimestamp(file_path.stat().st_mtime)


def _json_serializer(value: object) -> str:
    return orjson.dumps(value).decode("utf-8")


class VipRepository:
    def __init__(self, database_url: str):
        self.engine = create_engine(
            database_url,
            future=True,
            json_serializer=_json_serializer,
        )

    # ── vip_product_daily (环比/罗盘) ──────────────────────────────

    def import_daily(self, file_path: Path, *, replace_existing: bool = True) -> dict[str, object]:
        filename = file_path.stem
        report_type = _report_type_from_filename(filename)
        period = _period_from_filename(filename)

        if report_type is None or period is None:
            return {"imported": 0, "message": f"无法识别报表类型/周期: {filename}"}

        rows = self._read_excel(file_path, VIP_DAILY_COLUMN_ALIASES)
        if not rows:
            return {"imported": 0, "message": "无数据行"}

        date_range = ""
        for row in rows:
            row["report_type"] = report_type
            row["period"] = period
            if not date_range and row.get("date"):
                date_range = str(row["date"])
            row["date_range"] = date_range
            start_date, end_date = parse_date_range(row.get("date"))
            row["report_start_date"] = start_date
            row["report_end_date"] = end_date

        if replace_existing:
            from sqlalchemy import delete as sa_delete

            with self.engine.begin() as conn:
                conn.execute(
                    sa_delete(VIP_DAILY_TABLE).where(
                        VIP_DAILY_TABLE.c.report_type == report_type,
                        VIP_DAILY_TABLE.c.period == period,
                    )
                )
                self._batch_insert(VIP_DAILY_TABLE, rows, conn=conn)
        else:
            update_cols = [
                c for c in rows[0]
                if c not in ("id", "report_type", "period", "goods_id", "date")
            ]
            self._upsert(
                VIP_DAILY_TABLE,
                rows,
                ["report_type", "period", "goods_id", "date"],
                update_cols,
            )

        if report_type == "罗盘" and period == "1d":
            fallback_snapshot_date = _snapshot_date_from_path(file_path)
            snapshot_by_key: dict[tuple[date, str], dict] = {}
            for row in rows:
                goods_id = str(row.get("goods_id") or "").strip()
                if not goods_id:
                    continue
                snapshot_date = row.get("report_end_date") or fallback_snapshot_date
                if not isinstance(snapshot_date, date):
                    continue
                snapshot_by_key[(snapshot_date, goods_id)] = {
                    **row,
                    "snapshot_date": snapshot_date,
                }
            snapshot_rows = list(snapshot_by_key.values())
            if snapshot_rows:
                snapshot_update_cols = [
                    c for c in snapshot_rows[0]
                    if c not in ("id", "snapshot_date", "goods_id")
                ]
                self._upsert(
                    VIP_DAILY_SNAPSHOT_TABLE,
                    snapshot_rows,
                    ["snapshot_date", "goods_id"],
                    snapshot_update_cols,
                )

        return {
            "imported": len(rows),
            "message": f"{filename}: {len(rows)} 条 ({report_type} / {period})",
        }

    # ── vip_product_realtime (实时商品) ─────────────────────────────

    def import_realtime(self, file_path: Path) -> dict[str, object]:
        rows = self._read_excel(file_path, VIP_REALTIME_COLUMN_ALIASES)
        if not rows:
            return {"imported": 0, "message": "无数据行"}

        seen: dict[str, int] = {}
        deduped: list[dict] = []
        for row in rows:
            key = row.get("goods_id")
            if key is None:
                deduped.append(row)
            elif key in seen:
                deduped[seen[key]] = row
            else:
                seen[key] = len(deduped)
                deduped.append(row)

        update_cols = [c for c in deduped[0] if c not in ("id", "goods_id")]

        self._upsert(VIP_REALTIME_TABLE, deduped, ["goods_id"], update_cols)

        return {
            "imported": len(deduped),
            "message": f"{file_path.name}: {len(deduped)} 条 (原始 {len(rows)} 行)",
        }

    # ── vip_product_ops (常态商品运营) ──────────────────────────────

    def import_ops(
        self,
        file_path: Path,
        snapshot_date: date | None = None,
        *,
        replace_existing: bool = True,
    ) -> dict[str, object]:
        rows = self._read_excel(file_path, VIP_OPS_COLUMN_ALIASES, duplicate_policy="first")
        if not rows:
            return {"imported": 0, "message": "无数据行"}
        snapshot_date = snapshot_date or _snapshot_date_from_path(file_path)

        seen: dict[str, int] = {}
        deduped: list[dict] = []
        for row in rows:
            key = row.get("goods_id")
            if key is None:
                deduped.append(row)
            elif key in seen:
                deduped[seen[key]] = row
            else:
                seen[key] = len(deduped)
                deduped.append(row)

        if replace_existing:
            from sqlalchemy import delete as sa_delete

            with self.engine.begin() as conn:
                conn.execute(sa_delete(VIP_OPS_TABLE))
                self._batch_insert(VIP_OPS_TABLE, deduped, conn=conn)
        else:
            update_cols = [
                c for c in deduped[0]
                if c not in ("id", "goods_id")
            ]
            self._upsert(VIP_OPS_TABLE, deduped, ["goods_id"], update_cols)

        snapshot_rows = [{**row, "snapshot_date": snapshot_date} for row in deduped if row.get("goods_id")]
        if snapshot_rows:
            snapshot_update_cols = [
                c for c in snapshot_rows[0]
                if c not in ("id", "snapshot_date", "goods_id")
            ]
            self._upsert(VIP_OPS_SNAPSHOT_TABLE, snapshot_rows, ["snapshot_date", "goods_id"], snapshot_update_cols)

        return {
            "imported": len(deduped),
            "message": f"{file_path.name}: {len(deduped)} 条 (原始 {len(rows)} 行, 快照 {snapshot_date.isoformat()})",
        }

    # ── vip_product_price (物价信息) ──────────────────────────────

    def import_price(self, file_path: Path) -> dict[str, object]:
        """Import 物价信息 Excel. Header is on row 4 (1-indexed)."""
        self._ensure_price_history_schema()
        batch_date = file_path.parent.name
        source_date_value = parse_date(batch_date)
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        assert ws is not None
        sheet_title = ws.title
        rows_iter = ws.iter_rows(values_only=True)

        # Skip first 3 rows (title, blank, blank), header on row 4
        for _ in range(3):
            next(rows_iter, None)

        header_row = next(rows_iter, None)
        if header_row is None:
            wb.close()
            return {"imported": 0, "message": "未找到表头（第4行）"}

        headers = [str(h).strip() if h else "" for h in header_row]
        column_map = {h: JST_PRICE_COLUMN_ALIASES[h] for h in headers if h in JST_PRICE_COLUMN_ALIASES}

        rows_to_upsert: list[dict] = []
        for row_num, row in enumerate(rows_iter, start=4):  # data starts at row 5
            record: dict[str, object] = {
                "source_date": batch_date,
                "source_date_value": source_date_value,
                "source_workbook": file_path.stem,
                "source_sheet": sheet_title,
                "source_row_number": str(row_num),
            }
            raw: dict[str, object] = {}
            for idx, h in enumerate(headers):
                value = row[idx] if idx < len(row) else None
                raw[h] = value
                col = column_map.get(h)
                if col:
                    record[col] = str(value).strip() if value is not None else None
            record["raw_payload"] = raw
            rows_to_upsert.append(record)

        wb.close()

        if not rows_to_upsert:
            return {"imported": 0, "message": "无数据行"}

        # Dedup by goods_full_name: keep last occurrence (later row wins)
        seen: dict[str, int] = {}
        for idx, row in enumerate(rows_to_upsert):
            name = str(row.get("goods_full_name", ""))
            if name:
                seen[name] = idx

        deduped: list[dict] = []
        kept: set[int] = set()
        for idx, row in enumerate(rows_to_upsert):
            if not str(row.get("goods_full_name", "")):
                deduped.append(row)
                kept.add(idx)
        for idx in sorted(seen.values()):
            if idx not in kept:
                deduped.append(rows_to_upsert[idx])

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            conn.execute(
                sa_delete(JST_PRICE_TABLE)
                .where(JST_PRICE_TABLE.c.source_date == batch_date)
            )
            self._batch_insert(JST_PRICE_TABLE, deduped, conn=conn)

        return {
            "imported": len(deduped),
            "source_date": batch_date,
            "message": f"{file_path.name}: {len(deduped)} 条 (原始 {len(rows_to_upsert)} 行)",
        }

    # ── import_all: 按日期目录全量导入 ─────────────────────────────

    def import_all(self, dir_path: Path, *, replace_existing: bool = True) -> dict[str, object]:
        if not dir_path.exists():
            return {"success": False, "message": f"目录不存在: {dir_path}"}

        batch_date = dir_path.name
        results: list[dict] = []
        daily_files: list[Path] = []

        for ext in (".xlsx", ".xlsm", ".xls"):
            for file_path in sorted(dir_path.glob(f"*{ext}")):
                if file_path.name.startswith("~$"):
                    continue
                filename = file_path.stem

                if "实时商品" in filename:
                    results.append(self.import_realtime(file_path))
                elif "常态商品运营" in filename:
                    results.append(
                        self.import_ops(
                            file_path,
                            parse_date(batch_date),
                            replace_existing=replace_existing,
                        )
                    )
                elif "合并" in filename and "物价" in filename:
                    results.append(self.import_price(file_path))
                elif _report_type_from_filename(filename) and _period_from_filename(filename):
                    daily_files.append(file_path)

        for file_path in daily_files:
            results.append(self.import_daily(file_path, replace_existing=replace_existing))

        total = sum(r.get("imported", 0) for r in results)
        return {
            "success": True,
            "batch_date": batch_date,
            "total_imported": total,
            "details": results,
        }

    # ── jst_monthly_orders (月聚水潭订单) ─────────────────────────

    def import_monthly_order(self, file_path: Path) -> dict[str, object]:
        wb = load_workbook(str(file_path), data_only=True)
        ws = wb.active
        assert ws is not None
        sheet_title = ws.title
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            wb.close()
            return {"imported": 0, "message": "无表头"}

        aliases = JST_MONTHLY_ORDERS_COLUMN_ALIASES
        headers = [str(h).strip() if h else "" for h in header_row]
        column_map = {h: aliases[h] for h in headers if h in aliases}

        rows: list[dict] = []
        for row_num, row in enumerate(iterator, start=2):
            record: dict[str, object] = {
                "source_workbook": file_path.stem,
                "source_sheet": sheet_title,
                "source_row_number": str(row_num),
            }
            raw: dict[str, object] = {}
            for idx, h in enumerate(headers):
                value = row[idx] if idx < len(row) else None
                raw[h] = value
                col = column_map.get(h)
                if col:
                    record[col] = str(value).strip() if value is not None else None
            record["order_time_at"] = parse_datetime(record.get("order_time"))
            record["ship_date_value"] = parse_date(record.get("ship_date"))
            record["raw_payload"] = raw
            rows.append(record)

        wb.close()

        if not rows:
            return {"imported": 0, "message": "无数据行"}

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            conn.execute(sa_delete(JST_MONTHLY_ORDERS_TABLE))
            self._batch_insert(JST_MONTHLY_ORDERS_TABLE, rows, conn=conn)

        return {
            "imported": len(rows),
            "message": f"{file_path.name}: {len(rows)} 条",
        }

    # ── jst_size_stock (尺码库存) ──────────────────────────────────

    def import_size_stock(self, file_path: Path, *, snapshot_date: date | None = None) -> dict[str, object]:
        snapshot_date = snapshot_date or date.today()
        JST_SIZE_STOCK_SNAPSHOT_TABLE.create(self.engine, checkfirst=True)
        wb = load_workbook(str(file_path), data_only=True)
        ws = wb["尺码表"]
        sheet_title = ws.title

        # Row 2: 列标签, 0, 220, 225, ..., 285, (空白), 总计
        header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        # Find size columns (220 to 285) and their indices
        size_cols: list[tuple[int, str]] = []
        for idx, val in enumerate(header_row):
            s = str(val).strip() if val else ""
            if s in {"220", "225", "230", "235", "240", "245", "250", "255", "260", "265", "270", "275", "280", "285"}:
                size_cols.append((idx, s))

        rows: list[dict] = []
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if row[0] is None:
                continue
            product_code = str(row[0]).strip()
            # Skip summary/blank rows
            if product_code in ("总计", "(空白)", ""):
                continue

            for col_idx, size_val in size_cols:
                qty = row[col_idx] if col_idx < len(row) else None
                if qty is None:
                    continue
                record: dict[str, object] = {
                    "source_workbook": file_path.stem,
                    "source_sheet": sheet_title,
                    "source_row_number": str(row_num),
                    "raw_payload": {},
                    "product_code": product_code,
                    "size": size_val,
                    "stock_qty": int(qty),
                }
                rows.append(record)

        wb.close()

        if not rows:
            return {"imported": 0, "message": "无数据行"}

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            conn.execute(sa_delete(JST_SIZE_STOCK_TABLE))
            self._batch_insert(JST_SIZE_STOCK_TABLE, rows, conn=conn)
            snapshot_rows = [{**row, "snapshot_date": snapshot_date} for row in rows]
            self._upsert(
                JST_SIZE_STOCK_SNAPSHOT_TABLE,
                snapshot_rows,
                ["snapshot_date", "product_code", "size"],
                [column.name for column in JST_SIZE_STOCK_SNAPSHOT_TABLE.columns if column.name not in {"id", "created_at", "snapshot_date", "product_code", "size"}],
            )

        return {
            "imported": len(rows),
            "message": f"{file_path.name}: {len(rows)} 条",
        }

    # ── jst_stock_summary (商品库存 Sheet4) ───────────────────────

    def import_stock_summary(self, file_path: Path, *, snapshot_date: date | None = None) -> dict[str, object]:
        snapshot_date = snapshot_date or date.today()
        JST_STOCK_SUMMARY_SNAPSHOT_TABLE.create(self.engine, checkfirst=True)
        wb = load_workbook(str(file_path), data_only=True, read_only=True)
        ws = wb["Sheet4"]
        sheet_title = ws.title

        rows: list[dict] = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 5:
                continue
            product_code = str(row[0]).strip() if row[0] else ""
            if product_code in ("", "总计", "(空白)", "行标签", "合计"):
                continue
            record: dict[str, object] = {
                "stock_date": file_path.parent.name,
                "stock_date_value": parse_date(file_path.parent.name),
                "source_workbook": file_path.stem,
                "source_sheet": sheet_title,
                "source_row_number": str(row_num),
                "raw_payload": {},
                "product_code": product_code,
                "defect_stock_qty": int(row[1]) if row[1] is not None else None,
                "purchase_in_transit_qty": int(row[2]) if row[2] is not None else None,
                "off_shelf_qty": int(row[3]) if row[3] is not None else None,
                "order_occupy_qty": int(row[4]) if row[4] is not None else None,
            }
            rows.append(record)

        wb.close()

        if not rows:
            return {"imported": 0, "message": "无数据行"}

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            conn.execute(sa_delete(JST_STOCK_SUMMARY_TABLE))
            self._batch_insert(JST_STOCK_SUMMARY_TABLE, rows, conn=conn)
            snapshot_column_names = {column.name for column in JST_STOCK_SUMMARY_SNAPSHOT_TABLE.columns}
            snapshot_rows = [
                {
                    **{key: value for key, value in row.items() if key in snapshot_column_names},
                    "snapshot_date": snapshot_date,
                }
                for row in rows
            ]
            self._upsert(
                JST_STOCK_SUMMARY_SNAPSHOT_TABLE,
                snapshot_rows,
                ["snapshot_date", "product_code"],
                [column.name for column in JST_STOCK_SUMMARY_SNAPSHOT_TABLE.columns if column.name not in {"id", "created_at", "snapshot_date", "product_code"}],
            )

        return {
            "imported": len(rows),
            "message": f"{file_path.name}: {len(rows)} 条",
        }

    # ── jst_purchase_diff (采购差异) ───────────────────────────────

    def import_purchase_diff(self, file_path: Path) -> dict[str, object]:
        wb = load_workbook(str(file_path), data_only=True)
        ws = wb["Sheet8"]
        sheet_title = ws.title

        rows: list[dict] = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Only use Col C-D (index 2-3)
            if len(row) < 4:
                continue
            product_code = str(row[2]).strip() if row[2] else ""
            if product_code in ("列标签", "", "(空白)"):
                continue
            diff = row[3]
            record: dict[str, object] = {
                "source_workbook": file_path.stem,
                "source_sheet": sheet_title,
                "source_row_number": str(row_num),
                "raw_payload": {},
                "product_code": product_code,
                "difference_count": int(diff) if diff is not None else None,
            }
            rows.append(record)

        wb.close()

        if not rows:
            return {"imported": 0, "message": "无数据行"}

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            conn.execute(sa_delete(JST_PURCHASE_DIFF_TABLE))
            self._batch_insert(JST_PURCHASE_DIFF_TABLE, rows, conn=conn)

        return {
            "imported": len(rows),
            "message": f"{file_path.name}: {len(rows)} 条",
        }

    # ── jst_product_profiles (聚水潭商品资料表) ───────────────────

    def import_product_profiles(self, source_root: Path) -> dict[str, object]:
        if not source_root.exists():
            return {"imported": 0, "read_rows": 0, "message": f"目录不存在: {source_root}"}

        files: list[Path] = []
        for ext in (".xlsx", ".xlsm", ".xls"):
            files.extend(path for path in source_root.rglob(f"*{ext}") if not path.name.startswith("~$"))

        rows: list[dict[str, object]] = []
        skipped_sheets: list[str] = []
        for file_path in sorted(files):
            try:
                file_rows, file_skipped = self._read_product_profile_file(file_path)
            except Exception as exc:
                skipped_sheets.append(f"{file_path.name}: {type(exc).__name__}: {exc}")
                continue
            rows.extend(file_rows)
            skipped_sheets.extend(file_skipped)

        if not rows:
            return {
                "imported": 0,
                "read_rows": 0,
                "files": len(files),
                "skipped_sheets": skipped_sheets,
                "message": "无可导入数据",
            }

        deduped_by_code: dict[str, dict[str, object]] = {}
        for row in rows:
            product_code = str(row.get("product_code") or "").strip()
            if product_code:
                deduped_by_code[product_code] = row
        payload = list(deduped_by_code.values())

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            JST_PRODUCT_PROFILE_TABLE.create(conn, checkfirst=True)
            conn.execute(sa_delete(JST_PRODUCT_PROFILE_TABLE))
            self._batch_insert(JST_PRODUCT_PROFILE_TABLE, payload, conn=conn)

        return {
            "imported": len(payload),
            "read_rows": len(rows),
            "files": len(files),
            "skipped_sheets": skipped_sheets,
            "message": f"聚水潭商品资料表导入完成: {len(payload)} 条 (原始 {len(rows)} 行, 文件 {len(files)} 个)",
        }

    # ── jst_aftersale_returns (售后退货退款) ───────────────────────

    def import_aftersale_returns(self, file_path: Path) -> dict[str, object]:
        if not file_path.exists():
            return {"imported": 0, "read_rows": 0, "message": f"文件不存在: {file_path}"}

        wb = load_workbook(str(file_path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        sheet_title = ws.title
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            wb.close()
            return {"imported": 0, "read_rows": 0, "message": "无表头"}

        aliases = JST_AFTERSALE_RETURN_COLUMN_ALIASES
        headers = [self._normalize_header(value) for value in header_row]
        column_map = {
            idx: aliases[header]
            for idx, header in enumerate(headers)
            if header in aliases
        }
        mapped_columns = set(column_map.values())
        missing = sorted({"original_goods_code", "returned_qty"} - mapped_columns)
        if "order_date" not in mapped_columns and "order_time" not in mapped_columns:
            missing.append("order_date/order_time")
        if missing:
            wb.close()
            return {
                "imported": 0,
                "read_rows": 0,
                "message": f"售后表缺少字段: {', '.join(missing)}",
            }

        rows: list[dict[str, object]] = []
        read_rows = 0
        skipped_rows = 0
        for row_num, row in enumerate(iterator, start=2):
            read_rows += 1
            record: dict[str, object] = {
                "source_workbook": file_path.stem,
                "source_sheet": sheet_title,
                "source_row_number": str(row_num),
            }
            raw: dict[str, object] = {}
            original_values: dict[str, object] = {}
            for idx, header in enumerate(headers):
                value = row[idx] if idx < len(row) else None
                raw[header] = self._json_cell_value(value)
                col = column_map.get(idx)
                if not col:
                    continue
                original_values[col] = value
                if col == "returned_qty":
                    record[col] = self._cell_int(value)
                else:
                    record[col] = self._cell_text(value)

            original_code = str(record.get("original_goods_code") or "").strip()
            returned_qty = self._cell_int(record.get("returned_qty"))
            if not original_code or returned_qty <= 0:
                skipped_rows += 1
                continue
            record["original_goods_code"] = original_code
            record["returned_qty"] = returned_qty
            record["order_date_value"] = self._parse_excel_date(original_values.get("order_date") or record.get("order_date"))
            record["order_time_value"] = self._parse_excel_date(original_values.get("order_time") or record.get("order_time"))
            record["raw_payload"] = raw
            rows.append(record)

        wb.close()

        from sqlalchemy import delete as sa_delete

        with self.engine.begin() as conn:
            JST_AFTERSALE_RETURN_TABLE.create(conn, checkfirst=True)
            self._ensure_aftersale_return_schema(conn)
            conn.execute(sa_delete(JST_AFTERSALE_RETURN_TABLE))
            if rows:
                self._batch_insert(JST_AFTERSALE_RETURN_TABLE, rows, conn=conn)

        return {
            "imported": len(rows),
            "read_rows": read_rows,
            "skipped_rows": skipped_rows,
            "source_file": str(file_path),
            "message": f"{file_path.name}: {len(rows)} 条 (读取 {read_rows} 行, 跳过 {skipped_rows} 行)",
        }

    # ── Helpers ────────────────────────────────────────────────────

    @staticmethod
    def _ensure_aftersale_return_schema(conn) -> None:
        conn.execute(text("ALTER TABLE jst_aftersale_returns ADD COLUMN IF NOT EXISTS order_time TEXT"))
        conn.execute(text("ALTER TABLE jst_aftersale_returns ADD COLUMN IF NOT EXISTS platform_site TEXT"))
        conn.execute(text("ALTER TABLE jst_aftersale_returns ADD COLUMN IF NOT EXISTS shop_name TEXT"))
        conn.execute(text("ALTER TABLE jst_aftersale_returns ADD COLUMN IF NOT EXISTS online_order_id TEXT"))
        conn.execute(text("ALTER TABLE jst_aftersale_returns ADD COLUMN IF NOT EXISTS order_time_value DATE"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_jst_aftersale_returns_order_time ON jst_aftersale_returns (order_time_value)"))

    def _ensure_price_history_schema(self) -> None:
        JST_PRICE_TABLE.create(self.engine, checkfirst=True)
        with self.engine.begin() as conn:
            conn.execute(text("ALTER TABLE jst_product_price ADD COLUMN IF NOT EXISTS source_date TEXT"))
            conn.execute(text("ALTER TABLE jst_product_price ADD COLUMN IF NOT EXISTS source_date_value DATE"))
            conn.execute(
                text(
                    """
                    update jst_product_price
                    set
                        source_date = coalesce(source_date, to_char(current_date, 'YYYY-MM-DD')),
                        source_date_value = coalesce(source_date_value, current_date)
                    where source_date is null
                       or source_date = ''
                       or source_date_value is null
                    """
                )
            )
            conn.execute(text("ALTER TABLE jst_product_price ALTER COLUMN source_date SET NOT NULL"))
            conn.execute(
                text(
                    """
                    do $$
                    begin
                        if exists (
                            select 1
                            from pg_constraint
                            where conname = 'uq_jst_price_code_name'
                        ) then
                            alter table jst_product_price drop constraint uq_jst_price_code_name;
                        end if;
                    end $$;
                    """
                )
            )
            conn.execute(
                text(
                    """
                    create unique index if not exists uq_jst_price_date_code_name
                    on jst_product_price (source_date, goods_code, goods_full_name)
                    """
                )
            )
            conn.execute(
                text(
                    """
                    create index if not exists idx_jst_price_source_date_value
                    on jst_product_price (source_date_value)
                    """
                )
            )

    def _upsert(self, table, rows: list[dict], key_cols: list[str], update_cols: list[str], chunk_size: int = 500) -> None:
        """Upsert rows using ON CONFLICT DO UPDATE."""
        for i in range(0, len(rows), chunk_size):
            chunk = rows[i:i + chunk_size]
            set_ = {col: getattr(pg_insert(table).excluded, col) for col in update_cols}
            set_["updated_at"] = sa_func.date_trunc('minute', sa_func.now())
            stmt = pg_insert(table).values(chunk).on_conflict_do_update(
                index_elements=key_cols,
                set_=set_,
            )
            with self.engine.begin() as conn:
                conn.execute(stmt)

    def _batch_insert(self, table, rows: list[dict], chunk_size: int = 500, conn=None) -> None:
        from sqlalchemy import insert

        if conn is not None:
            for i in range(0, len(rows), chunk_size):
                conn.execute(insert(table), rows[i:i + chunk_size])
            return

        with self.engine.begin() as connection:
            for i in range(0, len(rows), chunk_size):
                connection.execute(insert(table), rows[i:i + chunk_size])

    def _read_excel(self, file_path: Path, aliases: dict[str, str], duplicate_policy: str = "last") -> list[dict]:
        """Read an Excel file, map headers to canonical columns, return rows as dicts."""
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        assert ws is not None
        sheet_title = ws.title
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            wb.close()
            return []

        headers = [str(h).strip() if h else "" for h in header_row]
        column_map = {h: aliases[h] for h in headers if h in aliases}

        rows: list[dict] = []
        for row_num, row in enumerate(iterator, start=2):
            record: dict[str, object] = {
                "source_workbook": file_path.stem,
                "source_sheet": sheet_title,
                "source_row_number": str(row_num),
            }
            raw: dict[str, object] = {}
            for idx, h in enumerate(headers):
                value = row[idx] if idx < len(row) else None
                raw[h] = value
                col = column_map.get(h)
                if col:
                    if duplicate_policy == "first" and col in record:
                        continue
                    record[col] = str(value).strip() if value is not None else None
            record["raw_payload"] = raw
            rows.append(record)

        wb.close()
        return rows

    def _read_product_profile_file(self, file_path: Path) -> tuple[list[dict[str, object]], list[str]]:
        if file_path.suffix.lower() == ".xls":
            return self._read_product_profile_xls(file_path)

        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)
        except (BadZipFile, InvalidFileException):
            return self._read_product_profile_xls(file_path)
        aliases = JST_PRODUCT_PROFILE_COLUMN_ALIASES
        rows: list[dict[str, object]] = []
        skipped_sheets: list[str] = []

        try:
            for ws in wb.worksheets:
                header_info = self._find_product_profile_header(ws, aliases)
                if header_info is None:
                    skipped_sheets.append(f"{file_path.name}/{ws.title}: 未找到必要表头")
                    continue
                header_row_number, headers, column_map = header_info
                for row_num, row in enumerate(ws.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
                    record: dict[str, object] = {
                        "source_workbook": file_path.stem,
                        "source_sheet": ws.title,
                        "source_row_number": str(row_num),
                    }
                    raw: dict[str, object] = {}
                    for idx, header in enumerate(headers):
                        value = row[idx] if idx < len(row) else None
                        raw[header] = value
                        col = column_map.get(idx)
                        if col:
                            record[col] = self._cell_text(value)
                    if not record.get("product_code"):
                        continue
                    record["raw_payload"] = raw
                    rows.append(record)
        finally:
            wb.close()

        return rows, skipped_sheets

    @staticmethod
    def _find_product_profile_header(ws, aliases: dict[str, str]) -> tuple[int, list[str], dict[int, str]] | None:
        required = {"product_code", "style_code", "color_name", "size_barcode"}
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_number > 30:
                break
            headers = [VipRepository._normalize_header(value) for value in row]
            column_map = {
                idx: aliases[header]
                for idx, header in enumerate(headers)
                if header in aliases
            }
            if required.issubset(set(column_map.values())):
                return row_number, headers, column_map
        return None

    def _read_product_profile_xls(self, file_path: Path) -> tuple[list[dict[str, object]], list[str]]:
        workbook = xlrd.open_workbook(str(file_path))
        aliases = JST_PRODUCT_PROFILE_COLUMN_ALIASES
        rows: list[dict[str, object]] = []
        skipped_sheets: list[str] = []

        for sheet in workbook.sheets():
            header_info = self._find_product_profile_header_xls(sheet, aliases)
            if header_info is None:
                skipped_sheets.append(f"{file_path.name}/{sheet.name}: 未找到必要表头")
                continue
            header_row_index, headers, column_map = header_info
            for row_index in range(header_row_index + 1, sheet.nrows):
                values = sheet.row_values(row_index)
                record: dict[str, object] = {
                    "source_workbook": file_path.stem,
                    "source_sheet": sheet.name,
                    "source_row_number": str(row_index + 1),
                }
                raw: dict[str, object] = {}
                for idx, header in enumerate(headers):
                    value = values[idx] if idx < len(values) else None
                    raw[header] = value
                    col = column_map.get(idx)
                    if col:
                        record[col] = self._cell_text(value)
                if not record.get("product_code"):
                    continue
                record["raw_payload"] = raw
                rows.append(record)

        return rows, skipped_sheets

    @staticmethod
    def _find_product_profile_header_xls(sheet, aliases: dict[str, str]) -> tuple[int, list[str], dict[int, str]] | None:
        required = {"product_code", "style_code", "color_name", "size_barcode"}
        for row_index in range(min(sheet.nrows, 30)):
            headers = [VipRepository._normalize_header(value) for value in sheet.row_values(row_index)]
            column_map = {
                idx: aliases[header]
                for idx, header in enumerate(headers)
                if header in aliases
            }
            if required.issubset(set(column_map.values())):
                return row_index, headers, column_map
        return None

    @staticmethod
    def _normalize_header(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip().replace("\n", "").replace("\r", "")

    @staticmethod
    def _cell_text(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text_value = str(value).strip()
        return text_value or None

    @staticmethod
    def _cell_int(value: object) -> int:
        if value in (None, ""):
            return 0
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        text_value = str(value).strip().replace(",", "")
        if not text_value:
            return 0
        try:
            return int(float(text_value))
        except ValueError:
            return 0

    @staticmethod
    def _parse_excel_date(value: object) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text_value = str(value).strip() if value is not None else ""
        if not text_value:
            return None
        for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(text_value, fmt).date()
            except ValueError:
                continue
        return parse_date(text_value)

    @staticmethod
    def _json_cell_value(value: object) -> object:
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return value
